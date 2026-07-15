#!/usr/bin/env python3
"""
TENDER_FINDER Agent #2 — Future Projects pipeline (first working version)
================================================================

Что делает:
  1. Тянет реальные данные из открытого civic-API (источник на выбор).
  2. Гоняет их через Claude для скоринга под scope TENDER_FINDER (или офлайн-эвристику).
  3. Пишет результат в Excel + CSV + markdown-отчёт, с дедупом против прошлого прогона.

Источники:
  --source vancouver  -> Vancouver Open Data (ПУБЛИЧНЫЙ, работает с любой машины)
  --source surrey      -> Surrey FutureWorks GIS (ГОТОВ, но сначала пройди Тест №1
                          из карты источников — сервер за firewall'ом для дата-центров)
  --source langley_tol -> Township of Langley dev-applications (ArcGIS Hub, публичный)
  --source maple_ridge -> Maple Ridge active dev-applications (ArcGIS Hub, публичный)

Запуск:
  # быстрый офлайн-тест на встроенных данных (без сети и без API-ключа):
  python tenderfinder_agent2.py --demo

  # реальный прогон по Vancouver, со скорингом через Claude:
  export ANTHROPIC_API_KEY=sk-ant-...
  python tenderfinder_agent2.py --source vancouver --limit 60

  # когда Surrey откроется:
  python tenderfinder_agent2.py --source surrey --limit 200

  # муниципальные ArcGIS-хабы (работают с любой нормальной сети):
  python tenderfinder_agent2.py --source langley_tol --limit 200
  python tenderfinder_agent2.py --source maple_ridge --limit 200

Зависимости:
  pip install openpyxl        (всё остальное — стандартная библиотека)

LEGACY NOTICE: this entry point is frozen and keeps its original built-in
keyword lists. The launcher and connector sweep use config/keywords.xlsx.
"""

import argparse
import csv
import json
import os
import sys
import urllib.parse
import urllib.request
from datetime import datetime, date

# ────────────────────────────────────────────────────────────────────────────
# КОНФИГ — редактируй это под TENDER_FINDER. Раз в квартал обновляй keywords/clients.
# ────────────────────────────────────────────────────────────────────────────

SCOPE_STATEMENT = (
    "Example Civil Contractor — civil/earthwork-подрядчик (Langley/Maple Ridge, BC). "
    "БЕРЁМ: subdivision servicing, site servicing, excavation, underground utilities "
    "(water main, storm sewer, sanitary sewer), bedding gravel, manholes, "
    "footings/foundations, roadworks, structural/site concrete, bridges. "
    "Приоритет — subdivisions с чётким scope по Lower Mainland "
    "(ядро: Surrey, Langley, Maple Ridge). "
    "НЕ БЕРЁМ: вертикальные здания, general contracting с размытым scope."
)

# Слова, повышающие релевантность (для офлайн-эвристики и как подсказка модели).
POSITIVE_KEYWORDS = [
    "subdivision", "servicing", "service agreement", "site servicing", "excavat",
    "water main", "watermain", "storm", "sanitary", "sewer", "drainage",
    "utility", "utilities", "road", "roadwork", "grading", "site prep",
    "earthwork", "manhole", "bridge", "civil", "lot grading", "site concrete",
    "curb", "gutter", "detention pond", "forcemain", "trunk main",
]

# Слова, понижающие релевантность (vertical / building-focused).
NEGATIVE_KEYWORDS = [
    "tenant improvement", "interior", "renovation", "high-rise interior",
    "mechanical", "electrical", "hvac", "roofing", "facade", "cladding",
    "unit alteration", "suite", "kitchen", "bathroom",
]

# Известные клиенты TENDER_FINDER — совпадение по застройщику/owner повышает приоритет.
TENDER_FINDER_CLIENTS = [
    "polygon", "beedie", "wesgroup", "bird", "chandos", "turner",
    "stuart olson", "dp world", "anthem", "city of surrey", "city of langley",
    "township of langley", "maple ridge", "translink", "metro vancouver",
    "semiahmoo", "ministry of transportation", "vancouver airport",
]

# Модель для скоринга. Haiku — дёшево и быстро для объёма; Sonnet — для сложных.
CLAUDE_MODEL = "claude-haiku-4-5-20251001"
CLAUDE_MODEL_FALLBACK = "claude-sonnet-4-6"

OUTPUT_DIR = "."
XLSX_PATH = os.path.join(OUTPUT_DIR, "tenderfinder_future_projects.xlsx")
CSV_PATH = os.path.join(OUTPUT_DIR, "tenderfinder_future_projects.csv")
REPORT_PATH = os.path.join(OUTPUT_DIR, "tenderfinder_morning_report.md")

# Итоговая схема строки (порядок колонок в Excel/CSV).
COLUMNS = [
    "project_id", "project_name", "owner_developer", "type", "horizon",
    "location", "scope_summary", "why_fits_tenderfinder", "relevance_score",
    "estimated_tender_window", "action", "source", "date_found", "source_link",
]

# ────────────────────────────────────────────────────────────────────────────
# ИСТОЧНИК 1 — Vancouver Open Data (публичный, работает сейчас)
# ────────────────────────────────────────────────────────────────────────────

VANCOUVER_BASE = (
    "https://opendata.vancouver.ca/api/explore/v2.1/"
    "catalog/datasets/issued-building-permits/records"
)
VANCOUVER_DATASET_URL = (
    "https://opendata.vancouver.ca/explore/dataset/issued-building-permits/"
)


def fetch_vancouver(limit=60):
    """Тянет недавние building permits. ODS v2.1 отдаёт max 100 на запрос — пагинируем."""
    records, offset, page = [], 0, 100
    # Берём только новое строительство и снос/site work — там civil-сигнал, а не интерьеры.
    where = "typeofwork in ('New Building','Addition / Alteration','Demolition / Deconstruction')"
    while len(records) < limit:
        params = {
            "where": where,
            "order_by": "issuedate DESC",
            "limit": min(page, limit - len(records)),
            "offset": offset,
        }
        url = VANCOUVER_BASE + "?" + urllib.parse.urlencode(params)
        data = _http_get_json(url)
        batch = data.get("results", [])
        if not batch:
            break
        for r in batch:
            records.append(_normalize_vancouver(r))
        offset += len(batch)
        if len(batch) < params["limit"]:
            break
    return records


def _normalize_vancouver(r):
    """Vancouver-запись -> универсальный формат. Поля берём защищённо (.get)."""
    pid = r.get("permitnumber") or r.get("permit_number") or "VAN-?"
    desc = (r.get("projectdescription") or r.get("specificusecategory")
            or r.get("propertyuse") or "")
    text_parts = [
        str(r.get("typeofwork", "")), str(r.get("specificusecategory", "")),
        str(r.get("propertyuse", "")), str(desc), str(r.get("address", "")),
    ]
    return {
        "source": "Vancouver Building Permits",
        "source_link": VANCOUVER_DATASET_URL,
        "id": pid,
        "raw": r,
        "text_blob": " ".join(text_parts).lower(),
        "location": r.get("geolocalarea") or r.get("address") or "Vancouver",
        "date_found": str(date.today()),
    }


# ────────────────────────────────────────────────────────────────────────────
# ИСТОЧНИК 2 — Surrey FutureWorks GIS (готов; включить после Теста №1)
# ────────────────────────────────────────────────────────────────────────────

SURREY_FUTUREWORKS = (
    "https://gisservices.surrey.ca/arcgis/rest/services/FutureWorks/MapServer"
)
# Слои, релевантные TENDER_FINDER (см. карту источников): утилиты + дороги + дренаж.
SURREY_LAYERS = {
    0: "Drainage", 8: "Roads", 9: "Sanitary", 10: "Sidewalk",
    11: "Storm", 15: "Turnaround", 16: "Walkway", 17: "Water",
}


def fetch_surrey(limit=200):
    """Тянет запланированные работы Surrey по релевантным слоям.
    ВНИМАНИЕ: сервер за firewall'ом — заработает только если Тест №1 вернул JSON
    (т.е. ты запускаешь скрипт из канадской/офисной сети, а не из дата-центра)."""
    records = []
    per_layer = max(10, limit // len(SURREY_LAYERS))
    for layer_id, layer_name in SURREY_LAYERS.items():
        params = {
            "where": "1=1", "outFields": "*", "returnGeometry": "false",
            "resultRecordCount": per_layer, "f": "json",
        }
        url = f"{SURREY_FUTUREWORKS}/{layer_id}/query?" + urllib.parse.urlencode(params)
        try:
            data = _http_get_json(url)
        except Exception as e:
            print(f"  [!] Surrey слой {layer_id} ({layer_name}) недоступен: {e}")
            continue
        for f in data.get("features", []):
            records.append(_normalize_surrey(f, layer_id, layer_name))
        if len(records) >= limit:
            break
    return records[:limit]


def _normalize_surrey(feature, layer_id, layer_name):
    attrs = feature.get("attributes", {})
    pid = (attrs.get("OBJECTID") or attrs.get("FID") or "SUR-?")
    text_blob = " ".join(str(v) for v in attrs.values() if v is not None).lower()
    # Слой сам по себе — сильный сигнал типа работ.
    text_blob += f" futureworks {layer_name.lower()}"
    return {
        "source": f"Surrey FutureWorks / {layer_name}",
        "source_link": f"{SURREY_FUTUREWORKS}/{layer_id}",
        "id": f"FW-{layer_id}-{pid}",
        "raw": attrs,
        "text_blob": text_blob,
        "location": attrs.get("NEIGHBOURHOOD") or attrs.get("AREA") or "Surrey",
        "date_found": str(date.today()),
    }



# ────────────────────────────────────────────────────────────────────────────
# ИСТОЧНИК 3/4 — ArcGIS Hub муниципалитетов (Township of Langley, Maple Ridge)
# Публичные ArcGIS-хабы, работают с любой нормальной машины (не за firewall'ом,
# в отличие от внутреннего Surrey GIS). Структура слоёв одинаковая — поэтому
# один общий загрузчик + defensive-нормализатор (имена полей у всех разные).
# ────────────────────────────────────────────────────────────────────────────

# Township of Langley — Development Activity Status Table (ArcGIS Hub).
TOL_ITEM_ID = "aea97e65c9db4dad8242783c96e6b70c"     # Hub item (слой _1)
TOL_LAYER = 1
TOL_DATASET_URL = ("https://data-tol.opendata.arcgis.com/datasets/"
                   "development-activity-status-table")

# Maple Ridge — Active Development Applications (ArcGIS Hub).
MR_SLUG = "mapleridge::active-development-applications"   # резолвим в service URL
MR_LAYER = 0
MR_DATASET_URL = ("https://opengov2-mapleridge.opendata.arcgis.com/datasets/"
                  "active-development-applications")

# ── ЗАКРЕПЛЁННЫЕ endpoint'ы (проверены живым прогоном 22-06-2026) ──────────────
# "Prove then pin": разрешённые при первом прогоне URL фиксируем, чтобы повторный
# discover не "уплыл" на чужой слой (как случилось с Surrey -> Subdivision Markers).
PINNED_SERVICE = {
    # Township of Langley — Development Activity Status Table (782 строк, проверено)
    "langley_tol": ("https://services5.arcgis.com/frpHL0Fv8koQRVWY/arcgis/rest/"
                    "services/Development_Activity_Status_Table/FeatureServer", 1),
    # Maple Ridge — Active Development Applications (879 строк, проверено)
    "maple_ridge": ("https://geoservices.mapleridge.ca/server/rest/services/"
                    "DataCatalog/PlanningDevelopment/MapServer", 1),
    # Surrey — ВЕРНЫЙ слой dev-applications (НЕ Subdivision Markers!). Проверить,
    # что возвращает заявки (поля dev file/type/status/address), затем оставить.
    "surrey_public": ("https://services5.arcgis.com/YRpe0VKTJytZSSIB/arcgis/rest/"
                      "services/Development%20Applications/FeatureServer", 0),
}

# Слои-ловушки: имена, которые discover путает с "development applications", но
# для лидогенерации бесполезны (оверлеи/areas/маркеры/разрешения/геометрия).
_LAYER_DENYLIST = [
    "hazard", "dpa", "permit area", "development permit area", "marker",
    "watering", "lawn", "alr", "agricultural land reserve", "reserve",
    "neighbourhood plan", "np_", "asset", "facility", "watercourse",
    "flood", "slope", "creek", "wildfire", "streamside", "zoning",
]
# Минимум осмысленных (не-геометрия/не-OID) полей со значениями, иначе слой пустой.
_MIN_USEFUL_FIELDS = 3
_GEOM_OID_FIELDS = {"objectid", "fid", "globalid", "shape", "shape__area",
                    "shape__length", "shape.starea()", "shape.stlength()",
                    "se_anno_cad_data", "geom", "geo_point_2d"}


def _layer_is_denylisted(name):
    n = (name or "").lower()
    return any(bad in n for bad in _LAYER_DENYLIST)


def _attrs_are_rich_enough(attrs):
    """True, если в записи есть >= _MIN_USEFUL_FIELDS непустых смысловых полей
    (а не только OBJECTID + геометрия). Отсекает ALR/маркеры с пустыми атрибутами."""
    useful = [k for k, v in attrs.items()
              if k.lower() not in _GEOM_OID_FIELDS and v not in (None, "")]
    return len(useful) >= _MIN_USEFUL_FIELDS

# Поля dev-application слоёв называются по-разному в каждом муниципалитете,
# поэтому ищем значение по списку вероятных алиасов (регистр/подчёркивания игнор).
_DEVAPP_FIELD_ALIASES = {
    "app_no":    ["applicationno", "appno", "file", "fileno", "filenumber",
                  "application", "applicationnumber", "folder", "folderno",
                  "permitnumber", "devappno", "projectno", "casenumber",
                  "foldernumber", "projectnumber", "referencefile", "devfile"],
    "app_type":  ["applicationtype", "apptype", "type", "applicationtypedesc",
                  "category", "subtype", "projecttype", "permittype"],
    "status":    ["status", "stage", "applicationstatus", "currentstatus",
                  "processstage", "statusdesc", "milestone", "state",
                  "folderstatus", "statusdescription", "taskstatus", "amandastatus"],
    "address":   ["address", "civicaddress", "fulladdress", "location",
                  "siteaddress", "propertyaddress", "addr", "streetaddress",
                  "street", "house"],
    "applicant": ["applicant", "applicantname", "owner", "ownername", "developer",
                  "agent", "company", "proponent"],
    "desc":      ["description", "projectdescription", "details", "purpose",
                  "proposal", "summary", "comments", "projectname", "name",
                  "workproposed", "subtype", "tasktype"],
}


def _resolve_arcgis_service_url(item_id=None, slug=None):
    """Определяет реальный FeatureServer/FeatureLayer URL слоя.
    item_id -> ArcGIS item info; slug -> ArcGIS Hub v3 datasets API.
    Возвращает базовый service URL (без хвоста /{layer}/query)."""
    if item_id:
        info = _http_get_json(
            f"https://www.arcgis.com/sharing/rest/content/items/{item_id}?f=json")
        url = info.get("url")
        if url:
            return url.rstrip("/")
    if slug:
        data = _http_get_json(
            "https://hub.arcgis.com/api/v3/datasets"
            f"?filter[slug]={urllib.parse.quote(slug)}"
            "&fields[datasets]=url,name")
        items = data.get("data", [])
        if items:
            url = (items[0].get("attributes", {}) or {}).get("url")
            if url:
                url = url.rstrip("/")
                return url.rsplit("/", 1)[0] if url[-1:].isdigit() else url
    raise RuntimeError("Не удалось определить service URL (проверь item_id/slug).")


def _pick_field(attrs_lower, aliases):
    for a in aliases:
        if a in attrs_lower and attrs_lower[a] not in (None, ""):
            return attrs_lower[a]
    return ""


def _normalize_arcgis_devapp(attrs, muni, source_name, source_link):
    """ArcGIS dev-application запись -> универсальный формат.
    Имена полей варьируются между муниципалитетами -> пробуем алиасы, а text_blob
    строим из ВСЕХ значений (как в Surrey) для устойчивого скоринга."""
    al = {str(k).lower().replace("_", "").replace(" ", ""): v
          for k, v in attrs.items()}
    app_no = (_pick_field(al, _DEVAPP_FIELD_ALIASES["app_no"])
              or attrs.get("OBJECTID") or attrs.get("FID") or "?")
    app_type = _pick_field(al, _DEVAPP_FIELD_ALIASES["app_type"])
    status = _pick_field(al, _DEVAPP_FIELD_ALIASES["status"])
    address = _pick_field(al, _DEVAPP_FIELD_ALIASES["address"])
    applicant = _pick_field(al, _DEVAPP_FIELD_ALIASES["applicant"])
    desc = _pick_field(al, _DEVAPP_FIELD_ALIASES["desc"])

    text_blob = " ".join(str(v) for v in attrs.values() if v is not None).lower()
    text_blob += (f" {muni.lower()} development application "
                  f"{str(app_type).lower()} {str(status).lower()} {str(desc).lower()}")

    muni_tag = muni.upper().replace(" ", "-")
    return {
        "source": source_name,
        "source_link": source_link,
        "id": f"{muni_tag}-{app_no}",
        "raw": attrs,
        "text_blob": text_blob,
        "location": address or muni,
        "date_found": str(date.today()),
        # необязательные подсказки для отчёта/маппинга в Future_Projects
        "_app_type": app_type, "_status": status, "_applicant": applicant,
        "_desc": desc, "_address": address, "_municipality": muni,
    }


def _fetch_arcgis_devapps(service_url, layer, muni, source_name, source_link,
                          limit=200):
    """Тянет dev-application слой через стандартный ArcGIS REST /query
    с пагинацией. Работает с любого нормального сетевого доступа."""
    records, offset, page = [], 0, 1000
    base = f"{service_url}/{layer}/query"
    while len(records) < limit:
        params = {
            "where": "1=1", "outFields": "*", "returnGeometry": "false",
            "resultOffset": offset,
            "resultRecordCount": min(page, limit - len(records)),
            "f": "json",
        }
        url = base + "?" + urllib.parse.urlencode(params)
        data = _http_get_json(url)
        feats = data.get("features", [])
        if not feats:
            break
        # gate: первый батч проверяем на "богатость" атрибутов — если слой
        # отдаёт только геометрию/OBJECTID (как ALR-полигоны Port Coquitlam),
        # это не заявки -> прекращаем и сигналим.
        if offset == 0 and feats and not _attrs_are_rich_enough(
                feats[0].get("attributes", {})):
            raise RuntimeError(
                f"Слой '{source_name}' отдаёт только геометрию/OID без атрибутов "
                f"— это не development applications. Проверь правильность слоя.")
        for f in feats:
            records.append(_normalize_arcgis_devapp(
                f.get("attributes", {}), muni, source_name, source_link))
        got = len(feats)
        offset += got
        if got < params["resultRecordCount"]:
            break
    return records[:limit]


def fetch_langley_tol(limit=200):
    """Township of Langley — Development Activity Status Table (ArcGIS Hub).
    Самый чистый структурный источник в ядре географии TENDER_FINDER."""
    service, layer = PINNED_SERVICE["langley_tol"]
    try:
        return _fetch_arcgis_devapps(
            service, layer, "Township of Langley",
            "Township of Langley — Development Activity", TOL_DATASET_URL, limit)
    except Exception as e:                       # pinned устарел -> авто-резолв
        print(f"  (pinned TOL не сработал: {e}; резолвлю заново)")
        service = _resolve_arcgis_service_url(item_id=TOL_ITEM_ID)
        return _fetch_arcgis_devapps(
            service, TOL_LAYER, "Township of Langley",
            "Township of Langley — Development Activity", TOL_DATASET_URL, limit)


def fetch_maple_ridge(limit=200):
    """Maple Ridge — Active Development Applications (ArcGIS Hub).
    Слой содержит стадию процесса -> готовый future-signal."""
    service, layer = PINNED_SERVICE["maple_ridge"]
    try:
        return _fetch_arcgis_devapps(
            service, layer, "Maple Ridge",
            "Maple Ridge — Active Development Applications", MR_DATASET_URL, limit)
    except Exception as e:                       # pinned устарел -> авто-резолв
        print(f"  (pinned Maple Ridge не сработал: {e}; резолвлю по slug)")
        service = _resolve_arcgis_service_url(slug=MR_SLUG)
        return _fetch_arcgis_devapps(
            service, MR_LAYER, "Maple Ridge",
            "Maple Ridge — Active Development Applications", MR_DATASET_URL, limit)


# ────────────────────────────────────────────────────────────────────────────
# СКОРИНГ — Claude, если есть ключ; иначе офлайн-эвристика
# ────────────────────────────────────────────────────────────────────────────

def score_records(records, use_claude=True):
    if use_claude and os.environ.get("ANTHROPIC_API_KEY"):
        try:
            return score_with_claude(records)
        except Exception as e:
            print(f"  [!] Claude недоступен ({e}); падаю на офлайн-эвристику.")
    return [heuristic_score(r) for r in records]


def heuristic_score(rec):
    """Дешёвый офлайн-скоринг по ключевым словам. Не замена Claude — пре-фильтр."""
    blob = rec["text_blob"]
    score = 35
    for kw in POSITIVE_KEYWORDS:
        if kw in blob:
            score += 9
    for kw in NEGATIVE_KEYWORDS:
        if kw in blob:
            score -= 12
    for client in TENDER_FINDER_CLIENTS:
        if client in blob:
            score += 8
            break
    score = max(0, min(100, score))
    action = "BID_READY" if score >= 75 else "MONITOR" if score >= 50 else "SKIP"
    raw = rec["raw"]
    name = (str(raw.get("projectdescription") or raw.get("specificusecategory")
            or rec["source"]))[:80]
    return {
        "project_id": rec["id"],
        "project_name": name,
        "owner_developer": str(raw.get("applicant") or raw.get("APPLICANT") or "—"),
        "type": str(raw.get("typeofwork") or rec["source"].split("/")[-1].strip()),
        "horizon": "B",  # эвристика не различает горизонт — ставит консервативно
        "location": rec["location"],
        "scope_summary": name,
        "why_fits_tenderfinder": "(эвристика по ключевым словам — нужен Claude для разбора)",
        "relevance_score": score,
        "estimated_tender_window": "—",
        "action": action,
        "source": rec["source"],
        "date_found": rec["date_found"],
        "source_link": rec["source_link"],
    }


def score_with_claude(records, batch_size=15):
    """Батчами шлём записи в Claude, получаем структурированные оценённые строки."""
    scored = []
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        payload_records = [
            {"id": r["id"], "source": r["source"], "location": r["location"],
             "data": r["raw"]}
            for r in batch
        ]
        prompt = _build_scoring_prompt(payload_records)
        text = _call_claude(prompt)
        scored.extend(_parse_scoring_response(text, batch))
        print(f"  · оценено {min(i + batch_size, len(records))}/{len(records)}")
    return scored


def _build_scoring_prompt(payload_records):
    return f"""Ты — аналитик тендерной разведки для Example Civil Contractor, BC civil contractor.

SCOPE КОМПАНИИ:
{SCOPE_STATEMENT}

Известные клиенты (совпадение повышает приоритет): {', '.join(TENDER_FINDER_CLIENTS)}

ВХОДНЫЕ ДАННЫЕ (записи из civic open-data):
{json.dumps(payload_records, ensure_ascii=False, indent=1)}

ЗАДАЧА: для КАЖДОЙ записи верни объект в JSON-массиве с полями:
- project_id (используй "id" из входа)
- project_name (краткое название, 5-9 слов, англ.)
- owner_developer (заказчик/застройщик, либо "—")
- type (subdivision / utility-capital / road / drainage / building / mixed-civil)
- horizon ("A" = 0-3 мес до тендера / "B" = 3-18 мес)
- location
- scope_summary (1 предложение: что конкретно делается)
- why_fits_tenderfinder (1 предложение: почему TENDER_FINDER стоит обратить внимание, либо почему нет)
- relevance_score (0-100; 100 = идеальный subdivision servicing со всеми утилитами;
  vertical building с размытым scope = 20-40)
- estimated_tender_window (например "Q3 2026" или "—")
- action ("BID_READY" >=75 / "MONITOR" 50-74 / "SKIP" <50)

Верни ТОЛЬКО JSON-массив. Без markdown, без пояснений."""


def _call_claude(prompt, model=None):
    api_key = os.environ["ANTHROPIC_API_KEY"]
    body = json.dumps({
        "model": model or CLAUDE_MODEL,
        "max_tokens": 4000,
        "messages": [{"role": "user", "content": prompt}],
    }).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={
            "content-type": "application/json",
            "anthropic-version": "2023-06-01",
            "x-api-key": api_key,
        },
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        result = json.loads(resp.read())
    return result["content"][0]["text"]


def _parse_scoring_response(text, batch):
    text = text.strip()
    if text.startswith("```"):
        text = text.split("```", 2)[1]
        if text.startswith("json"):
            text = text[4:]
    text = text.strip()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        print("  [!] не смог распарсить ответ модели для батча — пропускаю")
        return []
    out = []
    by_id = {r["id"]: r for r in batch}
    for item in parsed:
        src = by_id.get(item.get("project_id"), {})
        item.setdefault("source", src.get("source", ""))
        item.setdefault("source_link", src.get("source_link", ""))
        item.setdefault("date_found", src.get("date_found", str(date.today())))
        out.append({col: item.get(col, "") for col in COLUMNS})
    return out


# ────────────────────────────────────────────────────────────────────────────
# ДЕДУП + ВЫВОД
# ────────────────────────────────────────────────────────────────────────────

def load_existing_ids(csv_path):
    if not os.path.exists(csv_path):
        return set()
    with open(csv_path, newline="", encoding="utf-8") as f:
        return {row["project_id"] for row in csv.DictReader(f)}


def write_outputs(rows):
    rows.sort(key=lambda r: r.get("relevance_score", 0), reverse=True)
    _write_csv(rows)
    _write_xlsx(rows)
    _write_report(rows)


def _write_csv(rows):
    new_file = not os.path.exists(CSV_PATH)
    with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        if new_file:
            w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in COLUMNS})


def _write_xlsx(rows):
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("  [i] openpyxl не установлен — пропускаю Excel (pip install openpyxl).")
        return
    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Future Projects"
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5496")
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
        action = r.get("action", "")
        if action == "BID_READY":
            ws.cell(ws.max_row, COLUMNS.index("action") + 1).fill = green
        elif action == "MONITOR":
            ws.cell(ws.max_row, COLUMNS.index("action") + 1).fill = yellow
    wb.save(XLSX_PATH)


def _write_report(rows):
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    bid_ready = [r for r in rows if r.get("action") == "BID_READY"]
    monitor = [r for r in rows if r.get("action") == "MONITOR"]
    lines = [
        f"# TENDER_FINDER — утренний отчёт по будущим проектам",
        f"_{today} · добавлено {len(rows)} новых записей_\n",
        f"**BID_READY: {len(bid_ready)}  ·  MONITOR: {len(monitor)}**\n",
    ]
    if bid_ready:
        lines.append("## 🟢 BID_READY — высокий приоритет\n")
        for r in bid_ready[:15]:
            lines.append(
                f"- **{r['project_name']}** ({r['relevance_score']}) — "
                f"{r['owner_developer']} · {r['location']} · "
                f"тендер {r['estimated_tender_window']}\n  "
                f"_{r['why_fits_tenderfinder']}_  \n  `{r['source']}`"
            )
    if monitor:
        lines.append("\n## 🟡 MONITOR — наблюдать\n")
        for r in monitor[:15]:
            lines.append(
                f"- {r['project_name']} ({r['relevance_score']}) — "
                f"{r['location']} · {r['source']}"
            )
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ────────────────────────────────────────────────────────────────────────────
# HTTP helper
# ────────────────────────────────────────────────────────────────────────────

def _http_get_json(url):
    req = urllib.request.Request(url, headers={"User-Agent": "TENDER_FINDER-Agent2/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read())


# ────────────────────────────────────────────────────────────────────────────
# DEMO — встроенные данные, чтобы проверить весь конвейер без сети и ключа
# ────────────────────────────────────────────────────────────────────────────

def demo_records():
    samples = [
        {"typeofwork": "New Building", "specificusecategory": "Subdivision servicing — water main, storm, sanitary, 48-lot",
         "propertyuse": "Dwelling Uses", "address": "8200 Cambie St", "applicant": "Polygon Homes",
         "permitnumber": "BP-2026-1042", "geolocalarea": "Marpole"},
        {"typeofwork": "New Building", "specificusecategory": "High-rise tower interior fit-out",
         "propertyuse": "Dwelling Uses", "address": "1500 Robson St", "applicant": "Westbank",
         "permitnumber": "BP-2026-1077", "geolocalarea": "West End"},
        {"typeofwork": "Demolition / Deconstruction", "specificusecategory": "Site clearing and excavation for road extension",
         "propertyuse": "Transportation", "address": "3000 Grandview Hwy", "applicant": "City of Vancouver",
         "permitnumber": "BP-2026-0991", "geolocalarea": "Renfrew"},
        {"typeofwork": "Addition / Alteration", "specificusecategory": "Kitchen renovation single family",
         "propertyuse": "Dwelling Uses", "address": "22 W 12th Ave", "applicant": "—",
         "permitnumber": "BP-2026-1180", "geolocalarea": "Mount Pleasant"},
        {"typeofwork": "New Building", "specificusecategory": "Storm drainage and sanitary sewer servicing, industrial subdivision",
         "propertyuse": "Industrial", "address": "8800 Glenlyon Pkwy", "applicant": "Beedie",
         "permitnumber": "BP-2026-1203", "geolocalarea": "Riverside"},
    ]
    return [_normalize_vancouver(s) for s in samples]


# ────────────────────────────────────────────────────────────────────────────
# MAIN
# ────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="TENDER_FINDER Agent #2 — Future Projects pipeline")
    ap.add_argument("--source",
                    choices=["vancouver", "surrey", "langley_tol", "maple_ridge"],
                    default="vancouver")
    ap.add_argument("--limit", type=int, default=60)
    ap.add_argument("--demo", action="store_true",
                    help="офлайн-прогон на встроенных данных (без сети и API-ключа)")
    ap.add_argument("--no-claude", action="store_true",
                    help="принудительно офлайн-эвристика вместо Claude")
    args = ap.parse_args()

    print(
        "WARNING: tenderfinder_agent2.py is a frozen legacy entry point and "
        "still uses built-in keyword lists; use the main launcher for keywords.xlsx."
    )
    print("TENDER_FINDER Agent #2 — Future Projects\n" + "=" * 40)

    # 1. Сбор
    if args.demo:
        print("Режим DEMO — встроенные данные.")
        records = demo_records()
    elif args.source == "vancouver":
        print(f"Тяну Vancouver Open Data (limit={args.limit})...")
        records = fetch_vancouver(args.limit)
    elif args.source == "surrey":
        print(f"Тяну Surrey FutureWorks (limit={args.limit})...")
        print("  (если всё в ошибках — Тест №1 не пройден, сервер за firewall'ом)")
        records = fetch_surrey(args.limit)
    elif args.source == "langley_tol":
        print(f"Тяну Township of Langley — Development Activity (limit={args.limit})...")
        records = fetch_langley_tol(args.limit)
    elif args.source == "maple_ridge":
        print(f"Тяну Maple Ridge — Active Development Applications (limit={args.limit})...")
        records = fetch_maple_ridge(args.limit)
    else:
        print(f"Неизвестный источник: {args.source}")
        sys.exit(1)

    print(f"Собрано записей: {len(records)}")
    if not records:
        print("Нечего обрабатывать. Выход.")
        sys.exit(0)

    # 2. Дедуп против прошлого прогона
    seen = load_existing_ids(CSV_PATH)
    fresh = [r for r in records if r["id"] not in seen]
    print(f"Новых (после дедупа): {len(fresh)} из {len(records)}")
    if not fresh:
        print("Все записи уже были. Выход.")
        sys.exit(0)

    # 3. Скоринг
    use_claude = not (args.demo or args.no_claude)
    print("Скоринг..." + (" (Claude)" if use_claude and os.environ.get("ANTHROPIC_API_KEY")
                          else " (офлайн-эвристика)"))
    scored = score_records(fresh, use_claude=use_claude)

    # 4. Вывод
    write_outputs(scored)
    bid = sum(1 for r in scored if r.get("action") == "BID_READY")
    print("\n" + "=" * 40)
    print(f"Готово. Записей: {len(scored)} | BID_READY: {bid}")
    print(f"  Excel:  {XLSX_PATH}")
    print(f"  CSV:    {CSV_PATH}")
    print(f"  Отчёт:  {REPORT_PATH}")


if __name__ == "__main__":
    main()
