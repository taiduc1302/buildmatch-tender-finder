#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tenderfinder_bulk_io.py — Bulk_Intake_Raw writer for TENDER_FINDER controlled runner
====================================================================
Appends raw/normalized acquisition rows to a dedicated workbook sheet.
This module intentionally does not write Future_Projects or Rejected_Archive.
"""

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from tenderfinder_excel_safety import append_untrusted_row

BULK_SHEET = "Bulk_Intake_Raw"

BULK_COLUMNS = [
    "run_id",
    "run_timestamp",
    "source_id",
    "source_name",
    "tier",
    "municipality",
    "access_status",
    "fetch_type",
    "resolved_endpoint",
    "records_pulled",
    "record_index",
    "classification",
    "output_route",
    "status",
    "richness",
    "project_id",
    "application_no",
    "address",
    "owner_applicant",
    "application_type_stage",
    "scope_summary",
    "fit_score",
    "source_url",
    "raw_json",
    "raw_file_path",
    "error",
    "notes",
]


def ensure_bulk_sheet(wb):
    """Return Bulk_Intake_Raw sheet, creating it if missing."""
    if BULK_SHEET in wb.sheetnames:
        ws = wb[BULK_SHEET]
        if ws.max_row < 1:
            ws.append(BULK_COLUMNS)
    else:
        ws = wb.create_sheet(BULK_SHEET)
        ws.append(BULK_COLUMNS)
        ws.freeze_panes = "A2"
    for c in range(1, len(BULK_COLUMNS) + 1):
        cell = ws.cell(1, c)
        cell.value = BULK_COLUMNS[c - 1]
        cell.font = Font(bold=True)
        width = max(12, min(36, len(BULK_COLUMNS[c - 1]) + 4))
        ws.column_dimensions[get_column_letter(c)].width = width
    return ws


def append_bulk_rows(wb, rows):
    """Append bulk rows to Bulk_Intake_Raw and return append stats."""
    ws = ensure_bulk_sheet(wb)
    before = max(0, ws.max_row - 1)
    for row in rows:
        append_untrusted_row(ws, [row.get(col, "") for col in BULK_COLUMNS])
    after = max(0, ws.max_row - 1)
    return {"appended": after - before, "final_rows": after}
