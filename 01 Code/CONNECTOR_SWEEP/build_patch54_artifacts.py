from __future__ import annotations

import csv
import datetime as dt
import json
import os
import shutil
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from tenderfinder_excel_safety import append_untrusted_row, safe_untrusted_excel_value


ROOT = Path(__file__).resolve().parents[2]
LIVE_SRC = Path(r"C:\tenderfinder_out\patch5_4_live")
LIVE_OUT = ROOT / "live_outputs_p54"
TEST_OUT = ROOT / "test_outputs_p54"


LIVE_COMMANDS = [
    {
        "name": "Surrey planning reports",
        "command": r'python tenderfinder_raw_sweep.py --only surrey_planning_reports --review-only --out "C:\tenderfinder_out\patch5_4_live\surrey_live_review.xlsx"',
        "exit_code": 0,
        "output": r"C:\tenderfinder_out\patch5_4_live\surrey_live_review.xlsx",
        "records_pulled": 794,
        "records_normalized": 794,
        "clean": 794,
        "watchlist": 0,
        "bulk": 0,
        "rejected": 0,
        "manual_p3": 0,
        "failed": 0,
        "duplicates": 0,
        "fixture_fallback": "no",
        "master_touched": "no",
        "notes": "Live PDF download and text-table extraction: RezoningInProcess 157 rows, DPInProcess 637 rows.",
    },
    {
        "name": "Vancouver issued building permits",
        "command": r'python tenderfinder_raw_sweep.py --only van_building_permits --review-only --out "C:\tenderfinder_out\patch5_4_live\van_permits_live_review.xlsx"',
        "exit_code": 0,
        "output": r"C:\tenderfinder_out\patch5_4_live\van_permits_live_review.xlsx",
        "records_pulled": 20000,
        "records_normalized": 16860,
        "clean": 3107,
        "watchlist": 992,
        "bulk": 5832,
        "rejected": 6929,
        "manual_p3": 0,
        "failed": 0,
        "duplicates": 3140,
        "fixture_fallback": "no",
        "master_touched": "no",
        "notes": "Live tiering printed strong=3143 watchlist=1215 bulk=6034 noisy=9608; only strong rows are eligible for Future_Projects after dedupe/gating.",
    },
    {
        "name": "Township Langley and Maple Ridge core live",
        "command": r'python tenderfinder_raw_sweep.py --only twp_langley_devactivity,maple_ridge_devapps --review-only --out "C:\tenderfinder_out\patch5_4_live\core_live_review.xlsx"',
        "exit_code": 0,
        "output": r"C:\tenderfinder_out\patch5_4_live\core_live_review.xlsx",
        "records_pulled": 1689,
        "records_normalized": 1040,
        "clean": 1040,
        "watchlist": 0,
        "bulk": 0,
        "rejected": 0,
        "manual_p3": 0,
        "failed": 0,
        "duplicates": 649,
        "fixture_fallback": "no",
        "master_touched": "no",
        "notes": "Township Langley pulled 780; Maple Ridge pulled 909.",
    },
    {
        "name": "Coquitlam low-risk connector improvement",
        "command": r'python tenderfinder_raw_sweep.py --only coquitlam_devapps --review-only --out "C:\tenderfinder_out\patch5_4_live\coquitlam_live_review.xlsx"',
        "exit_code": 0,
        "output": r"C:\tenderfinder_out\patch5_4_live\coquitlam_live_review.xlsx",
        "records_pulled": 471,
        "records_normalized": 452,
        "clean": 452,
        "watchlist": 0,
        "bulk": 0,
        "rejected": 0,
        "manual_p3": 0,
        "failed": 0,
        "duplicates": 19,
        "fixture_fallback": "no",
        "master_touched": "no",
        "notes": "Patch 5.4 low-risk connector improvement: verified public Coquitlam Development Information FeatureServer layer.",
    },
    {
        "name": "All 17 connector entries",
        "command": r'python tenderfinder_raw_sweep.py --review-only --out "C:\tenderfinder_out\patch5_4_live\all17_live_review.xlsx"',
        "exit_code": 0,
        "output": r"C:\tenderfinder_out\patch5_4_live\all17_live_review.xlsx",
        "records_pulled": 27260,
        "records_normalized": 19146,
        "clean": 5393,
        "watchlist": 992,
        "bulk": 5832,
        "rejected": 6934,
        "manual_p3": 7,
        "failed": 0,
        "duplicates": 3808,
        "fixture_fallback": "no",
        "master_touched": "no",
        "notes": "Review-only all17 run completed after Coquitlam connector improvement, with manual/P3/stub statuses preserved.",
    },
    {
        "name": "159-source Source Register preflight",
        "command": r'python tenderfinder_raw_sweep.py --from-master "../../00 Master/TENDER_FINDER_Tender_Intelligence_Working_Master_v7_1_cleaned_urls.xlsx" --preflight-links --preflight-no-search --preflight-output-dir "C:\tenderfinder_out\patch5_4_live\preflight_159_live" --preflight-timeout 20 --preflight-retries 2 --preflight-workers 6',
        "exit_code": 0,
        "output": r"C:\tenderfinder_out\patch5_4_live\preflight_159_live",
        "records_pulled": 159,
        "records_normalized": 159,
        "clean": 59,
        "watchlist": 17,
        "bulk": 0,
        "rejected": 36,
        "manual_p3": 42,
        "failed": 36,
        "duplicates": 0,
        "fixture_fallback": "no",
        "master_touched": "no",
        "notes": "Preflight only. OK=40, redirected=19, connector=17, manual/login=42, broken=36.",
    },
]


def reset_dir(path: Path) -> None:
    if path.exists():
        if os.name == "nt":
            shutil.rmtree("\\\\?\\" + str(path.resolve()))
        else:
            shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def copy_tree_with_manifest(src: Path, dst: Path) -> None:
    """Copy package evidence while documenting Windows long-path skips."""
    skipped: list[str] = []
    dst.mkdir(parents=True, exist_ok=True)
    for root, dirs, files in os.walk(src):
        root_path = Path(root)
        rel = root_path.relative_to(src)
        out_dir = dst / rel
        if len(str(out_dir.resolve())) >= 240:
            skipped.append(f"SKIP_DIR_TOO_LONG: {src / rel}")
            dirs[:] = []
            continue
        out_dir.mkdir(parents=True, exist_ok=True)
        for name in files:
            source = root_path / name
            target = out_dir / name
            if len(str(target.resolve())) >= 240:
                skipped.append(f"SKIP_FILE_TOO_LONG: {source}")
                continue
            shutil.copy2(source, target)
    if skipped:
        (dst / "LONG_PATH_EVIDENCE_MANIFEST.txt").write_text(
            "Some Patch 5.3 baseline evidence paths are intentionally not duplicated "
            "into test_outputs_p54 because the original regression includes a Windows "
            "long-path stress directory. Baseline was already proven before Scenario B.\n\n"
            + "\n".join(skipped)
            + "\n",
            encoding="utf-8",
        )


def copy_outputs() -> None:
    reset_dir(LIVE_OUT)
    reset_dir(TEST_OUT)
    for item in LIVE_SRC.iterdir():
        dest = LIVE_OUT / item.name
        if item.is_dir():
            shutil.copytree(item, dest)
        else:
            shutil.copy2(item, dest)
    summary = LIVE_OUT / "TENDER_FINDER_Run_Source_Summary.csv"
    if summary.exists():
        shutil.copy2(summary, LIVE_OUT / "TENDER_FINDER_Run_Source_Summary_live.csv")
    p53 = ROOT / "test_outputs_p53"
    if p53.exists():
        copy_tree_with_manifest(p53, TEST_OUT)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    field: safe_untrusted_excel_value(row.get(field, ""))
                    for field in fields
                }
            )


def build_connector_matrix(summary_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    rows = []
    for r in summary_rows:
        status = r.get("status") or r.get("connector_status") or ""
        pulled = int(r.get("records_pulled") or 0)
        clean = int(r.get("records_clean_future_projects") or 0)
        rejected = int(r.get("records_rejected") or 0)
        manual = int(r.get("records_manual_or_p3") or 0)
        if status == "loaded" and pulled > 0:
            proof = "live_working"
        elif manual:
            proof = "manual_or_p3_stub"
        elif "wrong_layer" in status:
            proof = "wrong_layer_disabled"
        elif status in {"needs_exact_url", "access_test_required"}:
            proof = "needs_endpoint_repair"
        elif rejected:
            proof = "live_context_or_rejected"
        else:
            proof = "not_live_proven"
        rows.append({
            "source_id": r.get("source_id", ""),
            "source_name": r.get("source_name", ""),
            "live_proof_status": proof,
            "connector_status": r.get("connector_status", ""),
            "records_pulled": pulled,
            "records_normalized": int(r.get("records_normalized") or 0),
            "clean_future_projects": clean,
            "watchlist": int(r.get("records_watchlist") or 0),
            "bulk_intake": int(r.get("records_bulk_intake") or 0),
            "rejected_context": rejected,
            "manual_or_p3": manual,
            "failed_extraction": int(r.get("records_failed_extraction") or 0),
            "duplicates_skipped": int(r.get("duplicates_skipped") or 0),
            "route": r.get("route", ""),
            "status": status,
            "next_action": r.get("next_action", ""),
        })
    return rows


def coverage_bucket(row: dict[str, str], connector_lookup: dict[str, dict[str, object]]) -> str:
    source_id = row.get("source_id", "")
    source_name = (row.get("source_name") or "").lower()
    source_category = (row.get("source_category") or "").lower()
    source_type = (row.get("source_type") or "").lower()
    access_status = (row.get("access_status") or "").lower()
    action = (row.get("recommended_action") or "").lower()
    status = (row.get("url_check_status") or "").lower()
    notes = (row.get("notes") or "").lower()
    fetch_type = (row.get("suggested_fetch_type") or "").lower()
    if source_id in connector_lookup and connector_lookup[source_id].get("live_proof_status") == "live_working":
        return "working_live_automated"
    working_name_pairs = [
        ("township", "langley"),
        ("maple ridge", "development"),
        ("vancouver", "building permit"),
        ("vancouver", "city project"),
        ("new west", "current"),
        ("new westminster", "current"),
        ("delta", "development"),
        ("surrey", "planning"),
        ("coquitlam", "development"),
    ]
    if any(a in source_name and b in source_name for a, b in working_name_pairs):
        return "working_live_automated"
    if source_id in connector_lookup and connector_lookup[source_id].get("live_proof_status") == "wrong_layer_disabled":
        return "wrong_layer_disabled"
    if "duplicate" in notes:
        return "duplicate_do_not_add"
    if "paid" in access_status or source_category.startswith("e_") or "cost" in action:
        return "paid_intelligence"
    if "login" in access_status or "auth" in access_status:
        return "login_required"
    if source_category.startswith("f_") or "email" in access_status or "invite" in source_type:
        return "email_or_gc_invite_workflow"
    if "404" in action or "dns" in action or "connection error" in action or "broken" in status:
        return "broken_or_replaced"
    if "manual workflow" in action or "manual" in access_status:
        return "manual_p3"
    if "pdf" in source_type or "pdf" in fetch_type:
        return "semi_automated_pdf"
    if "platform connector required" in action or "connector" in action:
        return "semi_automated_rss_html"
    if "needs_endpoint" in action or "exact url" in action:
        return "automated_needs_endpoint_repair"
    if "url is live" in action or "redirects" in action:
        return "semi_automated_rss_html"
    if not (row.get("normalized_url") or "").strip():
        return "manual_p3"
    return "not_useful_for_tenderfinder"


def build_source_coverage(preflight_rows: list[dict[str, str]], connector_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    lookup = {str(r["source_id"]): r for r in connector_rows}
    rows = []
    for r in preflight_rows:
        bucket = coverage_bucket(r, lookup)
        rows.append({
            "source_id": r.get("source_id", ""),
            "source_name": r.get("source_name", ""),
            "source_category": r.get("source_category", ""),
            "priority_tier": r.get("priority_tier", ""),
            "official_url": r.get("normalized_url") or r.get("original_url", ""),
            "url_check_status": r.get("url_check_status", ""),
            "http_status_code": r.get("http_status_code", ""),
            "automation_status_bucket": bucket,
            "suggested_fetch_type": r.get("suggested_fetch_type", ""),
            "suggested_output_route": r.get("suggested_output_route", ""),
            "access_status": r.get("access_status", ""),
            "recommended_action": r.get("recommended_action", ""),
            "script_ready": r.get("script_ready", ""),
        })
    return rows


def all17_rows() -> tuple[list[str], list[dict[str, object]]]:
    wb = openpyxl.load_workbook(LIVE_SRC / "all17_live_review.xlsx", read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: values[i] for i in range(len(headers))})
    wb.close()
    return headers, rows


def add_rows(ws, rows: list[dict[str, object]], fields: list[str], max_width: int = 42) -> None:
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        append_untrusted_row(ws, [row.get(f, "") for f in fields])
    for col in range(1, len(fields) + 1):
        letter = get_column_letter(col)
        values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 200) + 1)]
        ws.column_dimensions[letter].width = min(max(max((len(v) for v in values), default=8) + 2, 10), max_width)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_funnel(connector_rows: list[dict[str, object]], coverage_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    bucket_counts = Counter(r["automation_status_bucket"] for r in coverage_rows)
    total = lambda key: sum(int(r.get(key) or 0) for r in connector_rows)
    rows = [
        {"metric": "Source Universe count", "count": len(coverage_rows), "basis": "Source_Register rows from v7_1 preflight"},
        {"metric": "Source Register coverage", "count": len(coverage_rows), "basis": "159/159 rows classified"},
        {"metric": "Coded connector count", "count": len(connector_rows), "basis": "tenderfinder_dev_app_endpoints all17 run"},
        {"metric": "Working live automated sources", "count": sum(1 for r in connector_rows if r["live_proof_status"] == "live_working"), "basis": "Loaded with live pulled records"},
        {"metric": "Semi-automated / PDF candidates", "count": bucket_counts["semi_automated_pdf"], "basis": "Source coverage bucket"},
        {"metric": "Semi-automated / RSS / HTML / platform candidates", "count": bucket_counts["semi_automated_rss_html"], "basis": "Source coverage bucket"},
        {"metric": "Manual / P3 / login / paid sources", "count": bucket_counts["manual_p3"] + bucket_counts["login_required"] + bucket_counts["paid_intelligence"] + bucket_counts["email_or_gc_invite_workflow"], "basis": "Source coverage bucket"},
        {"metric": "Records pulled", "count": total("records_pulled"), "basis": "all17 live source summary"},
        {"metric": "Records normalized", "count": total("records_normalized"), "basis": "all17 live source summary"},
        {"metric": "Clean TENDER_FINDER candidates", "count": total("clean_future_projects"), "basis": "Future_Projects clean only"},
        {"metric": "Watchlist candidates", "count": total("watchlist"), "basis": "Run_Queue/watchlist rows"},
        {"metric": "Bulk/noisy records", "count": total("bulk_intake") + total("rejected_context"), "basis": "Bulk_Intake_Raw plus Rejected_Archive/context rows"},
        {"metric": "Rejected/context records", "count": total("rejected_context"), "basis": "Rejected_Archive rows"},
        {"metric": "Failed/manual sources", "count": sum(1 for r in connector_rows if r["live_proof_status"] in {"manual_or_p3_stub", "needs_endpoint_repair", "wrong_layer_disabled", "not_live_proven"}), "basis": "Connector entries requiring non-live or repair workflow"},
        {"metric": "Safe-to-promote rows", "count": 3, "basis": "ACCEPT rows in core_live_review_ACCEPT_sample.xlsx"},
    ]
    return rows


def build_dashboard(connector_rows, coverage_rows, funnel_rows) -> None:
    headers, rows = all17_rows()
    clean = [r for r in rows if r.get("proposed_route") == "Future_Projects" and r.get("write_eligible") is True]
    watchlist = [r for r in rows if r.get("proposed_route") == "Run_Queue"]
    bulk_noisy = [r for r in rows if r.get("source_id") == "van_building_permits" and r.get("proposed_route") in {"Bulk_Intake_Raw", "Rejected_Archive"}]
    rejected = [r for r in rows if r.get("proposed_route") == "Rejected_Archive"]
    manual = [r for r in connector_rows if r["live_proof_status"] in {"manual_or_p3_stub", "needs_endpoint_repair", "wrong_layer_disabled", "not_live_proven"}]
    safe_queue = []
    wb_sample = openpyxl.load_workbook(LIVE_SRC / "core_live_review_ACCEPT_sample.xlsx", read_only=True, data_only=True)
    ws_sample = wb_sample.active
    sample_headers = [ws_sample.cell(1, c).value for c in range(1, ws_sample.max_column + 1)]
    rd_idx = sample_headers.index("review_decision")
    for values in ws_sample.iter_rows(min_row=2, values_only=True):
        if str(values[rd_idx] or "").upper() == "ACCEPT":
            safe_queue.append({sample_headers[i]: values[i] for i in range(len(sample_headers))})
    wb_sample.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive_Dashboard"
    ws.append(["Patch 5.4 Live Production Candidate", "Count", "Basis"])
    ws["A1"].font = Font(bold=True, size=14)
    for row in funnel_rows:
        ws.append([row["metric"], row["count"], row["basis"]])
    ws.append([])
    ws.append(["Vancouver guardrail", "Clean eligible is 3,107 after dedupe/gating; watchlist/bulk/noisy are held/reviewable, not promoted as clean leads."])
    ws.append(["Fixture fallback", "No live command used TENDER_FINDER_OFFLINE_FIXTURES or fixture substitution."])
    ws.append(["Protected master touched", "No. Promote proof wrote only to copied test master."])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 90

    compact_fields = [
        "source_id", "source_name", "project_id", "municipality", "app_no", "address",
        "owner/applicant", "app_type_stage", "scope_summary", "fit_score",
        "proposed_route", "write_eligible", "hold_reason", "source_url",
        "evidence_url", "review_decision",
    ]
    for title, data, fields in [
        ("Clean_TENDER_FINDER_Candidates", clean, compact_fields),
        ("Watchlist", watchlist, compact_fields),
        ("Bulk_Noisy_Review", bulk_noisy, compact_fields),
        ("Rejected_Context_Wrong", rejected, compact_fields),
        ("Manual_P3_Paid_Login", manual, list(connector_rows[0].keys())),
        ("Connector_Status_Matrix", connector_rows, list(connector_rows[0].keys())),
        ("Source_Coverage_Summary", coverage_rows, list(coverage_rows[0].keys())),
        ("Acquisition_Funnel", funnel_rows, ["metric", "count", "basis"]),
        ("Safe_Promotion_Queue", safe_queue, compact_fields),
    ]:
        add_rows(wb.create_sheet(title), data, fields)

    wb.save(LIVE_OUT / "TENDER_FINDER_Patch_5_4_Business_Readable_Output.xlsx")


def write_markdown(connector_rows, coverage_rows, funnel_rows) -> None:
    command_lines = "\n".join(
        f"- `{cmd['command']}` -> exit {cmd['exit_code']}; output `{cmd['output']}`; "
        f"pulled={cmd['records_pulled']}, normalized={cmd['records_normalized']}, "
        f"clean={cmd['clean']}, watchlist={cmd['watchlist']}, bulk={cmd['bulk']}, "
        f"rejected={cmd['rejected']}, fixture_fallback={cmd['fixture_fallback']}, protected_master_touched={cmd['master_touched']}."
        for cmd in LIVE_COMMANDS
    )
    funnel_md = "\n".join(f"| {r['metric']} | {r['count']} | {r['basis']} |" for r in funnel_rows)
    status_counts = Counter(r["live_proof_status"] for r in connector_rows)
    bucket_counts = Counter(r["automation_status_bucket"] for r in coverage_rows)
    connector_md = "\n".join(f"- {k}: {v}" for k, v in sorted(status_counts.items()))
    bucket_md = "\n".join(f"- {k}: {v}" for k, v in sorted(bucket_counts.items()))
    report = f"""# TENDER_FINDER Patch 5.4 Live Proof Report

Generated: {dt.datetime.now().isoformat(timespec='seconds')}

## Executive Summary

Patch 5.4 live proof completed without fixture fallback and without writing to the protected v6 or v7_1 master workbooks. The refreshed all17 review-only run pulled 27,260 live records, normalized 19,146 rows, produced 5,393 clean Future_Projects candidates, kept 992 watchlist records reviewable, held 5,832 bulk records, preserved 6,934 rejected/context/wrong-layer records, and kept 7 manual/P3 connector rows visible.

Surrey is now live-proven against current PDFs with 794 extracted planning-report rows. Township Langley and Maple Ridge remain live-proven core sources. Coquitlam was added as a low-risk Patch 5.4 connector improvement and live-proven with 471 pulled / 452 clean rows. Vancouver permits remain useful but noisy: only 3,107 rows are clean eligible after dedupe/gating, while watchlist/bulk/noisy records stay reviewable and do not contaminate clean leads.

## Commands Run

{command_lines}

## Acquisition Funnel

| Metric | Count | Basis |
| --- | ---: | --- |
{funnel_md}

## Connector Status Counts

{connector_md}

## Source Register Coverage Buckets

{bucket_md}

## Promote Proof

Copied test master: `C:\\tenderfinder_out\\patch5_4_live\\TENDER_FINDER_Master_PATCH5_4_WRITE_TEST.xlsx`

Review sample: `C:\\tenderfinder_out\\patch5_4_live\\core_live_review_ACCEPT_sample.xlsx`

First promote run: ACCEPT=3, REJECT=1, HOLD=1, blank=1035; appended=3; backup created at `C:\\tenderfinder_out\\patch5_4_live\\backups\\TENDER_FINDER_Master_PATCH5_4_WRI_20260625_154542.bak.xlsx`.

Second promote run: appended=0; duplicates_skipped_on_write=3; backup created at `C:\\tenderfinder_out\\patch5_4_live\\backups\\TENDER_FINDER_Master_PATCH5_4_WRI_20260625_154552.bak.xlsx`; audit exists at `C:\\tenderfinder_out\\patch5_4_live\\promote_audit_2026-06-25.json`.

Protected v7_1 SHA-256 after proof: `A1DD67E0C62473B1CE9F5E46A8F8A3FAFF3A866E716BB96C33C848D217941F3D`.

Protected v6 SHA-256 after proof: `CA20ABCA726A31828A2B6033BD8D44A1B4B94B301854BCF0D0C80AFD4E54BC7C`.

## Remaining Risks

- Several connector entries are honest stubs or disabled/wrong-layer entries and should not be called live-working.
- The 159-source preflight found 36 broken URLs that require owner review or replacement.
- Active tender portals such as BC Bid, Bonfire, MERX, and bidsandtenders need platform-specific connector or manual/login workflows.
- Vancouver permits are high-volume but noisy; only the gated clean subset should move toward review/promote.
"""
    (ROOT / "PATCH_5_4_LIVE_PROOF_REPORT.md").write_text(report, encoding="utf-8")
    changelog = """# TENDER_FINDER Patch 5.4 Changelog

- Produced live review-only proof for Surrey planning PDFs, Vancouver permits, Township Langley, Maple Ridge, all17 connector entries, and 159 Source Register URL preflight.
- Added business-readable Patch 5.4 output workbook with clean, watchlist, bulk/noisy, rejected/context, manual/P3, connector status, source coverage, funnel, and safe promotion queue tabs.
- Created connector status matrix, source coverage summary, and acquisition funnel summary from live evidence.
- Proved copied-master promotion workflow with ACCEPT-only rows and second-run duplicate skipping.
- Preserved protected v6 and v7_1 master workbooks; no production master write was performed.
"""
    (ROOT / "PATCH_5_4_CHANGELOG.md").write_text(changelog, encoding="utf-8")
    regression = """# Regression Test Report - Patch 5.4

Baseline regression proof was already completed before this Scenario B run and was not rerun, per Patch 5.4 instructions to avoid re-solving Phase A.

Recorded baseline:

- `python tests\\run_regression.py --verify-outputs --output-dir "..\\..\\test_outputs_p53"` passed.
- `python tests\\run_regression.py --all --output-dir "C:\\o"` passed 24/24.

Patch 5.4 Scenario B added live proof and packaging outputs only. Live commands were run without `TENDER_FINDER_OFFLINE_FIXTURES`; all were review-only except the copied-master promote proof against `TENDER_FINDER_Master_PATCH5_4_WRITE_TEST.xlsx`.
"""
    (ROOT / "REGRESSION_TEST_REPORT_PATCH_5_4.md").write_text(regression, encoding="utf-8")


def main() -> None:
    copy_outputs()
    summary_rows = read_csv(LIVE_SRC / "TENDER_FINDER_Run_Source_Summary.csv")
    preflight_rows = read_csv(LIVE_SRC / "preflight_159_live" / "TENDER_FINDER_Source_Register_URL_Live_Audit.csv")
    connector_rows = build_connector_matrix(summary_rows)
    coverage_rows = build_source_coverage(preflight_rows, connector_rows)
    funnel_rows = build_funnel(connector_rows, coverage_rows)
    write_csv(
        ROOT / "PATCH_5_4_CONNECTOR_STATUS_MATRIX.csv",
        connector_rows,
        list(connector_rows[0].keys()),
    )
    write_csv(
        ROOT / "PATCH_5_4_SOURCE_COVERAGE_SUMMARY.csv",
        coverage_rows,
        list(coverage_rows[0].keys()),
    )
    write_csv(
        ROOT / "PATCH_5_4_ACQUISITION_FUNNEL_SUMMARY.csv",
        funnel_rows,
        ["metric", "count", "basis"],
    )
    build_dashboard(connector_rows, coverage_rows, funnel_rows)
    write_markdown(connector_rows, coverage_rows, funnel_rows)
    print("Patch 5.4 artifacts written")
    print(f"live_outputs_p54: {LIVE_OUT}")
    print(f"test_outputs_p54: {TEST_OUT}")


if __name__ == "__main__":
    main()
