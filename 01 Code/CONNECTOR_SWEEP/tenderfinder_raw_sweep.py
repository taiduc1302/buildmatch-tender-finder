#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TENDER_FINDER controlled all-source probe / load runner
===============================================
Patched from the original collect-only raw sweep into a SAFE, controlled runner
that bridges the master workbook + connector registry, probes sources, rejects
wrong layers, quarantines bad returns, separates Future Projects from Active
Tenders, and writes only useful structured leads into the existing master
workbook (v7) schema.

Target architecture:
  Source_Register / Run_Queue
    -> technical endpoint registry (tenderfinder_dev_app_endpoints.csv)
    -> safe probe / sweep dispatcher (fetch_type routes)
    -> classification + scoring (tenderfinder_guards)
    -> dedup / update (tenderfinder_master_io)
    -> Future_Projects / Active_Tenders / Rejected_Archive / Run_Log / Source_QA

SAFETY (enforced here):
  * No silent fallback. A missing slug/layer is reported, never substituted.
  * ArcGIS layer denylist + positive scoring + attribute-richness gate run
    BEFORE anything loads (tenderfinder_guards).
  * Wrong-layer / thin / trailing returns are quarantined, never crash the run.
  * Raw outputs (json/csv/logs) are always saved, even for rejected sources.
  * Paid / login sources are skipped (never scraped).
  * Writes go to v7 only; v6 is never overwritten; a backup precedes every write.

USAGE  (run from 01 Code/CONNECTOR_SWEEP/):
  python tenderfinder_raw_sweep.py --list
  python tenderfinder_raw_sweep.py --from-master <master.xlsx> --preflight-links --preflight-no-search
  python tenderfinder_raw_sweep.py --sync-registry --from-master <v7.xlsx>
  python tenderfinder_raw_sweep.py --from-master <v7.xlsx> --probe --out <probe.xlsx>
  python tenderfinder_raw_sweep.py --from-master <v7.xlsx> --tier 1 --max-records 500 --write-master <v7.xlsx>
  python tenderfinder_raw_sweep.py --from-master <v7.xlsx> --tier 1 --dry-run
  python tenderfinder_raw_sweep.py --only langley_tol,maple_ridge --probe

DEPENDENCIES
  Required: openpyxl
  Optional: requests (falls back to urllib)
  Local:    tenderfinder_guards.py, tenderfinder_master_io.py, tenderfinder_source_registry.py
"""

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
import urllib.error

try:
    import requests  # type: ignore
    _HAVE_REQUESTS = True
except Exception:
    _HAVE_REQUESTS = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
except Exception:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl", file=sys.stderr)
    sys.exit(2)

import tenderfinder_guards as G
import tenderfinder_source_registry as REG


# ---------------------------------------------------------------------------
# Patch 5.3 (P0.3): Windows-safe console logging
# ---------------------------------------------------------------------------
def init_windows_safe_console():
    """Prevent UnicodeEncodeError crashes on Windows consoles (cp1252).

    The live run on Windows crashed with:
        UnicodeEncodeError: 'charmap' codec can't encode character '\\u2192'
    whenever an arrow/em-dash/check-mark reached a cp1252 console. We reconfigure
    stdout/stderr to UTF-8 with errors='replace' so the console can never crash,
    while file logs (opened explicitly as UTF-8) keep full fidelity. This is a
    no-op safety wrapper on platforms that are already UTF-8.
    """
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if stream is None:
            continue
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")  # Py3.7+
        except Exception:
            try:
                import io
                wrapped = io.TextIOWrapper(stream.buffer, encoding="utf-8",
                                           errors="replace", line_buffering=True)
                setattr(sys, stream_name, wrapped)
            except Exception:
                pass  # last resort: leave stream as-is; ASCII-only strings still safe


# Apply immediately at import so any module-level prints are also safe.
init_windows_safe_console()


UA = "TENDER_FINDER-tender-intel/2.0 (civil contractor lead screen; contact estimating@example.com)"
REST_LAYER_RE = re.compile(r"https?://[^\s\"'<>]+?/(?:FeatureServer|MapServer)/\d+", re.I)

# Fetch types that are NEVER auto-fetched (queue/manual/active-tender/paid).
QUEUE_FETCH_TYPES = {"web_page", "pdf", "council_agenda", "capital_plan",
                     "rss", "email_gc_invite", "manual_only"}
SKIP_FETCH_TYPES = {"paid_login", "paid_or_login"}
ARCGIS_FETCH_TYPES = {"arcgis_hub_item", "arcgis_hub_discover",
                      "arcgis_map_discover", "arcgis_mapserver",
                      "arcgis_rest_layer", "arcgis_feature_service"}
# Fetch types with their own dedicated connector modules (Patch 5.0)
CUSTOM_FETCH_TYPES = {"surrey_planning_reports"}


# ---------------------------------------------------------------------------
# Patch 5.0 verified-package fallback fixtures
# ---------------------------------------------------------------------------
def _looks_like_network_blocked_error(exc_or_text):
    txt = str(exc_or_text or "").lower()
    needles = [
        "temporary failure in name resolution", "name resolution", "dns", "gaierror",
        "failed to establish a new connection", "connectionerror", "max retries exceeded",
        "network is unreachable", "connection timed out", "read timed out",
    ]
    return any(n in txt for n in needles)


def _offline_fixtures_enabled():
    """Patch 5.3: explicit opt-in for offline regression fixtures.

    The packaged regression suite runs in network-restricted sandboxes whose
    egress proxies may return DNS failures OR HTTP 403/Forbidden. When
    TENDER_FINDER_OFFLINE_FIXTURES is set (only the regression harness sets it), any
    fetch failure may fall back to packaged fixtures so reproducible
    runner/output/promotion behavior can be verified without live network.

    This is DELIBERATELY off by default. On a real machine a government 403 is
    surfaced honestly as FORBIDDEN_BUT_LIKELY_VALID and is NEVER silently
    replaced with fixture data.
    """
    return str(os.environ.get("TENDER_FINDER_OFFLINE_FIXTURES", "")).strip().lower() in ("1", "true", "yes")


def _should_use_fixture_fallback(exc_or_text):
    """True when a fetch failure should fall back to packaged fixtures."""
    if _looks_like_network_blocked_error(exc_or_text):
        return True
    if _offline_fixtures_enabled():
        txt = str(exc_or_text or "").lower()
        sandbox_signals = ["403", "forbidden", "proxy", "egress", "blocked",
                           "denied", "407", "tunnel", "ssl", "certificate"]
        return any(s in txt for s in sandbox_signals) or bool(txt)
    return False


def _fixture_csv_path(source_id):
    mapping = {
        "surrey_planning_reports": "surrey_planning_reports_review_fixture.csv",
        "twp_langley_devactivity": "twp_langley_devactivity_review_fixture.csv",
        "maple_ridge_devapps": "maple_ridge_devapps_review_fixture.csv",
    }
    name = mapping.get(str(source_id or ""))
    if not name:
        return None
    return os.path.join(os.path.dirname(__file__), "tests", "fixtures", name)


def _load_fixture_leads_for_connector(conn, res, original_error=""):
    """Load explicit regression fixtures only when sandbox/live network is blocked.

    This is not a live-source success claim. It makes the packaged regression
    runner reproducible from a fresh unzip when public municipal endpoints are
    unreachable from the execution sandbox.
    """
    cid = conn.get("source_id") or conn.get("id")
    path = _fixture_csv_path(cid)
    if not path or not os.path.exists(path):
        return None
    import csv as _csv
    leads = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in _csv.DictReader(f):
            if not any((row.get(k) or "").strip() for k in row):
                continue
            leads.append({
                "project_id": row.get("project_id") or row.get("app_no") or "",
                "date_found": row.get("date_found") or dt.date.today().isoformat(),
                "source": row.get("source_name") or conn.get("name") or cid,
                "source_url": row.get("source_url") or row.get("evidence_url") or conn.get("last_good_endpoint") or conn.get("endpoint") or "",
                "evidence_url": row.get("evidence_url") or row.get("source_url") or "",
                "title": row.get("app_type_stage") or row.get("scope_summary") or row.get("project_id") or "",
                "owner": row.get("owner/applicant") or row.get("owner") or "",
                "municipality": row.get("municipality") or conn.get("municipality") or "",
                "app_no": row.get("app_no") or row.get("application_no") or row.get("project_id") or "",
                "app_type_stage": row.get("app_type_stage") or "",
                "scope_summary": row.get("scope_summary") or "",
                "fit_score": row.get("fit_score") or 50,
                "fit_reason": row.get("fit_reason") or "Packaged regression fixture used after network/DNS failure.",
                "verification": "Needs Review",
                "horizon": "",
                "est_value": "Not available",
                "next_milestone": "",
                "assigned_to": "",
                "notes": "Regression fixture fallback. Rerun live on a networked TENDER_FINDER machine for live-source proof.",
                "address": row.get("address") or "",
                "proposed_route": row.get("proposed_route") or "Future_Projects",
                "write_eligible": True,
                "hold_reason": row.get("hold_reason") or "",
                "raw_hash": row.get("raw_hash") or "",
                "dedupe_key": row.get("dedupe_key") or "",
            })
    if not leads:
        return None
    res["records_pulled"] = len(leads)
    res["records_normalized"] = len(leads)
    res["raw_records"] = []
    res["leads"] = leads
    res["rejected"] = []
    res["status"] = "fixture_fallback_loaded"
    res["classification"] = G.CLASS_DEV_LEAD
    res["output_route"] = "Future_Projects"
    res["richness"] = "fixture fallback after network/DNS failure"
    res["error"] = ""
    res["next_action"] = "fixture fallback used; rerun live on a networked machine for live-source proof"
    res["_fixture_fallback"] = True
    res["_fixture_error"] = str(original_error or "")
    print(f"    [fixture-fallback] {cid}: loaded {len(leads)} packaged regression rows after network/DNS failure")
    return res

PROTECTED_V6_FILENAME = "TENDER_FINDER_Tender_Intelligence_Working_Master_v6.xlsx"


def _fail(message, code=2):
    print(f"ERROR: {message}", file=sys.stderr)
    sys.exit(code)


def _resolve_path(path):
    return os.path.abspath(os.path.normpath(path))


def _validate_write_master_target(path):
    """Fail closed unless the write target is clearly a copied v7/test workbook."""
    resolved = _resolve_path(path)
    filename = os.path.basename(resolved)
    filename_l = filename.lower()
    protected_l = PROTECTED_V6_FILENAME.lower()

    if filename_l == protected_l or protected_l in resolved.lower():
        _fail(f"Refusing to write to protected v6 master workbook: {resolved}")
    if "working_master_v6" in filename_l or "_v6" in filename_l:
        _fail(f"Refusing write target that appears to be v6/protected: {resolved}")
    if "v7" not in filename_l and "test" not in filename_l and "bulk" not in filename_l:
        _fail("Refusing write target that does not look like a copied v7/test/bulk workbook: "
              f"{resolved}")
    return resolved


def _num_or_none(value):
    try:
        if value in (None, ""):
            return None
        return float(value)
    except Exception:
        return None


_FUTURE_PROJECT_ROUTES = {"Future_Projects", "Future_Projects_Needs_Review"}


def _connector_pre_routed(lead):
    """True if a connector has already made a non-default routing decision for
    this lead (e.g. the Vancouver permit tier filter or Surrey fit gate).

    Patch 5.3 P0.2: global write-gates must NEVER override these per-connector
    decisions, because doing so silently re-routed bulk/noisy/watchlist
    Vancouver permits into Future_Projects and inflated the eligible count.
    """
    if lead.get("write_eligible") is False:
        return True
    route = (lead.get("proposed_route") or "").strip()
    if route and route not in _FUTURE_PROJECT_ROUTES:
        return True
    if (lead.get("hold_reason") or "").strip():
        return True
    return False


def _apply_write_gates(results, min_fit_score=None, max_write_per_source=0):
    """Mark normalized leads as write-eligible or held for review.

    This protects broader writes by filtering at the record level after layer
    classification and normalization, but before any master workbook operation.

    Patch 5.3 (P0.2): this function now RESPECTS per-connector routing. If a
    connector already held a lead or routed it somewhere other than
    Future_Projects (Vancouver permit watchlist/bulk/noisy, Surrey low-fit,
    etc.), the gate leaves that decision intact. The optional thresholds below
    can only TIGHTEN (move an otherwise-eligible lead to held); they can never
    loosen a held lead back into Future_Projects.
    """
    for r in results:
        eligible_for_source = 0
        leads = r.get("leads") or []
        for lead in leads:
            # 1) Respect connector-level routing. Do not resurrect held leads.
            if _connector_pre_routed(lead):
                lead["write_eligible"] = False
                if not lead.get("proposed_route"):
                    lead["proposed_route"] = "Run_Queue"
                if not lead.get("hold_reason"):
                    lead["hold_reason"] = "held_by_connector_routing"
                # keep any routing_decision/reason the connector set
                lead.setdefault("routing_decision", lead.get("proposed_route"))
                lead.setdefault("routing_reason", lead.get("hold_reason"))
                continue

            # 2) Lead was a clean Future_Projects candidate. Apply optional gates.
            hold_reasons = []
            fit = _num_or_none(lead.get("fit_score"))
            if min_fit_score is not None:
                if fit is None:
                    hold_reasons.append("missing_fit_score")
                elif fit < min_fit_score:
                    hold_reasons.append(f"fit_score_below_min_{min_fit_score:g}")
            if not hold_reasons and max_write_per_source and eligible_for_source >= max_write_per_source:
                hold_reasons.append(f"max_write_per_source_{max_write_per_source}")

            if hold_reasons:
                lead["write_eligible"] = False
                lead["proposed_route"] = "Run_Queue"
                lead["hold_reason"] = "; ".join(hold_reasons)
                lead["routing_decision"] = "Run_Queue"
                lead["routing_reason"] = "; ".join(hold_reasons)
            else:
                lead["write_eligible"] = True
                lead["proposed_route"] = lead.get("proposed_route") or "Future_Projects"
                lead["hold_reason"] = ""
                lead["routing_decision"] = lead.get("proposed_route") or "Future_Projects"
                lead.setdefault("routing_reason", "passed_write_gates")
                eligible_for_source += 1

        r["records_normalized"] = len(leads)
        r["records_eligible"] = sum(1 for lead in leads if lead.get("write_eligible"))
        r["records_held_for_review"] = len(leads) - r["records_eligible"]
    return results


def _eligible_leads(results):
    return [lead for r in results for lead in (r.get("leads") or [])
            if lead.get("write_eligible")]


def _held_leads(results):
    return [lead for r in results for lead in (r.get("leads") or [])
            if not lead.get("write_eligible")]


def _summary_counts(results, eligible_leads, rejected):
    return {
        "records_pulled": sum(int(r.get("records_pulled") or 0) for r in results),
        "records_normalized": sum(int(r.get("records_normalized") or 0) for r in results),
        "records_eligible": len(eligible_leads),
        "records_held_for_review": sum(int(r.get("records_held_for_review") or 0) for r in results),
        "records_rejected": len(rejected),
    }


def _excel_safe_text(value, limit=32000):
    """Return a string safe for an Excel cell."""
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", " ", text)
    if len(text) > limit:
        return text[:limit - 40] + " ... [truncated]"
    return text


def _json_safe_blob(value, limit=32000):
    try:
        text = json.dumps(value, ensure_ascii=False, default=str, separators=(",", ":"))
    except Exception:
        text = str(value)
    return _excel_safe_text(text, limit=limit)


def _result_source_url(result, attrs=None):
    if attrs:
        picked = G.pick_field(attrs, "url")
        if picked:
            return str(picked)
    resolved = result.get("resolved_endpoint") or ""
    if resolved:
        return str(resolved)
    resolved_list = result.get("resolved") or []
    if resolved_list:
        return str(resolved_list[0])
    return ""


def _bulk_row_from_attrs(result, attrs, record_index, run_id, run_timestamp):
    """Normalize any fetched record into a Bulk_Intake_Raw row.

    This is intentionally looser than Future_Projects normalization. It keeps
    raw rejected/thin/context records too, so the workbook can be used as a
    harvest/debug dump without polluting Future_Projects.
    """
    attrs = attrs or {}
    muni = result.get("municipality") or result.get("source_name") or ""
    app_no = G.pick_field(attrs, "app_no") or attrs.get("OBJECTID") or attrs.get("FID") or ""
    app_type = G.pick_field(attrs, "app_type")
    status = G.pick_field(attrs, "status")
    address = G.pick_field(attrs, "address")
    applicant = G.pick_field(attrs, "applicant")
    desc = G.pick_field(attrs, "desc")
    source_url = _result_source_url(result, attrs)

    if app_no:
        pid = f"{G.norm_municipality(muni).upper()}-{G.norm_appno(app_no)}"
    elif address:
        pid = f"{G.norm_municipality(muni).upper()}-{G.norm_address(address)[:20].replace(' ', '-').upper()}"
    else:
        pid = f"{result.get('source_id')}-{record_index}"

    blob = " ".join(str(v) for v in attrs.values() if v is not None)
    blob += f" {app_type or ''} {status or ''} {desc or ''} {applicant or ''}"
    fit = G.score_civil_fit(blob, muni) if blob.strip() else ""
    type_stage = " — ".join([str(x) for x in (app_type, status) if x]) or ""

    return {
        "run_id": run_id,
        "run_timestamp": run_timestamp,
        "source_id": result.get("source_id"),
        "source_name": result.get("source_name"),
        "tier": result.get("tier"),
        "municipality": muni,
        "access_status": result.get("access_status"),
        "fetch_type": result.get("fetch_type"),
        "resolved_endpoint": result.get("resolved_endpoint") or " | ".join(result.get("resolved") or []),
        "records_pulled": result.get("records_pulled") or 0,
        "record_index": record_index,
        "classification": result.get("classification"),
        "output_route": result.get("output_route"),
        "status": result.get("status"),
        "richness": result.get("richness"),
        "project_id": pid,
        "application_no": str(app_no),
        "address": str(address or ""),
        "owner_applicant": str(applicant or ""),
        "application_type_stage": type_stage,
        "scope_summary": str(desc or "")[:500],
        "fit_score": fit,
        "source_url": source_url,
        "raw_json": _json_safe_blob(attrs),
        "raw_file_path": result.get("raw_json_path") or "",
        "error": result.get("error") or "",
        "notes": result.get("next_action") or "",
    }


def _bulk_stub_row(result, run_id, run_timestamp):
    """Create one Bulk_Intake_Raw row for manual/P3/access/error sources."""
    return {
        "run_id": run_id,
        "run_timestamp": run_timestamp,
        "source_id": result.get("source_id"),
        "source_name": result.get("source_name"),
        "tier": result.get("tier"),
        "municipality": result.get("municipality") or "",
        "access_status": result.get("access_status"),
        "fetch_type": result.get("fetch_type"),
        "resolved_endpoint": result.get("resolved_endpoint") or " | ".join(result.get("resolved") or []),
        "records_pulled": result.get("records_pulled") or 0,
        "record_index": 0,
        "classification": result.get("classification"),
        "output_route": result.get("output_route"),
        "status": result.get("status"),
        "richness": result.get("richness"),
        "project_id": result.get("source_id"),
        "application_no": "",
        "address": "",
        "owner_applicant": "",
        "application_type_stage": "",
        "scope_summary": result.get("next_action") or result.get("richness") or "",
        "fit_score": "",
        "source_url": _result_source_url(result),
        "raw_json": "",
        "raw_file_path": result.get("raw_json_path") or "",
        "error": result.get("error") or "",
        "notes": result.get("next_action") or result.get("error") or "bulk intake stub; no pull performed",
    }


def build_bulk_rows(results, run_id, run_timestamp):
    rows = []
    per_source = {}
    for result in results:
        start = len(rows)
        raw_records = result.get("raw_records") or []
        if raw_records:
            for i, attrs in enumerate(raw_records, start=1):
                rows.append(_bulk_row_from_attrs(result, attrs, i, run_id, run_timestamp))
        else:
            rows.append(_bulk_stub_row(result, run_id, run_timestamp))
        per_source[result.get("source_id")] = len(rows) - start
    return rows, per_source


def _bulk_access_filter(rows, args):
    """Keep bulk mode broad but explicit for trailing/wrong-layer dumps."""
    out = []
    for c in rows:
        access = (c.get("access_status") or "").strip().lower()
        if access == "trailing_context" and not args.include_trailing_context:
            continue
        if access == "disabled_wrong_layer" and not args.include_wrong_layer:
            continue
        out.append(c)
    return out


# ============================================================================
# HTTP helpers
# ============================================================================
def http_get(url, timeout=90):
    if _HAVE_REQUESTS:
        r = requests.get(url, headers={"User-Agent": UA}, timeout=timeout)
        r.raise_for_status()
        return r.text
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def http_get_json(url, timeout=90):
    return json.loads(http_get(url, timeout=timeout))


# ============================================================================
# Endpoint resolution (no silent fallback)
# ============================================================================
def resolve_hub_item(item_id, layer_index):
    meta = http_get_json(f"https://www.arcgis.com/sharing/rest/content/items/{item_id}?f=json")
    base = (meta.get("url") or "").rstrip("/")
    if not base:
        raise RuntimeError("endpoint_not_found: item has no service URL")
    idx = layer_index if (layer_index not in (None, "")) else 0
    name = meta.get("title") or item_id
    return name, f"{base}/{idx}"


def discover_hub_candidates(hub_base, keywords):
    """Return ranked, denylist-filtered (title, layer_url) candidates from the
    ArcGIS Hub DCAT feed. Only explicit REST layer URLs are used — NO base+/0
    blind fallback (that caused wrong-layer pollution)."""
    feed_url = hub_base.rstrip("/") + "/api/feed/dcat-us/1.1.json"
    feed = http_get_json(feed_url)
    cands = []
    for ds in feed.get("dataset", []):
        title = ds.get("title") or ""
        tl = title.lower()
        if keywords and not any(k in tl for k in keywords):
            continue
        blob = json.dumps(ds)
        for u in dict.fromkeys(REST_LAYER_RE.findall(blob)):
            cands.append((title, u))
    return G.rank_candidate_layers(cands)   # denylist + positive scoring


def discover_map_candidates(mapserver_url, keywords):
    meta = http_get_json(mapserver_url.rstrip("/") + "?f=json")
    cands = []
    for lyr in meta.get("layers", []):
        name = lyr.get("name") or ""
        nl = name.lower()
        if keywords and not any(k in nl for k in keywords):
            continue
        cands.append((name, f"{mapserver_url.rstrip('/')}/{lyr['id']}"))
    return G.rank_candidate_layers(cands)


# ============================================================================
# Data pulls (return RAW attribute dicts so the richness gate can sample them)
# ============================================================================
def query_arcgis_layer(layer_url, max_records, page=250, pause=0.4,
                       max_retries=3, retry_backoff=2.0):
    """Pull records from an ArcGIS FeatureServer/MapServer layer.

    Patch 5.0 changes vs previous:
    - page size reduced from 1000 → 250 (prevents 504 Gateway Timeout on
      servers like Maple Ridge that time out on large result sets)
    - per-batch retry with exponential backoff on 504/connection errors
    - partial results are returned with a warning log rather than failing silently
    - batch progress is logged at DEBUG level for diagnostics
    """
    rows, offset = [], 0
    batch_num = 0
    partial_warning = False

    while True:
        batch_num += 1
        params = {"where": "1=1", "outFields": "*", "returnGeometry": "false",
                  "resultOffset": offset, "resultRecordCount": page, "f": "json"}
        url = layer_url.rstrip("/") + "/query?" + urllib.parse.urlencode(params)

        batch = None
        for attempt in range(1, max_retries + 1):
            try:
                data = http_get_json(url, timeout=60)
                if isinstance(data, dict) and data.get("error"):
                    raise RuntimeError(f"arcgis_error: {data['error'].get('message', data['error'])}")
                feats = data.get("features", []) if isinstance(data, dict) else []
                batch = [f.get("attributes", {}) for f in feats]
                break  # success
            except Exception as e:
                err_str = str(e)
                is_504 = (
                    "504" in err_str
                    or "503" in err_str
                    or "Gateway" in err_str
                    or "Service Unavailable" in err_str
                    or "timeout" in err_str.lower()
                )
                if attempt < max_retries and is_504:
                    retry_waits = (1, 3, 8)
                    wait = retry_waits[min(attempt - 1, len(retry_waits) - 1)]
                    print(f"    [arcgis] batch={batch_num} offset={offset} 503/504/timeout "
                          f"(attempt {attempt}/{max_retries}), retrying in {wait:.0f}s...")
                    time.sleep(wait)
                else:
                    if rows:
                        partial_warning = True
                        print(f"    [arcgis] PARTIAL: batch={batch_num} failed after "
                              f"{attempt} attempts: {e}. Returning {len(rows)} rows collected so far.")
                        break  # return what we have
                    else:
                        raise  # nothing yet, surface the error

        if batch is None:  # partial_warning path: give up on this batch
            break

        rows.extend(batch)
        print(f"    [arcgis] batch={batch_num} offset={offset} fetched={len(batch)} total={len(rows)}")

        if len(batch) < page or len(rows) >= max_records:
            break
        offset += page
        time.sleep(pause)

    if partial_warning:
        print(f"    [arcgis] WARNING: partial result only - {len(rows)} rows returned. "
              f"Some batches failed. Check logs for 504 details.")

    return rows[:max_records]


def query_ods(base, slug, max_records):
    """Pull an Opendatasoft v2.1 dataset. NO keyword fallback: if the slug does
    not exist, raise slug_not_found (the old auto-search caused wrong datasets)."""
    # verify the exact slug exists first
    try:
        http_get_json(f"{base}/api/explore/v2.1/catalog/datasets/{slug}")
    except Exception:
        raise RuntimeError(f"slug_not_found: ODS dataset '{slug}' not found "
                           f"(no silent fallback)")
    url = f"{base}/api/explore/v2.1/catalog/datasets/{slug}/exports/json"
    data = http_get_json(url)
    if isinstance(data, list):
        return data[:max_records]
    if isinstance(data, dict) and isinstance(data.get("results"), list):
        return data["results"][:max_records]
    return []


# ============================================================================
# Normalisation: raw attribute dict -> Future_Projects lead dict
# ============================================================================
def normalize_lead(conn, layer_name, attrs):
    muni = conn.get("municipality") or conn.get("name") or ""
    app_no = G.pick_field(attrs, "app_no") or attrs.get("OBJECTID") or attrs.get("FID") or "?"
    app_type = G.pick_field(attrs, "app_type")
    status = G.pick_field(attrs, "status")
    address = G.pick_field(attrs, "address")
    applicant = G.pick_field(attrs, "applicant")
    desc = G.pick_field(attrs, "desc")
    url = G.pick_field(attrs, "url") or conn.get("last_good_endpoint") or conn.get("endpoint")

    muni_tag = G.norm_municipality(muni).upper()
    pid = f"{muni_tag}-{G.norm_appno(app_no)}" if app_no not in ("", "?") else \
          f"{muni_tag}-{G.norm_address(address)[:20].replace(' ', '-').upper()}"

    blob = " ".join(str(v) for v in attrs.values() if v is not None)
    blob += f" {app_type} {status} {desc} {applicant}"
    fit = G.score_civil_fit(blob, muni)

    type_stage = " — ".join([str(x) for x in (app_type, status) if x]) or layer_name
    scope = (str(desc) or str(app_type) or "").strip()[:300]

    return {
        "project_id": pid,
        "date_found": dt.date.today().isoformat(),
        "source": conn.get("name") or conn.get("source_id"),
        "source_url": str(url),
        "title": (str(desc)[:120] or f"{muni} {app_type}".strip() or pid),
        "owner": str(applicant) if applicant else "Not available",
        "municipality": muni,
        "app_no": str(app_no),
        "app_type_stage": type_stage,
        "scope_summary": scope,
        "expected_civil": "",     # left for human / PROMPT_DEV_APP_SCORING pass
        "fit_score": fit,
        "verification": "Needs Review",   # AI-extracted -> never 'Verified'
        "horizon": "",
        "est_value": "Not available",
        "next_milestone": status or "",
        "assigned_to": "",
        "notes": f"Auto-extracted from {layer_name}. Owner/value need manual 2nd pass.",
        "address": str(address),         # dedup helper (not a FP column)
        "_raw": attrs,
    }


# ============================================================================
# Connector runner — probe or pull one connector, with all gates
# ============================================================================
def run_connector(conn, max_records, probe=False, raw_dir=None):
    """Returns (result) dict with classification + payloads.
    Never raises for data problems — failures are captured as status."""
    cid = conn.get("source_id") or conn.get("id")
    ft = (conn.get("fetch_type") or "").strip()
    endpoint = (conn.get("last_good_endpoint") or conn.get("endpoint") or "").strip()
    kw = [k.strip().lower() for k in (conn.get("layer_keywords") or "").split(";") if k.strip()]
    access = (conn.get("access_status") or "").strip().lower()
    
    res = {"source_id": cid, "source_name": conn.get("name"), "fetch_type": ft,
           "tier": conn.get("priority_tier") or (conn.get("tier") or "").split()[0],
           "municipality": conn.get("municipality"),
           "access_status": conn.get("access_status"),
           "resolved": [], "resolved_endpoint": "", "raw_records": [],
           "raw_json_path": "", "records_pulled": 0, "classification": None,
           "output_route": None, "status": "", "richness": "", "error": "",
           "next_action": "", "leads": [], "rejected": []}

    # ---- Patch 5.3: offline fixture short-circuit (regression harness only) -
    # When TENDER_FINDER_OFFLINE_FIXTURES is set we NEVER touch the network. Fixture-
    # backed connectors load their fixture immediately (no live fetch), which
    # keeps the regression suite fast and deterministic even in sandboxes whose
    # egress hangs (TCP connects but never responds) instead of failing fast --
    # the cause of the suite appearing to hang after the Maple Ridge step.
    # Non-fixture connectors fall through to the normal fast access-status
    # checks below; only an actual network fetch is skipped (see guard before
    # the fetch dispatch). Production never sets this flag.
    if not probe and _offline_fixtures_enabled():
        fb = _load_fixture_leads_for_connector(conn, res, "offline_fixtures_enabled")
        if fb:
            fb["_fixture_fallback"] = True
            fb["status"] = "fixture_fallback_loaded"
            return fb

    # ---- access_status check FIRST (blocks before fetch_type routing) ------
    if access == "manual_p3_only":
        res["status"] = "p3_extract_required"
        res["classification"] = G.CLASS_P3
        res["output_route"] = "Run_Queue"
        res["next_action"] = f"run P3 extractor ({conn.get('prompt_type') or 'PROMPT_WEB_PDF_EXTRACT'})"
        return res
    if access == "access_test_required":
        res["status"] = "access_test_required"
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = "run access test (e.g. TENDER_FINDER office network)"
        return res
    if access == "paid_or_login_skip":
        res["status"] = "paid_or_login_skip"
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = "skip unless paid/login access configured"
        return res
    if access == "disabled_wrong_layer":
        res["status"] = "disabled_wrong_layer"
        res["classification"] = G.CLASS_WRONG
        res["output_route"] = "Rejected_Archive"
        res["next_action"] = "re-pin to correct dev-app layer"
        res["rejected"] = [{
            "project_id": cid, "source": conn.get("name"), "source_url": endpoint,
            "title": "DISABLED: wrong layer", "municipality": conn.get("municipality"),
            "scope_summary": "Source disabled due to previous wrong-layer detection.",
            "reason": "disabled_wrong_layer", "verification": "Rejected",
            "notes": "Re-verify and re-pin the endpoint before re-enabling.",
        }]
        return res
    if access in ("needs_exact_url", "endpoint_stale"):
        res["status"] = "needs_exact_url"
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = "verify/update endpoint manually; no silent fallback allowed"
        return res

    # ---- non-fetch fetch types: queue / skip (legacy fallback) -----------
    if endpoint.startswith("DISABLED") or ft == "" and not endpoint:
        res["status"] = "access_test_required"
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = "verify endpoint / run access test"
        return res
    if ft in SKIP_FETCH_TYPES:
        res["status"] = "paid_or_login_skip"
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = "skip unless paid/login access configured"
        return res
    if ft in QUEUE_FETCH_TYPES:
        res["status"] = "p3_extract_required"
        res["classification"] = G.CLASS_P3
        res["output_route"] = "Run_Queue"
        res["next_action"] = f"run P3 extractor ({conn.get('prompt_type') or 'PROMPT_WEB_PDF_EXTRACT'})"
        return res

    # ---- Patch 5.3: offline network-skip for non-fixture fetch connectors --
    # Reached only when offline mode is set and this connector has no fixture
    # AND would otherwise make a live network call. Skip the fetch (no hang).
    if not probe and _offline_fixtures_enabled():
        res["status"] = "offline_no_fixture"
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = ("offline mode (TENDER_FINDER_OFFLINE_FIXTURES): no live fetch attempted; "
                              "rerun live on a networked machine for this source")
        res["_fixture_fallback"] = True
        return res

    # ---- custom connector modules (Patch 5.0) ---------------------------------
    if ft in CUSTOM_FETCH_TYPES:
        try:
            from tenderfinder_surrey_inprocess import run_surrey_connector
            custom_res = run_surrey_connector(conn, max_records, probe=probe,
                                              result_template=res, debug_dir=raw_dir)
            if (not custom_res.get("leads")) and _should_use_fixture_fallback(custom_res.get("error") or custom_res.get("next_action") or ""):
                fb = _load_fixture_leads_for_connector(conn, custom_res, custom_res.get("error") or custom_res.get("next_action"))
                if fb:
                    return fb
            return custom_res
        except ImportError:
            res["error"] = "tenderfinder_surrey_inprocess.py not found in CONNECTOR_SWEEP/"
            res["status"] = "connector_module_missing"
            res["classification"] = G.CLASS_MANUAL
            res["output_route"] = "Run_Queue"
            res["next_action"] = "ensure tenderfinder_surrey_inprocess.py is present"
            return res
        except Exception as e:
            if _should_use_fixture_fallback(e):
                fb = _load_fixture_leads_for_connector(conn, res, e)
                if fb:
                    return fb
            res["error"] = str(e)
            res["status"] = "custom_fetch_failed"
            res["classification"] = G.CLASS_MANUAL
            res["output_route"] = "Run_Queue"
            res["next_action"] = "review custom connector failure"
            return res

    # ---- resolve endpoint --------------------------------------------------
    candidates = []   # list of (layer_name, layer_url)
    try:
        if ft in ("arcgis_hub_item",):
            # If last_good_endpoint is already a concrete FeatureServer/MapServer URL,
            # use it directly; do NOT pass it to resolve_hub_item (which expects item ID).
            good_ep = conn.get("last_good_endpoint", "").strip()
            if good_ep and ("/FeatureServer/" in good_ep or "/MapServer/" in good_ep):
                # Direct pinned URL; skip hub resolution
                name = conn.get("name") or "pinned_layer"
                candidates = [(name, good_ep)]
            else:
                # Resolve from item ID
                name, url = resolve_hub_item(endpoint, conn.get("layer_index"))
                candidates = [(name, url)]
        elif ft in ("arcgis_rest_layer", "arcgis_feature_service", "arcgis_mapserver"):
            candidates = [(conn.get("name") or "layer", endpoint)]
        elif ft == "arcgis_hub_discover":
            candidates = discover_hub_candidates(endpoint, kw)
            if not candidates:
                raise RuntimeError("layer_not_confirmed: no non-denylisted dev-app "
                                   "layer matched in hub DCAT feed")
        elif ft == "arcgis_map_discover":
            candidates = discover_map_candidates(endpoint, kw)
            if not candidates:
                raise RuntimeError("layer_not_confirmed: no non-denylisted dev-app "
                                   "sublayer matched in MapServer")
        elif ft == "ods_v21" or ft == "ods_api":
            base = "https://opendata.vancouver.ca"
            candidates = [("ODS:" + endpoint, (base, endpoint))]   # special form
        else:
            raise RuntimeError(f"unknown_fetch_type: '{ft}'")
    except Exception as e:
        res["status"] = "endpoint_not_found" if "not_found" not in str(e) else str(e).split(":")[0]
        res["error"] = str(e)
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = "verify endpoint manually"
        return res

    for name, url in candidates[:1]:   # take best-ranked candidate only
        if isinstance(url, tuple):
            resolved_text = f"{url[0]}/api/explore/v2.1/catalog/datasets/{url[1]}"
        else:
            resolved_text = str(url)
        res["resolved_endpoint"] = resolved_text
        res["resolved"].append(f"{name} -> {resolved_text}")

    if probe:
        res["status"] = "resolved_ok"
        res["next_action"] = "ready for --tier load"
        return res

    # ---- pull + richness gate + classify ----------------------------------
    name, url = candidates[0]
    try:
        if isinstance(url, tuple):     # ODS
            base, slug = url
            recs = query_ods(base, slug, max_records)
        else:
            recs = query_arcgis_layer(url, max_records)
    except Exception as e:
        if _should_use_fixture_fallback(e):
            fb = _load_fixture_leads_for_connector(conn, res, e)
            if fb:
                return fb
        res["error"] = str(e)
        res["status"] = str(e).split(":")[0] if ":" in str(e) else "fetch_failed"
        res["classification"] = G.CLASS_MANUAL
        res["output_route"] = "Run_Queue"
        res["next_action"] = "verify endpoint manually (no silent fallback)"
        return res

    res["records_pulled"] = len(recs)
    res["raw_records"] = recs

    # always save raw, even if rejected
    if raw_dir and recs:
        raw_paths = _save_raw(raw_dir, cid, recs)
        res["raw_json_path"] = raw_paths.get("json", "")

    richness = G.sample_richness(recs)
    res["richness"] = f"{richness['reason']} (avg {richness['avg_useful_fields']} fields)"
    classification = G.classify_layer_return(name, richness)
    res["classification"] = classification
    res["output_route"] = G.route_for(classification)

    if classification in (G.CLASS_DEV_LEAD, G.CLASS_CAPITAL):
        res["leads"] = [normalize_lead(conn, name, r) for r in recs]
        # Apply Vancouver building permit fit filter (Patch 5.0)
        apply_van_permit_filter(res)
        # In-run dedupe
        deduped, skipped = dedupe_leads(res["leads"])
        res["leads"] = deduped
        res["_dupes_skipped"] = skipped
        if skipped:
            print(f"    [dedupe] skipped {skipped} duplicate app_no/address within this run")
        res["status"] = "loaded"
        res["next_action"] = "review in Future_Projects; run PROMPT_DEV_APP_SCORING"
    else:
        # quarantine: build compact rejected rows, do NOT enter Future_Projects
        reason = {G.CLASS_WRONG: "wrong_layer", G.CLASS_THIN: "schema_too_thin",
                  G.CLASS_TRAILING: "trailing_permit", G.CLASS_CONTEXT: "context_only"}\
                 .get(classification, classification)
        res["rejected"] = [{
            "project_id": cid, "source": conn.get("name"),
            "source_url": url if isinstance(url, str) else url[0],
            "title": name, "municipality": conn.get("municipality"),
            "scope_summary": f"{len(recs)} records; {richness['reason']}",
            "reason": reason, "verification": "Rejected",
            "notes": f"Quarantined by guard: {classification}. Raw kept for QA.",
        }]
        res["status"] = reason
        res["next_action"] = ("re-pin to a real dev-app layer" if classification == G.CLASS_WRONG
                              else "keep as context/trailing only" if classification in (G.CLASS_TRAILING, G.CLASS_CONTEXT)
                              else "verify layer schema manually")
    return res


def _save_raw(raw_dir, cid, recs):
    for sub in ("json", "csv", "logs"):
        os.makedirs(os.path.join(raw_dir, sub), exist_ok=True)
    json_path = os.path.join(raw_dir, "json", f"{cid}.json")
    csv_path = os.path.join(raw_dir, "csv", f"{cid}.csv")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(recs, f, ensure_ascii=False, indent=2, default=str)
    # flat CSV of attribute keys for quick eyeballing
    keys = sorted({k for r in recs for k in r.keys()})
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        for r in recs:
            w.writerow({k: r.get(k, "") for k in keys})
    return {"json": json_path, "csv": csv_path}


# ============================================================================
# Vancouver Building Permit fit filter (Patch 5.0)
# ============================================================================
_VAN_PERMIT_HIGH_SIGNALS = [
    "subdivision", "site servicing", "servicing", "excavat", "earthwork",
    "grading", "roadwork", "road work", "civil", "watermain", "water main",
    "sanitary", "storm sewer", "drainage", "utility", "utilities",
    "rezoning", "development permit", "land development", "infrastructure",
    "multifamily", "multi-family", "industrial", "institutional",
    "municipal", "large parcel",
]
_VAN_PERMIT_LOW_SIGNALS = [
    "tenant improvement", "interior", "renovation", "reno",
    "small residential", "single family", "single-family",
    "sign permit", "plumbing only", "electrical only", "roofing",
    "small repair", "suite alteration", "kitchen", "bathroom",
]

def _van_permit_fit_tier(attrs):
    """Return 'strong', 'watchlist', 'bulk', or 'noisy' for a Van permit row."""
    blob = " ".join(str(v) for v in attrs.values() if v).lower()
    high = sum(1 for s in _VAN_PERMIT_HIGH_SIGNALS if s in blob)
    low  = sum(1 for s in _VAN_PERMIT_LOW_SIGNALS  if s in blob)
    if high >= 2 or (high >= 1 and low == 0):
        return "strong"
    if high >= 1:
        return "watchlist"
    if low >= 1:
        return "noisy"
    return "bulk"

def apply_van_permit_filter(result):
    """Re-route Vancouver building permit records: strong→Future_Projects,
    others→Bulk_Intake_Raw.  Mutates result leads/rejected in place."""
    if result.get("source_id") != "van_building_permits":
        return
    leads = result.get("leads") or []
    if not leads:
        return
    strong, watchlist, bulk, noisy = [], [], [], []
    for lead in leads:
        raw = lead.get("_raw") or {}
        tier = _van_permit_fit_tier(raw)
        lead["van_permit_fit_tier"] = tier
        lead["fit_reason"] = f"van_permit:{tier}"
        if tier == "strong":
            # Strong rows are the ONLY Vancouver permits that may enter
            # Future_Projects. Mark them explicitly so downstream write-gates
            # respect this per-connector routing decision (Patch 5.3 P0.2).
            lead["proposed_route"] = "Future_Projects"
            lead["hold_reason"] = ""
            lead["write_eligible"] = True
            lead["routing_decision"] = "Future_Projects"
            lead["routing_reason"] = "van_permit_strong_civil_signals"
            strong.append(lead)
        elif tier == "watchlist":
            lead["proposed_route"] = "Run_Queue"
            lead["hold_reason"] = "van_permit_watchlist"
            lead["write_eligible"] = False
            lead["routing_decision"] = "Run_Queue"
            lead["routing_reason"] = "van_permit_watchlist_single_civil_signal"
            watchlist.append(lead)
        elif tier == "bulk":
            lead["proposed_route"] = "Bulk_Intake_Raw"
            lead["hold_reason"] = "van_permit_bulk_intake"
            lead["write_eligible"] = False
            lead["routing_decision"] = "Bulk_Intake_Raw"
            lead["routing_reason"] = "van_permit_bulk_no_civil_signal"
            bulk.append(lead)
        else:  # noisy
            lead["proposed_route"] = "Rejected_Archive"
            lead["hold_reason"] = "van_permit_noisy_interior_ti"
            lead["write_eligible"] = False
            lead["routing_decision"] = "Rejected_Archive"
            lead["routing_reason"] = "van_permit_noisy_interior_ti_or_minor"
            noisy.append(lead)

    result["leads"] = strong + watchlist + bulk + noisy
    result["van_permit_strong"] = len(strong)
    result["van_permit_watchlist"] = len(watchlist)
    result["van_permit_bulk"] = len(bulk)
    result["van_permit_noisy"] = len(noisy)
    print(f"    [van_permit_filter] strong={len(strong)} watchlist={len(watchlist)} "
          f"bulk={len(bulk)} noisy={len(noisy)}")


# ============================================================================
# Dedupe (Patch 5.0): prevent repeated project_id from being appended twice
# ============================================================================
def dedupe_leads(leads):
    """Remove duplicate leads within a single run by (source_id + norm_app_no)
    then by (municipality + norm_address + app_type_stage) as fallback.
    Returns (deduped_list, duplicates_skipped_count)."""
    seen = {}
    out = []
    skipped = 0
    for lead in leads:
        # Primary key
        sid = lead.get("source") or lead.get("source_id") or ""
        app = G.norm_appno(lead.get("app_no") or "")
        addr = G.norm_address(lead.get("address") or "")
        muni = G.norm_municipality(lead.get("municipality") or "")
        key1 = (sid, app) if app and app != "?" else None
        key2 = (muni, addr, (lead.get("app_type_stage") or "")[:30]) if addr else None

        matched = False
        for key in filter(None, [key1, key2]):
            if key in seen:
                seen[key]["last_seen_date"] = dt.date.today().isoformat()
                skipped += 1
                matched = True
                break
        if not matched:
            out.append(lead)
            for key in filter(None, [key1, key2]):
                seen[key] = lead
    return out, skipped


# ============================================================================
# Source-level run summary (Patch 5.0)
# ============================================================================
def _categorize_result(r):
    """Patch 5.3 (P0.4): produce an honest per-connector breakdown that never
    collapses bulk/noisy/manual/failed records into one big eligible count.

    Returns a dict of integer counts keyed by routing category. Counts are
    derived from the FINAL (post-gate, post-dedupe) lead routing decisions so
    that the summary matches what would actually be written to the master.
    """
    leads = r.get("leads") or []
    status = (r.get("status") or "").lower()
    route_of = lambda l: (l.get("routing_decision") or l.get("proposed_route") or "Future_Projects")

    clean_fp = sum(1 for l in leads
                   if l.get("write_eligible") and route_of(l) in _FUTURE_PROJECT_ROUTES)
    watchlist = sum(1 for l in leads
                    if not l.get("write_eligible") and route_of(l) in ("Run_Queue", "Watchlist"))
    bulk = sum(1 for l in leads if route_of(l) == "Bulk_Intake_Raw")
    rejected_routed = sum(1 for l in leads if route_of(l) == "Rejected_Archive")
    rejected = rejected_routed + len(r.get("rejected") or [])

    # Connectors that pulled nothing because they need an exact URL / manual
    # extraction / access — these are P3/manual, not silent zeros.
    manual_states = ("needs_exact_url", "access_test_required", "p3_extract_required",
                     "manual_review_needed", "needs_manual", "manual_p3_only")
    failed_states = ("no_records_extracted", "fetch_failed", "parse_failed",
                     "extraction_failed")
    is_manual = (not leads) and any(s in status for s in manual_states)
    is_failed = (not leads) and any(s in status for s in failed_states)

    return {
        "records_clean_future_projects": clean_fp,
        "records_watchlist": watchlist,
        "records_bulk_intake": bulk,
        "records_rejected": rejected,
        "records_manual_or_p3": 1 if is_manual else 0,
        "records_failed_extraction": 1 if is_failed else 0,
        "duplicates_skipped": int(r.get("_dupes_skipped") or 0),
    }


def write_source_summary(results, out_dir, run_id):
    """Write one row per selected connector/source with acceptance columns."""
    rows = []
    for r in results:
        leads = r.get("leads") or []
        cat = _categorize_result(r)
        # Eligible == clean Future_Projects only. Bulk/noisy/watchlist excluded.
        fp_rows = cat["records_clean_future_projects"]
        bulk_rows = cat["records_bulk_intake"]
        rq_rows = sum(1 for l in leads
                      if (l.get("routing_decision") or l.get("proposed_route")) in ("Run_Queue", "Watchlist"))
        rej_rows = cat["records_rejected"]
        if r.get("output_route") == "Run_Queue" and not leads:
            rq_rows = 1
        warnings = []
        if r.get("_fixture_fallback"):
            warnings.append("fixture fallback used; rerun live on a networked machine for live-source proof")
        if r.get("_fixture_error"):
            warnings.append("live error: " + str(r.get("_fixture_error"))[:300])
        if r.get("error"):
            warnings.append(str(r.get("error"))[:300])
        next_action = r.get("next_action") or ""
        rows.append({
            "run_id": run_id,
            "connector_name": r.get("source_id") or "",
            "source_id": r.get("source_id") or "",
            "source_name": r.get("source_name") or "",
            "connector_status": r.get("status") or "",
            "resolved_status": r.get("status") or "",
            "probe_status": r.get("status") or "",
            "pull_status": r.get("status") or "",
            "records_pulled": r.get("records_pulled") or 0,
            "records_normalized": r.get("records_normalized") or len(leads),
            # ---- honest separated counts (Patch 5.3 P0.4) ----
            "records_clean_future_projects": cat["records_clean_future_projects"],
            "records_watchlist": cat["records_watchlist"],
            "records_bulk_intake": cat["records_bulk_intake"],
            "records_manual_or_p3": cat["records_manual_or_p3"],
            "records_failed_extraction": cat["records_failed_extraction"],
            # ---- backward-compatible counts ----
            "records_eligible": fp_rows,
            "records_written": 0,
            "records_rejected": rej_rows,
            "records_held_for_review": (len(leads) - fp_rows) + len(r.get("rejected") or []),
            "future_projects_rows": fp_rows,
            "bulk_intake_rows": bulk_rows,
            "run_queue_rows": rq_rows,
            "rejected_rows": rej_rows,
            "duplicates_skipped": cat["duplicates_skipped"],
            "route": r.get("output_route") or "",
            "proposed_route": r.get("output_route") or "",
            "status": r.get("status") or "",
            "warnings": "; ".join(warnings),
            "errors": r.get("error") or "",
            "warning": "; ".join(warnings),
            "next_action": next_action,
            "output_file": "",
        })
    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    csv_path = os.path.join(out_dir, "TENDER_FINDER_Run_Source_Summary.csv")
    fieldnames = [
        "run_id", "connector_name", "source_id", "source_name",
        "connector_status", "resolved_status", "probe_status", "pull_status",
        "records_pulled", "records_normalized",
        "records_clean_future_projects", "records_watchlist", "records_bulk_intake",
        "records_manual_or_p3", "records_failed_extraction",
        "records_eligible", "records_written",
        "records_rejected", "records_held_for_review", "future_projects_rows",
        "bulk_intake_rows", "run_queue_rows", "rejected_rows", "duplicates_skipped",
        "route", "proposed_route", "status",
        "warnings", "errors", "warning", "next_action", "output_file",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"\n  Source summary  -> {csv_path}  ({len(rows)} connector rows)")
    return csv_path


# ============================================================================
# Connector selection (from master or CSV) + filters
# ============================================================================
def select_connectors(args, connectors):
    rows = list(connectors.values())
    if args.only:
        wanted = {x.strip() for x in args.only.split(",")}
        # accept either source_id or a short alias contained in source_id
        rows = [c for c in rows if (c.get("source_id") in wanted
                or any(w in (c.get("source_id") or "") for w in wanted))]
    if getattr(args, "bulk_intake", False) and not args.only and not args.tier and not args.category:
        allowed_tiers = {"1"}
        if args.include_tier2:
            allowed_tiers.add("2")
        if args.include_tier3:
            allowed_tiers.add("3")
        rows = [c for c in rows if str(c.get("priority_tier") or (c.get("tier") or "").split()[0]) in allowed_tiers]
    if args.tier:
        rows = [c for c in rows if str(c.get("priority_tier") or
                                       (c.get("tier") or "").split()[0]) == str(args.tier)]
    if args.category:
        rows = [c for c in rows if (c.get("category") or "").lower().startswith(args.category.lower())]
    if args.skip_paid:
        rows = [c for c in rows if "paid" not in (c.get("access_status") or "").lower()
                and not (c.get("category") or "").lower().startswith("e —")]
    if args.skip_login:
        rows = [c for c in rows if "login" not in (c.get("access_status") or "").lower()]
    # when loading (not probing), control broad vs normal loading separately
    if getattr(args, "bulk_intake", False) and not args.probe:
        rows = _bulk_access_filter(rows, args)
    elif (args.tier or args.category) and not args.probe:
        loadable = {"ready_for_load", "ready_for_probe", "trailing_context"}
        rows = [c for c in rows if (c.get("access_status") or "") in loadable
                or (c.get("fetch_type") or "") in QUEUE_FETCH_TYPES]
    return rows


def _run_promote_reviewed(args):
    """Promote ACCEPT rows from a human-reviewed workbook into the master.

    Patch 5.0: Safer middle workflow between --review-only and --write-master.

    Workflow:
      1. Generate review workbook:  --review-only --out review.xlsx
      2. Human opens review.xlsx, fills review_decision column:
             ACCEPT | REJECT | HOLD | NEEDS_MORE_INFO
      3. Run promote:
             python tenderfinder_raw_sweep.py --promote-reviewed review_marked.xlsx --write-master TENDER_FINDER_Master_WRITE_TEST.xlsx

    Rules:
      - Only ACCEPT rows are written
      - Backup is created before any write
      - Audit log is written
      - Dedupe before append
      - Never overwrites human-reviewed rows silently
    """
    review_path = _resolve_path(args.promote_reviewed)
    if not os.path.exists(review_path):
        _fail(f"--promote-reviewed: file not found: {review_path}")
    if not args.write_master:
        _fail("--promote-reviewed requires --write-master <target workbook>")

    resolved_write_target = _validate_write_master_target(args.write_master)
    print(f"\nPROMOTE-REVIEWED: reading {review_path}")

    wb = openpyxl.load_workbook(review_path, data_only=True)
    ws = wb.active
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    if "review_decision" not in hdr:
        _fail("review_decision column not found in review workbook. "
              "Fill in ACCEPT/REJECT/HOLD/NEEDS_MORE_INFO in the review_decision column first.")

    rd_col = hdr.index("review_decision") + 1

    # Collect ACCEPT rows
    accept_leads = []
    counts = {"ACCEPT": 0, "REJECT": 0, "HOLD": 0, "NEEDS_MORE_INFO": 0, "blank": 0}

    for r in range(2, ws.max_row + 1):
        decision = str(ws.cell(r, rd_col).value or "").strip().upper()
        if not decision:
            counts["blank"] += 1
            continue
        counts[decision] = counts.get(decision, 0) + 1
        if decision != "ACCEPT":
            continue
        # Build a lead dict from this row
        lead = {hdr[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        # Map review workbook columns to Future_Projects schema
        accept_leads.append({
            "project_id": lead.get("project_id") or "",
            "date_found": lead.get("date_found") or dt.date.today().isoformat(),
            "source": lead.get("source_name") or lead.get("source_id") or "",
            "source_url": lead.get("source_url") or "",
            "title": lead.get("app_type_stage") or lead.get("scope_summary") or lead.get("project_id") or "",
            "owner": lead.get("owner/applicant") or "Not available",
            "municipality": lead.get("municipality") or "",
            "app_no": lead.get("app_no") or "",
            "app_type_stage": lead.get("app_type_stage") or "",
            "scope_summary": lead.get("scope_summary") or "",
            "fit_score": lead.get("fit_score") or "",
            "verification": "Human-Accepted",
            "horizon": "",
            "est_value": "Not available",
            "next_milestone": "",
            "assigned_to": "",
            "notes": f"Promoted from review workbook via --promote-reviewed. Original decision: ACCEPT.",
            "address": lead.get("address") or "",
        })

    print(f"  Review decisions: ACCEPT={counts['ACCEPT']} REJECT={counts['REJECT']} "
          f"HOLD={counts['HOLD']} NEEDS_MORE_INFO={counts.get('NEEDS_MORE_INFO',0)} blank={counts['blank']}")
    print(f"  Rows to promote: {len(accept_leads)}")

    if not accept_leads:
        print("  Nothing to promote. Run with --review-only first, fill in review_decision, then re-run.")
        return

    # Dedupe accepted rows before writing
    deduped, skipped = dedupe_leads(accept_leads)
    if skipped:
        print(f"  Dedupe: {skipped} duplicate rows suppressed before write")

    print(f"  Resolved write target: {resolved_write_target}")

    import tenderfinder_master_io as MIO
    v7 = MIO.ensure_v7(args.v6, resolved_write_target)
    v7 = _resolve_path(v7)
    backup_path = MIO.backup(v7)
    print(f"  Backup: {backup_path}")

    m = MIO.MasterIO(v7)
    fp_stats = m.write_future(deduped)
    m.save()

    run_id = f"PROMOTE-{dt.date.today().isoformat()}"
    # Write audit log
    audit_path = os.path.join(os.path.dirname(v7), f"promote_audit_{dt.date.today().isoformat()}.json")
    with open(audit_path, "w", encoding="utf-8") as f:
        json.dump({
            "run_id": run_id,
            "review_source": review_path,
            "write_target": resolved_write_target,
            "backup": backup_path,
            "decisions": counts,
            "rows_accepted": len(accept_leads),
            "rows_deduped_out": skipped,
            "rows_written": fp_stats.get("appended", 0),
            "rows_updated": fp_stats.get("updated_stage", 0),
            "duplicates_skipped_on_write": max(0, len(deduped) - (fp_stats.get("appended", 0) or 0) - (fp_stats.get("updated_stage", 0) or 0) - (fp_stats.get("enriched", 0) or 0)),
            "timestamp": dt.datetime.now().isoformat(),
        }, f, indent=2, default=str)

    print(f"\n  PROMOTE RESULT:")
    appended = fp_stats.get('appended', 0) or 0
    updated = fp_stats.get('updated_stage', 0) or 0
    enriched = fp_stats.get('enriched', 0) or 0
    duplicates_skipped_on_write = max(0, len(deduped) - appended - updated - enriched)
    print(f"    Future_Projects: appended={appended} "
          f"updated={updated} "
          f"collapsed_dupes={fp_stats.get('collapsed_dupes',0)}")
    print(f"    duplicates_skipped_on_write={duplicates_skipped_on_write}")
    print(f"    Audit log: {audit_path}")
    print("\nDone (promote-reviewed).\n")


def _write_demo_output(results, path, run_id):
    """Generate a clean demo workbook (Patch 5.0).

    Sheets:
      Top_Leads         — high-quality Future_Projects leads (not dominated by Van permits)
      Strong_Fit        — fit_score >= 60
      Watchlist         — fit_score 40-59 or watchlist-flagged
      Surrey_Leads      — surrey_planning_reports rows only
      Langley_Leads     — twp_langley_devactivity rows only
      Maple_Ridge_Leads — maple_ridge_devapps rows only
      Source_Status     — per-connector pull stats
      Run_Summary       — total counts
    """
    out_path = _resolve_path(path)
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    # Collect all leads
    all_leads = []
    for r in results:
        for lead in (r.get("leads") or []):
            lead["_source_id"] = r.get("source_id") or ""
            all_leads.append(lead)

    def _fit(lead):
        try:
            return int(lead.get("fit_score") or 0)
        except Exception:
            return 0

    # Filter sets
    top    = [l for l in all_leads if _fit(l) >= 50 and l.get("_source_id") != "van_building_permits"
              and l.get("write_eligible", True)][:200]
    strong = [l for l in all_leads if _fit(l) >= 60 and l.get("write_eligible", True)][:300]
    watch  = [l for l in all_leads if 40 <= _fit(l) < 60 or l.get("van_permit_fit_tier") == "watchlist"][:300]
    surrey = [l for l in all_leads if l.get("_source_id") == "surrey_planning_reports"]
    langley= [l for l in all_leads if l.get("_source_id") == "twp_langley_devactivity"]
    maple  = [l for l in all_leads if l.get("_source_id") == "maple_ridge_devapps"]

    # Lead columns for display
    lead_cols = ["project_id", "municipality", "app_no", "address",
                 "owner/applicant", "app_type_stage", "scope_summary",
                 "fit_score", "source_url"]

    def _write_lead_sheet(ws, leads, label):
        cols = ["_source_id"] + lead_cols + ["hold_reason"]
        ws.append(["source"] + lead_cols + ["hold_reason"])
        for c in range(1, len(cols) + 1):
            ws.cell(1, c).font = bold
        for lead in leads:
            row = [lead.get("_source_id") or lead.get("source") or ""]
            for col in lead_cols:
                row.append(_excel_safe_text(lead.get(col) or lead.get(col.replace("/", "_")) or ""))
            row.append(lead.get("hold_reason") or "")
            ws.append(row)
        ws.freeze_panes = "A2"
        ws.title = label
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

    # Build sheets
    ws_top = wb.active
    _write_lead_sheet(ws_top, top, "Top_Leads")
    for leads, label in [(strong, "Strong_Fit"), (watch, "Watchlist"),
                         (surrey, "Surrey_Leads"), (langley, "Langley_Leads"),
                         (maple, "Maple_Ridge_Leads")]:
        ws = wb.create_sheet(label)
        _write_lead_sheet(ws, leads, label)

    # Source_Status sheet
    ws_ss = wb.create_sheet("Source_Status")
    ss_cols = ["source_id", "source_name", "pull_status", "records_pulled",
               "future_projects", "rejected", "errors", "next_action"]
    ws_ss.append(ss_cols)
    for c in range(1, len(ss_cols) + 1):
        ws_ss.cell(1, c).font = bold
    for r in results:
        leads = r.get("leads") or []
        fp = sum(1 for l in leads if l.get("write_eligible", True) and
                 l.get("proposed_route", "Future_Projects") == "Future_Projects")
        ws_ss.append([
            r.get("source_id"), r.get("source_name"), r.get("status"),
            r.get("records_pulled") or 0, fp,
            len(r.get("rejected") or []), 1 if r.get("error") else 0,
            r.get("next_action") or "",
        ])
    ws_ss.freeze_panes = "A2"

    # Run_Summary sheet
    ws_rs = wb.create_sheet("Run_Summary")
    ws_rs.append(["Metric", "Value"])
    ws_rs.cell(1, 1).font = bold
    ws_rs.cell(1, 2).font = bold
    summary = [
        ("Run ID", run_id),
        ("Run Date", dt.date.today().isoformat()),
        ("Connectors Selected", len(results)),
        ("Records Pulled Total", sum(r.get("records_pulled") or 0 for r in results)),
        ("Top Leads (demo)", len(top)),
        ("Strong Fit Leads", len(strong)),
        ("Surrey Leads", len(surrey)),
        ("Langley Leads", len(langley)),
        ("Maple Ridge Leads", len(maple)),
        ("Sources with Errors", sum(1 for r in results if r.get("error"))),
    ]
    for row in summary:
        ws_rs.append(list(row))

    wb.save(out_path)
    print(f"  Demo workbook  -> {out_path}  "
          f"(top={len(top)}, surrey={len(surrey)}, langley={len(langley)}, maple={len(maple)})")
    return out_path


# ============================================================================
# Main
# ============================================================================
def load_config(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    out = {}
    for r in rows:
        sid = r.get("source_id") or r.get("id")
        if sid:
            out[sid] = r
    return out


def print_registry(connectors):
    print(f"\n{'SOURCE_ID':<26}{'TIER':<6}{'FETCH TYPE':<22}{'ACCESS STATUS':<22}ROUTE")
    print("-" * 100)
    for c in connectors.values():
        print(f"{(c.get('source_id') or ''):<26}{str(c.get('priority_tier') or ''):<6}"
              f"{(c.get('fetch_type') or ''):<22}{(c.get('access_status') or ''):<22}"
              f"{c.get('output_route') or ''}")
    print(f"\n{len(connectors)} connectors.\n")


def main():
    ap = argparse.ArgumentParser(description="TENDER_FINDER controlled all-source probe/load runner")
    ap.add_argument("--config", default="tenderfinder_dev_app_endpoints.csv")
    ap.add_argument("--from-master", default="", help="master v7 workbook to read Source_Register/Run_Queue")
    ap.add_argument("--write-master", default="", help="master v7 workbook to write leads into (created from v6 if needed)")
    ap.add_argument("--v6", default="../../00 Master/TENDER_FINDER_Tender_Intelligence_Working_Master_v6.xlsx",
                    help="source v6 used to create v7 when --write-master target is missing")
    ap.add_argument("--tier", type=int, default=0, help="load only this priority tier")
    ap.add_argument("--category", default="", help="filter by category prefix, e.g. 'B'")
    ap.add_argument("--only", default="", help="comma-separated source ids/aliases")
    ap.add_argument("--out", default="./tenderfinder_raw_out", help="output dir or probe/review xlsx path")
    ap.add_argument("--max-records", type=int, default=20000)
    ap.add_argument("--probe", action="store_true", help="resolve endpoints only, no pull")
    ap.add_argument("--dry-run", action="store_true", help="pull + classify but do not write master")
    ap.add_argument("--min-fit-score", type=float, default=None,
                    help="minimum fit_score required before a lead is eligible for master write")
    ap.add_argument("--max-write-per-source", type=int, default=0,
                    help="maximum eligible Future_Projects records per source; 0 means no cap")
    ap.add_argument("--review-only", action="store_true",
                    help="write a review output file and do not open, back up, or save the master workbook")
    ap.add_argument("--bulk-intake", action="store_true",
                    help="write broad raw/normalized harvest rows to Bulk_Intake_Raw instead of Future_Projects by default")
    ap.add_argument("--promote-to-future-projects", action="store_true",
                    help="with --bulk-intake, also promote eligible dev/capital leads to Future_Projects")
    ap.add_argument("--also-write-rejected", action="store_true",
                    help="with --bulk-intake, also append rejected/wrong-layer/thin summaries to Rejected_Archive")
    ap.add_argument("--include-tier2", action="store_true",
                    help="with --bulk-intake and no explicit filter, include Tier 2 sources")
    ap.add_argument("--include-tier3", action="store_true",
                    help="with --bulk-intake and no explicit filter, include Tier 3 sources")
    ap.add_argument("--include-trailing-context", action="store_true",
                    help="with --bulk-intake, include trailing/context permit sources")
    ap.add_argument("--include-wrong-layer", action="store_true",
                    help="with --bulk-intake, include disabled/wrong-layer sources as stubs or rejected candidates")
    ap.add_argument("--max-records-per-source", type=int, default=0,
                    help="with --bulk-intake, maximum pulled records per source; overrides --max-records when set")
    ap.add_argument("--skip-paid", action="store_true")
    ap.add_argument("--skip-login", action="store_true")
    ap.add_argument("--source-summary", action="store_true",
                    help="write TENDER_FINDER_Run_Source_Summary.csv to --out directory after run")
    ap.add_argument("--demo-output", action="store_true",
                    help="generate a filtered demo workbook (Top_Leads, Surrey, Langley, Maple_Ridge, "
                         "Source_Status sheets) into --out or --out directory")
    ap.add_argument("--promote-reviewed", default="",
                    help="path to a review workbook that has review_decision column filled in; "
                         "promotes ACCEPT rows into the --write-master target")
    ap.add_argument("--sync-registry", action="store_true", help="emit Source_Register<->connector sync report and exit")
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--preflight-links", action="store_true",
                    help="run Source_Register live-link preflight through tenderfinder_raw_sweep.py before scraping")
    ap.add_argument("--preflight-output-dir", default="./link_audit_out",
                    help="directory for integrated live-link audit outputs")
    ap.add_argument("--preflight-no-search", action="store_true",
                    help="disable replacement URL search during link preflight")
    ap.add_argument("--preflight-search-provider", default="tavily",
                    choices=["bing", "serpapi", "tavily", "google"],
                    help="search provider for replacement candidates during link preflight")
    ap.add_argument("--preflight-timeout", type=int, default=20,
                    help="HTTP timeout in seconds for link preflight")
    ap.add_argument("--preflight-retries", type=int, default=2,
                    help="retry count for link preflight")
    ap.add_argument("--preflight-workers", type=int, default=4,
                    help="worker threads for link preflight")
    ap.add_argument("--preflight-rate-limit", type=float, default=1.0,
                    help="minimum delay between link preflight requests")
    ap.add_argument("--preflight-fail-on-broken", action="store_true",
                    help="exit non-zero if critical broken links are found during preflight")
    ap.add_argument("--include-broken-sources", action="store_true",
                    help="after link preflight, allow BROKEN/FIX_URL_FIRST sources into selected connector run")
    args = ap.parse_args()

    preflight_result = None
    if args.promote_reviewed:
        _run_promote_reviewed(args)
        return
    if args.preflight_links:
        if not args.from_master:
            _fail("--preflight-links needs --from-master <Source_Register workbook>")
        import tenderfinder_link_preflight as LP
        preflight_result = LP.run_preflight_from_runner_args(args)
        if preflight_result.exit_code != 0:
            sys.exit(preflight_result.exit_code)
        preflight_only = not any([
            args.list, args.sync_registry, args.probe, args.bulk_intake,
            args.write_master, args.review_only, args.only, args.tier, args.category
        ])
        if preflight_only:
            print("\nPreflight-only run complete. No connector sweep or master write was run.\n")
            return

    connectors = load_config(args.config)

    if args.bulk_intake and args.review_only:
        _fail("Use either --bulk-intake or --review-only, not both.")
    if args.bulk_intake and not args.write_master and not args.dry_run:
        _fail("--bulk-intake real writes require --write-master <copied v7/test/bulk workbook>.")

    if args.list:
        print_registry(connectors)
        return

    if args.sync_registry:
        if not args.from_master:
            print("--sync-registry needs --from-master <workbook>", file=sys.stderr)
            sys.exit(2)
        rows, summary = REG.build_sync_report(args.from_master, args.config)
        out = args.out if args.out.endswith(".csv") else "sync_report.csv"
        REG.write_sync_report_csv(rows, out)
        print(f"\nSync report -> {out}  ({len(rows)} sources)")
        for k in sorted(summary):
            print(f"  {k:<24} {summary[k]}")
        return

    if args.write_master:
        resolved_write_target = _validate_write_master_target(args.write_master)
    else:
        resolved_write_target = ""

    selected = select_connectors(args, connectors)
    if preflight_result is not None:
        import tenderfinder_link_preflight as LP
        selected, skipped_by_preflight = LP.filter_connectors_by_preflight(
            selected, preflight_result, include_broken_sources=args.include_broken_sources
        )
        if skipped_by_preflight:
            print("\nLINK PREFLIGHT GATE: skipped broken/FIX_URL_FIRST sources before connector run:")
            for c in skipped_by_preflight:
                print(f"  - {c.get('source_id') or c.get('id')}: {c.get('name') or c.get('source_name')}")
            print("  Use --include-broken-sources to override this gate.\n")
    if not selected:
        print("No connectors match the given filters.", file=sys.stderr)
        return

    stamp = dt.date.today().isoformat()
    raw_dir = os.path.join("raw_runs", stamp)
    os.makedirs(raw_dir, exist_ok=True)

    if args.probe:
        mode = "PROBE"
    elif args.bulk_intake:
        mode = "BULK-INTAKE DRY-RUN" if args.dry_run else "BULK-INTAKE"
    elif args.review_only:
        mode = "REVIEW-ONLY"
    else:
        mode = "DRY-RUN" if args.dry_run else "LOAD"
    print(f"\nTENDER_FINDER controlled runner - {mode} - {dt.datetime.now().isoformat(timespec='seconds')}")
    print(f"Connectors selected: {len(selected)} | raw_dir: {raw_dir}\n" + "=" * 78)

    results = []
    for conn in selected:
        cid = conn.get("source_id")
        print(f"\n[{cid}]  {conn.get('name')}  (ft={conn.get('fetch_type')})")
        source_max_records = args.max_records_per_source if (args.bulk_intake and args.max_records_per_source) else args.max_records
        r = run_connector(conn, source_max_records, probe=args.probe, raw_dir=raw_dir)
        for line in r["resolved"]:
            print(f"    resolved: {line}")
        if r["error"]:
            print(f"    !! {r['error']}")
        if not args.probe:
            print(f"    pulled={r['records_pulled']}  class={r['classification']}  "
                  f"route={r['output_route']}  status={r['status']}")
        results.append(r)

    # ---- probe summary -----------------------------------------------------
    if args.probe:
        if args.out.endswith(".xlsx"):
            _write_probe_xlsx(results, args.out)
            print(f"\nProbe workbook -> {args.out}")
        ok = sum(1 for r in results if r["status"] in ("resolved_ok",))
        print("\n" + "=" * 78 + f"\nPROBE SUMMARY: resolved OK {ok}/{len(results)}")
        for r in results:
            tag = "OK " if r["status"] == "resolved_ok" else "-- "
            print(f"  {tag}{r['source_id']:<26} {r['status']}")
        _save_run_log(raw_dir, results)
        # Source summary even for probe runs
        run_id = f"PROBE-{dt.date.today().isoformat()}"
        out_dir_probe = os.path.dirname(os.path.abspath(args.out)) if args.out.endswith(".xlsx") else raw_dir
        write_source_summary(results, out_dir_probe, run_id)
        print("\nDone (probe).\n")
        return

    # ---- bulk intake: raw acquisition dump, separate from Future_Projects ---
    if args.bulk_intake:
        _apply_write_gates(results, args.min_fit_score, args.max_write_per_source)
        run_timestamp = dt.datetime.now().isoformat(timespec="seconds")
        run_id = "bulk_" + dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        bulk_rows, bulk_per_source = build_bulk_rows(results, run_id, run_timestamp)
        all_leads = _eligible_leads(results) if args.promote_to_future_projects else []
        all_reject = [x for r in results for x in r["rejected"]] if args.also_write_rejected else []

        selected_count = len(results)
        pulled_count = sum(int(r.get("records_pulled") or 0) for r in results)
        skipped_count = sum(1 for r in results if not (r.get("raw_records") or []))
        error_count = sum(1 for r in results if r.get("error"))

        print("\n" + "=" * 78 + "\nBULK INTAKE SUMMARY")
        print(f"  selected                    : {selected_count}")
        print(f"  pulled                      : {pulled_count}")
        print(f"  written_to_bulk_intake      : {0 if args.dry_run else len(bulk_rows)}")
        if args.dry_run:
            print(f"  would_write_to_bulk_intake  : {len(bulk_rows)}")
        print(f"  promoted_to_future_projects : {0 if not args.promote_to_future_projects else len(all_leads)}")
        print(f"  rejected_written            : {0 if not args.also_write_rejected else len(all_reject)}")
        print(f"  skipped                     : {skipped_count}")
        print(f"  errors                      : {error_count}")
        print("\nSOURCE-LEVEL BULK SUMMARY")
        for r in results:
            sid = r.get("source_id")
            print(f"  {sid:<28} selected=1 pulled={int(r.get('records_pulled') or 0)} "
                  f"written_to_bulk_intake={0 if args.dry_run else bulk_per_source.get(sid, 0)} "
                  f"promoted_to_future_projects={len([x for x in (r.get('leads') or []) if x.get('write_eligible')]) if args.promote_to_future_projects else 0} "
                  f"rejected_written={len(r.get('rejected') or []) if args.also_write_rejected else 0} "
                  f"skipped={0 if (r.get('raw_records') or []) else 1} "
                  f"errors={1 if r.get('error') else 0}")
        _save_run_log(raw_dir, results)

        if args.dry_run:
            print("\n  BULK DRY-RUN: no workbook opened, backed up, or saved.")
            print("\nDone.\n")
            return

        if not args.write_master:
            _fail("--bulk-intake real writes require --write-master <copied v7/test/bulk workbook>.")

        print(f"\n  [master] resolved write target -> {resolved_write_target}")
        import tenderfinder_master_io as MIO
        import tenderfinder_bulk_io as BIO
        v7 = MIO.ensure_v7(args.v6, resolved_write_target)
        v7 = _resolve_path(v7)
        backup_path = MIO.backup(v7)
        print(f"  [master] resolved backup path -> {backup_path}")
        m = MIO.MasterIO(v7)
        bulk_stats = BIO.append_bulk_rows(m.wb, bulk_rows)
        fp_stats = {"appended": 0, "updated_stage": 0, "enriched": 0, "collapsed_dupes": 0, "final_rows": 0}
        rj_stats = {"appended": 0}
        if args.promote_to_future_projects:
            fp_stats = m.write_future(all_leads)
        if args.also_write_rejected:
            rj_stats = m.append_rejected(all_reject)
        m.save()
        print("\n  BULK MASTER WRITE:")
        print(f"    Bulk_Intake_Raw: appended={bulk_stats['appended']} -> {bulk_stats['final_rows']} rows")
        print(f"    Future_Projects promoted: appended={fp_stats.get('appended', 0)} "
              f"updated={fp_stats.get('updated_stage', 0)} enriched={fp_stats.get('enriched', 0)} "
              f"collapsed_dupes={fp_stats.get('collapsed_dupes', 0)} -> {fp_stats.get('final_rows', 0)} rows")
        print(f"    Rejected_Archive: appended={rj_stats.get('appended', 0)}")
        print("\nDone.\n")
        return

    # ---- aggregate leads / rejects + record-level write gates --------------
    _apply_write_gates(results, args.min_fit_score, args.max_write_per_source)
    all_leads = _eligible_leads(results)
    held_leads = _held_leads(results)
    all_reject = [x for r in results for x in r["rejected"]]
    counts = _summary_counts(results, all_leads, all_reject)
    queued = [r for r in results if r["output_route"] == "Run_Queue"]

    # Patch 5.3 (P0.2/P0.4): aggregate the honest per-connector breakdown so the
    # console never presents bulk/noisy/watchlist permits as clean eligible leads.
    agg = {"records_clean_future_projects": 0, "records_watchlist": 0,
           "records_bulk_intake": 0, "records_rejected": 0,
           "records_manual_or_p3": 0, "records_failed_extraction": 0,
           "duplicates_skipped": 0}
    van_tiers = None
    for r in results:
        c = _categorize_result(r)
        for k in agg:
            agg[k] += c[k]
        if r.get("source_id") == "van_building_permits":
            van_tiers = (r.get("van_permit_strong", 0), r.get("van_permit_watchlist", 0),
                         r.get("van_permit_bulk", 0), r.get("van_permit_noisy", 0))

    print("\n" + "=" * 78 + "\nRUN SUMMARY")
    print(f"  records_pulled               : {counts['records_pulled']}")
    print(f"  records_normalized           : {counts['records_normalized']}")
    print(f"  records_clean_future_projects: {agg['records_clean_future_projects']}")
    print(f"  records_watchlist            : {agg['records_watchlist']}")
    print(f"  records_bulk_intake          : {agg['records_bulk_intake']}")
    print(f"  records_rejected             : {agg['records_rejected']}")
    print(f"  records_manual_or_p3         : {agg['records_manual_or_p3']}")
    print(f"  records_failed_extraction    : {agg['records_failed_extraction']}")
    print(f"  duplicates_skipped           : {agg['duplicates_skipped']}")
    print(f"  records_written              : 0")
    print(f"  records_held_for_review      : {counts['records_held_for_review']}")
    print(f"  P3/manual/queue sources      : {len(queued)}")
    if van_tiers is not None:
        print(f"  [van_building_permits tiers] strong={van_tiers[0]} watchlist={van_tiers[1]} "
              f"bulk={van_tiers[2]} noisy={van_tiers[3]} (only strong -> Future_Projects)")
    print(f"  Future_Projects eligible leads (clean only): {len(all_leads)}")
    if held_leads:
        print(f"  Held for review before master write: {len(held_leads)}")
    _save_run_log(raw_dir, results)

    # ---- review-only output; do not open/back up/save master ----------------
    if args.review_only:
        review_path = args.out if args.out else os.path.join(raw_dir, "review.xlsx")
        _write_review_output(results, review_path)
        # Source-level summary (always written with --source-summary, or when --review-only)
        out_dir_for_summary = os.path.dirname(os.path.abspath(review_path))
        run_id = f"RUN-{dt.date.today().isoformat()}"
        write_source_summary(results, out_dir_for_summary, run_id)
        if getattr(args, "demo_output", False):
            demo_path = os.path.join(out_dir_for_summary, "TENDER_FINDER_Demo_Top_Leads.xlsx")
            _write_demo_output(results, demo_path, run_id)
        print("\n  REVIEW-ONLY: no master workbook opened, backed up, or saved.")
        print(f"  Review output -> {review_path}")
        print("\nDone.\n")
        return

    if args.write_master:
        print(f"\n  [master] resolved write target -> {resolved_write_target}")

    # ---- write master ------------------------------------------------------
    if args.write_master and not args.dry_run:
        import tenderfinder_master_io as MIO
        v7 = MIO.ensure_v7(args.v6, resolved_write_target)
        v7 = _resolve_path(v7)
        backup_path = MIO.backup(v7)
        print(f"  [master] resolved backup path -> {backup_path}")
        m = MIO.MasterIO(v7)
        fp_stats = m.write_future(all_leads)
        rj_stats = m.append_rejected(all_reject)
        m.append_run_log([{
            "source_id": r["source_id"], "source_name": r["source_name"],
            "fetch_type": r["fetch_type"],
            "resolved_endpoint": (r["resolved"][0] if r["resolved"] else ""),
            "records_pulled": r["records_pulled"], "classification": r["classification"],
            "output_route": r["output_route"], "status": r["status"],
            "richness": r["richness"], "next_action": r["next_action"],
            "notes": r["error"],
        } for r in results])
        # source QA roll-up
        qa = []
        for r in results:
            qa.append({
                "source_id": r["source_id"], "source_name": r["source_name"],
                "records_pulled": r["records_pulled"],
                "records_loaded": r.get("records_eligible", 0),
                "records_rejected": len(r["rejected"]),
                "wrong_layer_count": 1 if r["status"] == "wrong_layer" else 0,
                "schema_too_thin_count": 1 if r["status"] == "schema_too_thin" else 0,
                "useful_leads_count": r.get("records_eligible", 0),
                "manual_review_count": r.get("records_held_for_review", 0),
                "next_action": r["next_action"],
            })
        m.upsert_source_qa(qa)
        # update Run_Queue status per source
        for r in results:
            m.update_run_queue(r["source_name"] or "",
                               status=r["status"],
                               notes=f"{r['classification']} | {r['next_action']}")
        m.save()
        records_written = len(all_leads) + len(all_reject)
        print("\n  MASTER WRITE:")
        print(f"    records_written: {records_written}")
        print(f"    Future_Projects: appended={fp_stats['appended']} "
              f"updated={fp_stats['updated_stage']} enriched={fp_stats['enriched']} "
              f"collapsed_dupes={fp_stats['collapsed_dupes']} -> {fp_stats['final_rows']} rows")
        print(f"    Rejected_Archive: appended={rj_stats['appended']}")
    elif args.dry_run:
        print("\n  DRY-RUN: no master write. Classified leads/rejects above.")
        if args.write_master:
            import tenderfinder_master_io as MIO
            if os.path.exists(resolved_write_target):
                m = MIO.MasterIO(resolved_write_target)
                print("    FP merge preview:", m.write_future(all_leads, dry_run=True))
            else:
                print("    FP merge preview skipped: target workbook does not exist; "
                      "dry-run will not create or copy workbooks.")

    # Always write source-level summary (Patch 5.0)
    run_id_final = f"RUN-{dt.date.today().isoformat()}"
    out_dir_final = os.path.dirname(os.path.abspath(args.out)) if args.out else raw_dir
    if not args.out.endswith(".xlsx"):
        out_dir_final = args.out
    if getattr(args, "source_summary", False) or True:  # always write in 5.0
        write_source_summary(results, out_dir_final, run_id_final)
    if getattr(args, "demo_output", False):
        demo_path = os.path.join(out_dir_final, "TENDER_FINDER_Demo_Top_Leads.xlsx")
        _write_demo_output(results, demo_path, run_id_final)

    print("\nDone.\n")


def _review_rows(results):
    rows = []
    for r in results:
        for lead in r.get("leads") or []:
            rows.append({
                "source_id": r.get("source_id"),
                "source_name": r.get("source_name"),
                "project_id": lead.get("project_id"),
                "municipality": lead.get("municipality"),
                "app_no": lead.get("app_no"),
                "address": lead.get("address"),
                "owner/applicant": lead.get("owner"),
                "app_type_stage": lead.get("app_type_stage"),
                "scope_summary": lead.get("scope_summary"),
                "fit_score": lead.get("fit_score"),
                "fit_reason": lead.get("fit_reason") or "",
                "proposed_route": lead.get("proposed_route") or "Future_Projects",
                "write_eligible": bool(lead.get("write_eligible")),
                "hold_reason": lead.get("hold_reason") or "",
                "source_url": lead.get("source_url"),
                "evidence_url": lead.get("evidence_url") or lead.get("source_url") or "",
                "date_found": lead.get("date_found") or "",
                "run_id": lead.get("run_id") or "",
                "raw_hash": lead.get("raw_hash") or "",
                "dedupe_key": lead.get("dedupe_key") or "",
                # Patch 5.0: human review decision column — fill in and run --promote-reviewed
                "review_decision": "",  # ACCEPT | REJECT | HOLD | NEEDS_MORE_INFO
            })
        for item in r.get("rejected") or []:
            rows.append({
                "source_id": r.get("source_id"),
                "source_name": r.get("source_name"),
                "project_id": item.get("project_id"),
                "municipality": item.get("municipality"),
                "app_no": "",
                "address": "",
                "owner/applicant": "",
                "app_type_stage": item.get("title"),
                "scope_summary": item.get("scope_summary"),
                "fit_score": item.get("fit_score"),
                "fit_reason": "",
                "proposed_route": "Rejected_Archive",
                "write_eligible": False,
                "hold_reason": item.get("reason") or r.get("status") or "rejected",
                "source_url": item.get("source_url"),
                "evidence_url": item.get("source_url") or "",
                "date_found": "",
                "run_id": "",
                "raw_hash": "",
                "dedupe_key": "",
                "review_decision": "REJECT",  # pre-populated for rejected rows
            })
        if not (r.get("leads") or r.get("rejected")):
            rows.append({
                "source_id": r.get("source_id"),
                "source_name": r.get("source_name"),
                "project_id": "",
                "municipality": "",
                "app_no": "",
                "address": "",
                "owner/applicant": "",
                "app_type_stage": "",
                "scope_summary": r.get("richness") or r.get("next_action") or "",
                "fit_score": "",
                "fit_reason": "",
                "proposed_route": r.get("output_route") or "Run_Queue",
                "write_eligible": False,
                "hold_reason": r.get("status") or r.get("error") or "no_normalized_records",
                "source_url": (r.get("resolved") or [""])[0],
                "evidence_url": (r.get("resolved") or [""])[0],
                "date_found": "",
                "run_id": "",
                "raw_hash": "",
                "dedupe_key": "",
                "review_decision": "",
            })
    return rows


def _write_review_output(results, path):
    cols = ["source_id", "source_name", "project_id", "municipality", "app_no",
            "address", "owner/applicant", "app_type_stage", "scope_summary",
            "fit_score", "fit_reason", "proposed_route", "write_eligible", "hold_reason",
            "source_url", "evidence_url", "date_found", "run_id", "raw_hash", "dedupe_key",
            "review_decision"]  # Patch 5.0: ACCEPT | REJECT | HOLD | NEEDS_MORE_INFO
    rows = _review_rows(results)
    out_path = _resolve_path(path)
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    if out_path.lower().endswith(".xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "review"
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(1, c).font = Font(bold=True)
        # Highlight the review_decision column header in yellow
        from openpyxl.styles import PatternFill
        ws.cell(1, len(cols)).fill = PatternFill("solid", fgColor="FFFF00")
        for row in rows:
            ws.append([_excel_safe_text(row.get(c, "")) for c in cols])
        ws.freeze_panes = "A2"
        # Add data validation for review_decision column
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            dv = DataValidation(type="list", formula1='"ACCEPT,REJECT,HOLD,NEEDS_MORE_INFO"', allow_blank=True)
            dv.error = "Use: ACCEPT, REJECT, HOLD, or NEEDS_MORE_INFO"
            dv.errorTitle = "Invalid Decision"
            dv.prompt = "Mark each row with ACCEPT, REJECT, HOLD, or NEEDS_MORE_INFO"
            dv.promptTitle = "Review Decision"
            rd_col = get_column_letter(len(cols))
            dv.sqref = f"{rd_col}2:{rd_col}10000"
            ws.add_data_validation(dv)
        except Exception:
            pass  # data validation is optional enhancement
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(c) + 4))
        wb.save(out_path)
    else:
        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for row in rows:
                w.writerow(row)


def _write_probe_xlsx(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "probe"
    cols = ["source_id", "source_name", "fetch_type", "status",
            "resolved", "classification", "output_route", "next_action", "error"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for r in results:
        ws.append([r["source_id"], r["source_name"], r["fetch_type"], r["status"],
                   " | ".join(r["resolved"]), r["classification"] or "",
                   r["output_route"] or "", r["next_action"], r["error"]])
    ws.freeze_panes = "A2"
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 24
    wb.save(path)


def _save_run_log(raw_dir, results):
    os.makedirs(os.path.join(raw_dir, "logs"), exist_ok=True)
    with open(os.path.join(raw_dir, "logs", "run_log.json"), "w", encoding="utf-8") as f:
        slim = [{k: v for k, v in r.items() if k not in ("leads", "rejected")} for r in results]
        json.dump(slim, f, ensure_ascii=False, indent=2, default=str)


if __name__ == "__main__":
    main()
