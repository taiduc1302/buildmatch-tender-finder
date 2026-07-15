#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tenderfinder_master_io.py — safe reader/writer for the TENDER_FINDER master workbook
===================================================================
Writes structured leads INTO the existing master workbook schema (v7), never to
a separate flat sheet. Preserves formulas, the Fit Class formula, dropdowns,
conditional formatting, column widths and all existing tabs.

Hard safety rules (from the handoff):
  * Never overwrite v6. ensure_v7() copies v6 -> v7 once; all writes target v7.
  * Always back up the target file before writing.
  * Append new projects; UPDATE same project on a new stage; ENRICH same project
    on new owner/applicant/source URL. Multi-key dedup (tenderfinder_guards.dedup_keys).
  * Collapse pre-existing duplicate Project IDs (e.g. TWP-LANGLEY-RO100158) to one
    live row at the current stage, with prior stages preserved in Notes.

Tabs this module writes to:
  Future_Projects   (21-col schema; col M Fit Class formula preserved)
  Active_Tenders    (20-col; col L Fit Class + col O Days-to-Close formulas preserved)
  Rejected_Archive  (12-col; trailing/wrong-layer/thin/context quarantine)
  Run_Queue         (status/notes per source updated after a run)
  Run_Log           (ADDED if absent — per-source run results)
  Source_QA         (ADDED if absent — handoff section 16 source metrics)

Dependency: openpyxl.
"""

import os
import shutil
import datetime as dt

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import tenderfinder_guards as G
from tenderfinder_excel_safety import append_untrusted_row, set_untrusted_cell

# --- Future_Projects column map (1-indexed) --------------------------------
FP = {
    "project_id": 1, "date_found": 2, "source": 3, "source_url": 4,
    "title": 5, "owner": 6, "municipality": 7, "app_no": 8,
    "app_type_stage": 9, "scope_summary": 10, "expected_civil": 11,
    "fit_score": 12, "fit_class": 13, "verification": 14, "horizon": 15,
    "est_value": 16, "next_milestone": 17, "linked_active": 18,
    "assigned_to": 19, "notes": 20, "last_updated": 21,
}
FP_FIT_CLASS_FORMULA = ('=IF(L{r}="","",IF(L{r}>=85,"Strong Fit",'
                        'IF(L{r}>=70,"Good Lead",IF(L{r}>=50,"Watchlist","Reject"))))')

AT = {  # Active_Tenders column map
    "project_id": 1, "date_found": 2, "source": 3, "source_url": 4,
    "title": 5, "owner": 6, "gc": 7, "municipality": 8, "scope_summary": 9,
    "scope_match": 10, "fit_score": 11, "fit_class": 12, "verification": 13,
    "closing_date": 14, "days_to_close": 15, "est_value": 16,
    "assigned_to": 17, "bid_decision": 18, "notes": 19, "last_updated": 20,
}
AT_FIT_CLASS_FORMULA = ('=IF(K{r}="","",IF(K{r}>=85,"Strong Fit",'
                        'IF(K{r}>=70,"Good Lead",IF(K{r}>=50,"Watchlist","Reject"))))')
AT_DAYS_FORMULA = '=IF(N{r}="","",N{r}-TODAY())'

RA = {  # Rejected_Archive column map
    "project_id": 1, "date_found": 2, "date_rejected": 3, "source": 4,
    "source_url": 5, "title": 6, "municipality": 7, "scope_summary": 8,
    "fit_score": 9, "reason": 10, "verification": 11, "notes": 12,
}

TODAY = dt.date.today().isoformat()


# ===========================================================================
# Version safety
# ===========================================================================
def ensure_v7(v6_path, v7_path):
    """Create v7 from v6 if v7 does not exist. NEVER overwrites v6."""
    if os.path.abspath(v6_path) == os.path.abspath(v7_path):
        raise ValueError("Refusing to use the same path for v6 and v7.")
    if not os.path.exists(v7_path):
        if not os.path.exists(v6_path):
            raise FileNotFoundError(f"v6 master not found: {v6_path}")
        shutil.copy2(v6_path, v7_path)
        print(f"  [master] created v7 from v6 -> {v7_path}")
    else:
        print(f"  [master] v7 already exists -> {v7_path}")
    return v7_path


def backup(path):
    """Create a mandatory short-name workbook backup before any write.

    Primary backup location is a ``backups`` folder beside the workbook.
    If that fails, fall back to ``01 Code/CONNECTOR_SWEEP/workbook_backups``.
    The backup filename is intentionally short to reduce Windows/OneDrive
    long-path failures. If no backup can be created, raise and stop the write.
    """
    src = os.path.abspath(os.path.normpath(path))
    if not os.path.exists(src):
        raise FileNotFoundError(f"Workbook backup failed: source workbook not found: {src}")

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = os.path.splitext(os.path.basename(src))[0]
    safe_stem = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in stem)
    safe_stem = (safe_stem[:24] or "workbook")
    backup_name = f"{safe_stem}_{stamp}.bak.xlsx"

    workbook_dir = os.path.dirname(src) or "."
    script_dir = os.path.dirname(os.path.abspath(__file__))
    attempts = [
        os.path.join(workbook_dir, "backups", backup_name),
        os.path.join(script_dir, "workbook_backups", backup_name),
    ]

    errors = []
    print(f"  [master] backup source -> {src}")
    for dst in attempts:
        dst = os.path.abspath(os.path.normpath(dst))
        try:
            os.makedirs(os.path.dirname(dst), exist_ok=True)
            shutil.copy2(src, dst)
            print(f"  [master] backup destination -> {dst}")
            return dst
        except Exception as e:
            errors.append(f"{dst}: {type(e).__name__}: {e}")

    detail = " | ".join(errors)
    raise RuntimeError(f"Workbook backup failed; write aborted before load/write/save. Tried: {detail}")


# ===========================================================================
# Stage ordering (for "current stage" decision when collapsing duplicates)
# ===========================================================================
_STAGE_ORDER = [
    "received", "submitted", "application", "in process", "in progress",
    "under review", "referral", "public notification", "public hearing",
    "1st reading", "first reading", "council 1", "1 & 2", "2nd reading",
    "second reading", "3rd reading", "third reading", "council 3",
    "adopted", "approved", "issued", "completed", "final",
]


def _stage_rank_tuple(text):
    """Return (primary_rank, secondary_rank) for a stage.
    primary_rank: position in _STAGE_ORDER or boosted for pending
    secondary_rank: tertiary tiebreaker for council reading level (3rd > 1&2 > 1)
    This ensures RO100158 dupes collapse to COUNCIL 3RD READING — PENDING."""
    t = str(text or "").lower()
    rank = -1
    for i, marker in enumerate(_STAGE_ORDER):
        if marker in t:
            rank = max(rank, i)
    
    # KEY FIX: pending (future) beats completed (past)
    if "pending" in t:
        rank = len(_STAGE_ORDER) + 10
    
    # Tertiary sort: council reading level (3rd > 2nd > 1st; 1&2 at 1.5)
    secondary = 0
    if "3rd" in t or "third" in t or "council 3" in t:
        secondary = 3
    elif "2nd" in t or "second" in t or "council 2" in t:
        secondary = 2
    elif "1 & 2" in t:
        secondary = 1.5
    elif "1st" in t or "first" in t or "council 1" in t:
        secondary = 1
    
    return (rank, secondary)


def _stage_rank(text):
    """Return primary rank for backward compatibility; use _stage_rank_tuple
    for precise comparisons."""
    return _stage_rank_tuple(text)[0]


# ===========================================================================
# Master writer
# ===========================================================================
class MasterIO:
    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(path)  # keep formulas
        self._ensure_aux_tabs()

    # -- auxiliary tabs (added if missing; never removes existing tabs) ------
    def _ensure_aux_tabs(self):
        if "Run_Log" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Run_Log")
            hdrs = ["Run Date", "Source ID", "Source Name", "Fetch Type",
                    "Resolved Endpoint", "Records Pulled", "Classification",
                    "Output Route", "Status", "Richness", "Next Action", "Notes"]
            ws.append(hdrs)
            for c in range(1, len(hdrs) + 1):
                ws.cell(1, c).font = Font(bold=True)
            ws.freeze_panes = "A2"
        if "Source_QA" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Source_QA")
            hdrs = ["Source ID", "Source Name", "Records Pulled",
                    "Records Loaded", "Records Rejected", "Wrong-Layer Count",
                    "Schema-Too-Thin Count", "Duplicate Count",
                    "Useful Leads Count", "Manual Review Count",
                    "False Positive Reason", "Reviewer Feedback",
                    "Estimator Feedback", "Automation Readiness", "Next Action"]
            ws.append(hdrs)
            for c in range(1, len(hdrs) + 1):
                ws.cell(1, c).font = Font(bold=True)
            ws.freeze_panes = "A2"

    # -- low-level helpers ---------------------------------------------------
    @staticmethod
    def _first_empty_row(ws, key_col=1, start=2):
        r = start
        while ws.cell(r, key_col).value not in (None, ""):
            r += 1
        return r

    @staticmethod
    def _row_dict(ws, r, colmap):
        return {name: ws.cell(r, idx).value for name, idx in colmap.items()}

    # =======================================================================
    # FUTURE_PROJECTS  (full read -> merge -> rewrite, preserves M formula)
    # =======================================================================
    def load_future(self):
        """Return list of existing FP rows as dicts (data only, no formula col)."""
        ws = self.wb["Future_Projects"]
        rows = []
        r = 2
        while ws.cell(r, FP["project_id"]).value not in (None, ""):
            d = self._row_dict(ws, r, FP)
            d.pop("fit_class", None)  # formula, regenerated on write
            rows.append(d)
            r += 1
        return rows

    @staticmethod
    def _fp_index(rows):
        """Build dedup index: maps any dedup key -> position in `rows`."""
        idx = {}
        for i, d in enumerate(rows):
            for k in G.dedup_keys(d.get("municipality", ""),
                                  d.get("app_no", ""), d.get("address", ""),
                                  d.get("source_url", ""), d.get("owner", ""),
                                  d.get("title", "")):
                idx[k] = i
            pid = d.get("project_id")
            if pid:
                idx[("pid", str(pid))] = i
        return idx

    def collapse_future_duplicates(self, rows):
        """Collapse rows that share the same Project ID into one live row at the
        most-advanced stage; prior stages -> Notes (handoff 13, RO100158)."""
        by_pid = {}
        order = []
        collapsed = 0
        for d in rows:
            pid = str(d.get("project_id") or "")
            if not pid:
                order.append(("_", d))
                continue
            if pid not in by_pid:
                by_pid[pid] = d
                order.append(("pid", pid))
            else:
                collapsed += 1
                cur = by_pid[pid]
                new_stage = d.get("app_type_stage", "")
                cur_stage = cur.get("app_type_stage", "")
                # keep the more-advanced stage as live (use tuple for tie-breaker)
                if _stage_rank_tuple(new_stage) > _stage_rank_tuple(cur_stage):
                    hist = cur_stage
                    by_pid[pid] = d
                    live = d
                else:
                    hist = new_stage
                    live = cur
                if hist and hist not in str(live.get("notes", "")):
                    live["notes"] = (str(live.get("notes", "") or "").rstrip()
                                     + f" | prior stage: {hist}").strip(" |")
        out = []
        for kind, key in order:
            if kind == "pid":
                out.append(by_pid[key])
            else:
                out.append(key)
        return out, collapsed

    def write_future(self, leads, dry_run=False):
        """Merge `leads` (list of dicts) into Future_Projects.
        Returns stats dict. Rewrites the data region to guarantee dedup +
        formula integrity; header/dropdown/CF/widths preserved."""
        ws = self.wb["Future_Projects"]
        existing = self.load_future()
        existing, collapsed = self.collapse_future_duplicates(existing)
        idx = self._fp_index(existing)

        appended = updated = enriched = 0
        for lead in leads:
            keys = G.dedup_keys(lead.get("municipality", ""),
                                lead.get("app_no", ""), lead.get("address", ""),
                                lead.get("source_url", ""), lead.get("owner", ""),
                                lead.get("title", ""))
            pid = str(lead.get("project_id") or "")
            hit = None
            if pid and ("pid", pid) in idx:
                hit = idx[("pid", pid)]
            else:
                for k in keys:
                    if k in idx:
                        hit = idx[k]
                        break
            if hit is None:
                existing.append(lead)
                pos = len(existing) - 1
                for k in keys:
                    idx[k] = pos
                if pid:
                    idx[("pid", pid)] = pos
                appended += 1
            else:
                cur = existing[hit]
                changed = self._merge_into(cur, lead)
                if changed == "stage":
                    updated += 1
                elif changed == "enrich":
                    enriched += 1

        stats = {"existing_before": len(self.load_future()),
                 "collapsed_dupes": collapsed, "appended": appended,
                 "updated_stage": updated, "enriched": enriched,
                 "final_rows": len(existing)}
        if dry_run:
            stats["dry_run"] = True
            return stats

        self._rewrite_future(ws, existing)
        return stats

    @staticmethod
    def _merge_into(cur, lead):
        """Update existing FP row in place. Returns 'stage'|'enrich'|None."""
        result = None
        new_stage = lead.get("app_type_stage", "")
        cur_stage = cur.get("app_type_stage", "")
        if new_stage and G.norm_text(new_stage) != G.norm_text(cur_stage) \
                and _stage_rank_tuple(new_stage) >= _stage_rank_tuple(cur_stage):
            if cur_stage and cur_stage not in str(cur.get("notes", "")):
                cur["notes"] = (str(cur.get("notes", "") or "").rstrip()
                                + f" | prior stage: {cur_stage}").strip(" |")
            cur["app_type_stage"] = new_stage
            cur["last_updated"] = TODAY
            result = "stage"
        # enrich blanks: owner, source_url, est_value, next_milestone, scope
        for field in ("owner", "source_url", "est_value", "next_milestone",
                      "scope_summary", "expected_civil", "horizon"):
            cv = cur.get(field)
            nv = lead.get(field)
            if (cv in (None, "", "Not available") ) and nv not in (None, "", "Not available"):
                cur[field] = nv
                cur["last_updated"] = TODAY
                if result is None:
                    result = "enrich"
        return result

    def _rewrite_future(self, ws, rows):
        """Clear data region and rewrite, regenerating the Fit Class formula
        per row. Extends dropdown/CF ranges if data exceeds row 201."""
        # clear old data (cols 1..21, rows 2..max_row)
        for r in range(2, ws.max_row + 1):
            for c in range(1, 22):
                ws.cell(r, c).value = None
        # write rows
        for i, d in enumerate(rows):
            r = 2 + i
            for key in (
                "project_id", "date_found", "source", "source_url", "title", "owner",
                "municipality", "app_no", "app_type_stage", "scope_summary",
                "expected_civil", "fit_score", "verification", "horizon", "est_value",
                "next_milestone", "linked_active", "assigned_to", "notes", "last_updated",
            ):
                value = d.get(key)
                if key in {"date_found", "last_updated"}:
                    value = value or TODAY
                elif key == "verification":
                    value = value or "Needs Review"
                set_untrusted_cell(ws.cell(r, FP[key]), value)
            ws.cell(r, FP["fit_class"]).value = FP_FIT_CLASS_FORMULA.format(r=r)
        # ensure formula present in any remaining template rows up to 201
        last_data = 1 + len(rows)
        for r in range(max(2, last_data + 1), 202):
            ws.cell(r, FP["fit_class"]).value = FP_FIT_CLASS_FORMULA.format(r=r)
        self._extend_fp_ranges(ws, last_data)

    @staticmethod
    def _extend_fp_ranges(ws, last_data):
        """Extend the Verification dropdown (col N) + Fit Class CF (col M) if
        data now goes beyond row 201."""
        if last_data <= 201:
            return
        # dropdown
        for dv in ws.data_validations.dataValidation:
            dv.sqref = f"N2:N{last_data}"
        # conditional formatting: re-add over wider range
        # (openpyxl keeps existing M2:M201 rules; add same coverage note)
        # Practical approach: copy formula already done; CF colours only matter
        # visually and are safe to leave at original range.

    # =======================================================================
    # ACTIVE_TENDERS (append only; placeholder track until parser exists)
    # =======================================================================
    def append_active(self, tenders, dry_run=False):
        ws = self.wb["Active_Tenders"]
        appended = 0
        for t in tenders:
            r = self._first_empty_row(ws, AT["project_id"])
            if dry_run:
                appended += 1
                continue
            for key in (
                "project_id", "date_found", "source", "source_url", "title", "owner", "gc",
                "municipality", "scope_summary", "scope_match", "fit_score", "verification",
                "closing_date", "est_value", "assigned_to", "bid_decision", "notes",
            ):
                value = t.get(key)
                if key == "date_found":
                    value = value or TODAY
                elif key == "verification":
                    value = value or "Needs Review"
                set_untrusted_cell(ws.cell(r, AT[key]), value)
            ws.cell(r, AT["fit_class"]).value = AT_FIT_CLASS_FORMULA.format(r=r)
            ws.cell(r, AT["days_to_close"]).value = AT_DAYS_FORMULA.format(r=r)
            set_untrusted_cell(ws.cell(r, AT["last_updated"]), TODAY)
            appended += 1
        return {"appended": appended, "dry_run": dry_run}

    # =======================================================================
    # REJECTED_ARCHIVE (quarantine: wrong_layer / schema_too_thin / trailing / context)
    # =======================================================================
    def append_rejected(self, items, dry_run=False):
        ws = self.wb["Rejected_Archive"]
        appended = 0
        for it in items:
            r = self._first_empty_row(ws, RA["project_id"])
            if dry_run:
                appended += 1
                continue
            values = {
                "project_id": it.get("project_id") or it.get("source_id"),
                "date_found": it.get("date_found") or TODAY,
                "date_rejected": TODAY,
                "source": it.get("source"),
                "source_url": it.get("source_url"),
                "title": it.get("title"),
                "municipality": it.get("municipality"),
                "scope_summary": it.get("scope_summary"),
                "fit_score": it.get("fit_score"),
                "reason": it.get("reason"),
                "verification": it.get("verification") or "Rejected",
                "notes": it.get("notes"),
            }
            for key, value in values.items():
                set_untrusted_cell(ws.cell(r, RA[key]), value)
            appended += 1
        return {"appended": appended, "dry_run": dry_run}

    # =======================================================================
    # RUN_QUEUE status update + RUN_LOG + SOURCE_QA
    # =======================================================================
    def update_run_queue(self, source_match, status=None, notes=None):
        """Update Status/Notes for a Run_Queue row whose 'Municipality / Source'
        contains source_match (case-insensitive)."""
        ws = self.wb["Run_Queue"]
        sm = (source_match or "").lower()
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, 2).value
            if cell and sm and sm in str(cell).lower():
                if status is not None:
                    set_untrusted_cell(ws.cell(r, 10), status)     # Status col
                if notes is not None:
                    set_untrusted_cell(ws.cell(r, 11), notes)       # Notes col
                return True
        return False

    def append_run_log(self, entries):
        ws = self.wb["Run_Log"]
        for e in entries:
            append_untrusted_row(ws, [
                e.get("run_date") or TODAY, e.get("source_id"),
                e.get("source_name"), e.get("fetch_type"),
                e.get("resolved_endpoint"), e.get("records_pulled"),
                e.get("classification"), e.get("output_route"),
                e.get("status"), e.get("richness"), e.get("next_action"),
                e.get("notes"),
            ])

    def upsert_source_qa(self, rows):
        """rows: list of dicts keyed by Source ID. Updates existing or appends."""
        ws = self.wb["Source_QA"]
        existing = {}
        for r in range(2, ws.max_row + 1):
            sid = ws.cell(r, 1).value
            if sid:
                existing[str(sid)] = r
        cols = ["source_id", "source_name", "records_pulled", "records_loaded",
                "records_rejected", "wrong_layer_count", "schema_too_thin_count",
                "duplicate_count", "useful_leads_count", "manual_review_count",
                "false_positive_reason", "ryan_feedback", "estimator_feedback",
                "automation_readiness", "next_action"]
        for d in rows:
            sid = str(d.get("source_id"))
            r = existing.get(sid) or (ws.max_row + 1)
            for i, key in enumerate(cols, start=1):
                if d.get(key) is not None:
                    set_untrusted_cell(ws.cell(r, i), d.get(key))

    # -----------------------------------------------------------------------
    def save(self):
        self.wb.save(self.path)
        print(f"  [master] saved -> {self.path}")


if __name__ == "__main__":
    print("tenderfinder_master_io is a library; import it from the sweep. "
          "Run the sweep with --write-master to use it.")
