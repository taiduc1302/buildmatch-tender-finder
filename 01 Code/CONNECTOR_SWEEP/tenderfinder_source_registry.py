#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Canonical editable source registry for TENDER_FINDER.

``config/sources.csv`` is the single runtime source of truth for both tender
and development tracks. This module strictly validates the registry, exposes
active rows to both collection paths, and provides atomic add/edit/toggle
operations used by the GUI.

The lower compatibility section can still reconcile a historical
``Source_Register`` worksheet with the canonical CSV for audit reports. Its
name aliases are reference-only; they do not define normal runtime sources.
"""

import csv
import os
import re
import shutil
import uuid
from datetime import datetime
from pathlib import Path

from tenderfinder_package_paths import detect_package_root, sources_config_path
from tenderfinder_runtime import user_settings_root
from tenderfinder_excel_safety import safe_untrusted_excel_value
from tenderfinder_url_safety import URLSafetyError, validate_public_url_syntax


REGISTRY_ENV_VAR = "TENDER_FINDER_SOURCES_CONFIG"
LEGACY_SOURCE_COLUMNS = (
    "source_id", "name", "track", "active", "adapter", "municipality", "url", "rss",
    "url_variants", "no_retry", "category", "tier", "platform", "fetch_type", "endpoint",
    "layer_index", "layer_keywords", "priority_tier", "access_status",
    "automation_feasibility", "output_route", "prompt_type", "status", "last_probe_status",
    "last_good_endpoint", "notes",
)
OPERATIONAL_COLUMNS = (
    "operational_status", "last_tested_at", "last_test_type", "last_test_result",
    "last_http_status", "last_error", "last_candidate_count", "last_normalized_count",
)
LIVE_TEST_QUERY_COLUMNS = ("test_query_where", "test_query_order_by")
SOURCE_COLUMNS = LEGACY_SOURCE_COLUMNS + OPERATIONAL_COLUMNS + LIVE_TEST_QUERY_COLUMNS
SOURCE_TRACKS = frozenset({"tender", "development"})
TENDER_ADAPTERS = frozenset({"public_listing", "bc_bid_browser"})
DEVELOPMENT_ADAPTERS = frozenset(
    {
        "arcgis_hub_discover", "arcgis_hub_item", "arcgis_map_discover",
        "arcgis_rest_layer", "ods_v21", "surrey_planning_reports",
    }
)
SOURCE_ID_RE = re.compile(r"^[a-z0-9][a-z0-9_]{2,63}$")
RESERVED_SOURCE_IDS = frozenset({"email_alert_intake"})
OPERATIONAL_STATUSES = frozenset(
    {
        "verified_live",
        "ready_for_live_test",
        "adapter_fixture_pass",
        "config_valid_only",
        "needs_configuration",
        "manual_only",
        "blocked",
        "wrong_source",
        "deprecated",
    }
)
RUNTIME_ELIGIBLE_STATUSES = frozenset(
    {"verified_live", "ready_for_live_test", "adapter_fixture_pass"}
)
LAST_TEST_TYPES = frozenset(
    {"", "imported_history", "configuration_validation", "offline_fixture", "live_source_test"}
)
PLACEHOLDER_MARKERS = (
    "disabled_pending",
    "placeholder",
    "replace_me",
    "needs_exact_url",
    "pending_office_network",
    "example.invalid",
)


class SourceRegistryError(RuntimeError):
    """Raised when the founder-editable runtime source registry is unsafe."""

    def __init__(self, path, errors):
        self.path = Path(path)
        self.errors = tuple(str(error) for error in errors)
        details = "\n".join(f"  - {error}" for error in self.errors)
        super().__init__(f"Source registry is invalid: {self.path}\n{details}")


def resolve_registry_path(path=None, root=None):
    if path:
        return Path(path).expanduser().resolve()
    override = os.environ.get(REGISTRY_ENV_VAR, "").strip()
    if override:
        return Path(override).expanduser().resolve()
    package_root = detect_package_root(Path(root).resolve() if root else None)
    return sources_config_path(package_root).resolve()


def _is_public_url(value):
    """Compatibility boolean for the no-network registry validation step."""
    try:
        validate_public_url_syntax(value)
    except URLSafetyError:
        return False
    return True


def _looks_like_url(value):
    text = str(value or "").strip().casefold()
    return text.startswith(("http://", "https://", "file:", "ftp:", "\\\\", "//"))


def _is_placeholder(value):
    text = str(value or "").strip().casefold()
    return bool(text) and any(marker in text for marker in PLACEHOLDER_MARKERS)


def _safe_nonnegative_integer(value, label):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = int(text)
    except ValueError as exc:
        raise ValueError(f"{label} must be a non-negative whole number") from exc
    if parsed < 0:
        raise ValueError(f"{label} must be a non-negative whole number")
    return parsed


def derive_operational_status(row):
    """Derive a conservative status for pre-stabilization CSV rows."""
    explicit = str(row.get("operational_status", "") or "").strip().casefold()
    if explicit:
        return explicit
    access = str(row.get("access_status", "") or "").strip().casefold()
    legacy_status = str(row.get("status", "") or "").strip().casefold()
    last_probe = str(row.get("last_probe_status", "") or "").strip().casefold()
    endpoint = str(row.get("endpoint", "") or "").strip()
    notes = str(row.get("notes", "") or "").casefold()
    wrong_markers = (
        "wrong_layer", "picked_permits", "building_permits_only", "not_found_infra",
        "not_found_ocp", "do not connect",
    )
    if "wrong_layer" in access or "wrong layer" in notes or any(marker in last_probe for marker in wrong_markers):
        return "wrong_source"
    if access in {"manual_p3_only", "paid_or_login_skip"} or (
        str(row.get("track", "")).casefold() == "tender"
        and ("membership likely" in notes or "login platform skipped" in notes)
    ):
        return "manual_only"
    if access in {"access_test_required", "gated_office_network"} or _is_placeholder(endpoint):
        return "blocked"
    if access in {"needs_exact_url", "endpoint_stale"} or "slug_to_verify" in legacy_status:
        return "needs_configuration"
    # Historical LIVE/CONFIRMED labels are retained in their legacy columns,
    # but stabilization does not silently promote them to current
    # ``verified_live``.  A controlled live test with structured metadata is
    # the only operation that assigns that status.
    live_markers = ("live", "gold_", "pulled_", "capital_context_", "trailing_issued_")
    if legacy_status in {"live", "confirmed"} or any(marker in last_probe for marker in live_markers):
        return "ready_for_live_test"
    if str(row.get("track", "")).casefold() == "tender":
        return "ready_for_live_test"
    if access in {"ready_for_load", "ready_for_probe", "trailing_context"}:
        return "ready_for_live_test"
    return "config_valid_only"


def _normalized_source_row(row):
    out = {
        str(column): str(value or "").strip()
        for column, value in dict(row or {}).items()
        if column is not None
    }
    for column in SOURCE_COLUMNS:
        out.setdefault(column, "")
    out["source_id"] = out["source_id"].casefold()
    out["track"] = out["track"].casefold()
    out["active"] = out["active"].upper()
    out["adapter"] = out["adapter"].casefold()
    out["no_retry"] = (out["no_retry"] or "N").upper()
    if not out["fetch_type"]:
        out["fetch_type"] = out["adapter"]
    if out["track"] == "development" and not out["endpoint"]:
        out["endpoint"] = out["url"]
    out["operational_status"] = derive_operational_status(out)
    return out


def _url_field_error(label, value):
    try:
        validate_public_url_syntax(value)
    except URLSafetyError as exc:
        return f"{label} is unsafe: {exc}"
    return ""


def _network_fields(row):
    fields = []
    track = str(row.get("track", "") or "").casefold()
    for label in ("url", "rss"):
        value = str(row.get(label, "") or "").strip()
        if value and (track == "tender" or _looks_like_url(value)):
            fields.append((label, value))
    fields.extend(
        ("url_variants", part.strip())
        for part in str(row.get("url_variants", "")).split("|")
        if part.strip()
    )
    for label in ("endpoint", "last_good_endpoint"):
        value = str(row.get(label, "") or "").strip()
        if value and _looks_like_url(value):
            fields.append((label, value))
    return fields


def _row_configuration_errors(row):
    errors = []
    status = row["operational_status"]
    if not row["name"]:
        errors.append("name is required")
    if row["track"] == "tender":
        if not row["adapter"]:
            errors.append("adapter is required")
        elif row["adapter"] not in TENDER_ADAPTERS:
            errors.append(
                "custom adapter required; supported tender adapters are "
                + ", ".join(sorted(TENDER_ADAPTERS))
            )
        if not row["url"]:
            errors.append("tender url is required")
    elif row["track"] == "development":
        if not row["adapter"]:
            errors.append("adapter is required")
        elif row["adapter"] not in DEVELOPMENT_ADAPTERS:
            errors.append(
                "custom adapter required; supported development adapters are "
                + ", ".join(sorted(DEVELOPMENT_ADAPTERS))
            )
        if not row["endpoint"]:
            errors.append("development endpoint is required")
    for label, value in _network_fields(row):
        error = _url_field_error(label, value)
        if error:
            errors.append(error)
    placeholder_fields = [
        label for label in ("url", "rss", "endpoint", "last_good_endpoint")
        if _is_placeholder(row.get(label, ""))
    ]
    if placeholder_fields and status in RUNTIME_ELIGIBLE_STATUSES:
        errors.append(
            "placeholder value is incompatible with runnable operational_status "
            f"({', '.join(placeholder_fields)})"
        )
    test_where = str(row.get("test_query_where") or "").strip()
    if len(test_where) > 500 or ";" in test_where or any(ord(ch) < 32 for ch in test_where):
        errors.append(
            "test_query_where must be at most 500 characters and contain no semicolon or control character"
        )
    test_order = str(row.get("test_query_order_by") or "").strip()
    if test_order and not re.fullmatch(
        r"[A-Za-z_][A-Za-z0-9_]*(?:\s+(?:ASC|DESC))?"
        r"(?:\s*,\s*[A-Za-z_][A-Za-z0-9_]*(?:\s+(?:ASC|DESC))?)*",
        test_order,
        flags=re.IGNORECASE,
    ):
        errors.append(
            "test_query_order_by must be a comma-separated field list with optional ASC or DESC"
        )
    return errors


def source_readiness_errors(row):
    """Return reasons why one row cannot be enabled or explicitly tested.

    Disabled rows are allowed to be incomplete drafts in the canonical CSV.
    This helper is the single readiness gate used when enabling/running them.
    """
    normalized = _normalized_source_row(row)
    errors = _row_configuration_errors(normalized)
    if normalized["active"] != "Y":
        errors.append("source is disabled by the founder-controlled active flag")
    if normalized["operational_status"] not in RUNTIME_ELIGIBLE_STATUSES:
        errors.append(f"operational_status={normalized['operational_status']}")
    return errors


def source_configuration_errors(row):
    """Return structural/configuration errors without asserting run readiness."""
    return tuple(_row_configuration_errors(_normalized_source_row(row)))


def source_skip_reason(row):
    errors = source_readiness_errors(row)
    return "; ".join(errors)


def source_is_runtime_eligible(row):
    return not source_readiness_errors(row)


def _validate_rows(rows, path):
    errors = []
    seen = set()
    for row_number, row in enumerate(rows, start=2):
        sid = row["source_id"]
        prefix = f"row {row_number}"
        if not SOURCE_ID_RE.fullmatch(sid):
            errors.append(f"{prefix}: source_id '{sid}' must use lowercase letters, numbers, and underscores")
        if sid in RESERVED_SOURCE_IDS:
            errors.append(f"{prefix}: source_id '{sid}' is reserved for an internal source")
        if sid in seen:
            errors.append(f"{prefix}: duplicate source_id '{sid}'")
        seen.add(sid)
        if row["track"] not in SOURCE_TRACKS:
            errors.append(f"{prefix}: track must be tender or development")
        if row["active"] not in {"Y", "N"}:
            errors.append(f"{prefix}: active must be Y or N")
        if row["no_retry"] not in {"Y", "N"}:
            errors.append(f"{prefix}: no_retry must be Y or N")
        if row["operational_status"] not in OPERATIONAL_STATUSES:
            errors.append(
                f"{prefix}: operational_status must be one of {', '.join(sorted(OPERATIONAL_STATUSES))}"
            )
        if row["last_test_type"] not in LAST_TEST_TYPES:
            errors.append(
                f"{prefix}: last_test_type must be one of {', '.join(sorted(LAST_TEST_TYPES - {''}))}"
            )
        try:
            http_status = _safe_nonnegative_integer(row["last_http_status"], "last_http_status")
            if http_status is not None and not 100 <= http_status <= 599:
                errors.append(f"{prefix}: last_http_status must be between 100 and 599")
            _safe_nonnegative_integer(row["last_candidate_count"], "last_candidate_count")
            _safe_nonnegative_integer(row["last_normalized_count"], "last_normalized_count")
        except ValueError as exc:
            errors.append(f"{prefix}: {exc}")
        if row["active"] == "Y":
            errors.extend(f"{prefix}: {message}" for message in _row_configuration_errors(row))
        else:
            # Disabled drafts may be incomplete, but every populated URL still
            # receives the no-network portion of the safety policy.
            for label, value in _network_fields(row):
                error = _url_field_error(label, value)
                if error:
                    errors.append(f"{prefix}: {error}")
    endpoint_owner = {}
    for row_number, row in enumerate(rows, start=2):
        if row["active"] != "Y" or row["operational_status"] not in RUNTIME_ELIGIBLE_STATUSES:
            continue
        values = []
        for label, value in _network_fields(row):
            if value:
                values.append((label, str(value).strip().casefold().rstrip("/")))
        for label, value in values:
            owner = endpoint_owner.get(value)
            if owner and owner[0] != row["source_id"]:
                errors.append(
                    f"row {row_number}: duplicate runnable endpoint in {label}; "
                    f"already used by source_id '{owner[0]}'"
                )
            else:
                endpoint_owner[value] = (row["source_id"], label)
    if errors:
        raise SourceRegistryError(path, errors)


def load_source_rows(path=None, *, root=None, active_only=False, track=None):
    """Load and strictly validate the one canonical runtime source registry."""
    resolved = resolve_registry_path(path, root)
    if not resolved.exists():
        raise SourceRegistryError(resolved, ["sources.csv is missing"])
    try:
        with resolved.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            headers = tuple(reader.fieldnames or ())
            missing = [column for column in LEGACY_SOURCE_COLUMNS if column not in headers]
            if missing:
                raise SourceRegistryError(resolved, [f"missing column(s): {', '.join(missing)}"])
            rows = [_normalized_source_row(row) for row in reader]
    except SourceRegistryError:
        raise
    except Exception as exc:
        raise SourceRegistryError(resolved, [f"could not read registry: {exc}"]) from exc
    _validate_rows(rows, resolved)
    if track is not None:
        normalized_track = str(track).strip().casefold()
        if normalized_track not in SOURCE_TRACKS:
            raise SourceRegistryError(resolved, [f"unknown track '{track}'"])
        rows = [row for row in rows if row["track"] == normalized_track]
    if active_only:
        rows = [row for row in rows if row["active"] == "Y"]
    return rows


def load_tender_sources(path=None, *, root=None, active_only=True):
    sources = []
    for row in load_source_rows(path, root=root, active_only=active_only, track="tender"):
        source = dict(row)
        source["url_variants"] = [part.strip() for part in row["url_variants"].split("|") if part.strip()]
        source["no_retry"] = row["no_retry"] == "Y"
        source["runtime_skip_reason"] = source_skip_reason(row)
        sources.append(source)
    return sources


def load_development_connectors(path=None, *, root=None, active_only=True):
    return {
        row["source_id"]: {**row, "runtime_skip_reason": source_skip_reason(row)}
        for row in load_source_rows(path, root=root, active_only=active_only, track="development")
    }


def registry_summary(path=None, *, root=None):
    resolved = resolve_registry_path(path, root)
    rows = load_source_rows(resolved)
    counts = {
        "path": str(resolved),
        "total": len(rows),
        "enabled": sum(row["active"] == "Y" for row in rows),
        "tender": sum(row["track"] == "tender" for row in rows),
        "development": sum(row["track"] == "development" for row in rows),
        "runtime_eligible": sum(source_is_runtime_eligible(row) for row in rows),
    }
    counts["active"] = counts["enabled"]  # compatibility for older manifests/tests
    for status in sorted(OPERATIONAL_STATUSES):
        counts[status] = sum(row["operational_status"] == status for row in rows)
    return counts


def source_registry_backup_root(*, root=None, create=True):
    package_root = detect_package_root(Path(root).resolve() if root else None)
    backup_root = user_settings_root(package_root=package_root, create=create) / "source_registry_backups"
    if create:
        backup_root.mkdir(parents=True, exist_ok=True)
    return backup_root


def _fieldnames_for_rows(rows):
    fieldnames = list(SOURCE_COLUMNS)
    for row in rows:
        for key in row:
            if key not in fieldnames and not str(key).startswith("_"):
                fieldnames.append(key)
    return fieldnames


def _timestamped_backup(resolved, *, root=None, backup_root=None):
    if not resolved.exists():
        return None
    destination_root = (
        Path(backup_root).expanduser().resolve()
        if backup_root is not None
        else source_registry_backup_root(root=root, create=True)
    )
    destination_root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S_%f")
    destination = destination_root / f"{resolved.stem}_{stamp}.csv.bak"
    shutil.copy2(resolved, destination)
    with destination.open("rb+") as handle:
        os.fsync(handle.fileno())
    return destination


def write_source_rows(rows, path=None, *, root=None, backup_root=None):
    """Validate and atomically replace a registry edited by the desktop GUI."""
    resolved = resolve_registry_path(path, root)
    normalized = [_normalized_source_row(row) for row in rows]
    _validate_rows(normalized, resolved)
    resolved.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = _fieldnames_for_rows(normalized)
    temporary = resolved.with_name(f".{resolved.name}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows({column: row.get(column, "") for column in fieldnames} for row in normalized)
            handle.flush()
            os.fsync(handle.fileno())
        _timestamped_backup(resolved, root=root, backup_root=backup_root)
        os.replace(temporary, resolved)
    finally:
        if temporary.exists():
            temporary.unlink()
    return resolved


def upsert_source(source, path=None, *, root=None, backup_root=None):
    rows = load_source_rows(path, root=root)
    supplied = {str(key) for key in dict(source or {}) if key is not None}
    incoming = _normalized_source_row(source)
    replaced = False
    for index, row in enumerate(rows):
        if row["source_id"] == incoming["source_id"]:
            merged = dict(row)
            for key in supplied:
                merged[key] = incoming.get(key, "")
            rows[index] = _normalized_source_row(merged)
            replaced = True
            break
    if not replaced:
        rows.append(incoming)
    return write_source_rows(rows, path, root=root, backup_root=backup_root)


def set_source_active(source_id, active, path=None, *, root=None, backup_root=None):
    rows = load_source_rows(path, root=root)
    wanted = str(source_id).strip().casefold()
    found = False
    for row in rows:
        if row["source_id"] == wanted:
            row["active"] = "Y" if bool(active) else "N"
            found = True
            break
    if not found:
        raise SourceRegistryError(resolve_registry_path(path, root), [f"source_id '{source_id}' was not found"])
    return write_source_rows(rows, path, root=root, backup_root=backup_root)


def update_source_test_metadata(
    source_id,
    *,
    operational_status=None,
    test_type,
    test_result,
    http_status="",
    error="",
    candidate_count=0,
    normalized_count=0,
    last_good_endpoint=None,
    tested_at=None,
    path=None,
    root=None,
    backup_root=None,
):
    rows = load_source_rows(path, root=root)
    wanted = str(source_id).strip().casefold()
    found = False
    for row in rows:
        if row["source_id"] != wanted:
            continue
        found = True
        if operational_status is not None:
            row["operational_status"] = str(operational_status).strip().casefold()
        row["last_tested_at"] = tested_at or datetime.now().astimezone().isoformat(timespec="seconds")
        row["last_test_type"] = str(test_type).strip().casefold()
        row["last_test_result"] = str(test_result or "").strip()
        row["last_http_status"] = str(http_status or "").strip()
        row["last_error"] = str(error or "").strip()[:1000]
        row["last_candidate_count"] = str(int(candidate_count or 0))
        row["last_normalized_count"] = str(int(normalized_count or 0))
        row["last_probe_status"] = row["last_test_result"]
        if last_good_endpoint is not None:
            row["last_good_endpoint"] = str(last_good_endpoint or "").strip()
        break
    if not found:
        raise SourceRegistryError(resolve_registry_path(path, root), [f"source_id '{source_id}' was not found"])
    return write_source_rows(rows, path, root=root, backup_root=backup_root)

# Compatibility-only map from historical Source_Register display names to
# canonical source IDs. Runtime collection never reads this list.
NAME_TO_CONNECTOR = {
    "township of langley development activity": "twp_langley_devactivity",
    "maple ridge active development application": "maple_ridge_devapps",
    "surrey development applications / open dat": "surrey_devapps",
    "surrey development applications v2": "surrey_devapps_v2",
    "surrey planning reports": "surrey_planning_reports",
    "surrey planning reports / in-process": "surrey_planning_reports",
    "city of surrey - planning reports": "surrey_planning_reports",
    "surrey gis / futureworks": "surrey_futureworks",
    "surrey capital / futureworks layers": "surrey_futureworks",
    "city of langley development portal": "city_langley_devapps",
    "vancouver issued building permits": "van_building_permits",
    "vancouver open data": "van_city_projects",
    "abbotsford development / open data": "abbotsford_devapps",
    "new westminster open data": "new_west_currentdev",
    "burnaby open data / capital projects": "burnaby_devapps",
    "coquitlam development information portal": "coquitlam_devapps",
    "delta current development applications": "delta_devapps",
    "district of north vancouver geoweb": "dnv_devapps",
}

# Category-driven defaults for sources WITHOUT a technical connector.
PAID_TOKENS = ("paid",)
LOGIN_TOKENS = ("login", "registration", "bceid", "ariba", "supplier portal")
ACTIVE_TENDER_CAT = "a —"          # category A — active tender portals
NEWS_CAT = "g —"                   # category G — news / early-signal
COUNCIL_CAT = "c —"                # category C — council / committee agendas
CAPITAL_CAT = "d —"                # category D — capital plans
GC_CAT = "f —"                     # category F — GC / developer invitations
PAID_CAT = "e —"                   # category E — paid intelligence


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def load_connectors(csv_path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = tuple(reader.fieldnames or ())
        rows = list(reader)
    if "track" in fieldnames:
        rows = load_source_rows(csv_path, active_only=False, track="development")
    # tolerant key access: source_id or id
    out = {}
    for r in rows:
        sid = r.get("source_id") or r.get("id")
        if sid:
            out[sid] = r
    return out


def load_source_register(master_path):
    import openpyxl
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb["Source_Register"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        rec = {hdr[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        out.append(rec)
    return out


def match_connector(source_name, connectors):
    """Return connector source_id for a Source_Register name, or None."""
    n = _norm(source_name)
    # longest alias first for specificity
    for alias in sorted(NAME_TO_CONNECTOR, key=len, reverse=True):
        if alias in n or n in alias:
            sid = NAME_TO_CONNECTOR[alias]
            if sid in connectors:
                return sid
    return None


def classify_source(sr_row, connector):
    """Decide the sync status + next action for one Source_Register row.
    sr_row    : dict from Source_Register
    connector : dict from connector CSV, or None
    Returns dict with sync fields."""
    cat = _norm(sr_row.get("Category"))
    cost = _norm(sr_row.get("Cost"))
    access = _norm(sr_row.get("Access Status / Confidence"))
    name = sr_row.get("Source Name")

    # ---- paid / login first (never scrape) -------------------------------
    if cat.startswith(PAID_CAT) or any(t in cost for t in PAID_TOKENS):
        return _mk("paid_or_login_skip", "skip unless paid access configured",
                   connector, sr_row)
    if any(t in access for t in LOGIN_TOKENS) and cat.startswith(ACTIVE_TENDER_CAT):
        return _mk("paid_or_login_skip",
                   "login/registration portal — use alerts, do not scrape",
                   connector, sr_row)

    # ---- active tender portals (separate track) --------------------------
    if cat.startswith(ACTIVE_TENDER_CAT):
        return _mk("not_automation_ready",
                   "Active-tender track: configure alerts in Task D Phase 0",
                   connector, sr_row, route="Active_Tenders",
                   prompt="PROMPT_ACTIVE_TENDER_PARSE")

    # ---- has a technical connector ---------------------------------------
    if connector:
        acc = _norm(connector.get("access_status"))
        if acc == "ready_for_load":
            return _mk("ready_for_load", "Tier load OK — run --tier with richness gate",
                       connector, sr_row)
        if acc == "ready_for_probe":
            return _mk("ready_for_probe", "probe + verify layer before load",
                       connector, sr_row)
        if acc == "needs_exact_url":
            return _mk("needs_exact_url",
                       "verify endpoint manually (no silent fallback)",
                       connector, sr_row)
        if acc == "disabled_wrong_layer":
            return _mk("disabled_wrong_layer",
                       "wrong layer — denylisted; re-pin to a real dev-app layer",
                       connector, sr_row, route="Rejected_Archive")
        if acc == "access_test_required":
            return _mk("access_test_required",
                       "run from TENDER_FINDER office network (Surrey Test #1)",
                       connector, sr_row)
        if acc == "trailing_context":
            return _mk("ready_for_probe",
                       "trailing/context only — load to Rejected_Archive/Context, not FP",
                       connector, sr_row, route="Rejected_Archive")
        if acc == "manual_p3_only":
            return _mk("manual_p3_only", "run P3 web/PDF extractor",
                       connector, sr_row, route="Run_Queue",
                       prompt="PROMPT_WEB_PDF_EXTRACT")
        # connector exists but status unknown
        return _mk("ready_for_probe", "probe to confirm layer",
                   connector, sr_row)

    # ---- no connector: route by category ---------------------------------
    if cat.startswith(COUNCIL_CAT):
        return _mk("manual_p3_only", "run P3 council-agenda extractor", None,
                   sr_row, route="Run_Queue", prompt="PROMPT_COUNCIL_CAPITAL")
    if cat.startswith(CAPITAL_CAT):
        return _mk("manual_p3_only", "run P3 capital-plan extractor", None,
                   sr_row, route="Run_Queue", prompt="PROMPT_COUNCIL_CAPITAL")
    if cat.startswith(GC_CAT):
        return _mk("not_automation_ready",
                   "GC invites: estimating@example.com auto-forward (Task D)", None,
                   sr_row, route="Active_Tenders", prompt="PROMPT_ACTIVE_TENDER_PARSE")
    if cat.startswith(NEWS_CAT):
        return _mk("manual_p3_only", "manual/LLM news scan (low priority)", None,
                   sr_row, route="Run_Queue", prompt="PROMPT_WEB_PDF_EXTRACT")
    # municipal dev-app source with no connector
    return _mk("missing_connector",
               "add endpoint to connector CSV, then probe", None, sr_row,
               route="Run_Queue", prompt="PROMPT_WEB_PDF_EXTRACT")


def _mk(status, next_action, connector, sr_row, route=None, prompt=None):
    fetch = connector.get("fetch_type") if connector else ""
    endpoint = (connector.get("last_good_endpoint") or connector.get("endpoint")) if connector else ""
    layer = connector.get("layer_index") if connector else ""
    auto = (connector.get("automation_feasibility") if connector
            else sr_row.get("Automation Feasibility")) or ""
    out_route = route or (connector.get("output_route") if connector else "") or ""
    ptype = prompt or (connector.get("prompt_type") if connector else "") or ""
    return {
        "source_register_status": sr_row.get("Access Status / Confidence") or "",
        "connector_registry_status": status,
        "fetch_type": fetch, "endpoint": endpoint, "layer": layer,
        "automation_feasibility": auto, "access_status": status,
        "output_route": out_route, "prompt_type": ptype,
        "next_action": next_action,
    }


def build_sync_report(master_path, csv_path):
    """Return (rows, summary) where rows match the handoff section-8 schema."""
    connectors = load_connectors(csv_path)
    sources = load_source_register(master_path)
    used_connectors = set()
    rows = []
    for sr in sources:
        sid = match_connector(sr.get("Source Name"), connectors)
        conn = connectors.get(sid) if sid else None
        if sid:
            used_connectors.add(sid)
        cls = classify_source(sr, conn)
        rows.append({
            "source_id": sid or _slug(sr.get("Source Name")),
            "source_name": sr.get("Source Name"),
            "category": sr.get("Category"),
            "priority_tier": sr.get("Priority Tier"),
            **cls,
            "notes": sr.get("Notes / Cleanup") or "",
        })
    # connectors not present in Source_Register
    for sid, conn in connectors.items():
        if sid not in used_connectors:
            rows.append({
                "source_id": sid, "source_name": conn.get("name"),
                "category": conn.get("category"),
                "priority_tier": conn.get("priority_tier"),
                "source_register_status": "NOT IN SOURCE_REGISTER",
                "connector_registry_status": conn.get("access_status") or "ready_for_probe",
                "fetch_type": conn.get("fetch_type"),
                "endpoint": conn.get("last_good_endpoint") or conn.get("endpoint"),
                "layer": conn.get("layer_index"),
                "automation_feasibility": conn.get("automation_feasibility"),
                "access_status": conn.get("access_status"),
                "output_route": conn.get("output_route"),
                "prompt_type": conn.get("prompt_type"),
                "next_action": "connector present but no business source row — review",
                "notes": conn.get("notes") or "",
            })
    summary = {}
    for r in rows:
        s = r["connector_registry_status"]
        summary[s] = summary.get(s, 0) + 1
    return rows, summary


def _slug(name):
    return re.sub(r"[^a-z0-9]+", "_", _norm(name)).strip("_")[:40]


SYNC_COLUMNS = [
    "source_id", "source_name", "category", "priority_tier",
    "source_register_status", "connector_registry_status", "fetch_type",
    "endpoint", "layer", "automation_feasibility", "access_status",
    "output_route", "prompt_type", "next_action", "notes",
]


def write_sync_report_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=SYNC_COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    column: safe_untrusted_excel_value(r.get(column, ""))
                    for column in SYNC_COLUMNS
                }
            )


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Reconcile a historical Source_Register sheet with the canonical source registry")
    ap.add_argument("--from-master", required=True)
    ap.add_argument("--config", default=str(resolve_registry_path()))
    ap.add_argument("--out", default="sync_report.csv")
    a = ap.parse_args()
    rows, summary = build_sync_report(a.from_master, a.config)
    write_sync_report_csv(rows, a.out)
    print(f"Sync report: {len(rows)} sources -> {a.out}")
    for k in sorted(summary):
        print(f"  {k:<24} {summary[k]}")
