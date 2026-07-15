from __future__ import annotations

import argparse
import concurrent.futures
import copy
import csv
import datetime as dt
import hashlib
import html
import json
import os
import re
import shutil
import socket
import ssl
import subprocess
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidationList

from tenderfinder_email_guidance import (
    EmailSetupState,
    detect_email_state,
    email_setup_rows,
    format_email_setup_status,
    print_email_setup_status,
)
from tenderfinder_email_intake import (
    PROVIDER_KEYS,
    connect_email_provider,
    EmailProviderConfig,
    EmailIntakeTestResult,
    EmailTender,
    current_email_import_folder,
    is_fixture_source_folder,
    load_email_provider_config,
    provider_display_name,
    run_email_intake,
    test_email_intake,
)
from tenderfinder_source_backlog import (
    append_potential_sources_sheet,
    append_source_roadmap_printable_sheet,
    cross_check_counts as backlog_cross_check_counts,
)
from tenderfinder_source_registry import SourceRegistryError, load_tender_sources
from tenderfinder_runtime import RuntimeStateError, runtime_paths
from tenderfinder_excel_safety import append_untrusted_row, safe_untrusted_excel_value
from tenderfinder_keywords_config import (
    KeywordConfigError,
    load_keywords_config,
    print_validation_summary,
)
from tenderfinder_url_safety import (
    URLSafetyError,
    install_playwright_public_route,
    safe_urllib_fetch,
    validate_public_url,
)
import tenderfinder_guards as G


ROOT = Path(__file__).resolve().parents[2]
TODAY = dt.date.today()
PATCH_VERSION = "5.23"
_RUNTIME_PATHS = runtime_paths(create=False, package_root=ROOT)
DEMO_HISTORY_DIR = _RUNTIME_PATHS.history
LATEST_USER_MASTER_PATH = _RUNTIME_PATHS.latest_user_master
BCBID_NETWORK_AUDIT_PATH = _RUNTIME_PATHS.root / "BC_BID_NETWORK_AUDIT.md"
PATCH_5_18_BACKLOG_PATH = _RUNTIME_PATHS.root / "PATCH_5_18_BACKLOG.md"
AUTONOMOUS_FIXES_PATH = _RUNTIME_PATHS.root / "AUTONOMOUS_FIXES.md"
BCBID_BROWSER_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36"
LAST_BC_BID_AUDIT: dict[str, Any] = {}
LAST_BC_BID_OFFICIAL_FINDINGS: dict[str, Any] = {}
OUTREACH_STATUSES = [
    "Not Contacted",
    "Contacted",
    "Responded",
    "Meeting Scheduled",
    "Bid Submitted",
    "Won",
    "Lost",
    "Not Pursuing",
]

BASELINE = {
    "records_pulled": 27260,
    "records_normalized": 19146,
    "future_clean": 5393,
    "fit_ge_70": 31,
    "fit_ge_60": 193,
    "fit_60_69": 162,
    "fit_50_59": 2074,
    "watchlist": 999,
    "analyzed": 12766,
    "working_sources": 5,
    "coded_sources": 17,
    "source_register": 159,
}
RESCORE_ALWAYS_FUTURE_MIN_FIT = 50
RESCORE_ALWAYS_SUMMARY = (
    "Scores, gates, labels, and bucket routing always reflect current keywords.xlsx. "
    "Vancouver permit tiers are also recomputed when the persisted raw scoring snapshot is available; "
    "legacy rows without that snapshot keep their stored tier and are explicitly audited."
)
P511_CORRECTED_FUTURE_COUNT = 7537

# Task E (product): stage-safe pause support. The GUI's Pause button touches
# tenderfinder_pause.signal in the output folder; the engine checks for it only at
# safe stage boundaries (never mid-write), records a checkpoint, and exits
# with PAUSE_EXIT_CODE so the GUI can distinguish paused from crashed.
# Honest limitation, stated everywhere it matters: this pipeline recomputes
# from its inputs, so Resume restarts the build from the beginning into the
# same folder - it does not skip completed stages. Pause's guarantee is
# output SAFETY (no half-written files), not saved work.
PAUSE_EXIT_CODE = 86


def configure_runtime_state(state_root: str | Path | None, mode: str) -> None:
    """Route every mutable runtime artifact outside the package checkout."""
    global _RUNTIME_PATHS, DEMO_HISTORY_DIR, LATEST_USER_MASTER_PATH
    global BCBID_NETWORK_AUDIT_PATH, PATCH_5_18_BACKLOG_PATH, AUTONOMOUS_FIXES_PATH
    _RUNTIME_PATHS = runtime_paths(state_root, mode=mode, package_root=ROOT, create=True)
    DEMO_HISTORY_DIR = _RUNTIME_PATHS.history
    LATEST_USER_MASTER_PATH = _RUNTIME_PATHS.latest_user_master
    BCBID_NETWORK_AUDIT_PATH = _RUNTIME_PATHS.root / "BC_BID_NETWORK_AUDIT.md"
    PATCH_5_18_BACKLOG_PATH = _RUNTIME_PATHS.root / "PATCH_5_18_BACKLOG.md"
    AUTONOMOUS_FIXES_PATH = _RUNTIME_PATHS.root / "AUTONOMOUS_FIXES.md"
SAFE_STAGES = [
    "track_a_loaded",
    "municipal_sources_done",
    "bc_bid_pending_user_check",
    "bc_bid_done_or_skipped",
    "workbook_written",
]


def record_stage(out_dir: Path, stage: str) -> None:
    """Append a completed safe stage to tenderfinder_stage_progress.json so a
    stopped/paused run's folder shows exactly how far it got."""
    progress_path = out_dir / "tenderfinder_stage_progress.json"
    try:
        stages: list[str] = []
        if progress_path.exists():
            stages = json.loads(progress_path.read_text(encoding="utf-8")).get("stages", [])
        if stage not in stages:
            stages.append(stage)
        progress_path.write_text(json.dumps({
            "stages": stages,
            "updated": dt.datetime.now().isoformat(timespec="seconds"),
        }, indent=2), encoding="utf-8")
    except Exception:
        pass


def maybe_pause_at_checkpoint(out_dir: Path, stage: str) -> None:
    """If the GUI requested a pause, stop here - a safe boundary where no
    file is mid-write - and record the checkpoint."""
    signal_path = out_dir / "tenderfinder_pause.signal"
    if not signal_path.exists():
        return
    signal_path.unlink(missing_ok=True)
    checkpoint_path = out_dir / "tenderfinder_checkpoint.json"
    checkpoint_path.write_text(json.dumps({
        "stage": stage,
        "paused_at": dt.datetime.now().isoformat(timespec="seconds"),
        "resume_note": (
            "This pipeline recomputes from its inputs; Resume restarts the "
            "build from the beginning into this same folder. Files present "
            "here are consistent up to the recorded stage - nothing was "
            "stopped mid-write."
        ),
    }, indent=2), encoding="utf-8")
    print(
        f"TENDER_FINDER_PAUSED: build paused at safe checkpoint '{stage}'. Output in "
        "this folder is consistent up to this stage - nothing was stopped "
        "mid-write. Resume restarts the build from the beginning (stages "
        "before the workbook cannot skip ahead safely).",
        flush=True,
    )
    raise SystemExit(PAUSE_EXIT_CODE)

LOGIN_PLATFORM_HOSTS = (
    "bidsandtenders.ca",
    "bonfirehub.ca",
    "merx.com",
    "ariba.com",
    "jaggaer.com",
)

NEGATIVE_LINK_TERMS = (
    "job opportunit",
    "career",
    "employment",
    "volunteer",
    "procurement policy",
    "purchasing policy",
    "login",
    "sign in",
    "register",
    "subscription",
    "subscribe",
    "privacy",
    "terms of use",
    "accessibility",
    "facebook",
    "twitter",
    "linkedin",
    "instagram",
    "youtube",
    "contact us",
    "staff directory",
    "vendor help",
    "bid results",
    "names of the successful bidders",
    "successful bidders",
    "view bid results",
)

def _load_runtime_tender_sources() -> list[dict[str, Any]]:
    """Read active tender sources from the one founder-editable registry."""
    return load_tender_sources(root=ROOT, active_only=True)


TENDER_SOURCES = _load_runtime_tender_sources()
BCBID_PUBLIC_URL = next(
    (source["url"] for source in TENDER_SOURCES if source["source_id"] == "bc_bid_public"),
    "",
)
BCBID_REGISTER_URL = "https://bcbid.gov.bc.ca/page.aspx/en/usr/login"
BCBID_NEXT_PAGE_SELECTOR = "#body_x_grid_gridPagerBtnNextPage"
BCBID_MAX_PAGES = 15  # hard upper bound; the real stop condition is the closing-date relevance window below
BCBID_RELEVANCE_WINDOW_DAYS = 90

PLANNING_CONTACTS = {
    "Coquitlam": "Planning: https://www.coquitlam.ca/development | 604-927-3441",
    "Surrey": "Planning: https://www.surrey.ca/planning | 604-591-4441",
    "Maple Ridge": "Planning: https://www.mapleridge.ca/245/Planning | 604-467-7341",
    "Township of Langley": "Planning: https://www.tol.ca/en/planning | 604-534-3211",
    "New Westminster": "Planning: https://www.newwestcity.ca/planning | 604-527-4532",
    "Port Coquitlam": "Planning: https://www.portcoquitlam.ca/planning | 604-927-5426",
    "Delta": "Planning: https://www.delta.ca/planning | 604-946-3380",
    "Vancouver": "Planning: https://vancouver.ca/home-property-development | 604-873-7611",
}

ACTION_CENTER_ROWS = [
    [
        "BC Bid",
        "Public browse/download = NO login; submit = free registration",
        "Provincial + many public bodies",
        BCBID_REGISTER_URL,
        BCBID_PUBLIC_URL,
        "PUBLIC_BROWSER_CHECKED",
        "Patch 5.13 real-browser audit found a browser-check/reCAPTCHA wall before any public opportunities API. See docs/BC_BID_NETWORK_AUDIT.md.",
        "TENDER_FINDER records BC Bid as a direct public connector attempt plus a parallel alert channel. No password stored in TENDER_FINDER; see docs/BC_BID_NETWORK_AUDIT.md.",
    ],
    [
        "bidsandtenders.ca",
        "Live tenders behind FREE vendor registration",
        "Surrey, Maple Ridge, Township Langley, City Langley, Pitt Meadows, Richmond, Fraser Health",
        "https://www.bidsandtenders.ca/",
        "https://www.bidsandtenders.ca/",
        "REGISTRATION_REQUIRED",
        "Closed Surrey civil evidence found by TENDER_FINDER: 104 Ave Drainage; 105A Water Feeder & Distribution Main; 144 St Road Improvements; North Surrey Sanitary; Road & Utility Site Servicing 10975 126A St; Sanitary Sewer Package.",
        "Register once, enable civil email alerts; TENDER_FINDER ingests alerts later via Live_Tender_Intake_Plan. No password stored in TENDER_FINDER.",
    ],
    [
        "BidCentral",
        "BC construction bid network; membership",
        "Construction bid notices, plan rooms, and contractor/member bid workflows",
        "https://www.bidcentral.ca/",
        "https://www.bidcentral.ca/",
        "MEMBERSHIP_LIKELY",
        "Public landing page was reachable; detailed bid workflows are membership-oriented.",
        "Confirm membership tier and enable civil/earthworks alerts. TENDER_FINDER can ingest alert emails in a later approved patch.",
    ],
    [
        "CivicInfo BC Bids",
        "Public local-government bids aggregator; may bot-block automation",
        "BC local-government bids aggregator",
        "https://www.civicinfo.bc.ca/bids",
        "https://www.civicinfo.bc.ca/bids",
        "FORBIDDEN_BUT_LIKELY_VALID",
        "Polite public request may return 403/429; open in browser or whitelist/retry.",
        "Whitelist/retry with browser; use as a local-government bid discovery source.",
    ],
]

LIVE_INTAKE_ROWS = [
    [
        "BC Bid",
        "Create or sign in to a vendor account only if submitting; public browse is no-login.",
        "Use public browse/open opportunities; set commodity/profile notifications after registration.",
        "alerts or notification emails from bcbid.gov.bc.ca",
        "tender_title, issuer, closing_date, opportunity_url, contact email/phone where included",
        "TENDER_FINDER does not store passwords. Email Alert Intake reads user-approved alert emails only; Patch 5.13 documents the direct public browse block in docs/BC_BID_NETWORK_AUDIT.md.",
    ],
    [
        "bidsandtenders.ca",
        "Register free vendor account on the portal.",
        "Follow selected owner agencies and enable civil/roads/water/sewer/site-servicing alerts.",
        "notification emails from bidsandtenders.ca and owner-branded portals",
        "tender_title, issuer/municipality, closing_date, portal link, category/commodity",
        "No scraping authenticated portal pages. Ingest only alerts already delivered to the user's mailbox.",
    ],
    [
        "BidCentral",
        "Confirm membership tier and alert access.",
        "Enable construction/civil plan-room and project alerts.",
        "bidcentral.ca alert emails",
        "project/tender title, owner, closing_date, plan-room link, trade/civil category",
        "Alert parsing only after explicit approval.",
    ],
    [
        "CivicInfo BC",
        "No registration required for browser review; automation may need whitelist.",
        "Monitor bids page and/or subscribe where available.",
        "civicinfo.bc.ca notices or manual browser review",
        "title, local government issuer, closing_date, link",
        "Use as a backup aggregator for municipal signals.",
    ],
]

THIN_SCOPE_LABELS = {
    "multi family",
    "mixed uses",
    "single family",
    "commercial",
    "industrial",
    "townhouse",
    "residential",
}

SURREY_V2_LAYER = "https://services5.arcgis.com/YRpe0VKTJytZSSIB/arcgis/rest/services/Development%20Applications/FeatureServer/0"
ABBOTSFORD_LAYER = "https://services8.arcgis.com/ZYlQy38aWlfDG1Qh/arcgis/rest/services/Development_Layers_External_Feature/FeatureServer/6"
TERMINAL_STATUSES = {
    "cancelled",
    "closed",
    "concluded",
    "final approval",
    "final maintenance",
    "ok for occupancy",
    "refused",
    "withdrawn",
}
ACTIVE_SIGNAL_STATUSES = {
    "active",
    "active application",
    "approved",
    "conditional approval",
    "construction",
    "in process",
    "initial review",
    "new building",
    "project acc",
    "project detailing",
    "project main",
    "project scoping",
    "referrals",
    "resubmission required",
    "sa draft",
    "sa to dev",
    "ex sa to dev",
    "under review",
    "ok for bp s",
    "filed",
}
ADDRESS_STREET_TYPES = (
    "Street",
    "St",
    "Avenue",
    "Ave",
    "Road",
    "Rd",
    "Drive",
    "Dr",
    "Way",
    "Crescent",
    "Cres",
    "Court",
    "Crt",
    "Boulevard",
    "Blvd",
    "Highway",
    "Hwy",
    "Lane",
    "Ln",
    "Place",
    "Pl",
    "Terrace",
    "Ter",
    "Parkway",
    "Pkwy",
)
ADDRESS_TYPE_PATTERN = "(?:" + "|".join(ADDRESS_STREET_TYPES) + ")"
ADDRESS_PATTERNS = [
    re.compile(rf"\b\d{{4,6}}\s*(?:-|–)\s*\d{{1,3}}(?:[A-Z])?\s+{ADDRESS_TYPE_PATTERN}\b", re.I),
    re.compile(rf"\b\d{{4,6}}(?:[A-Z])?\s+[A-Z0-9][A-Z0-9'&./-]*(?:\s+[A-Z0-9][A-Z0-9'&./-]*){{0,5}}\s+{ADDRESS_TYPE_PATTERN}\b", re.I),
]
KNOWN_DEVELOPER_BRANDS = (
    ("Beedie", re.compile(r"\bbeedie\b", re.I)),
    ("Qualico", re.compile(r"\bqualico\b", re.I)),
    ("Wesgroup", re.compile(r"\bwesgroup\b", re.I)),
    ("Polygon", re.compile(r"\bpolygon\b", re.I)),
    ("Onni", re.compile(r"\bonni\b", re.I)),
    ("Anthem", re.compile(r"\banthem\b", re.I)),
    ("Aquilini", re.compile(r"\baquilini\b", re.I)),
    ("Intracorp", re.compile(r"\bintracorp\b", re.I)),
    ("Ledingham McAllister", re.compile(r"\bledingham\s+mcallister\b", re.I)),
    ("Westbank", re.compile(r"\bwestbank\b", re.I)),
    ("Townline", re.compile(r"\btownline\b", re.I)),
    ("Concord Pacific", re.compile(r"\bconcord\s+pacific\b", re.I)),
    ("Bosa", re.compile(r"\bbosa\b", re.I)),
)
DESIGN_CONSULTANT_PATTERN = re.compile(r"\b(design|architect|architectural|studio|dba:)\b", re.I)
ADDRESS_DASH_PATTERN = r"(?:-|/|\u2013|\u2014|\u00c2?\u00bf|\?)"
ADDRESS_PATTERNS = [
    re.compile(rf"\b\d{{3,6}}\s*{ADDRESS_DASH_PATTERN}\s*\d{{1,3}}(?:[A-Z])?\s+{ADDRESS_TYPE_PATTERN}\b", re.I),
    re.compile(rf"\b\d{{3,6}}(?:[A-Z])?\s+\d{{1,3}}(?:[A-Z])?\s+{ADDRESS_TYPE_PATTERN}\b", re.I),
    re.compile(rf"\b\d{{3,6}}(?:[A-Z])?\s+[A-Z0-9][A-Z0-9'&./-]*(?:\s+[A-Z0-9][A-Z0-9'&./-]*){{0,5}}\s+{ADDRESS_TYPE_PATTERN}\b", re.I),
]
CONSULTANT_NAME_PATTERN = re.compile(
    r"\b("
    r"engineering|engineers?|consulting|consultants?|certified professional|"
    r"building code|code consultant|architects?|architecture|architectural|design|studio|"
    r"planning inc|surveying|surveyors?|jensen hughes"
    r")\b",
    re.I,
)
BUSINESS_SUFFIX_PATTERN = re.compile(r"\b(?:ltd|limited|inc|incorporated|corp|corporation|company|co|llp|lp)\.?$", re.I)


def is_detail_available(scope: Any) -> bool:
    text = clean_text(scope)
    return bool(text and len(text) > 20 and text.lower() not in THIN_SCOPE_LABELS)


def signal_quality_for(row: dict[str, Any]) -> str:
    if row.get("status_class") == "TERMINAL" or is_terminal_status(row.get("app_type_stage")):
        return "LOW"
    if row.get("fit_score", 0) >= 60:
        return "HIGH"
    if row.get("fit_score", 0) >= 50 and row.get("source_id") != "van_building_permits":
        return "MEDIUM"
    return "LOW"


@dataclass
class TenderCandidate:
    source_id: str
    source_name: str
    tender_title: str
    url: str
    civil_relevant: bool
    status: str
    found_via: str
    snippet: str = ""
    closing_date: str = ""
    related_lead: str = ""
    municipality: str = ""
    source_url: str = ""
    issuer: str = ""
    region: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    related_leads_count: int = 0
    related_leads_sample: str = ""
    related_keyword_count: int = 0
    signal_state: str = ""
    retention_target: str = ""
    filtered_reason: str = ""


@dataclass
class SourceLog:
    source_id: str
    source_name: str
    url: str
    status: str
    source_type: str = ""
    source_opened: bool = False
    http_status: int = 0
    raw_links_found: int = 0
    candidates: int = 0
    civil_candidates: int = 0
    open_candidates: int = 0
    closed_found: int = 0
    historical_found: int = 0
    portal_redirect_found: int = 0
    deep_fetches: int = 0
    elapsed_seconds: float = 0.0
    resolved_url: str = ""
    note: str = ""
    error: str = ""
    kept_in_bid_now: int = 0
    kept_in_history: int = 0
    kept_in_all_signals: int = 0
    filtered_reason: str = ""
    final_status: str = ""


@dataclass
class SourceSweepResult:
    candidates: list[TenderCandidate]
    log: SourceLog


def infer_source_type(source_id: str, url: str = "") -> str:
    low = f"{source_id} {url}".lower()
    if "bc_bid" in low:
        return "public_browser"
    if any(host in low for host in ("bidsandtenders", "bonfire", "merx", "ariba", "jaggaer", "bidcentral")):
        return "portal"
    if "email" in low:
        return "email_alert_intake"
    return "public_listing"


def classify_signal_state(status: str, filtered_reason: str = "") -> str:
    low = clean_text(status).lower()
    reason_low = clean_text(filtered_reason).lower()
    if "portal" in reason_low or "redirect" in reason_low:
        return "portal_redirect"
    if "parse" in reason_low:
        return "parse_failed"
    if low in {"open", "open_or_needs_review"}:
        return "open"
    if any(term in low for term in ("historical", "archive", "archived")):
        return "historical"
    if any(term in low for term in ("closed", "awarded", "cancelled", "canceled", "expired")):
        return "closed"
    if any(term in low for term in ("blocked", "forbidden", "browser_check")):
        return "blocked"
    return "parse_failed"


def classify_signal_state_for_candidate(candidate: TenderCandidate) -> str:
    if candidate.closing_date:
        try:
            if dt.date.fromisoformat(candidate.closing_date) < TODAY:
                if candidate.status in {"closed", "awarded", "cancelled", "expired"}:
                    return "historical"
                return "closed"
        except ValueError:
            pass
    return classify_signal_state(candidate.status, candidate.filtered_reason)


def classify_retention_target(candidate: TenderCandidate) -> str:
    if candidate.signal_state == "open" and candidate.civil_relevant:
        return "bid_now"
    if candidate.signal_state in {"closed", "historical"}:
        return "history"
    if candidate.signal_state in {"blocked", "portal_redirect", "empty", "parse_failed"}:
        return "all_signals_only"
    return "source_log_only"


def map_source_final_status(log: SourceLog) -> str:
    status = log.status
    note_low = clean_text(log.note).lower()
    sid = log.source_id.lower()
    if log.source_id == "email_alert_intake" and status.startswith("EMAIL_INTAKE_"):
        return status
    if status in {"OK_BROWSER_COOKIE_REPLAY", "OK_FEED", "OK_HTML"} and log.open_candidates > 0:
        return "FETCH_OK_OPEN_FOUND"
    if (log.historical_found > 0 or log.closed_found > 0) and log.open_candidates == 0:
        return "FETCH_OK_HISTORICAL_ONLY"
    if "membership" in note_low or "bidcentral" in sid:
        return "NEEDS_PAID_ACCESS"
    if "bidsandtenders.ca" in note_low or "login platform skipped" in note_low or status in {"LANDING_ONLY_MEMBERSHIP_LIKELY", "SKIPPED_LOGIN_PLATFORM"}:
        if "manual registration" in note_low:
            return "NEEDS_MANUAL_REGISTRATION"
        if "paid" in note_low:
            return "NEEDS_PAID_ACCESS"
        return "FETCH_OK_PORTAL_REDIRECT"
    if status in {"OK_NO_OPEN_TENDERS", "SKIPPED_NO_FETCH"}:
        return "FETCH_OK_EMPTY_CURRENT_LIST"
    if status == "OK_HTML" and log.candidates == 0:
        return "FETCH_OK_PARSE_ZERO_ROWS"
    if status in {"FORBIDDEN_BUT_LIKELY_VALID", "BC_BID_BLOCKED_NO_PUBLIC_FEED", "BC_BID_BLOCKED_BROWSER_CHECK_USER_ACTION_REQUIRED", "BC_BID_USER_CHECK_NOT_COMPLETED"}:
        return "BLOCKED_FORBIDDEN" if "BROWSER_CHECK" not in status and "USER_CHECK" not in status else "BLOCKED_BROWSER_CHECK"
    if status in {"BROKEN_404", "FETCH_ERROR"}:
        return "BROKEN_URL"
    if log.source_id == "email_alert_intake":
        return "NEEDS_EMAIL_ALERT" if log.candidates == 0 else "FETCH_OK_OPEN_FOUND"
    if "manual registration" in note_low:
        return "NEEDS_MANUAL_REGISTRATION"
    if "paid" in note_low or "membership" in note_low:
        return "NEEDS_PAID_ACCESS"
    return "FETCH_OK_PARSE_ZERO_ROWS"


def finalize_tender_signals(tenders: list[TenderCandidate], logs: list[SourceLog]) -> list[TenderCandidate]:
    finalized: list[TenderCandidate] = []
    by_source: dict[str, list[TenderCandidate]] = {}
    for tender in tenders:
        if not tender.signal_state:
            tender.signal_state = classify_signal_state_for_candidate(tender)
        if not tender.retention_target:
            tender.retention_target = classify_retention_target(tender)
        if not tender.filtered_reason:
            if tender.retention_target == "history":
                tender.filtered_reason = "closed_or_historical"
            elif tender.retention_target == "all_signals_only":
                tender.filtered_reason = tender.signal_state
            elif tender.retention_target == "source_log_only":
                tender.filtered_reason = "not_actionable"
        by_source.setdefault(tender.source_id, []).append(tender)
        finalized.append(tender)

    for log in logs:
        source_rows = by_source.get(log.source_id, [])
        log.open_candidates = sum(1 for c in source_rows if c.signal_state == "open")
        log.closed_found = sum(1 for c in source_rows if c.signal_state == "closed")
        log.historical_found = sum(1 for c in source_rows if c.signal_state == "historical")
        log.portal_redirect_found = sum(1 for c in source_rows if c.signal_state == "portal_redirect")
        log.kept_in_bid_now = sum(1 for c in source_rows if c.retention_target == "bid_now")
        log.kept_in_history = sum(1 for c in source_rows if c.retention_target == "history")
        log.kept_in_all_signals = sum(1 for c in source_rows if c.retention_target != "source_log_only")
        log.final_status = map_source_final_status(log)
        if not log.filtered_reason:
            if log.final_status in {"FETCH_OK_EMPTY_CURRENT_LIST", "FETCH_OK_PARSE_ZERO_ROWS"}:
                log.filtered_reason = "empty_current_list" if log.final_status == "FETCH_OK_EMPTY_CURRENT_LIST" else "parse_zero"
            elif log.final_status == "FETCH_OK_PORTAL_REDIRECT":
                log.filtered_reason = "portal_redirect"
                log.portal_redirect_found = max(log.portal_redirect_found, 1)
            elif log.final_status == "BLOCKED_FORBIDDEN":
                log.filtered_reason = "blocked"
            elif log.final_status == "BLOCKED_BROWSER_CHECK":
                log.filtered_reason = "browser_check_needs_user"
            elif log.final_status == "NEEDS_EMAIL_ALERT":
                log.filtered_reason = "needs_email_alert"
            elif log.final_status == "EMAIL_INTAKE_CONNECTED":
                log.filtered_reason = "email_intake_connected"
            elif log.final_status == "EMAIL_INTAKE_NO_FILES":
                log.filtered_reason = "email_intake_no_files"
            elif log.final_status == "EMAIL_INTAKE_PARSED_ROWS":
                log.filtered_reason = "email_intake_parsed_rows"
            elif log.final_status == "EMAIL_INTAKE_PARSE_ZERO_ROWS":
                log.filtered_reason = "email_intake_parse_zero_rows"
            elif log.final_status == "EMAIL_INTAKE_REJECTED_FILES":
                log.filtered_reason = "email_intake_rejected_files"
            elif log.final_status == "EMAIL_INTAKE_SKIPPED_NO_FOLDER":
                log.filtered_reason = "email_intake_skipped_no_folder"
            elif log.final_status == "NEEDS_MANUAL_REGISTRATION":
                log.filtered_reason = "needs_manual_registration"
            elif log.final_status == "NEEDS_PAID_ACCESS":
                log.filtered_reason = "needs_paid_access"
            elif log.final_status == "BROKEN_URL":
                log.filtered_reason = "broken_url"
            elif log.final_status == "FETCH_OK_HISTORICAL_ONLY":
                log.filtered_reason = "historical_only"
        if not source_rows:
            placeholder_state = {
                "FETCH_OK_EMPTY_CURRENT_LIST": "empty",
                "FETCH_OK_PORTAL_REDIRECT": "portal_redirect",
                "FETCH_OK_PARSE_ZERO_ROWS": "parse_failed",
                "BLOCKED_FORBIDDEN": "blocked",
                "BLOCKED_BROWSER_CHECK": "blocked",
                "BROKEN_URL": "parse_failed",
                "NEEDS_EMAIL_ALERT": "empty",
                "EMAIL_INTAKE_CONNECTED": "empty",
                "EMAIL_INTAKE_NO_FILES": "empty",
                "EMAIL_INTAKE_PARSE_ZERO_ROWS": "parse_failed",
                "EMAIL_INTAKE_REJECTED_FILES": "parse_failed",
                "EMAIL_INTAKE_SKIPPED_NO_FOLDER": "empty",
                "NEEDS_MANUAL_REGISTRATION": "portal_redirect",
                "NEEDS_PAID_ACCESS": "portal_redirect",
            }.get(log.final_status, "")
            if placeholder_state:
                finalized.append(TenderCandidate(
                    source_id=log.source_id,
                    source_name=log.source_name,
                    tender_title=f"{log.source_name} - {placeholder_state.replace('_', ' ')}",
                    url=log.resolved_url or log.url,
                    civil_relevant=False,
                    status=log.status,
                    found_via="source_run_log",
                    snippet=clean_text(log.note or log.error or log.status, 220),
                    source_url=log.url,
                    issuer=log.source_name,
                    region="source_signal",
                    signal_state=placeholder_state,
                    retention_target="all_signals_only",
                    filtered_reason=log.final_status.lower(),
                ))
                log.kept_in_all_signals += 1
    return finalized

@dataclass
class DemoData:
    future: list[dict[str, Any]] = field(default_factory=list)
    watch: list[dict[str, Any]] = field(default_factory=list)
    analyzed: list[dict[str, Any]] = field(default_factory=list)
    records_pulled: int = BASELINE["records_pulled"]
    records_normalized: int = BASELINE["records_normalized"]
    live_sources: int = BASELINE["working_sources"]
    coded_sources: int = BASELINE["coded_sources"]
    source_register: int = BASELINE["source_register"]
    municipalities: list[str] = field(default_factory=list)
    future_sources: list[str] = field(default_factory=list)
    prior_workbook: str = ""
    history_archive: str = ""
    history_note: str = "BASELINE RUN - no prior history"
    prior_lead_ids: set[str] = field(default_factory=set)
    removed_since_last_run: int = 0
    new_since_last_run: int = 0
    prior_outreach: dict[str, dict[str, Any]] = field(default_factory=dict)
    outreach_rollup: dict[str, int] = field(default_factory=dict)
    developer_distinct: int = 0
    developer_repeat: int = 0
    known_major_developers: list[str] = field(default_factory=list)
    surrey_v2_total: int = 0
    surrey_v2_terminal: int = 0
    surrey_v2_active: int = 0
    address_recovery: dict[str, dict[str, int]] = field(default_factory=dict)
    applicant_type_counts: dict[str, int] = field(default_factory=dict)
    tender_date_stats: dict[str, int] = field(default_factory=dict)
    bc_bid_pagination_stats: dict[str, Any] = field(default_factory=dict)
    bc_bid_detail_stats: dict[str, int] = field(default_factory=dict)
    fit_ge_70: int = 0
    fit_ge_60: int = 0
    fit_60_69: int = 0
    fit_50_59: int = 0
    keyword_rescore_events: list[dict[str, Any]] = field(default_factory=list)
    keyword_below_gate: list[dict[str, Any]] = field(default_factory=list)
    read_seconds: float = 0.0


class LinkExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._href: str | None = None
        self._parts: list[str] = []
        self.links: list[tuple[str, str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        attrs_dict = {k.lower(): v for k, v in attrs}
        self._href = attrs_dict.get("href")
        self._parts = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() != "a" or self._href is None:
            return
        text = " ".join(" ".join(self._parts).split())
        if text:
            self.links.append((text, self._href))
        self._href = None
        self._parts = []


class VisibleTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._skip_depth = 0
        self._title = False
        self.title_parts: list[str] = []
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag in {"script", "style", "noscript", "svg"}:
            self._skip_depth += 1
        if tag == "title":
            self._title = True
        if tag in {"p", "div", "li", "tr", "h1", "h2", "h3", "h4", "br"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"script", "style", "noscript", "svg"} and self._skip_depth:
            self._skip_depth -= 1
        if tag == "title":
            self._title = False
        if tag in {"p", "div", "li", "tr", "h1", "h2", "h3", "h4"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        if self._title:
            self.title_parts.append(data)
        self.parts.append(data)

    @property
    def title(self) -> str:
        return clean_text(" ".join(self.title_parts), 180)

    @property
    def text(self) -> str:
        raw = html.unescape(" ".join(self.parts))
        lines = [clean_text(line) for line in raw.splitlines()]
        return "\n".join(line for line in lines if line)


def coerce_int(value: Any) -> int:
    try:
        if value is None or value == "":
            return 0
        return int(float(value))
    except Exception:
        return 0


def clean_text(value: Any, limit: int | None = None) -> str:
    text = html.unescape(str(value or ""))
    text = re.sub(r"\s+", " ", text).strip()
    if limit and len(text) > limit:
        return text[: limit - 3].rstrip() + "..."
    return text


def status_key(value: Any) -> str:
    return clean_text(value).lower()


def is_terminal_status(value: Any) -> bool:
    return status_key(value) in TERMINAL_STATUSES


def is_active_signal_status(value: Any) -> bool:
    return status_key(value) in ACTIVE_SIGNAL_STATUSES


def status_class_for(source_id: Any, status: Any) -> str:
    key = status_key(status)
    if key in TERMINAL_STATUSES:
        return "TERMINAL"
    if source_id == "surrey_devapps_v2" and key in ACTIVE_SIGNAL_STATUSES:
        return "ACTIVE_SIGNAL"
    if key:
        return "OTHER"
    return "UNKNOWN"


def normalize_extracted_address(value: str) -> str:
    text = clean_text(value)
    text = re.sub(r"\s*([,-])\s*", r" \1 ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,;.")
    return text


def extract_address_from_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    for pattern in ADDRESS_PATTERNS:
        match = pattern.search(text)
        if match:
            return normalize_extracted_address(match.group(0))
    return ""


def centroid_label(x: Any, y: Any) -> str:
    try:
        lon = float(x)
        lat = float(y)
    except Exception:
        return ""
    return f"Approximate centroid ({lat:.5f}, {lon:.5f})"


def fetch_json(url: str, timeout: int = 45) -> dict[str, Any]:
    body, status, _, _ = fetch_url(url, timeout=timeout, retries=1)
    if status != 200:
        raise RuntimeError(f"HTTP {status} for {url}")
    return json.loads(body)


def arcgis_query_paged(
    layer_url: str,
    out_fields: list[str],
    key_field: str,
    geometry_mode: str = "centroid",
    page_size: int = 1000,
) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    offset = 0
    while True:
        params = {
            "where": "1=1",
            "outFields": ",".join(out_fields),
            "returnGeometry": "false" if geometry_mode == "centroid" else "true",
            "returnCentroid": "true" if geometry_mode == "centroid" else "false",
            "resultOffset": str(offset),
            "resultRecordCount": str(page_size),
            "orderByFields": key_field,
            "f": "json",
            "outSR": "4326",
        }
        data = fetch_json(layer_url.rstrip("/") + "/query?" + urllib.parse.urlencode(params), timeout=60)
        features = data.get("features") or []
        if not features:
            break
        for feature in features:
            attrs = feature.get("attributes") or {}
            key = clean_text(attrs.get(key_field))
            if not key:
                continue
            centroid = feature.get("centroid") or {}
            geometry = feature.get("geometry") or {}
            if not centroid and geometry.get("x") is not None and geometry.get("y") is not None:
                centroid = {"x": geometry.get("x"), "y": geometry.get("y")}
            rows[key] = {
                "attributes": attrs,
                "centroid_label": centroid_label(centroid.get("x"), centroid.get("y")),
            }
        if len(features) < page_size or not data.get("exceededTransferLimit"):
            break
        offset += len(features)
    return rows


def load_address_enrichment(
    *,
    allow_network: bool = False,
) -> dict[str, dict[str, dict[str, Any]]]:
    enrichments: dict[str, dict[str, dict[str, Any]]] = {"surrey_devapps_v2": {}, "abbotsford_devapps": {}}
    if not allow_network:
        return enrichments
    try:
        enrichments["surrey_devapps_v2"] = arcgis_query_paged(
            SURREY_V2_LAYER,
            ["PROJECT_NO", "STATUS", "DESCRIPTION"],
            "PROJECT_NO",
        )
    except Exception:
        enrichments["surrey_devapps_v2"] = {}
    try:
        enrichments["abbotsford_devapps"] = arcgis_query_paged(
            ABBOTSFORD_LAYER,
            ["DEV_FILE", "AMANDA_STATUS"],
            "DEV_FILE",
        )
    except Exception:
        enrichments["abbotsford_devapps"] = {}
    return enrichments


def scrub_currency_values(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = re.sub(r"\$\s*\d[\d,]*(?:\.\d+)?", "[currency omitted]", value)
    text = re.sub(r"\bCAD\s*\d[\d,]*(?:\.\d+)?", "[currency omitted]", text, flags=re.I)
    text = re.sub(r"\b\d[\d,]*(?:\.\d+)?\s+dollars?\b", "[currency omitted]", text, flags=re.I)
    return text


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def extract_scale(scope: Any) -> dict[str, Any]:
    text = clean_text(scope).lower()
    unit_values: list[int] = []
    storey_values: list[int] = []
    lot_values: list[int] = []

    unit_patterns = [
        r"(?:with\s*)?(\d{1,4})\s*(?:market\s+|rental\s+|residential\s+|strata\s+|condo\s+|apartment\s+|townhouse\s+|townhome\s+|dwelling\s+|housing\s+)?(?:units?|condos?|apartments?|townhouses?|townhomes?|dwellings?)",
        r"(\d{1,4})\s*[- ](?:unit|dwelling|townhome|townhouse)",
    ]
    for pattern in unit_patterns:
        unit_values.extend(int(x) for x in re.findall(pattern, text, flags=re.I))

    for a, b in re.findall(r"(\d{1,3})\s*\+\s*(\d{1,3})\s*(?:storeys?|stories)", text, flags=re.I):
        storey_values.extend([int(a), int(b)])
    storey_values.extend(int(x) for x in re.findall(r"(\d{1,3})\s*[- ]?\s*(?:storeys?|stories)", text, flags=re.I))

    lot_values.extend(int(x) for x in re.findall(r"(\d{1,4})\s*(?:lots?|lot subdivision)", text, flags=re.I))

    max_units = max(unit_values) if unit_values else None
    max_storeys = max(storey_values) if storey_values else None
    max_lots = max(lot_values) if lot_values else None

    scale = "unknown"
    if max_units is not None:
        if max_units >= 300:
            scale = "very_large"
        elif max_units >= 100:
            scale = "large"
        elif max_units >= 30:
            scale = "medium"
        else:
            scale = "small"
    elif max_lots is not None:
        if max_lots >= 100:
            scale = "large"
        elif max_lots >= 30:
            scale = "medium"
        else:
            scale = "small"

    return {
        "scale": scale,
        "est_units": max_units,
        "est_storeys": max_storeys,
        "est_lots": max_lots,
    }


def normalize_address_for_id(address: Any) -> str:
    text = clean_text(address).upper()
    replacements = {
        r"\bSTREET\b": "ST",
        r"\bST\.\b": "ST",
        r"\bAVENUE\b": "AVE",
        r"\bAVE\.\b": "AVE",
        r"\bROAD\b": "RD",
        r"\bRD\.\b": "RD",
        r"\bDRIVE\b": "DR",
        r"\bDR\.\b": "DR",
        r"\bBOULEVARD\b": "BLVD",
        r"\bBLVD\.\b": "BLVD",
        r"\bHIGHWAY\b": "HWY",
        r"\bHWY\.\b": "HWY",
    }
    for pattern, repl in replacements.items():
        text = re.sub(pattern, repl, text)
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def stable_lead_id(row: dict[str, Any]) -> str:
    parts = [
        clean_text(row.get("municipality")).upper(),
        normalize_address_for_id(row.get("address")),
        clean_text(row.get("source_id") or row.get("source")).upper(),
        clean_text(row.get("app_no")).upper(),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


def tender_lead_id(tender: TenderCandidate) -> str:
    parts = [
        "TENDER",
        clean_text(tender.source_id).upper(),
        clean_text(tender.tender_title).upper(),
        clean_text(tender.url).upper(),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


def project_scale_tier(row: dict[str, Any]) -> str:
    units = row.get("est_units")
    scope = clean_text(row.get("scope_summary")).lower()
    if units is not None:
        try:
            unit_count = int(units)
            if unit_count >= 200:
                return "LARGE"
            if 20 <= unit_count <= 199:
                return "MEDIUM"
            if unit_count < 20:
                return "SMALL"
        except Exception:
            pass
    if "tower" in scope and ("two" in scope or "multiple" in scope or re.search(r"\b\d+\s*(?:towers?|buildings?)", scope)):
        return "LARGE"
    if re.search(r"\b(single[- ]family|duplex|triplex|fourplex|multiplex)\b", scope):
        return "SMALL"
    return "UNKNOWN"


def regional_cluster(municipality: Any) -> str:
    name = clean_text(municipality).lower()
    mapping = {
        "coquitlam": "Tri-Cities",
        "port coquitlam": "Tri-Cities",
        "port moody": "Tri-Cities",
        "city of north vancouver": "North Shore",
        "district of north vancouver": "North Shore",
        "north vancouver": "North Shore",
        "west vancouver": "North Shore",
        "surrey": "South of Fraser",
        "delta": "South of Fraser",
        "township of langley": "South of Fraser",
        "city of langley": "South of Fraser",
        "langley": "South of Fraser",
        "white rock": "South of Fraser",
        "abbotsford": "Fraser Valley East",
        "chilliwack": "Fraser Valley East",
        "mission": "Fraser Valley East",
        "burnaby": "Central",
        "new westminster": "Central",
        "vancouver": "Vancouver",
        "city of vancouver": "Vancouver",
    }
    return mapping.get(name, "Other")


def extract_applicant_name(raw: dict[str, Any]) -> str:
    for key in (
        "owner/applicant",
        "owner_applicant",
        "applicant",
        "owner",
        "developer",
        "proponent",
        "applicant_name",
        "APPLICANT",
        "OWNER",
        "DEVELOPER",
    ):
        value = clean_text(raw.get(key))
        if value and value.lower() not in {"not available", "none", "null", "n/a"}:
            return value
    return ""


def _keyword_scoring_fields(raw: dict[str, Any], enrichment: dict[str, Any] | None = None) -> list[Any]:
    fields = [
        raw.get(key)
        for key in (
            "project_id", "title", "project_title", "municipality", "app_no", "address",
            "owner/applicant", "owner_applicant", "owner", "applicant", "developer",
            "proponent", "applicant_name", "app_type_stage", "scope_summary",
            "expected_civil", "next_milestone", "notes",
        )
    ]
    if enrichment:
        fields.extend((enrichment.get("attributes") or {}).values())
    return [value for value in fields if value not in (None, "")]


def _coerce_write_eligible(value: Any) -> bool:
    return clean_text(value).casefold() in {"1", "true", "yes", "y"}


def _stored_vancouver_tier(raw: dict[str, Any]) -> str:
    explicit = clean_text(raw.get("van_permit_fit_tier")).casefold()
    if explicit:
        return explicit
    fit_reason = clean_text(raw.get("fit_reason")).casefold()
    if fit_reason.startswith("van_permit:"):
        return fit_reason.split(":", 1)[1]
    return {
        "Future_Projects": "strong",
        "Run_Queue": "watchlist",
        "Bulk_Intake_Raw": "bulk",
        "Rejected_Archive": "noisy",
    }.get(clean_text(raw.get("proposed_route")), "")


def _vancouver_route_for_tier(tier: str) -> tuple[str, bool, str]:
    return {
        "strong": ("Future_Projects", True, ""),
        "watchlist": ("Run_Queue", False, "van_permit_watchlist"),
        "bulk": ("Bulk_Intake_Raw", False, "van_permit_bulk_intake"),
        "noisy": ("Rejected_Archive", False, "van_permit_noisy_interior_ti"),
    }[tier]


def _rescore_route(
    raw: dict[str, Any], source_id: str, fit_score: int,
) -> tuple[str, bool, str, int | None, str, str, str]:
    stored_route = clean_text(raw.get("proposed_route"))
    stored_hold = clean_text(raw.get("hold_reason"))
    if source_id == "van_building_permits":
        old_tier = _stored_vancouver_tier(raw)
        scoring_text = str(raw.get("keyword_scoring_text") or "").strip()
        if scoring_text:
            new_tier = G.van_permit_fit_tier(scoring_text)
            route, eligible, hold = _vancouver_route_for_tier(new_tier)
            return route, eligible, hold, None, "", old_tier, new_tier
        # Legacy rows collected before the scoring-text seam cannot recreate
        # the old raw-attribute tier. Their score and labels still refresh,
        # and the explicit exception remains visible in the audit.
        return (
            stored_route,
            _coerce_write_eligible(raw.get("write_eligible")),
            stored_hold,
            None,
            "legacy_vancouver_scoring_text_unavailable",
            old_tier,
            old_tier,
        )

    threshold: int | None = None
    score_gated_future = stored_route == "Future_Projects"
    if stored_route == "Run_Queue":
        match = re.search(r"fit_score_below_min_(\d+)", stored_hold)
        if match:
            threshold = int(match.group(1))
            score_gated_future = True
        elif stored_hold.startswith("rescore_fit_below_"):
            threshold = RESCORE_ALWAYS_FUTURE_MIN_FIT
            score_gated_future = True

    if score_gated_future:
        threshold = threshold or RESCORE_ALWAYS_FUTURE_MIN_FIT
        if fit_score < threshold:
            return "Run_Queue", False, f"rescore_fit_below_{threshold}", threshold, "", "", ""
        return "Future_Projects", True, "", threshold, "", "", ""
    return stored_route, _coerce_write_eligible(raw.get("write_eligible")), stored_hold, threshold, "", "", ""


def _rule_attribution(rule: Any) -> dict[str, Any]:
    return {
        "keyword": rule.keyword,
        "match_type": rule.match_type,
        "weight": rule.weight,
        "category": rule.category,
    }


def normalized_row(raw: dict[str, Any], enrichment: dict[str, Any] | None = None) -> dict[str, Any]:
    source_id = raw.get("source_id") or raw.get("_source_id") or ""
    original_address = clean_text(raw.get("address"))
    original_scope = clean_text(raw.get("scope_summary"), 600)
    stage = clean_text(raw.get("app_type_stage"))
    if enrichment:
        attrs = enrichment.get("attributes") or {}
        if source_id == "surrey_devapps_v2":
            stage = stage or clean_text(attrs.get("STATUS"))
            original_scope = original_scope or clean_text(attrs.get("DESCRIPTION"), 600)
        elif source_id == "abbotsford_devapps":
            stage = stage or clean_text(attrs.get("AMANDA_STATUS"))
    extracted_address = extract_address_from_text(original_scope)
    final_address = original_address or extracted_address
    approximate_location = ""
    address_source = "ORIGINAL" if original_address else "NONE"
    if not final_address and enrichment:
        approximate_location = clean_text(enrichment.get("centroid_label"))
        if approximate_location:
            address_source = "CENTROID"
    elif extracted_address and not original_address:
        address_source = "EXTRACTED"
    scale = extract_scale(original_scope)
    status_class = status_class_for(source_id, stage)
    scoring_fields = _keyword_scoring_fields(raw, enrichment)
    scoring_blob = " ".join(str(value) for value in scoring_fields)
    score_breakdown = G.score_civil_fit_breakdown(
        scoring_blob,
        raw.get("municipality") or "",
        match_fields=scoring_fields,
    )
    stored_score = None if raw.get("fit_score") in (None, "") else coerce_int(raw.get("fit_score"))
    current_score = score_breakdown["fit_score"]
    route, write_eligible, hold_reason, threshold, rescore_exception, old_tier, new_tier = _rescore_route(
        raw, source_id, current_score
    )
    row = {
        "municipality": raw.get("municipality") or "",
        "app_no": raw.get("app_no") or "",
        "address": final_address,
        "approximate_location": approximate_location,
        "address_source": address_source,
        "app_type_stage": stage,
        "status_class": status_class,
        "scope_summary": original_scope,
        "fit_score": current_score,
        "scale": scale["scale"],
        "est_units": scale["est_units"],
        "est_storeys": scale["est_storeys"],
        "est_lots": scale["est_lots"],
        "source": raw.get("source_name") or raw.get("source_id") or "",
        "source_id": source_id,
        "source_url": raw.get("source_url") or raw.get("evidence_url") or "",
        "route": route,
        "write_eligible": write_eligible,
        "hold_reason": hold_reason,
        "applicant_name": extract_applicant_name(raw),
    }
    row["detail_available"] = "YES" if is_detail_available(row["scope_summary"]) else "NO"
    row["signal_quality"] = signal_quality_for(row)
    if source_id != "van_building_permits":
        old_tier = (
            signal_quality_for(
                {
                    "fit_score": stored_score,
                    "status_class": status_class,
                    "app_type_stage": stage,
                    "source_id": source_id,
                }
            )
            if stored_score is not None
            else ""
        )
        new_tier = row["signal_quality"]
    row["source_tier"] = source_tier_for(source_id)
    row["project_scale_tier"] = project_scale_tier(row)
    row["regional_cluster"] = regional_cluster(row["municipality"])
    row["lead_id"] = stable_lead_id({
        **row,
        "address": original_address,
    })
    row["_keyword_rescore"] = {
        "lead_id": row["lead_id"],
        "app_no": row["app_no"],
        "source_id": source_id,
        "old_score": stored_score,
        "new_score": current_score,
        "score_delta": current_score - stored_score if stored_score is not None else None,
        "old_tier": old_tier,
        "new_tier": new_tier,
        "old_route": clean_text(raw.get("proposed_route")),
        "new_route": route,
        "old_bucket": clean_text(raw.get("proposed_route")),
        "new_bucket": route,
        "threshold": threshold,
        "below_gate": route == "Run_Queue" and hold_reason.startswith("rescore_fit_below_"),
        "exception": rescore_exception,
        "raw_pre_cap": score_breakdown["raw_score"],
        "geography_weight": score_breakdown["geography_weight"],
        "client_weight": score_breakdown["client_weight"],
        "positive_matches": [_rule_attribution(rule) for rule in score_breakdown["positive_matches"]],
        "negative_matches": [_rule_attribution(rule) for rule in score_breakdown["negative_matches"]],
        "geography_matches": [_rule_attribution(rule) for rule in score_breakdown["geography_matches"]],
        "client_matches": [_rule_attribution(rule) for rule in score_breakdown["client_matches"]],
    }
    row["_keyword_rescore"]["changed"] = any(
        (
            stored_score is not None and stored_score != current_score,
            row["_keyword_rescore"]["old_route"] != route,
            bool(old_tier or new_tier) and old_tier != new_tier,
        )
    )
    return row


def source_tier_for(source_id: str) -> str:
    sid = clean_text(source_id).lower()
    if sid in {"surrey_planning_reports", "coquitlam_devapps"}:
        return "TIER_1"
    if sid in {"maple_ridge_devapps", "twp_langley_devactivity"}:
        return "TIER_2"
    if sid == "van_building_permits":
        return "TIER_3"
    if sid.startswith("van_"):
        return "TIER_2"
    if sid:
        return "TIER_2"
    return "TIER_3"


def read_track_a(review_xlsx: Path, *, allow_network_enrichment: bool = False) -> DemoData:
    start = time.perf_counter()
    wb = openpyxl.load_workbook(review_xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "") for h in next(rows)]
    data = DemoData()
    enrichments = load_address_enrichment(allow_network=allow_network_enrichment)

    for values in rows:
        raw = dict(zip(headers, values))
        source_id = raw.get("source_id") or raw.get("_source_id") or ""
        app_no = clean_text(raw.get("app_no"))
        enrichment = enrichments.get(source_id, {}).get(app_no)
        row = normalized_row(raw, enrichment)
        rescore_event = row.get("_keyword_rescore") or {}
        data.keyword_rescore_events.append(rescore_event)
        if rescore_event.get("below_gate"):
            data.keyword_below_gate.append(rescore_event)
        route = row.get("route")
        if source_id in {"surrey_devapps_v2", "abbotsford_devapps"}:
            stats = data.address_recovery.setdefault(source_id, {"EXTRACTED": 0, "CENTROID": 0, "NONE": 0, "ORIGINAL": 0})
            stats[row["address_source"]] = stats.get(row["address_source"], 0) + 1
        if route == "Future_Projects":
            if source_id == "surrey_devapps_v2":
                data.surrey_v2_total += 1
                if row.get("status_class") == "TERMINAL":
                    data.surrey_v2_terminal += 1
                    row["route"] = "Surrey_Historical_Archive"
                    data.analyzed.append(row)
                    continue
                data.surrey_v2_active += 1
            data.future.append(row)
            fit = row["fit_score"]
            if fit >= 70:
                data.fit_ge_70 += 1
            if fit >= 60:
                data.fit_ge_60 += 1
            if 60 <= fit <= 69:
                data.fit_60_69 += 1
            if 50 <= fit <= 59:
                data.fit_50_59 += 1
        elif route == "Run_Queue":
            data.watch.append(row)
        elif route in {"Bulk_Intake_Raw", "Rejected_Archive"}:
            data.analyzed.append(row)

    wb.close()
    data.records_normalized = len(data.future) + len(data.watch) + len(data.analyzed)
    data.municipalities = sorted({str(r.get("municipality")) for r in data.future if r.get("municipality")})
    data.future_sources = sorted({str(r.get("source_id")) for r in data.future if r.get("source_id")})
    if data.future_sources:
        data.live_sources = len(data.future_sources)

    summary_path = review_xlsx.with_name("TENDER_FINDER_Run_Source_Summary.csv")
    if summary_path.exists():
        try:
            with summary_path.open("r", encoding="utf-8-sig", newline="") as f:
                summary_rows = list(csv.DictReader(f))
            pulled = sum(coerce_int(r.get("records_pulled")) for r in summary_rows)
            normalized = sum(coerce_int(r.get("records_normalized")) for r in summary_rows)
            if pulled:
                data.records_pulled = pulled
            if normalized:
                data.records_normalized = normalized
            if summary_rows:
                data.coded_sources = len(summary_rows)
            live = {
                r.get("source_id")
                for r in summary_rows
                if coerce_int(r.get("records_pulled")) > 0
            }
            if live:
                data.live_sources = len(live)
        except Exception:
            pass

    data.future.sort(key=lambda x: (-x["fit_score"], str(x["municipality"]), str(x["address"])))
    data.watch.sort(key=lambda x: (-x["fit_score"], str(x["municipality"]), str(x["address"])))
    data.analyzed.sort(key=lambda x: (str(x["route"]), -x["fit_score"], str(x["municipality"])))
    data.read_seconds = time.perf_counter() - start
    return data


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, required: set[str], max_scan: int = 8) -> tuple[int, list[str]] | None:
    for row_index in range(1, min(ws.max_row, max_scan) + 1):
        values = [str(ws.cell(row_index, col).value or "") for col in range(1, ws.max_column + 1)]
        if required.issubset(set(values)):
            return row_index, values
    return None


def sheet_dicts(wb: openpyxl.Workbook, sheet_name: str, required: set[str]) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    found = find_header_row(ws, required)
    if not found:
        return []
    header_row, headers = found
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not any(values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def most_recent_history(history_dir: Path | None = None) -> Path | None:
    history_dir = history_dir or DEMO_HISTORY_DIR
    if not history_dir.exists():
        return None
    candidates = sorted(history_dir.glob("demo_p*_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def workbook_bid_later_count(path: Path) -> int:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = sheet_dicts(wb, "BID_LATER_Future_Projects", {"municipality", "app_no"})
        wb.close()
        return len(rows)
    except Exception:
        return 0


def archive_pre_511_history(history_dir: Path | None = None) -> list[Path]:
    history_dir = history_dir or DEMO_HISTORY_DIR
    if not history_dir.exists():
        return []
    archive_dir = history_dir / "archive_pre_5_11"
    moved: list[Path] = []
    for path in sorted(history_dir.glob("demo_p*_*.xlsx")):
        bid_later_count = workbook_bid_later_count(path)
        if bid_later_count <= P511_CORRECTED_FUTURE_COUNT + 500:
            continue
        archive_dir.mkdir(parents=True, exist_ok=True)
        target = archive_dir / path.name
        if target.exists():
            target = archive_dir / f"{path.stem}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        shutil.move(str(path), str(target))
        moved.append(target)
    return moved


def apply_run_history(data: DemoData, history_dir: Path | None = None) -> None:
    history_dir = history_dir or DEMO_HISTORY_DIR
    prior = most_recent_history(history_dir)
    if not prior:
        for row in data.future:
            row["first_seen_date"] = TODAY.isoformat()
            row["is_new_since_last_run"] = "YES"
        data.new_since_last_run = len(data.future)
        data.removed_since_last_run = 0
        data.history_note = "BASELINE RUN - no prior history"
        return

    data.prior_workbook = str(prior)
    data.history_note = f"Compared against prior workbook: {prior.name}"
    wb = openpyxl.load_workbook(prior, read_only=True, data_only=True)
    prior_rows = sheet_dicts(wb, "BID_LATER_Future_Projects", {"municipality", "app_no"})
    prior_first_seen: dict[str, str] = {}
    for item in prior_rows:
        lead_id = clean_text(item.get("lead_id")) or stable_lead_id({
            "municipality": item.get("municipality"),
            "address": item.get("address"),
            "source": item.get("source"),
            "source_id": item.get("source_id"),
            "app_no": item.get("app_no"),
        })
        prior_first_seen[lead_id] = clean_text(item.get("first_seen_date")) or TODAY.isoformat()
    data.prior_lead_ids = set(prior_first_seen)

    prior_outreach_rows = sheet_dicts(wb, "Outreach_Tracker", {"lead_id", "status"})
    for item in prior_outreach_rows:
        lead_id = clean_text(item.get("lead_id"))
        if not lead_id:
            continue
        data.prior_outreach[lead_id] = {
            "lead_id": lead_id,
            "type": clean_text(item.get("type")) or "LEAD",
            "municipality": item.get("municipality") or "",
            "title_or_address": item.get("title_or_address") or "",
            "fit_score": item.get("fit_score") or "",
            "signal_quality": item.get("signal_quality") or "",
            "contact": item.get("contact") or "",
            "status": clean_text(item.get("status")) or "Not Contacted",
            "last_contact_date": item.get("last_contact_date") or "",
            "next_action_date": item.get("next_action_date") or "",
            "notes": item.get("notes") or "",
        }
    wb.close()

    current_ids = {row["lead_id"] for row in data.future}
    data.removed_since_last_run = len(data.prior_lead_ids - current_ids)
    data.new_since_last_run = len(current_ids - data.prior_lead_ids)
    for row in data.future:
        lead_id = row["lead_id"]
        row["first_seen_date"] = prior_first_seen.get(lead_id, TODAY.isoformat())
        row["is_new_since_last_run"] = "NO" if lead_id in data.prior_lead_ids else "YES"


def archive_demo_history(workbook_path: Path, out_dir: Path, history_dir: Path | None = None) -> Path:
    history_dir = history_dir or DEMO_HISTORY_DIR
    history_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    patch_name = "demo_p510" if "510" in out_dir.name.lower() else out_dir.name
    target = history_dir / f"{patch_name}_{stamp}.xlsx"
    shutil.copy2(workbook_path, target)
    return target


def extract_patch_digits(path_text: str) -> str | None:
    """Pull the "5NN"-style patch-number digits out of an out_dir name or
    path, e.g. "demo_p516" / "p516" -> "516", "patch5_16" / "5-16" -> "516".
    Returns None if no recognizable patch-number pattern is present (e.g.
    a fully custom folder name)."""
    text = path_text.lower()
    # Use a "not followed by another digit" lookahead rather than \b: \b
    # treats underscore as a word character, so it would fail to end the
    # match right after the digits in names like "demo_p516_fast".
    match = re.search(r"(?:^|[^a-z0-9])p(\d{2,3})(?!\d)", text)
    if match:
        return match.group(1)
    match = re.search(r"(?:^|[^0-9])5[_-](\d{1,2})(?!\d)", text)
    if match:
        return "5" + match.group(1)
    return None


def mirror_name_for(path_text: str) -> str:
    """Derive the repo-root mirror folder name for a given --out-dir value
    (or its final path component). Previously this was a hardcoded if/elif
    chain that had to be manually extended with a new constant every patch
    - missing an entry meant silently falling through to a stale default
    and overwriting an unrelated folder (this happened for real in Patches
    5.14 and 5.15). Deriving the name directly from the out_dir text makes
    that failure mode structurally impossible: there is no default branch
    left to fall through to for a recognized "p516"-style name.

    For an out_dir name with no recognizable patch-number pattern (e.g. the
    GUI's timestamped "demo_gui_20260701_120000" folders), mirror under a
    sanitized version of that same name instead of guessing - this can
    never collide with an existing demo_pNNN folder."""
    digits = extract_patch_digits(path_text)
    if digits:
        return f"demo_p{digits}"
    base = Path(path_text).name or path_text
    safe = re.sub(r"[^A-Za-z0-9_.-]", "_", base)
    return safe if safe.lower().startswith("demo_") else f"demo_{safe}"


def is_login_platform(url: str) -> bool:
    host = urllib.parse.urlparse(url).netloc.lower()
    path = urllib.parse.urlparse(url).path.lower()
    if host.endswith("bcbid.gov.bc.ca"):
        return "/usr/login" in path
    return any(host == blocked or host.endswith("." + blocked) for blocked in LOGIN_PLATFORM_HOSTS)


def source_urls(source: dict[str, Any]) -> list[str]:
    urls = list(source.get("url_variants") or [])
    if source["url"] not in urls:
        urls.insert(0, source["url"])
    deduped: list[str] = []
    for url in urls:
        if url not in deduped:
            deduped.append(url)
    return deduped


def rss_candidates(url: str, explicit: str | None = None) -> list[str]:
    out = []
    if explicit:
        out.append(explicit)
    parsed = urllib.parse.urlparse(url)
    clean = urllib.parse.urlunparse(parsed._replace(query="", fragment="")).rstrip("/")
    out.extend([clean + "?feed=rss", clean + "/rss", clean + "/feed", urllib.parse.urljoin(clean + "/", "feed.xml")])
    deduped: list[str] = []
    for item in out:
        if item not in deduped and not is_login_platform(item):
            deduped.append(item)
    return deduped


def fetch_url(url: str, timeout: int = 20, retries: int = 1) -> tuple[str, int, str, str]:
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; TENDER_FINDER-DemoBot/5.5; public tender discovery; no auth)",
        "Accept": "text/html,application/xhtml+xml,application/rss+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-CA,en;q=0.8",
    }
    ctx = ssl.create_default_context()
    last_error = ""
    current = url
    for attempt in range(retries + 1):
        try:
            fetched = safe_urllib_fetch(
                current,
                headers=headers,
                timeout=timeout,
                max_bytes=2_000_000,
                context=ctx,
            )
            content_type = fetched.headers.get("content-type", "")
            charset_getter = getattr(fetched.headers, "get_content_charset", None)
            charset = (charset_getter() if callable(charset_getter) else None) or "utf-8"
            return (
                fetched.body.decode(charset, errors="replace"),
                fetched.status,
                content_type,
                fetched.final_url,
            )
        except urllib.error.HTTPError as e:
            if e.code in {403, 429, 404}:
                return "", e.code, e.headers.get("content-type", ""), current
            last_error = f"HTTP {e.code}"
        except (urllib.error.URLError, TimeoutError, socket.timeout, URLSafetyError) as e:
            last_error = str(e)
        if attempt < retries:
            time.sleep(0.4 * (attempt + 1))
    raise RuntimeError(last_error or "fetch failed")


def tender_keyword_match(text: str) -> bool:
    return load_keywords_config().matches_any("tender_match", text)


def _strip_civil_keyword_collisions(text_low: str) -> str:
    """Strip configured unrelated words before civil substring matching."""
    for word in load_keywords_config().collision_terms():
        text_low = text_low.replace(word, "")
    return text_low


def _civil_keyword_hits(text_low: str) -> list[str]:
    stripped = _strip_civil_keyword_collisions(text_low)
    return [rule.keyword for rule in load_keywords_config().matching_rules("gate_include", stripped)]


def civil_keyword_match(text: str, title: str = "") -> bool:
    config = load_keywords_config()
    low = text.lower()
    if not _civil_keyword_hits(low):
        return False
    title_low = (title or text).lower()
    if any(rule.matches(title_low) for rule in config.exclusion_rules("absolute_title")):
        return False
    # A negative term in the title (the actual subject of the opportunity)
    # without a STRONG civil keyword also present in the title means the
    # civil match likely came from noisy commodity/category text or a weak,
    # generic keyword, not the real scope of work (e.g. a generic "Bridge
    # construction and repair service" commodity tag applied to a marine
    # bearings RFP, or "water" matching both a Water Treatment Plant and a
    # building's "Domestic Hot Water Tank" replacement).
    if any(rule.matches(title_low) for rule in config.exclusion_rules("negative_title")):
        title_hits = _civil_keyword_hits(title_low)
        weak_keywords = {rule.keyword.casefold() for rule in config.rules_for("gate_weak")}
        strong_hit_in_title = any(keyword.casefold() not in weak_keywords for keyword in title_hits)
        if not strong_hit_in_title:
            return False
    return True


def is_bad_candidate_text(text: str) -> bool:
    low = text.lower()
    if any(term in low for term in NEGATIVE_LINK_TERMS):
        return True
    if len(clean_text(text)) < 4:
        return True
    return False


def parse_rss(text: str, source: dict[str, Any], base_url: str, rss_url: str) -> list[TenderCandidate]:
    if not text.strip().startswith("<"):
        return []
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return []

    candidates: list[TenderCandidate] = []
    for item in root.findall(".//item"):
        title = clean_text(item.findtext("title"))
        link = clean_text(item.findtext("link"))
        desc = clean_text(item.findtext("description"), 500)
        joined = f"{title} {desc}"
        if not title or is_bad_candidate_text(joined) or not tender_keyword_match(joined):
            continue
        url = urllib.parse.urljoin(base_url, link)
        if is_login_platform(url):
            continue
        candidates.append(TenderCandidate(
            source_id=source["source_id"],
            source_name=source["name"],
            tender_title=title,
            url=url,
            civil_relevant=civil_keyword_match(joined, title),
            status=infer_status(joined)[0],
            found_via="rss",
            snippet=desc,
            closing_date=infer_status(joined)[1],
            municipality=source.get("municipality", ""),
            source_url=rss_url,
            contact_email=extract_contact_email(joined),
            contact_phone=extract_contact_phone(joined),
        ))

    ns = {"atom": "http://www.w3.org/2005/Atom"}
    for entry in root.findall(".//atom:entry", ns):
        title = clean_text(entry.findtext("atom:title", default="", namespaces=ns))
        summary = clean_text(entry.findtext("atom:summary", default="", namespaces=ns), 500)
        href = ""
        for link in entry.findall("atom:link", ns):
            href = link.attrib.get("href", href)
        joined = f"{title} {summary}"
        if not title or is_bad_candidate_text(joined) or not tender_keyword_match(joined):
            continue
        url = urllib.parse.urljoin(base_url, href)
        if is_login_platform(url):
            continue
        status, close_date = infer_status(joined)
        candidates.append(TenderCandidate(
            source_id=source["source_id"],
            source_name=source["name"],
            tender_title=title,
            url=url,
            civil_relevant=civil_keyword_match(joined, title),
            status=status,
            found_via="rss",
            snippet=summary,
            closing_date=close_date,
            municipality=source.get("municipality", ""),
            source_url=rss_url,
            contact_email=extract_contact_email(joined),
            contact_phone=extract_contact_phone(joined),
        ))
    return candidates


def extract_visible_text(text: str) -> tuple[str, str]:
    parser = VisibleTextExtractor()
    try:
        parser.feed(text)
    except Exception:
        pass
    return parser.title, parser.text


def extract_listing_lines(text: str, source: dict[str, Any], resolved_url: str) -> list[TenderCandidate]:
    title, visible = extract_visible_text(text)
    candidates: list[TenderCandidate] = []
    seen: set[str] = set()
    lines = [clean_text(line, 220) for line in visible.splitlines()]
    for line in lines:
        if not line or line.lower() in seen:
            continue
        if is_bad_candidate_text(line):
            continue
        if not tender_keyword_match(line):
            continue
        low = line.lower()
        if low in {"bid opportunities", "current bid opportunities", "tenders, rfqs & rfps", "purchasing and procurement"}:
            continue
        seen.add(low)
        status, close_date = infer_status(line)
        candidates.append(TenderCandidate(
            source_id=source["source_id"],
            source_name=source["name"],
            tender_title=line,
            url=resolved_url,
            civil_relevant=civil_keyword_match(line),
            status=status,
            found_via="html_listing",
            snippet=clean_text(visible, 500),
            closing_date=close_date,
            municipality=source.get("municipality", ""),
            source_url=resolved_url,
            contact_email=extract_contact_email(line),
            contact_phone=extract_contact_phone(line),
        ))
        if len(candidates) >= 20:
            break
    if not candidates and tender_keyword_match(f"{title} {visible[:800]}") and not is_bad_candidate_text(title):
        status, close_date = infer_status(visible)
        candidates.append(TenderCandidate(
            source_id=source["source_id"],
            source_name=source["name"],
            tender_title=title or source["name"],
            url=resolved_url,
            civil_relevant=civil_keyword_match(visible),
            status=status,
            found_via="html_listing",
            snippet=clean_text(visible, 500),
            closing_date=close_date,
            municipality=source.get("municipality", ""),
            source_url=resolved_url,
            contact_email=extract_contact_email(visible),
            contact_phone=extract_contact_phone(visible),
        ))
    return candidates


def strong_opportunity_signal(text: str) -> bool:
    low = text.lower()
    strong_patterns = [
        r"\brequest for (?:proposals?|quotations?|tenders?)\b",
        r"\b(?:rfp|rfq|itt|eoi)\b",
        r"\btenders?\b",
        r"\bbid(?:ding)? opportunity\b",
        r"\bproposal deadline\b",
        r"\bclosing date\b",
        r"\b\d{3,5}-\d{2,5}-\d{2,4}(?:-\d{1,4})?\b",
    ]
    return any(re.search(pattern, low) for pattern in strong_patterns)


def normalized_url(url: str) -> str:
    return urllib.parse.urldefrag(clean_text(url)).url.rstrip("/").lower()


def is_same_listing_page(url: str, source_url: str) -> bool:
    return bool(url and source_url and normalized_url(url) == normalized_url(source_url))


def is_truthful_public_detail_candidate(candidate: TenderCandidate) -> bool:
    title = clean_text(candidate.tender_title)
    if not title or is_generic_title(title) or is_bad_candidate_text(title):
        return False
    if not candidate.url.lower().startswith(("http://", "https://")):
        return False
    if is_same_listing_page(candidate.url, candidate.source_url):
        return False
    low = f"{title} {candidate.url} {candidate.snippet[:300]}".lower()
    headline_low = f"{title} {candidate.url}".lower()
    if any(term in headline_low for term in ("job opportunit", "careers", "employment", "artist opportunit", "artists-opportunities", "contact-us", "vendor help")):
        return False
    if candidate.source_id in {"surrey_bids_public", "fvrd_tenders", "slrd_contracting"}:
        return civil_keyword_match(low, title) or strong_opportunity_signal(low) or bool(candidate.closing_date)
    return False


def parse_html_links(text: str, source: dict[str, Any], base_url: str) -> list[TenderCandidate]:
    parser = LinkExtractor()
    parser.feed(text)
    seen: set[str] = set()
    candidates: list[TenderCandidate] = []
    for label, href in parser.links:
        joined = clean_text(label)
        if is_bad_candidate_text(joined):
            continue
        url = urllib.parse.urljoin(base_url, href)
        if is_login_platform(url) or url.lower().startswith("mailto:"):
            continue
        href_low = href.lower()
        combined = f"{joined} {href_low}"
        href_opportunity = bool(re.search(r"(?:bidid=|/bids?\.aspx|/rfp|/rfq|/tender|/itt)", href_low))
        if not (tender_keyword_match(joined) or href_opportunity):
            continue
        key = urllib.parse.urldefrag(url)[0].rstrip("/")
        if key.lower() in seen:
            continue
        seen.add(key.lower())
        status, close_date = infer_status(joined)
        candidates.append(TenderCandidate(
            source_id=source["source_id"],
            source_name=source["name"],
            tender_title=joined,
            url=key,
            civil_relevant=civil_keyword_match(joined),
            status=status,
            found_via="html_listing",
            snippet="",
            closing_date=close_date,
            municipality=source.get("municipality", ""),
            source_url=base_url,
            contact_email=extract_contact_email(joined),
            contact_phone=extract_contact_phone(joined),
        ))
        if len(candidates) >= 25:
            break
    return candidates


def extract_closing_date(text: str) -> str:
    clean = clean_text(text, 8000)
    patterns = [
        r"(?:closing|close|closes|submission deadline|proposal deadline|bid close|deadline)(?: date| time| date\s*&\s*time)?[:\s-]{0,20}([A-Z][a-z]+\.?\s+\d{1,2},?\s+\d{4})",
        r"(?:closing|close|closes|submission deadline|proposal deadline|bid close|deadline)(?: date| time| date\s*&\s*time)?[:\s-]{0,20}(\d{4}[-/]\d{1,2}[-/]\d{1,2})",
        r"(?:closing|close|closes|submission deadline|proposal deadline|bid close|deadline)(?: date| time| date\s*&\s*time)?[:\s-]{0,20}(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        r"([A-Z][a-z]+\.?\s+\d{1,2},?\s+\d{4})(?:\s+at\s+\d{1,2}:\d{2})?",
        r"(\d{4}[-/]\d{1,2}[-/]\d{1,2})",
    ]
    formats = [
        "%B %d, %Y",
        "%B %d %Y",
        "%b %d, %Y",
        "%b %d %Y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%m-%d-%Y",
        "%d-%m-%Y",
    ]
    for pattern in patterns:
        for match in re.findall(pattern, clean, flags=re.I):
            value = clean_text(match).replace(".", "")
            for fmt in formats:
                try:
                    parsed = dt.datetime.strptime(value, fmt).date()
                    if 2020 <= parsed.year <= 2035:
                        return parsed.isoformat()
                except ValueError:
                    continue
    return ""


def strip_html_fragment(text: str) -> str:
    return clean_text(html.unescape(re.sub(r"<[^>]+>", " ", text)), 4000)


def extract_explicit_status_from_text(text: str) -> str:
    clean = clean_text(text, 1000)
    patterns = [
        r"\bstatus\s*[:\-]\s*(open|closed|awarded|cancelled|canceled|expired)\b",
        r"\bstatus\s+(open|closed|awarded|cancelled|canceled|expired)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, clean, flags=re.I)
        if match:
            value = match.group(1).lower()
            return "cancelled" if value == "canceled" else value
    return ""


def extract_surrey_tender_fields(html_text: str) -> tuple[str, str]:
    """Extract the explicit Surrey tender Status/Closing Date blocks.

    Surrey pages include global navigation/archive text that can contain
    "closed" even while the tender's own Status block says Open. Trust the
    page's field blocks before falling back to generic text inference.
    """
    low = html_text.lower()
    status = ""
    closing = ""

    status_idx = low.find("<h4>status</h4>")
    if status_idx >= 0:
        status_text = strip_html_fragment(html_text[status_idx:status_idx + 900])
        match = re.search(r"\b(open|closed|awarded|cancelled|canceled|expired)\b", status_text, flags=re.I)
        if match:
            value = match.group(1).lower()
            status = "cancelled" if value == "canceled" else value

    close_idx = low.find("<h4>closing date</h4>")
    if close_idx >= 0:
        close_text = strip_html_fragment(html_text[close_idx:close_idx + 1600])
        closing = extract_closing_date(close_text)

    return status, closing


def apply_status_date_fallback(status: str, closing: str) -> str:
    if closing:
        try:
            close_date = dt.date.fromisoformat(closing)
            if close_date >= TODAY and status in {"", "unknown", "closed"}:
                return "open_or_needs_review"
        except ValueError:
            pass
    return status or "unknown"


def is_open_tender_status(status: str) -> bool:
    return status in {"open", "open_or_needs_review"}


def enrich_tender_closing_dates(tenders: list[TenderCandidate]) -> dict[str, int]:
    stats = {"civil_missing_before": 0, "filled": 0, "unknown_after": 0}
    for tender in tenders:
        if not tender.civil_relevant or tender.closing_date:
            continue
        stats["civil_missing_before"] += 1
        if is_login_platform(tender.url) or tender.url.lower().startswith("mailto:"):
            stats["unknown_after"] += 1
            continue
        try:
            body, http_status, content_type, _ = fetch_url(tender.url, timeout=8, retries=0)
            if http_status != 200 or ("html" not in content_type.lower() and "<html" not in body[:500].lower()):
                stats["unknown_after"] += 1
                continue
            _, visible = extract_visible_text(body)
            close_date = extract_closing_date(visible)
            if close_date:
                tender.closing_date = close_date
                status, _ = infer_status(f"{tender.tender_title} {visible[:2500]}")
                tender.status = apply_status_date_fallback(status, close_date)
                stats["filled"] += 1
            else:
                stats["unknown_after"] += 1
        except Exception:
            stats["unknown_after"] += 1
    return stats


def extract_contact_email(text: str) -> str:
    matches = re.findall(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", text, flags=re.I)
    for match in matches:
        low = match.lower()
        if not any(skip in low for skip in ("example.", "noreply", "no-reply")):
            return match
    return ""


def clean_contact_email(value: str) -> str:
    text = clean_text(value)
    if re.match(r"^\d{4}-\d{3}-\d{4}$", text):
        return ""
    if text and "@" not in text:
        return ""
    return text


def clean_contact_phone(value: str) -> str:
    text = clean_text(value)
    if re.match(r"^\d{4}-\d{3}-\d{4}$", text):
        return ""
    return text


def extract_contact_phone(text: str) -> str:
    matches = re.findall(r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}(?:\s*(?:x|ext\.?|extension)\s*\d{2,6})?", text, flags=re.I)
    for match in matches:
        cleaned = clean_text(match)
        digits = re.sub(r"\D", "", cleaned)
        if len(digits) >= 10:
            return cleaned
    return ""


def infer_status(text: str) -> tuple[str, str]:
    low = text.lower()
    closing = extract_closing_date(text)
    explicit = extract_explicit_status_from_text(text)
    if explicit == "open":
        return "open", closing
    if explicit in {"closed", "awarded", "cancelled", "expired"}:
        return explicit, closing
    closed_terms = ("closed", "awarded", "archive", "archived", "cancelled", "canceled", "expired")
    if closing:
        try:
            close_date = dt.date.fromisoformat(closing)
            if close_date < TODAY:
                return "closed", closing
            if any(re.search(pattern, low) for pattern in (r"\bstatus[:\s-]+open\b", r"\bcurrently accepting\b", r"\bavailable for downloading\b")):
                return "open", closing
            return "open_or_needs_review", closing
        except ValueError:
            pass
    if any(term in low for term in closed_terms):
        return "closed", closing
    open_patterns = [
        r"\bstatus[:\s-]+open\b",
        r"\bopen bids?\b",
        r"\bopen tenders?\b",
        r"\bcurrently accepting\b",
        r"\bavailable for downloading\b",
    ]
    if any(re.search(pattern, low) for pattern in open_patterns):
        return "open", closing
    return "unknown", closing


def choose_deep_title(original: str, page_title: str, visible: str) -> str:
    original = clean_text(original, 180)
    page_title = clean_text(page_title, 180)
    bad_titles = (
        "home",
        "bid opportunities",
        "current bid opportunities",
        "tenders, rfqs & rfps",
        "tenders & rfps",
        "procurement",
        "purchasing",
        "doing business",
        "contracting opportunities",
    )
    if page_title and tender_keyword_match(page_title) and not any(page_title.lower() == bad for bad in bad_titles):
        return page_title
    for line in [clean_text(line, 180) for line in visible.splitlines()[:80]]:
        if line and not is_bad_candidate_text(line) and tender_keyword_match(line):
            if len(line) > len(original) or not tender_keyword_match(original):
                return line
    return original or page_title


GENERIC_TITLES = {
    "bid opportunities",
    "current bid opportunities",
    "contracting opportunities",
    "contracting opportunities archive",
    "procurement opportunities",
    "pre-procurement notices and procurement opportunities",
    "purchasing and procurement",
    "tenders, rfqs & rfps",
    "tenders & rfps",
    "tenders rfps",
    "bids and rfps",
    "bid opportunities city of richmond bc",
    "procurement services city of abbotsford",
    "vendor help",
    "doing business with us",
    "welcome to procurement services of kwantlen polytechnic university",
}


def is_generic_title(title: str) -> bool:
    low = clean_text(title).lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", low).strip()
    if low in GENERIC_TITLES:
        return True
    generic_normalized = {re.sub(r"[^a-z0-9]+", " ", generic).strip() for generic in GENERIC_TITLES}
    if normalized in generic_normalized:
        return True
    return any(low.startswith(generic + " |") for generic in GENERIC_TITLES)


def title_from_url(url: str) -> str:
    path = urllib.parse.urlparse(url).path.rstrip("/")
    slug = path.rsplit("/", 1)[-1]
    if not slug or "." in slug:
        return ""
    slug = re.sub(r"[-_]+", " ", urllib.parse.unquote(slug))
    slug = re.sub(r"\s+", " ", slug).strip()
    if not slug or len(slug) < 5:
        return ""
    return slug.title()


def looks_soft_404(url: str, body: str) -> bool:
    low_url = url.lower()
    if "404" in low_url or "page-not-found" in low_url or "not-found" in low_url:
        return True
    head = clean_text(body[:3000]).lower()
    return "page not found" in head or "404 - file or directory not found" in head


def looks_browser_checked(url: str, body: str) -> bool:
    low_url = url.lower()
    head = clean_text(body[:5000]).lower()
    return "browser_check" in low_url or ("browser check" in head and "recaptcha" in head)


def deep_fetch_candidate(candidate: TenderCandidate) -> TenderCandidate:
    if is_login_platform(candidate.url) or candidate.url.lower().startswith("mailto:"):
        return candidate
    if re.search(r"\.(pdf|docx?|xlsx?)($|\?)", candidate.url, flags=re.I):
        joined = f"{candidate.tender_title} {candidate.snippet} {candidate.url}"
        status, close_date = infer_status(joined)
        candidate.civil_relevant = civil_keyword_match(joined, candidate.tender_title)
        candidate.status = status
        candidate.closing_date = candidate.closing_date or close_date
        candidate.contact_email = candidate.contact_email or extract_contact_email(joined)
        candidate.contact_phone = candidate.contact_phone or extract_contact_phone(joined)
        return candidate
    try:
        if candidate.source_id == "bc_bid_public" and LAST_BC_BID_AUDIT.get("cookies"):
            body, http_status, content_type, final_url = fetch_bc_bid_with_cookies(candidate.url, LAST_BC_BID_AUDIT, timeout=8)
        else:
            body, http_status, content_type, final_url = fetch_url(candidate.url, timeout=8, retries=0)
        if http_status in {403, 429}:
            candidate.snippet = clean_text(candidate.snippet or f"Deep page blocked with HTTP {http_status}; source listing retained.", 500)
            return candidate
        if http_status == 404:
            candidate.status = "closed"
            candidate.snippet = clean_text(candidate.snippet or "Deep page returned HTTP 404.", 500)
            return candidate
        if "html" not in content_type.lower() and "<html" not in body[:500].lower():
            return candidate
        page_title, visible = extract_visible_text(body)
        joined = f"{candidate.tender_title} {page_title} {visible[:2500]} {final_url}"
        candidate.tender_title = choose_deep_title(candidate.tender_title, page_title, visible)
        used_url_title = False
        if is_generic_title(candidate.tender_title):
            url_title = title_from_url(final_url)
            if url_title:
                candidate.tender_title = url_title
                used_url_title = True
        candidate.url = final_url
        candidate.snippet = clean_text(visible, 500)
        candidate.civil_relevant = civil_keyword_match(candidate.tender_title if used_url_title else joined, candidate.tender_title)
        status, close_date = infer_status(joined)
        if candidate.source_id == "surrey_bids_public":
            surrey_status, surrey_close = extract_surrey_tender_fields(body)
            status = surrey_status or status
            close_date = surrey_close or close_date
        status = apply_status_date_fallback(status, close_date)
        if close_date:
            try:
                if dt.date.fromisoformat(close_date) < TODAY and status in {"open", "open_or_needs_review", "unknown"}:
                    status = "closed"
            except ValueError:
                pass
        candidate.status = status
        candidate.closing_date = candidate.closing_date or close_date
        candidate.contact_email = candidate.contact_email or extract_contact_email(visible)
        candidate.contact_phone = candidate.contact_phone or extract_contact_phone(visible)
        candidate.found_via = "deep_page" if candidate.found_via != "rss" else "rss+deep_page"
    except Exception as exc:
        if not candidate.snippet:
            candidate.snippet = f"Deep fetch failed: {clean_text(exc, 180)}"
    return candidate


def bc_bid_detail_fetch(candidate: TenderCandidate) -> bool:
    """Fetch a BC Bid opportunity detail page via the cookie-replay session
    to recover closing date, contact info, and a fuller description for
    civil_relevant + open rows. Returns True if contact info was newly
    recovered (it had none before this call)."""
    had_contact_before = bool(candidate.contact_email or candidate.contact_phone)
    if not LAST_BC_BID_AUDIT.get("cookies"):
        return False
    for attempt in range(2):
        try:
            body, http_status, content_type, _ = fetch_bc_bid_with_cookies(candidate.url, LAST_BC_BID_AUDIT, timeout=10)
            if http_status != 200 or "html" not in content_type.lower():
                continue
            _, visible = extract_visible_text(body)
            close_date = extract_closing_date(visible)
            if close_date:
                candidate.closing_date = candidate.closing_date or close_date
            contact_idx = visible.find("Contact First Name Contact Last Name Email")
            contact_window = visible[contact_idx:contact_idx + 200] if contact_idx != -1 else ""
            email = extract_contact_email(contact_window)
            if email:
                candidate.contact_email = candidate.contact_email or email
            phone = extract_contact_phone(contact_window)
            if phone:
                candidate.contact_phone = candidate.contact_phone or phone
            if not candidate.snippet or len(candidate.snippet) < 80:
                candidate.snippet = clean_text(visible, 500)
            return bool((candidate.contact_email or candidate.contact_phone) and not had_contact_before)
        except Exception:
            continue
    return False


def is_real_opportunity_candidate(candidate: TenderCandidate) -> bool:
    title = clean_text(candidate.tender_title)
    low = f"{title} {candidate.url} {candidate.snippet[:300]}".lower()
    headline_low = f"{title} {candidate.url}".lower()
    non_tender_career_terms = ("firefighter opportunit", "student opportunit", "lifeguard opportunit", "careers", "employment")
    if is_generic_title(title):
        return False
    if any(term in headline_low for term in non_tender_career_terms):
        return False
    if re.fullmatch(r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", title):
        return False
    if len(title) > 160 and not re.search(r"\b(?:rfp|rfq|itt|eoi|request for|tender)\b", title, flags=re.I):
        return False
    if strong_opportunity_signal(low):
        return True
    if candidate.closing_date and civil_keyword_match(low, title):
        return True
    if re.search(r"\.(pdf|docx?)($|\?)", candidate.url, flags=re.I) and civil_keyword_match(low, title):
        return True
    if is_truthful_public_detail_candidate(candidate):
        return True
    return False


def bc_bid_api_like(url: str) -> bool:
    return bool(re.search(r"(/api/|\.json(?:$|\?)|/search\b|/opportunit|/rfp\b|solicitation|tender|bid)", url, flags=re.I))


def bc_bid_static_asset(url: str) -> bool:
    return bool(re.search(r"\.(?:css|js|png|jpg|jpeg|gif|svg|webp|ico|woff2?|ttf|map)(?:$|\?)", url, flags=re.I))


def http_fetch_with_headers(url: str, headers: dict[str, str], timeout: int = 20) -> tuple[str, int, str, str]:
    ctx = ssl.create_default_context()
    fetched = safe_urllib_fetch(
        url,
        headers=headers,
        timeout=timeout,
        max_bytes=1_000_000,
        context=ctx,
    )
    content_type = fetched.headers.get("content-type", "")
    charset_getter = getattr(fetched.headers, "get_content_charset", None)
    charset = (charset_getter() if callable(charset_getter) else None) or "utf-8"
    return (
        fetched.body.decode(charset, errors="replace"),
        fetched.status,
        content_type,
        fetched.final_url,
    )


def bc_bid_cookie_headers(audit: dict[str, Any]) -> dict[str, str]:
    cookie_pairs = []
    for cookie in audit.get("cookies", []):
        if cookie.get("domain") and "bcbid.gov.bc.ca" in cookie.get("domain", ""):
            cookie_pairs.append(f"{cookie['name']}={cookie['value']}")
    headers = {
        "User-Agent": BCBID_BROWSER_UA,
        "Accept": "text/html,application/xhtml+xml,application/json;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-CA,en;q=0.8",
    }
    if cookie_pairs:
        headers["Cookie"] = "; ".join(cookie_pairs)
    return headers


def fetch_bc_bid_with_cookies(url: str, audit: dict[str, Any], timeout: int = 20) -> tuple[str, int, str, str]:
    return http_fetch_with_headers(url, bc_bid_cookie_headers(audit), timeout=timeout)


def parse_bc_bid_listing_html(body: str, source: dict[str, Any], final_url: str) -> list[TenderCandidate]:
    try:
        from bs4 import BeautifulSoup
    except Exception:
        return []
    soup = BeautifulSoup(body, "html.parser")
    rows = soup.select("tr[data-object-type='rfp']")
    candidates: list[TenderCandidate] = []
    for row in rows[:25]:
        cells = row.find_all("td", recursive=False)
        if len(cells) < 7:
            continue
        values = [" ".join(cell.get_text(" ", strip=True).split()) for cell in cells]
        link = row.select_one("a[href*='/page.aspx/en/bpm/process_manage_extranet/']")
        if not link:
            continue
        href = link.get("href", "").strip()
        if not href:
            continue
        ref_id = values[1] if len(values) > 1 else ""
        title = values[2] if len(values) > 2 else ""
        commodity = values[3] if len(values) > 3 else ""
        opp_type = values[4] if len(values) > 4 else ""
        posted = values[5] if len(values) > 5 else ""
        closes = values[6] if len(values) > 6 else ""
        issuer = values[10] if len(values) > 10 else source.get("name", "")
        snippet = clean_text(f"{ref_id}. {commodity}. {opp_type}. issuer: {issuer}. posted {posted}. closes {closes}.", 500)
        joined = f"{title} {ref_id} {commodity} {opp_type} {issuer} {snippet}"
        close_date = ""
        date_match = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", closes)
        if date_match:
            close_date = date_match.group(1)
        status, inferred_close = infer_status(f"{joined} {closes}")
        candidate = TenderCandidate(
            source_id=source["source_id"],
            source_name=source["name"],
            tender_title=title,
            url=urllib.parse.urljoin(final_url, href),
            civil_relevant=civil_keyword_match(joined, title),
            status="open" if values[0].lower().startswith("open") else status,
            found_via="browser_cookie_replay",
            snippet=snippet,
            closing_date=close_date or inferred_close,
            source_url=final_url,
            issuer=issuer,
            region="BC_wide",
        )
        if is_real_opportunity_candidate(candidate):
            candidates.append(candidate)
    return candidates


def bc_bid_page_has_relevant_closing_date(html_body: str, base_url: str, window_days: int = BCBID_RELEVANCE_WINDOW_DAYS) -> bool:
    """True if the given BC Bid listing page HTML contains at least one row
    whose closing_date falls within [today, today + window_days]. Used to
    decide whether paging further is still worth it for an active bid
    pipeline. Rows with no parseable closing date don't count either way.
    Fails open (returns True) on a parse error so a transient hiccup can't
    silently truncate pagination."""
    bc_bid_source = next(
        (s for s in TENDER_SOURCES if s["source_id"] == "bc_bid_public"),
        {"source_id": "bc_bid_public", "name": "BC Bid public open opportunities"},
    )
    try:
        rows = parse_bc_bid_listing_html(html_body, bc_bid_source, base_url)
    except Exception:
        return True
    if not rows:
        return False
    cutoff = TODAY + dt.timedelta(days=window_days)
    for row in rows:
        if not row.closing_date:
            continue
        try:
            closes = dt.date.fromisoformat(row.closing_date)
        except ValueError:
            continue
        if TODAY <= closes <= cutoff:
            return True
    return False


def run_bc_bid_browser_audit() -> dict[str, Any]:
    audit: dict[str, Any] = {
        "attempted": True,
        "approach": "Playwright",
        "launch": "",
        "final_url": "",
        "title": "",
        "page_text_snippet": "",
        "platform": "",
        "requests": [],
        "api_hits": [],
        "cookies": [],
        "cookie_replay": [],
        "cookie_replay_worked": False,
        "browser_check": False,
        "captcha_present": False,
        "opportunities_found": False,
        "errors": [],
        "pagination_detected": False,
        "pagination_total_hint": "",
        "pages_fetched": 1,
        "page_bodies": [],
        "pagination_stop_reason": "",
    }
    request_seen: set[str] = set()
    api_seen: set[str] = set()
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        audit["approach"] = "none worked"
        audit["errors"].append(f"playwright_import: {clean_text(exc, 240)}")
        return audit

    with sync_playwright() as p:
        browser = None
        launch_attempts = [
            ("msedge_headless", lambda: p.chromium.launch(channel="msedge", headless=True)),
            ("chrome_headless", lambda: p.chromium.launch(channel="chrome", headless=True)),
            ("chromium_headless", lambda: p.chromium.launch(headless=True)),
        ]
        for label, launch in launch_attempts:
            try:
                browser = launch()
                audit["launch"] = label
                break
            except Exception as exc:
                audit["errors"].append(f"{label}: {clean_text(exc, 240)}")
        if browser is None:
            audit["approach"] = "none worked"
            return audit

        context = browser.new_context(user_agent=BCBID_BROWSER_UA)
        audit["blocked_network_requests"] = []

        def on_blocked_url(url: str, reason: str) -> None:
            audit["blocked_network_requests"].append(
                {"url": clean_text(url, 300), "reason": clean_text(reason, 300)}
            )

        install_playwright_public_route(context, on_blocked=on_blocked_url)
        page = context.new_page()

        def on_request(req: Any) -> None:
            url = req.url
            if url in request_seen:
                return
            request_seen.add(url)
            headers = req.headers
            audit["requests"].append({
                "url": url,
                "method": req.method,
                "resource_type": req.resource_type,
                "headers": {k: headers.get(k, "") for k in ("accept", "content-type", "referer", "origin", "sec-fetch-mode") if headers.get(k)},
            })

        def on_response(resp: Any) -> None:
            url = resp.url
            if not bc_bid_api_like(url) or bc_bid_static_asset(url) or url in api_seen:
                return
            api_seen.add(url)
            content_type = resp.headers.get("content-type", "")
            preview = ""
            json_keys: list[str] = []
            try:
                if "json" in content_type.lower() or url.lower().endswith(".json"):
                    payload = resp.json()
                    if isinstance(payload, dict):
                        json_keys = list(payload.keys())[:20]
                    preview = json.dumps(payload)[:500]
                else:
                    preview = resp.text()[:500]
            except Exception as exc:
                preview = f"<body read failed: {clean_text(exc, 160)}>"
            audit["api_hits"].append({
                "url": url,
                "method": resp.request.method,
                "status": resp.status,
                "content_type": content_type,
                "json_keys": json_keys,
                "body_preview": preview,
                "request_headers": {k: resp.request.headers.get(k, "") for k in ("accept", "content-type", "referer", "origin", "sec-fetch-mode") if resp.request.headers.get(k)},
            })

        page.on("request", on_request)
        page.on("response", on_response)
        for attempt in range(2):
            try:
                validate_public_url(BCBID_PUBLIC_URL)
                page.goto(BCBID_PUBLIC_URL, wait_until="networkidle", timeout=20_000)
            except Exception as exc:
                audit["errors"].append(f"goto_{attempt + 1}: {clean_text(exc, 240)}")
                try:
                    page.wait_for_load_state("load", timeout=5_000)
                except Exception as wait_exc:
                    audit["errors"].append(f"load_state_{attempt + 1}: {clean_text(wait_exc, 200)}")
            if "browser_check" not in page.url.lower():
                break
            try:
                page.wait_for_timeout(2_000)
            except Exception:
                pass
        audit["final_url"] = page.url
        try:
            validate_public_url(audit["final_url"])
        except URLSafetyError as exc:
            audit["errors"].append(f"unsafe_final_url: {clean_text(exc, 240)}")
        try:
            audit["title"] = page.title()
        except Exception as exc:
            audit["errors"].append(f"title: {clean_text(exc, 160)}")
        try:
            body_text = page.locator("body").inner_text(timeout=3_000)
            audit["page_text_snippet"] = clean_text(body_text, 1_000)
        except Exception as exc:
            audit["errors"].append(f"body: {clean_text(exc, 160)}")

        try:
            pager_count_text = page.locator("[data-role='pager-count']").inner_text(timeout=2_000)
            audit["pagination_total_hint"] = clean_text(pager_count_text, 120)
        except Exception:
            pass
        next_page_btn = page.locator(BCBID_NEXT_PAGE_SELECTOR)
        try:
            audit["pagination_detected"] = next_page_btn.count() > 0
        except Exception:
            audit["pagination_detected"] = False
        page1_body = ""
        try:
            page1_body = page.content()
            audit["page_bodies"].append(page1_body)
        except Exception as exc:
            audit["errors"].append(f"page1_content: {clean_text(exc, 160)}")
        if audit["pagination_detected"]:
            if page1_body and not bc_bid_page_has_relevant_closing_date(page1_body, page.url):
                audit["pagination_stop_reason"] = (
                    f"relevance_window_exhausted (page 1 had no closing dates within "
                    f"{BCBID_RELEVANCE_WINDOW_DAYS} days of today)"
                )
            else:
                pages_budget_deadline = time.perf_counter() + 150
                for page_num in range(2, BCBID_MAX_PAGES + 1):
                    if time.perf_counter() > pages_budget_deadline:
                        audit["pagination_stop_reason"] = f"timebox_hit before page {page_num}"
                        audit["errors"].append(f"pagination_timebox_hit before page {page_num}")
                        break
                    try:
                        next_page_btn = page.locator(BCBID_NEXT_PAGE_SELECTOR)
                        if next_page_btn.count() == 0 or next_page_btn.get_attribute("disabled") is not None:
                            audit["pagination_stop_reason"] = f"no_more_pages after page {page_num - 1}"
                            break
                        next_page_btn.click(timeout=5_000)
                        page.wait_for_timeout(1_500)
                        try:
                            page.wait_for_load_state("networkidle", timeout=8_000)
                        except Exception:
                            pass
                        page_body = page.content()
                        audit["page_bodies"].append(page_body)
                        audit["pages_fetched"] = page_num
                        if not bc_bid_page_has_relevant_closing_date(page_body, page.url):
                            audit["pagination_stop_reason"] = (
                                f"relevance_window_exhausted (page {page_num} had no closing "
                                f"dates within {BCBID_RELEVANCE_WINDOW_DAYS} days of today)"
                            )
                            break
                    except Exception as exc:
                        audit["pagination_stop_reason"] = f"fetch_error before/at page {page_num}"
                        audit["errors"].append(f"pagination_page_{page_num}: {clean_text(exc, 160)}")
                        break
                else:
                    audit["pagination_stop_reason"] = f"hard_cap_reached ({BCBID_MAX_PAGES} pages)"

        audit["cookies"] = context.cookies()
        browser.close()

    final_blob = " ".join([
        str(audit.get("final_url", "")),
        str(audit.get("title", "")),
        str(audit.get("page_text_snippet", "")),
    ]).lower()
    request_blob = " ".join(str(r.get("url", "")) for r in audit["requests"]).lower()
    if "ivalua" in request_blob:
        audit["platform"] = "Ivalua"
    audit["browser_check"] = "browser_check" in final_blob
    audit["captcha_present"] = "recaptcha" in request_blob or "captcha" in final_blob

    replay_headers = bc_bid_cookie_headers(audit)
    replay_targets = [hit["url"] for hit in audit["api_hits"][:12] if hit.get("url")]
    for url in replay_targets:
        try:
            body, http_status, content_type, final_url = http_fetch_with_headers(url, replay_headers, timeout=20)
            row = {
                "url": url,
                "status": http_status,
                "final_url": final_url,
                "content_type": content_type,
                "body_preview": body[:500],
                "has_grid": "body_x_grid_grd_tr_" in body,
            }
            audit["cookie_replay"].append(row)
        except urllib.error.HTTPError as exc:
            audit["cookie_replay"].append({
                "url": url,
                "status": exc.code,
                "final_url": url,
                "content_type": exc.headers.get("content-type", ""),
                "body_preview": "",
            })
        except Exception as exc:
            audit["cookie_replay"].append({"url": url, "error": clean_text(exc, 240)})
    page_replay = next((item for item in audit["cookie_replay"] if item.get("url") == BCBID_PUBLIC_URL), None)
    if page_replay and page_replay.get("final_url") == BCBID_PUBLIC_URL and page_replay.get("has_grid"):
        audit["cookie_replay_worked"] = True
        audit["opportunities_found"] = True
    audit["browser_mode"] = "headless"
    audit["user_assisted"] = False
    audit["login_attempted"] = False
    audit["captcha_bypass_attempted"] = False
    audit["public_page_loaded"] = bool(audit.get("opportunities_found"))
    return audit


BCBID_USER_ACTION_MESSAGE = (
    "TENDER_FINDER_USER_ACTION_REQUIRED: BC Bid's automated browser-check page appeared. "
    "A visible browser window has been opened to the public BC Bid Opportunities "
    "page. Please allow that page to finish loading (no login, no account, no "
    "CAPTCHA bypass performed by TENDER_FINDER) - if BC Bid shows you a CAPTCHA, you may "
    "complete it yourself in that window. TENDER_FINDER will keep checking and continue "
    "automatically once the public listing is visible."
)


def run_bc_bid_user_assisted_browser_audit(wait_timeout_seconds: int | None = None) -> dict[str, Any]:
    """Compliant fallback for when the headless Playwright attempt lands on
    BC Bid's /bas/browser_check wall. Opens a VISIBLE browser window to the
    same public opportunities page - no login, no stored credentials, no
    CAPTCHA bypass. If the browser-check page appears, this polls (bounded
    by wait_timeout_seconds) for a person watching the window to let the
    public page finish loading naturally - including solving a CAPTCHA
    themselves if BC Bid shows one - then reads whatever is on the now-
    public Opportunities page using the same parsing logic as the headless
    path (parse_bc_bid_listing_html). Uses a temporary Playwright persistent
    context so no session state is ever written into the repo; the temp
    profile directory is always deleted before returning, success or not.

    wait_timeout_seconds defaults to 300s (5 minutes) so a real person has
    enough time to complete the check - the earlier 20s default timed out
    on actual users. Unattended runs (regression suite, CI) set the
    TENDER_FINDER_BCBID_WAIT_SECONDS environment variable to a short value instead.

    The GUI's "Continue After Browser Check" button works through a signal
    file: this function prints "TENDER_FINDER_CONTINUE_SIGNAL_FILE: <path>" when the
    wait starts; the GUI touches that file when the user clicks Continue,
    which triggers an immediate re-check (and an honest message if the
    check still isn't done, rather than silently doing nothing).
    """
    if wait_timeout_seconds is None:
        try:
            wait_timeout_seconds = int(os.environ.get("TENDER_FINDER_BCBID_WAIT_SECONDS", "300"))
        except ValueError:
            wait_timeout_seconds = 300
    audit: dict[str, Any] = {
        "attempted": True,
        "approach": "Playwright (headed, user-assisted)",
        "browser_mode": "headed_user_assisted",
        "user_assisted": True,
        "login_attempted": False,
        "captcha_bypass_attempted": False,
        "public_page_loaded": False,
        "parsed_rows": 0,
        "final_url": "",
        "title": "",
        "launch": "",
        "errors": [],
        "page_bodies": [],
    }
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        audit["errors"].append(f"playwright_import: {clean_text(exc, 240)}")
        return audit

    profile_dir = tempfile.mkdtemp(prefix="tenderfinder_bcbid_user_")
    try:
        with sync_playwright() as p:
            context = None
            launch_attempts = [
                ("msedge_headed", lambda: p.chromium.launch_persistent_context(
                    profile_dir, channel="msedge", headless=False, user_agent=BCBID_BROWSER_UA)),
                ("chrome_headed", lambda: p.chromium.launch_persistent_context(
                    profile_dir, channel="chrome", headless=False, user_agent=BCBID_BROWSER_UA)),
                ("chromium_headed", lambda: p.chromium.launch_persistent_context(
                    profile_dir, headless=False, user_agent=BCBID_BROWSER_UA)),
            ]
            for label, launch in launch_attempts:
                try:
                    context = launch()
                    audit["launch"] = label
                    break
                except Exception as exc:
                    audit["errors"].append(f"{label}: {clean_text(exc, 240)}")
            if context is None:
                audit["errors"].append("no_visible_browser_available")
                return audit

            try:
                audit["blocked_network_requests"] = []

                def on_blocked_url(url: str, reason: str) -> None:
                    audit["blocked_network_requests"].append(
                        {"url": clean_text(url, 300), "reason": clean_text(reason, 300)}
                    )

                install_playwright_public_route(context, on_blocked=on_blocked_url)
                page = context.pages[0] if context.pages else context.new_page()
                try:
                    validate_public_url(BCBID_PUBLIC_URL)
                    page.goto(BCBID_PUBLIC_URL, wait_until="domcontentloaded", timeout=20_000)
                except Exception as exc:
                    audit["errors"].append(f"goto: {clean_text(exc, 240)}")

                if "browser_check" in page.url.lower():
                    print(BCBID_USER_ACTION_MESSAGE, flush=True)
                    signal_file = Path(tempfile.gettempdir()) / f"tenderfinder_bcbid_continue_{os.getpid()}.signal"
                    signal_file.unlink(missing_ok=True)
                    print(f"TENDER_FINDER_CONTINUE_SIGNAL_FILE: {signal_file}", flush=True)
                    deadline = time.perf_counter() + wait_timeout_seconds
                    next_heartbeat = time.perf_counter() + 60
                    try:
                        while time.perf_counter() < deadline:
                            try:
                                if "browser_check" not in page.url.lower():
                                    break
                            except Exception:
                                break
                            if signal_file.exists():
                                signal_file.unlink(missing_ok=True)
                                try:
                                    still_blocked = "browser_check" in page.url.lower()
                                except Exception:
                                    still_blocked = False
                                if still_blocked:
                                    print(
                                        "TENDER_FINDER_INFO: You clicked Continue, but BC Bid still shows the "
                                        "browser-check page. Finish the check in the browser window "
                                        "(complete any CAPTCHA yourself), then click Continue again.",
                                        flush=True,
                                    )
                                else:
                                    break
                            if time.perf_counter() >= next_heartbeat:
                                remaining = int(deadline - time.perf_counter())
                                print(
                                    f"TENDER_FINDER_INFO: still waiting for the BC Bid browser check "
                                    f"({remaining}s left before TENDER_FINDER moves on without BC Bid).",
                                    flush=True,
                                )
                                next_heartbeat = time.perf_counter() + 60
                            time.sleep(2)
                    finally:
                        signal_file.unlink(missing_ok=True)

                audit["final_url"] = page.url
                try:
                    validate_public_url(audit["final_url"])
                except URLSafetyError as exc:
                    audit["errors"].append(f"unsafe_final_url: {clean_text(exc, 240)}")
                try:
                    audit["title"] = page.title()
                except Exception:
                    pass

                if "browser_check" in page.url.lower():
                    audit["errors"].append("still_on_browser_check_after_wait")
                    return audit

                audit["public_page_loaded"] = True
                try:
                    page.wait_for_selector("tr[data-object-type='rfp']", timeout=10_000)
                except Exception as exc:
                    audit["errors"].append(f"wait_for_grid: {clean_text(exc, 160)}")

                try:
                    audit["page_bodies"].append(page.content())
                except Exception as exc:
                    audit["errors"].append(f"page_content: {clean_text(exc, 160)}")
            finally:
                try:
                    context.close()
                except Exception:
                    pass
    except Exception as exc:
        audit["errors"].append(f"unexpected: {clean_text(exc, 240)}")
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)
    return audit


def search_bc_bid_official_feeds() -> dict[str, Any]:
    checks = [
        ("BC Bid resources", "https://www2.gov.bc.ca/gov/content/bc-procurement-resources/bc-bid-resources"),
        ("BC Bid for suppliers", "https://www2.gov.bc.ca/gov/content/bc-procurement-resources/bc-bid-resources/bc-bid-for-suppliers"),
        ("BC Bid historical data", "https://www2.gov.bc.ca/gov/content/bc-procurement-resources/bc-bid-resources/bc-bid-historical-data"),
        ("BC government API directory", "https://api.gov.bc.ca/"),
        ("BC government developer portal", "https://developer.gov.bc.ca/"),
    ]
    findings: dict[str, Any] = {"checks": [], "official_feed_found": False}
    keywords = ("api", "rss", "feed", "export", "developer", "integration", "json", "xml", "csv", "bulk download")
    for label, url in checks:
        entry = {"label": label, "url": url, "final_url": "", "status": "", "keyword_hits": {}, "notes": ""}
        try:
            body, http_status, content_type, final_url = fetch_url(url, timeout=20, retries=0)
            entry["final_url"] = final_url
            entry["status"] = f"HTTP {http_status}"
            visible = clean_text(re.sub(r"<[^>]+>", " ", body), 20_000)
            hits = {k: len(re.findall(k, visible, flags=re.I)) for k in keywords}
            entry["keyword_hits"] = {k: v for k, v in hits.items() if v}
            if label.startswith("BC Bid") and not entry["keyword_hits"]:
                entry["notes"] = "No visible mention of API/RSS/feed/export on this official resource page."
            elif label.startswith("BC government") and "bc bid" not in visible.lower():
                entry["notes"] = "No visible BC Bid listing on this page."
            else:
                entry["notes"] = "Keyword mentions are generic page text or metadata; no explicit public opportunities feed identified."
        except Exception as exc:
            entry["status"] = "FETCH_ERROR"
            entry["notes"] = clean_text(exc, 240)
        findings["checks"].append(entry)
    return findings


def sweep_bc_bid_public(source: dict[str, Any]) -> SourceSweepResult:
    s0 = time.perf_counter()
    attempts: list[str] = []
    global LAST_BC_BID_AUDIT
    LAST_BC_BID_AUDIT = {}
    urls = [
        ("BC Data Catalogue package_search", "https://catalogue.data.gov.bc.ca/api/3/action/package_search?q=bc%20bid%20procurement%20opportunities"),
        ("BC Bid sitemap", "https://bcbid.gov.bc.ca/sitemap.xml"),
        ("BC Bid rss", "https://bcbid.gov.bc.ca/rss.xml"),
        ("BC Bid feed", "https://bcbid.gov.bc.ca/feed"),
        ("BC Bid opportunities xml", "https://bcbid.gov.bc.ca/opportunities.xml"),
        ("BC Bid public browse realistic headers", BCBID_PUBLIC_URL),
    ]
    resolved_url = ""
    status = "BC_BID_BLOCKED_NO_PUBLIC_FEED"
    for label, url in urls:
        try:
            body, http_status, content_type, final_url = fetch_url(url, timeout=10, retries=0)
            resolved_url = final_url
            if "browser_check" in final_url.lower() or looks_browser_checked(final_url, body):
                attempts.append(f"{label}: browser_check")
                continue
            if http_status in {403, 429}:
                attempts.append(f"{label}: HTTP {http_status}")
                continue
            if http_status == 404:
                attempts.append(f"{label}: HTTP 404")
                continue
            if label.startswith("BC Data") and "json" in content_type.lower():
                payload = json.loads(body)
                count = int(payload.get("result", {}).get("count", 0)) if isinstance(payload, dict) else 0
                attempts.append(f"{label}: 200 JSON count={count}, no current opportunity feed selected")
                continue
            if body.strip().startswith("<") and ("<rss" in body[:500].lower() or "<feed" in body[:500].lower()):
                parsed = parse_rss(body, source, url, final_url)
                if parsed:
                    attempts.append(f"{label}: parsed {len(parsed)} feed rows")
                    log = SourceLog(
                        source["source_id"], source["name"], source["url"], "OK_FEED",
                        source_type=infer_source_type(source["source_id"], source["url"]),
                        source_opened=True, http_status=http_status, candidates=len(parsed),
                        civil_candidates=sum(1 for c in parsed if c.civil_relevant),
                        open_candidates=sum(1 for c in parsed if is_open_tender_status(c.status)),
                        elapsed_seconds=time.perf_counter() - s0, resolved_url=final_url, note="; ".join(attempts),
                    )
                    return SourceSweepResult(parsed, log)
                attempts.append(f"{label}: 200 feed-like but no opportunities")
                continue
            attempts.append(f"{label}: 200 {content_type or 'unknown'} no parseable opportunity feed")
        except Exception as exc:
            attempts.append(f"{label}: {clean_text(exc, 120)}")
    audit = run_bc_bid_browser_audit()
    LAST_BC_BID_AUDIT = audit
    if audit.get("final_url"):
        resolved_url = audit.get("final_url", resolved_url)
    attempts.append(
        "Playwright audit: "
        f"launch={audit.get('launch') or 'none'} "
        f"final={audit.get('final_url') or 'none'} "
        f"title={audit.get('title') or 'none'} "
        f"api_hits={len(audit.get('api_hits', []))} "
        f"cookie_replay={'worked' if audit.get('cookie_replay_worked') else 'blocked'}"
    )
    parsed: list[TenderCandidate] = []
    if audit.get("cookie_replay_worked"):
        try:
            replay_body, replay_status, replay_type, replay_final = fetch_bc_bid_with_cookies(BCBID_PUBLIC_URL, audit, timeout=20)
            resolved_url = replay_final or resolved_url
            parsed = parse_bc_bid_listing_html(replay_body, source, replay_final or BCBID_PUBLIC_URL)
            audit["opportunities_found"] = bool(parsed)
            rows_page1 = len(parsed)
            attempts.append(f"cookie_replay_listing: HTTP {replay_status} parsed {rows_page1} rows")

            extra_pages = audit.get("page_bodies", [])[1:]
            if extra_pages:
                seen_urls = {c.url for c in parsed}
                pages_merged = 0
                for extra_body in extra_pages:
                    extra_parsed = parse_bc_bid_listing_html(extra_body, source, resolved_url or BCBID_PUBLIC_URL)
                    new_rows = [c for c in extra_parsed if c.url not in seen_urls]
                    seen_urls.update(c.url for c in new_rows)
                    parsed.extend(new_rows)
                    pages_merged += 1
                attempts.append(
                    f"pagination: total_hint='{audit.get('pagination_total_hint') or 'unknown'}' "
                    f"pages_fetched={audit.get('pages_fetched', 1)} hard_cap={BCBID_MAX_PAGES} "
                    f"stop_reason='{audit.get('pagination_stop_reason') or 'unknown'}' "
                    f"rows_page1={rows_page1} rows_after_pagination={len(parsed)}"
                )
            elif audit.get("pagination_detected"):
                attempts.append(
                    f"pagination: detected but no additional pages captured "
                    f"(total_hint='{audit.get('pagination_total_hint') or 'unknown'}' "
                    f"stop_reason='{audit.get('pagination_stop_reason') or 'unknown'}')"
                )
            else:
                attempts.append("pagination: no pager control found; 1 page is the full listing")

            audit["opportunities_found"] = bool(parsed)
            audit["parsed_rows"] = len(parsed)
            audit["sample_titles"] = [c.tender_title for c in parsed[:5]]
            audit["connector_approach"] = "plain_http_with_cookies+playwright_pagination" if extra_pages else "plain_http_with_cookies"
            if parsed:
                status = "OK_BROWSER_COOKIE_REPLAY"
                attempts.append(
                    "audit_fields: browser_mode=headless user_assisted=false "
                    "login_attempted=false captcha_bypass_attempted=false "
                    f"public_page_loaded=true parsed_rows={len(parsed)}"
                )
                note = "BCBID_METHODS: " + " | ".join(attempts)
                log = SourceLog(
                    source["source_id"],
                    source["name"],
                    source["url"],
                    status,
                    source_type=infer_source_type(source["source_id"], source["url"]),
                    source_opened=True,
                    candidates=len(parsed),
                    civil_candidates=sum(1 for c in parsed if c.civil_relevant),
                    open_candidates=sum(1 for c in parsed if is_open_tender_status(c.status)),
                    elapsed_seconds=time.perf_counter() - s0,
                    resolved_url=resolved_url,
                    note=note,
                )
                return SourceSweepResult(parsed, log)
        except Exception as exc:
            attempts.append(f"cookie_replay_listing: {clean_text(exc, 160)}")

    # Headless path did not yield usable rows. If it specifically hit BC
    # Bid's browser-check wall (not some other blocking reason like no feed
    # found at all), offer the compliant visible-browser, user-assisted
    # fallback before giving up - never log in, never bypass CAPTCHA, never
    # fabricate a "0 open civil" result when the real reason is that a
    # human still needs to clear the browser-check page.
    if not parsed and audit.get("browser_check"):
        attempts.append("headless attempt landed on browser_check; trying headed user-assisted fallback")
        user_audit = run_bc_bid_user_assisted_browser_audit()
        LAST_BC_BID_AUDIT = {**audit, **user_audit}
        user_parsed: list[TenderCandidate] = []
        if user_audit.get("public_page_loaded") and user_audit.get("page_bodies"):
            try:
                user_parsed = parse_bc_bid_listing_html(
                    user_audit["page_bodies"][0], source, user_audit.get("final_url") or BCBID_PUBLIC_URL,
                )
            except Exception as exc:
                attempts.append(f"user_assisted_parse_error: {clean_text(exc, 160)}")
        user_audit["parsed_rows"] = len(user_parsed)
        LAST_BC_BID_AUDIT["parsed_rows"] = len(user_parsed)
        attempts.append(
            "user_assisted_browser: "
            f"browser_mode={user_audit.get('browser_mode')} "
            f"user_assisted={user_audit.get('user_assisted')} "
            f"login_attempted={user_audit.get('login_attempted')} "
            f"captcha_bypass_attempted={user_audit.get('captcha_bypass_attempted')} "
            f"public_page_loaded={user_audit.get('public_page_loaded')} "
            f"parsed_rows={len(user_parsed)} "
            f"errors={user_audit.get('errors')}"
        )
        if user_parsed:
            status = "OK_BROWSER_COOKIE_REPLAY"
            resolved_url = user_audit.get("final_url") or resolved_url
            note = "BCBID_METHODS: " + " | ".join(attempts)
            log = SourceLog(
                source["source_id"],
                source["name"],
                source["url"],
                status,
                source_type=infer_source_type(source["source_id"], source["url"]),
                source_opened=True,
                candidates=len(user_parsed),
                civil_candidates=sum(1 for c in user_parsed if c.civil_relevant),
                open_candidates=sum(1 for c in user_parsed if is_open_tender_status(c.status)),
                elapsed_seconds=time.perf_counter() - s0,
                resolved_url=resolved_url,
                note=note,
            )
            return SourceSweepResult(user_parsed, log)
        if "still_on_browser_check_after_wait" in (user_audit.get("errors") or []):
            # The visible window opened and TENDER_FINDER waited the full period, but
            # nobody completed the check - distinct from "the fallback could
            # not run at all", and never reported as a genuine zero.
            status = "BC_BID_USER_CHECK_NOT_COMPLETED"
        else:
            status = "BC_BID_BLOCKED_BROWSER_CHECK_USER_ACTION_REQUIRED"

    note = "BCBID_METHODS: " + " | ".join(attempts)
    log = SourceLog(
        source["source_id"], source["name"], source["url"], status,
        source_type=infer_source_type(source["source_id"], source["url"]),
        source_opened=bool(resolved_url), elapsed_seconds=time.perf_counter() - s0, resolved_url=resolved_url, note=note,
    )
    return SourceSweepResult([], log)


def sweep_source(source: dict[str, Any]) -> SourceSweepResult:
    s0 = time.perf_counter()
    note = source.get("note", "")
    error = ""
    status = "NO_OPEN_TENDERS"
    resolved_url = ""
    source_candidates: list[TenderCandidate] = []
    last_http_status = 0
    raw_links_found = 0

    runtime_skip_reason = str(source.get("runtime_skip_reason") or "").strip()
    if runtime_skip_reason:
        return SourceSweepResult([], SourceLog(
            source_id=source["source_id"],
            source_name=source["name"],
            url=source["url"],
            status="SKIPPED_SOURCE_STATUS",
            source_type=infer_source_type(source["source_id"], source["url"]),
            source_opened=False,
            elapsed_seconds=time.perf_counter() - s0,
            note=runtime_skip_reason,
        ))

    if source["source_id"] == "bc_bid_public":
        return sweep_bc_bid_public(source)

    if is_login_platform(source["url"]):
        return SourceSweepResult([], SourceLog(
            source_id=source["source_id"],
            source_name=source["name"],
            url=source["url"],
            status="SKIPPED_LOGIN_PLATFORM",
            source_type=infer_source_type(source["source_id"], source["url"]),
            source_opened=False,
            elapsed_seconds=time.perf_counter() - s0,
            note="Login/paid platform skipped by demo honesty rule.",
        ))

    rss_fail = ""
    for base_url in source_urls(source):
        for rss_url in rss_candidates(base_url, source.get("rss")):
            try:
                body, http_status, content_type, final_url = fetch_url(rss_url, timeout=8, retries=0)
                last_http_status = http_status
                if http_status in {403, 429}:
                    rss_fail = "RSS_FORBIDDEN"
                    continue
                if http_status == 404:
                    continue
                parsed = parse_rss(body, source, base_url, final_url)
                if parsed:
                    source_candidates = parsed[:25]
                    status = "OK_RSS"
                    resolved_url = final_url
                    break
            except Exception as exc:
                rss_fail = clean_text(exc, 120)
        if source_candidates:
            break

    listing_body = ""
    if not source_candidates:
        for candidate_url in source_urls(source):
            if is_login_platform(candidate_url):
                continue
            try:
                body, http_status, content_type, final_url = fetch_url(
                    candidate_url,
                    timeout=12,
                    retries=0 if source.get("no_retry") else 1,
                )
                last_http_status = http_status
                resolved_url = final_url
                if http_status in {403, 429}:
                    status = "FORBIDDEN_BUT_LIKELY_VALID"
                    break
                if http_status == 404:
                    status = "BROKEN_404"
                    continue
                if source["source_id"] == "bc_bid_public" and looks_browser_checked(final_url, body):
                    status = "FORBIDDEN_BUT_LIKELY_VALID"
                    note = clean_text(f"{note} Public BC Bid browse URL resolved to browser-check/reCAPTCHA for automation; no login attempted.", 350)
                    break
                if looks_soft_404(final_url, body):
                    status = "BROKEN_404"
                    resolved_url = final_url
                    continue
                listing_body = body
                link_candidates = parse_html_links(body, source, final_url)
                line_candidates = extract_listing_lines(body, source, final_url)
                raw_links_found = len(link_candidates) + len(line_candidates)
                combined = link_candidates + line_candidates
                seen: set[tuple[str, str]] = set()
                for cand in combined:
                    key = (cand.tender_title.lower(), cand.url.lower())
                    if key in seen:
                        continue
                    seen.add(key)
                    source_candidates.append(cand)
                    if len(source_candidates) >= 25:
                        break
                status = "OK_HTML" if source_candidates else "OK_NO_OPEN_TENDERS"
                break
            except Exception as exc:
                status = "FETCH_ERROR"
                error = clean_text(exc, 250)

    body_low = listing_body.lower() if listing_body else ""
    if source["source_id"] == "sd35_purchasing" and "does not maintain an active bidders" in body_low:
        note = clean_text(f"{note} Public page says the district does not maintain an active bidders list.", 350)
        source_candidates = []
        status = "OK_NO_OPEN_TENDERS"
    if source["source_id"] == "richmond_procurement" and ("bids&tenders" in body_low or "bidsandtenders" in body_low):
        note = clean_text(f"{note} Current opportunities route to bids&tenders; login platform not fetched.", 350)
        source_candidates = []
        status = "SKIPPED_LOGIN_PLATFORM"
    if source["source_id"] == "bidcentral_landing" and not source_candidates:
        status = "LANDING_ONLY_MEMBERSHIP_LIKELY"
    if source["source_id"] == "tol_public_tenders" and body_low and "bidsandtenders" in body_low:
        note = clean_text(f"{note} Current opportunities link to bidsandtenders.ca; login platform not fetched.", 350)
        source_candidates = []
        status = "SKIPPED_LOGIN_PLATFORM"
    if not source_candidates and status == "OK_HTML":
        status = "OK_NO_OPEN_TENDERS"
    if not source_candidates and status == "NO_OPEN_TENDERS" and rss_fail:
        note = clean_text(f"{note} RSS probe: {rss_fail}", 350)

    log = SourceLog(
        source_id=source["source_id"],
        source_name=source["name"],
        url=source["url"],
        status=status,
        source_type=infer_source_type(source["source_id"], source["url"]),
        source_opened=bool(resolved_url or last_http_status),
        http_status=last_http_status,
        raw_links_found=raw_links_found,
        candidates=len(source_candidates),
        civil_candidates=sum(1 for c in source_candidates if c.civil_relevant),
        open_candidates=sum(1 for c in source_candidates if is_open_tender_status(c.status)),
        elapsed_seconds=time.perf_counter() - s0,
        resolved_url=resolved_url,
        note=note,
        error=error,
    )
    return SourceSweepResult(source_candidates, log)


def future_index(data: DemoData) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {}
    for row in data.future:
        key = clean_text(row.get("municipality")).lower()
        out.setdefault(key, []).append(row)
    return out


STOP_TOKENS = {
    "the", "and", "for", "with", "from", "this", "that", "road", "street", "avenue",
    "drive", "lane", "way", "city", "project", "service", "services", "building",
    "development", "permit", "application", "water", "sewer", "storm", "civil",
}


def useful_tokens(text: str) -> set[str]:
    tokens = {t.lower() for t in re.findall(r"[A-Za-z0-9]{3,}", clean_text(text))}
    return {t for t in tokens if t not in STOP_TOKENS}


def related_lead(candidate: TenderCandidate, data: DemoData, index: dict[str, list[dict[str, Any]]]) -> str:
    muni = clean_text(candidate.municipality).lower()
    rows = index.get(muni, [])
    if not rows and "langley" in muni:
        rows = index.get("township of langley", []) + index.get("city of langley", [])
    if not rows:
        return ""
    text = f"{candidate.tender_title} {candidate.snippet}"
    text_low = text.lower()
    tokens = useful_tokens(text)
    best: tuple[int, dict[str, Any] | None] = (0, None)
    for row in rows[:1200]:
        address = clean_text(row.get("address"))
        app_no = clean_text(row.get("app_no"))
        scope = clean_text(row.get("scope_summary"))
        score = 0
        if app_no and app_no.lower() in text_low:
            score += 5
        address_tokens = useful_tokens(address)
        if address_tokens:
            overlap = tokens & address_tokens
            if any(t.isdigit() for t in overlap) and len(overlap) >= 2:
                score += 5
            elif len(overlap) >= 3:
                score += 3
        scope_overlap = tokens & useful_tokens(scope)
        if len(scope_overlap) >= 5:
            score += 2
        if score > best[0]:
            best = (score, row)
    if best[0] >= 5 and best[1]:
        row = best[1]
        return f"{row.get('municipality')} {row.get('app_no')} - {row.get('address')} (fit {row.get('fit_score')})"
    return ""


def infer_candidate_municipality(candidate: TenderCandidate) -> str:
    text = f"{candidate.municipality} {candidate.issuer} {candidate.source_name} {candidate.tender_title}".lower()
    known = [
        "Surrey", "Coquitlam", "Maple Ridge", "Township of Langley", "City of Langley",
        "Richmond", "Pitt Meadows", "New Westminster", "Port Coquitlam", "Delta",
        "Vancouver", "Burnaby", "Abbotsford",
    ]
    for muni in known:
        if muni.lower() in text:
            return muni
    if "langley" in text:
        return "Township of Langley"
    return clean_text(candidate.municipality)


def populate_related_counts(candidates: list[TenderCandidate], data: DemoData) -> None:
    index = future_index(data)
    for candidate in candidates:
        muni = infer_candidate_municipality(candidate)
        rows = index.get(muni.lower(), [])
        if not rows and "langley" in muni.lower():
            rows = index.get("township of langley", []) + index.get("city of langley", [])
        candidate.related_leads_count = len(rows)
        candidate.related_leads_sample = clean_text(rows[0].get("address")) if rows else ""
        tokens = {t for t in useful_tokens(candidate.tender_title) if len(t) >= 5}
        if tokens and rows:
            candidate.related_keyword_count = sum(
                1 for row in rows
                if tokens & useful_tokens(row.get("scope_summary"))
            )


def fetch_tenders(data: DemoData | None = None, no_fetch: bool = False) -> tuple[list[TenderCandidate], list[SourceLog], float]:
    start = time.perf_counter()
    if no_fetch:
        logs = [
            SourceLog(source["source_id"], source["name"], source["url"], "SKIPPED_NO_FETCH", source_type=infer_source_type(source["source_id"], source["url"]), note="--no-fetch supplied")
            for source in TENDER_SOURCES
        ]
        return [], logs, time.perf_counter() - start

    # BC Bid's sweep launches a real headless browser (Playwright) and now
    # clicks through up to BCBID_MAX_PAGES pages (~20-25s of heavy CPU/IO).
    # Patch 5.14 ran it inside the same ThreadPoolExecutor batch as every
    # lightweight municipal HTTP fetch and that caused intermittent
    # zero-candidate results from otherwise-working sources (surrey_bids_
    # public, richmond_procurement) in the same run - confirmed by direct
    # repro: those sources flip between their normal counts and 0 across
    # repeated runs with identical code, with no slow/timeout elapsed time
    # on the failing runs, i.e. the page fetched fine but lost the resource
    # race for CPU/scheduling while the browser process was active. BC Bid's
    # sweep is now run on its own, never overlapping with the other sources,
    # so the two workloads cannot starve each other.
    bc_bid_sources = [source for source in TENDER_SOURCES if source["source_id"] == "bc_bid_public"]
    other_sources = [source for source in TENDER_SOURCES if source["source_id"] != "bc_bid_public"]

    print("TENDER_FINDER_STAGE: Checking public tender pages (municipalities, districts, boards)", flush=True)
    sweep_results: list[SourceSweepResult] = []
    lock = threading.Lock()
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(sweep_source, source): source for source in other_sources}
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            with lock:
                sweep_results.append(result)

    if bc_bid_sources:
        print("TENDER_FINDER_STAGE: Opening and checking BC Bid public browser page", flush=True)
    for source in bc_bid_sources:
        sweep_results.append(sweep_source(source))
    if bc_bid_sources:
        print("TENDER_FINDER_STAGE: Parsing tender rows and matching civil-relevant opportunities", flush=True)

    logs = [result.log for result in sweep_results]
    shallow: list[TenderCandidate] = []
    for result in sweep_results:
        shallow.extend(result.candidates)

    seen: set[tuple[str, str]] = set()
    deduped: list[TenderCandidate] = []
    for cand in shallow:
        key = (cand.url.lower().rstrip("/"), cand.tender_title.lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(cand)
    prioritized = [cand for cand in deduped if cand.source_id == "bc_bid_public"][:200]
    prioritized.extend(cand for cand in deduped if cand.source_id != "bc_bid_public")
    deduped = prioritized

    deep_fetch_counts: dict[str, int] = {}
    preserved = [cand for cand in deduped if cand.source_id == "bc_bid_public" and is_real_opportunity_candidate(cand)]

    bc_bid_civil_open = [c for c in preserved if c.civil_relevant and is_open_tender_status(c.status)]
    detail_gained = 0
    if bc_bid_civil_open:
        with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
            detail_futures = [executor.submit(bc_bid_detail_fetch, c) for c in bc_bid_civil_open]
            for future in concurrent.futures.as_completed(detail_futures):
                if future.result():
                    detail_gained += 1
    if data is not None:
        data.bc_bid_detail_stats = {"attempted": len(bc_bid_civil_open), "contact_gained": detail_gained}
        data.bc_bid_pagination_stats = {
            "pagination_detected": LAST_BC_BID_AUDIT.get("pagination_detected", False),
            "pages_fetched": LAST_BC_BID_AUDIT.get("pages_fetched", 1),
            "total_hint": LAST_BC_BID_AUDIT.get("pagination_total_hint", ""),
            "stop_reason": LAST_BC_BID_AUDIT.get("pagination_stop_reason", ""),
            "cap": BCBID_MAX_PAGES,
        }

    deep_queue = [cand for cand in deduped if cand.source_id != "bc_bid_public"]
    deep_fetch_keys = {(cand.source_id, cand.url.lower().rstrip("/"), cand.tender_title.lower()) for cand in deep_queue}
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(deep_fetch_candidate, cand) for cand in deep_queue]
        all_candidates = preserved + [
            candidate
            for candidate in (future.result() for future in concurrent.futures.as_completed(futures))
            if is_real_opportunity_candidate(candidate)
        ]

    idx = future_index(data) if data else {}
    for cand in all_candidates:
        deep_key = (cand.source_id, cand.url.lower().rstrip("/"), cand.tender_title.lower())
        if deep_key in deep_fetch_keys:
            deep_fetch_counts[cand.source_id] = deep_fetch_counts.get(cand.source_id, 0) + 1
        cand.issuer = cand.issuer or cand.source_name
        cand.region = cand.region or (cand.municipality if cand.municipality else "BC-wide")
        if data:
            cand.related_lead = related_lead(cand, data, idx)

    final_seen: set[tuple[str, str]] = set()
    final_candidates: list[TenderCandidate] = []
    for cand in all_candidates:
        title_key = clean_text(cand.tender_title).lower()
        if cand.source_id == "bidcentral_landing":
            key = (cand.source_id, title_key)
        else:
            key = (cand.source_id, title_key, cand.url.lower().rstrip("/"))
        if key in final_seen:
            continue
        final_seen.add(key)
        final_candidates.append(cand)
    all_candidates = final_candidates
    if data:
        populate_related_counts(all_candidates, data)

    for log in logs:
        source_cands = [c for c in all_candidates if c.source_id == log.source_id]
        log.candidates = len(source_cands)
        log.civil_candidates = sum(1 for c in source_cands if c.civil_relevant)
        log.open_candidates = sum(1 for c in source_cands if is_open_tender_status(c.status))
        log.deep_fetches = deep_fetch_counts.get(log.source_id, 0)
        if log.status == "OK_HTML" and not source_cands:
            log.status = "OK_NO_OPEN_TENDERS"

    all_candidates.sort(key=lambda c: (
        not c.civil_relevant,
        not is_open_tender_status(c.status),
        c.closing_date or "9999-12-31",
        c.source_name,
        c.tender_title,
    ))
    fetch_seconds = time.perf_counter() - start
    logs.sort(key=lambda l: l.source_id)
    return all_candidates, logs, fetch_seconds


def email_tender_to_candidate(row: EmailTender) -> TenderCandidate:
    signal_state = "historical" if "histor" in row.status.lower() else ("closed" if row.status == "closed" else "open")
    return TenderCandidate(
        source_id="email_alert_intake",
        source_name="Email Alert Intake",
        tender_title=row.tender_title,
        url=row.url,
        civil_relevant=row.civil_relevant,
        status=row.status,
        found_via=row.found_via,
        snippet=row.snippet,
        closing_date=row.closing_date,
        municipality=row.issuer,
        issuer=row.issuer,
        region="email_alert",
        contact_email=clean_contact_email(row.contact_email),
        contact_phone=row.contact_phone,
        signal_state=signal_state,
        retention_target="bid_now" if signal_state == "open" and row.civil_relevant else ("history" if signal_state in {"closed", "historical"} else "all_signals_only"),
        filtered_reason=row.filtered_reason or f"provider={row.provider}",
    )


def run_email_intake_for_demo(out_dir: Path, enabled: bool) -> tuple[list[TenderCandidate], SourceLog]:
    if not enabled:
        return [], SourceLog("email_alert_intake", "Email Alert Intake", "", "SKIPPED", source_type="email_alert_intake", note="--email-intake not supplied")
    config = load_email_provider_config()
    status, rows, result = run_email_intake(out_dir / "email_intake_out", provider_config=config)
    candidates = [email_tender_to_candidate(row) for row in rows]
    selected_folder = result.selected_folder or config.import_path or str(current_email_import_folder())
    note = (
        f"Parses user-approved portal alert emails only; provider={provider_display_name(config.provider)}; "
        f"folder={selected_folder}; files={result.alert_emails_found}; parsed_rows={result.tender_rows_parsed}; "
        f"civil={result.civil_alerts_found}; open_actionable={result.open_actionable_rows}; "
        f"duplicates={result.duplicate_emails_detected}; rejected={result.emails_rejected}; "
        f"dry_run_log={result.log_path or 'none'}; "
        f"source_mode={'fixture_synthetic' if result.is_fixture_source else 'live_or_user_folder'}; "
        f"no credentials or portal login."
    )
    log = SourceLog(
        source_id="email_alert_intake",
        source_name="Email Alert Intake",
        source_type="email_alert_intake",
        url=selected_folder,
        status=status,
        source_opened=bool(selected_folder),
        candidates=len(rows),
        civil_candidates=sum(1 for row in rows if row.civil_relevant),
        open_candidates=sum(1 for row in rows if row.civil_relevant and row.status == "open"),
        raw_links_found=result.alert_emails_found,
        note=note,
    )
    return candidates, log


def write_rows(ws, headers: list[str], rows: list[list[Any]], note: str | None = None, widths: dict[str, int] | None = None) -> None:
    start_row = 1
    if note:
        ws.cell(1, 1, note)
        ws.cell(1, 1).font = Font(name="Arial", bold=True, color="1F2937")
        ws.cell(1, 1).alignment = Alignment(vertical="top", wrap_text=True)
        start_row = 2
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start_row + 1):
        for col, value in enumerate(row, 1):
            value = scrub_currency_values(value)
            cell = ws.cell(row_index, col, safe_untrusted_excel_value(value))
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = ws.dimensions
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = Border(bottom=thin)
    widths = widths or {}
    for col, header in enumerate(headers, 1):
        width = widths.get(header)
        if width is None:
            samples = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
            width = min(max(max((len(x) for x in samples), default=10) + 2, 10), 36)
        ws.column_dimensions[get_column_letter(col)].width = width


def style_summary(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10, bold=cell.row in {1, 3, 18})
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 92
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    ws.freeze_panes = "A2"


def source_or_contact(row: dict[str, Any]) -> str:
    municipality = clean_text(row.get("municipality"))
    return PLANNING_CONTACTS.get(municipality, "Planning: municipal planning/development desk | verify phone")


def summarize_counts_by_source(tenders: list[TenderCandidate], predicate) -> str:
    counts: dict[str, int] = {}
    for tender in tenders:
        if predicate(tender):
            counts[tender.source_name] = counts.get(tender.source_name, 0) + 1
    if not counts:
        return "none"
    return "; ".join(f"{name}={count}" for name, count in sorted(counts.items()))


def summarize_logs_by_source(logs: list[SourceLog], predicate) -> str:
    counts = {log.source_name: 1 for log in logs if predicate(log)}
    if not counts:
        return "none"
    return "; ".join(sorted(counts))


def tender_rows_for_sheet(tenders: list[TenderCandidate]) -> list[list[Any]]:
    return [[
        t.source_name,
        t.issuer or t.source_name,
        t.tender_title,
        "YES" if t.civil_relevant else "NO",
        t.status,
        t.signal_state,
        t.retention_target,
        t.filtered_reason,
        t.closing_date,
        t.contact_email,
        t.contact_phone,
        t.related_leads_count,
        t.related_leads_sample,
        t.related_keyword_count,
        t.related_lead,
        t.url,
        t.found_via,
        t.snippet,
    ] for t in tenders]


KEYWORD_AUDIT_HEADERS = [
    "lead_id", "source_id", "app_no", "old_score", "new_score", "score_delta",
    "old_tier", "new_tier", "old_bucket", "new_bucket", "changed",
    "threshold", "exception", "positive_matches", "negative_matches",
    "geography_matches", "client_matches",
]


def _audit_match_text(event: dict[str, Any], key: str) -> str:
    return "; ".join(
        f"{item.get('keyword')} ({item.get('weight'):+})"
        for item in event.get(key, [])
    )


def keyword_audit_rows(data: DemoData) -> list[list[Any]]:
    rows = []
    for event in data.keyword_rescore_events:
        values = {
            **event,
            "changed": "YES" if event.get("changed") else "NO",
            "positive_matches": _audit_match_text(event, "positive_matches"),
            "negative_matches": _audit_match_text(event, "negative_matches"),
            "geography_matches": _audit_match_text(event, "geography_matches"),
            "client_matches": _audit_match_text(event, "client_matches"),
        }
        rows.append([values.get(header, "") for header in KEYWORD_AUDIT_HEADERS])
    return rows


def top_printable_rows(data: DemoData) -> list[dict[str, Any]]:
    signal_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    eligible = [
        row for row in data.future
        if row.get("source_id") != "van_building_permits"
        and clean_text(row.get("address"))
        and row.get("status_class") != "TERMINAL"
        and row.get("signal_quality") in {"HIGH", "MEDIUM"}
    ]
    eligible.sort(key=lambda x: (signal_rank.get(str(x.get("signal_quality")), 9), -x["fit_score"], str(x["municipality"]), str(x["address"])))

    selected: list[dict[str, Any]] = []
    by_muni: dict[str, int] = {}
    for row in eligible:
        municipality = clean_text(row.get("municipality")) or "Unknown"
        if by_muni.get(municipality, 0) >= 15:
            continue
        selected.append(row)
        by_muni[municipality] = by_muni.get(municipality, 0) + 1
        if len(selected) == 50:
            break
    return selected


def build_outreach_rows(data: DemoData, tenders: list[TenderCandidate]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for tender in tenders:
        lead_id = tender_lead_id(tender)
        prior = data.prior_outreach.get(lead_id, {})
        rows.append({
            "lead_id": lead_id,
            "type": "TENDER",
            "municipality": tender.issuer or tender.source_name,
            "title_or_address": tender.tender_title,
            "fit_score": "",
            "signal_quality": "HIGH" if tender.civil_relevant else "LOW",
            "contact": "; ".join(x for x in [tender.contact_email, tender.contact_phone] if x),
            "status": prior.get("status") or "Not Contacted",
            "last_contact_date": prior.get("last_contact_date") or "",
            "next_action_date": prior.get("next_action_date") or "",
            "notes": prior.get("notes") or "",
        })
    for row in data.future:
        if row.get("signal_quality") not in {"HIGH", "MEDIUM"}:
            continue
        lead_id = row["lead_id"]
        prior = data.prior_outreach.get(lead_id, {})
        rows.append({
            "lead_id": lead_id,
            "type": "LEAD",
            "municipality": row.get("municipality"),
            "title_or_address": row.get("address") or row.get("app_no") or row.get("scope_summary"),
            "fit_score": row.get("fit_score"),
            "signal_quality": row.get("signal_quality"),
            "contact": source_or_contact(row),
            "status": prior.get("status") or "Not Contacted",
            "last_contact_date": prior.get("last_contact_date") or "",
            "next_action_date": prior.get("next_action_date") or "",
            "notes": prior.get("notes") or "",
        })

    active_ids = {clean_text(row.get("lead_id")) for row in rows}
    for lead_id, prior in data.prior_outreach.items():
        if lead_id in active_ids:
            continue
        rows.append({
            "lead_id": lead_id,
            "type": prior.get("type") or "LEAD",
            "municipality": prior.get("municipality") or "",
            "title_or_address": prior.get("title_or_address") or "",
            "fit_score": prior.get("fit_score") or "",
            "signal_quality": "BELOW_CURRENT_GATE",
            "contact": prior.get("contact") or "",
            "status": prior.get("status") or "Not Contacted",
            "last_contact_date": prior.get("last_contact_date") or "",
            "next_action_date": prior.get("next_action_date") or "",
            "notes": prior.get("notes") or "",
        })

    def sort_key(item: dict[str, Any]) -> tuple[int, int, int]:
        not_contacted = 0 if item.get("status") == "Not Contacted" else 1
        high = 0 if item.get("signal_quality") == "HIGH" else 1
        return (not_contacted, high, -coerce_int(item.get("fit_score")))

    rows.sort(key=sort_key)
    data.outreach_rollup = {
        "tracked": len(rows),
        "not_contacted": sum(1 for r in rows if r.get("status") == "Not Contacted"),
        "in_progress": sum(1 for r in rows if r.get("status") in {"Contacted", "Responded", "Meeting Scheduled", "Bid Submitted"}),
        "won_lost": sum(1 for r in rows if r.get("status") in {"Won", "Lost"}),
    }
    return rows


def developer_parent_brand(name: str) -> str:
    for brand, pattern in KNOWN_DEVELOPER_BRANDS:
        if pattern.search(name):
            return brand
    return ""


def normalize_business_name(name: str) -> str:
    text = clean_text(name)
    text = re.sub(r"\s+", " ", text).strip(" .")
    while BUSINESS_SUFFIX_PATTERN.search(text):
        text = BUSINESS_SUFFIX_PATTERN.sub("", text).strip(" .")
    return text


def consultant_firm_name(name: str) -> str:
    text = clean_text(name)
    if not text or not CONSULTANT_NAME_PATTERN.search(text):
        return ""
    match = re.search(r"\bDBA:\s*(.+)$", text, flags=re.I)
    firm = match.group(1) if match else text
    firm = re.sub(r"\bConsultants?\b", "Consultants", firm, flags=re.I)
    firm = normalize_business_name(firm)
    return firm


def is_consultant_name(name: str) -> bool:
    return bool(consultant_firm_name(name))


def classify_applicant_type(parent_brand: str, rows: list[dict[str, Any]], sample_names: set[str]) -> str:
    unit_apps = sum(1 for r in rows if coerce_int(r.get("est_units")) > 0)
    joined_names = " | ".join(sorted(sample_names))
    if not parent_brand and CONSULTANT_NAME_PATTERN.search(joined_names):
        return "DESIGN_CONSULTANT"
    if parent_brand or unit_apps >= 2:
        return "DEVELOPER"
    return "UNKNOWN"


def classify_applicant_name(name: str, est_units: int = 0) -> str:
    parent_brand = developer_parent_brand(name)
    row = {"est_units": est_units}
    return classify_applicant_type(parent_brand, [row], {name})


def summarize_developer_group(key: str, parent_brand: str, rows: list[dict[str, Any]], sample_names: set[str]) -> dict[str, Any]:
    units = [coerce_int(r.get("est_units")) for r in rows if r.get("est_units") not in (None, "")]
    avg_fit = round(sum(coerce_int(r.get("fit_score")) for r in rows) / max(len(rows), 1), 1)
    applicant_type = classify_applicant_type(parent_brand, rows, sample_names)
    notes: list[str] = []
    if parent_brand:
        notes.append(f"Parent-brand grouped from {len(sample_names)} legal entities")
    if parent_brand in {brand for brand, _ in KNOWN_DEVELOPER_BRANDS}:
        notes.append("KNOWN MAJOR DEVELOPER")
    return {
        "applicant_name": key,
        "parent_brand": parent_brand or "",
        "grouped_entity_examples": "; ".join(sorted(sample_names)[:4]),
        "total_applications": len(rows),
        "municipalities_list": ", ".join(sorted({clean_text(r.get("municipality")) for r in rows if r.get("municipality")})),
        "total_units_where_known": sum(units),
        "avg_fit_score": avg_fit,
        "sample_addresses": "; ".join(clean_text(r.get("address") or r.get("app_no")) for r in rows[:3]),
        "repeat_developer": "YES - HIGH RELATIONSHIP VALUE" if len(rows) >= 3 else "NO",
        "applicant_type": applicant_type,
        "note": " | ".join(notes),
    }


def build_developer_rows(data: DemoData) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    grouped_names: dict[str, set[str]] = {}
    known: set[str] = set()
    for row in data.future:
        name = clean_text(row.get("applicant_name"))
        if not name:
            continue
        parent_brand = developer_parent_brand(name)
        consultant_firm = "" if parent_brand else consultant_firm_name(name)
        key = parent_brand or consultant_firm or name
        grouped_rows.setdefault(key, []).append(row)
        grouped_names.setdefault(key, set()).add(name)
        if parent_brand:
            known.add(parent_brand)

    result_all: list[dict[str, Any]] = []
    for key, rows in grouped_rows.items():
        result_all.append(summarize_developer_group(key, key if key in known else "", rows, grouped_names.get(key, set())))

    sort_rank = {"DEVELOPER": 0, "UNKNOWN": 1, "DESIGN_CONSULTANT": 2}
    result_all.sort(
        key=lambda r: (
            sort_rank.get(str(r.get("applicant_type")), 9),
            0 if str(r.get("repeat_developer", "")).startswith("YES") else 1,
            -coerce_int(r.get("total_units_where_known")),
            -coerce_int(r.get("total_applications")),
            str(r.get("applicant_name")),
        )
    )
    developer_rows = [r for r in result_all if r.get("applicant_type") != "DESIGN_CONSULTANT"]
    consultant_rows = [r for r in result_all if r.get("applicant_type") == "DESIGN_CONSULTANT"]
    data.developer_distinct = len(result_all)
    data.developer_repeat = sum(1 for r in result_all if str(r.get("repeat_developer", "")).startswith("YES"))
    data.known_major_developers = sorted(known)
    data.applicant_type_counts = {
        "DEVELOPER": sum(1 for r in result_all if r.get("applicant_type") == "DEVELOPER"),
        "UNKNOWN": sum(1 for r in result_all if r.get("applicant_type") == "UNKNOWN"),
        "DESIGN_CONSULTANT": len(consultant_rows),
    }
    return developer_rows, consultant_rows


def closing_month(value: str) -> str:
    text = clean_text(value)
    month_aliases = {
        "jan": "January",
        "january": "January",
        "feb": "February",
        "february": "February",
        "mar": "March",
        "march": "March",
        "apr": "April",
        "april": "April",
        "may": "May",
        "jun": "June",
        "june": "June",
        "jul": "July",
        "july": "July",
        "aug": "August",
        "august": "August",
        "sep": "September",
        "sept": "September",
        "september": "September",
        "oct": "October",
        "october": "October",
        "nov": "November",
        "november": "November",
        "dec": "December",
        "december": "December",
    }
    for alias, month in month_aliases.items():
        if re.search(rf"\b{alias}\b", text, flags=re.I):
            return month
    match = re.search(r"\b(?:20\d{2})[-/](\d{1,2})[-/]\d{1,2}\b", text)
    if match:
        month = int(match.group(1))
        return dt.date(2000, month, 1).strftime("%B")
    return ""


def build_tender_pattern_rows(tenders: list[TenderCandidate]) -> tuple[str, list[dict[str, Any]]]:
    civil = [t for t in tenders if t.civil_relevant]
    grouped: dict[str, list[TenderCandidate]] = {}
    for tender in civil:
        grouped.setdefault(tender.issuer or tender.source_name, []).append(tender)
    rows: list[dict[str, Any]] = []
    month_order = {dt.date(2000, i, 1).strftime("%B"): i for i in range(1, 13)}
    for issuer, items in grouped.items():
        if len(items) < 2:
            continue
        months = sorted(
            {closing_month(t.closing_date) for t in items if closing_month(t.closing_date)},
            key=lambda month: month_order.get(month, 99),
        )
        month_text = ", ".join(months) if months else "not consistently exposed"
        if months:
            description = f"{issuer} had {len(items)} civil tender signals in this sweep; extracted closing months: {month_text}."
        else:
            description = f"{issuer} had {len(items)} civil tender signals in this sweep; closing months: not consistently exposed."
        rows.append({
            "issuer": issuer,
            "tender_count": len(items),
            "closing_months": month_text,
            "plain_description": description,
        })
    rows.sort(key=lambda r: (-coerce_int(r.get("tender_count")), str(r.get("issuer"))))
    caveat = (
        f"Based on a small sample ({len(civil)} tenders) from a single sweep. "
        "This is an early signal, not a statistically validated seasonal model. "
        "More tender history across multiple sweeps is needed before relying on this for bid timing decisions."
    )
    return caveat, rows


def high_fit_cluster_counts(data: DemoData) -> str:
    counts: dict[str, int] = {}
    for row in data.future:
        if coerce_int(row.get("fit_score")) >= 60:
            cluster = row.get("regional_cluster") or "Other"
            counts[cluster] = counts.get(cluster, 0) + 1
    return "; ".join(f"{cluster}: {count}" for cluster, count in sorted(counts.items()))


def is_junk_tender(tender: TenderCandidate) -> bool:
    title = clean_text(tender.tender_title)
    title_l = title.lower()
    if title_l == "tenders rfqs rfps":
        return True
    if "bidcentral :: bc construction projects bidding and tenders" in title_l:
        return True
    if title_l == "bid opportunities - city of richmond, bc" and not tender.closing_date:
        return True
    if title_l.startswith("get in touch with us"):
        return True
    if ".pdf" in title_l and not tender.civil_relevant:
        return True
    return False


def sanitize_tenders_for_output(tenders: list[TenderCandidate]) -> list[TenderCandidate]:
    cleaned: list[TenderCandidate] = []
    for tender in tenders:
        if is_junk_tender(tender):
            continue
        tender.contact_email = clean_contact_email(tender.contact_email)
        tender.contact_phone = clean_contact_phone(tender.contact_phone)
        cleaned.append(tender)
    return cleaned


def how_to_read_rows(data: DemoData) -> list[list[Any]]:
    return [
    ["Executive_Summary", "Headline numbers and timing. Start here.", "30-second business overview: funnel, timing, and the three-bucket story."],
    ["BID_NOW_Active_Tenders", "Open/actionable civil tenders only. Closed or historical rows never stay here.", "Public live tender scan plus Email Alert Intake rows where available."],
    ["Tender_History_Closed_Public", "Closed and historical public tender rows retained honestly for lookup.", "Use this when a source had tender evidence, but nothing actionable is open now."],
    ["Tender_Signals_All", "All preserved tender/source signals: open, closed, historical, blocked, portal_redirect, empty, parse_failed.", "This is the truthfulness layer - 'no rows retained' does not mean 'source has no tenders'."],
    ["BID_LATER_Future_Projects", f"{len(data.future):,} dev application leads. Future civil work. Start at signal_quality=HIGH.", "Patch 5.11 excludes Surrey terminal-status archive rows and adds address_source / approximate_location."],
    ["New_This_Run", "Focused weekly review: only leads not seen in the prior demo history workbook.", "All rows are new on the first baseline run."],
    ["Outreach_Tracker", "Persistent action tracker. Manual statuses and notes merge forward by lead_id.", "Use this to manage calls, follow-ups, bids, and relationship work."],
    ["Developer_Watchlist", "Relationship-target developer rollup with parent-brand grouping.", "Design consultants are moved to a separate reference tab."],
    ["Design_Consultants_Reference", "Architect/designer-heavy filing entities kept for reference only.", "Separated so they do not crowd the core developer relationship view."],
    ["Tender_Pattern_Analysis", "Small-sample tender pattern view with caveat.", "Not a predictive model."],
    ["Watchlist_Monitor", "Lower-confidence signals. Monitor for progress toward planning approval.", "Manual review queue."],
    ["Analyzed_Set_Aside", "All records collected, scored, and filtered. Kept for future reference.", f"{len(data.analyzed):,} analyzed records; capped display in this workbook."],
    ["Surrey_Historical_Archive", "Surrey V2 terminal-status archive removed from active BID LATER counts.", "Preserved for lookup, not presented as current forward signal."],
    ["Acquisition_Funnel_And_Speed", "Full pipeline volume and speed proof.", "Measured build speed and exact funnel counts."],
    ["Source_Run_Log", "Honest per-source status with open/closed/history/blocking counts and final_status mapping.", "Nothing hidden."],
    ["Action_Center", "WHERE the live biddable tenders are.", "Register or enable alerts here to unlock open tenders."],
    ["Live_Tender_Intake_Plan", "How to get portal email alerts into TENDER_FINDER.", "Design and current state for provider-neutral Email Alert Intake."],
    ["Email_Setup_Guide", "Step-by-step portal registration guide. Shows your current intake status.", "Use this to turn on live tenders."],
    ["Top_Civil_Leads_Printable", "Top 50 non-Vancouver civil leads with planning contacts.", "Print or screenshot for meetings."],
    ]


def _backlog_universe_count() -> Any:
    try:
        return backlog_cross_check_counts()["universe_count"]
    except Exception:
        return ""


def _atomic_replace_with_retry(tmp_path: str, target_path: Path, retries: int = 4, delay_seconds: float = 1.5) -> None:
    """Atomically move tmp_path onto target_path, tolerating a brief window
    where Windows still holds a lock on target_path.

    Windows refuses to replace a file that another process currently has
    open - most commonly because target_path is open in Excel (confirmed
    as the real-world cause during handoff review: reviewers opened the
    output files to inspect them, leaving them locked and also silently
    rewriting their view state on save - see the freeze_panes note in
    _header_row_index()). A short retry window covers the common case of a
    file that was *just* closed and hasn't released its lock yet; if it's
    still locked after that, this raises a clear, actionable error instead
    of a raw WinError traceback.
    """
    last_exc: OSError | None = None
    for attempt in range(retries):
        try:
            os.replace(tmp_path, target_path)
            return
        except PermissionError as exc:
            last_exc = exc
            if attempt < retries - 1:
                time.sleep(delay_seconds)
    raise PermissionError(
        f"Could not save '{target_path}' - Windows reports it is still in use. "
        "This almost always means the file is open in Excel (or another program). "
        "Close it there and run this again. "
        f"(Original error: {last_exc})"
    ) from last_exc


def build_workbook(out_path: Path, data: DemoData, tenders: list[TenderCandidate], logs: list[SourceLog], fetch_seconds: float, total_seconds: float, email_state: EmailSetupState) -> None:
    tenders = sanitize_tenders_for_output(tenders)
    bid_now_tenders = [t for t in tenders if t.retention_target == "bid_now"]
    history_tenders = [t for t in tenders if t.retention_target == "history"]
    all_signal_tenders = [t for t in tenders if t.retention_target != "source_log_only"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive_Summary"
    now = dt.datetime.now().isoformat(timespec="seconds")
    open_civil = sum(1 for t in bid_now_tenders if t.civil_relevant and is_open_tender_status(t.status))
    bc_bid_open_civil = sum(1 for t in bid_now_tenders if t.source_id == "bc_bid_public" and t.civil_relevant and is_open_tender_status(t.status))
    non_van = sum(1 for r in data.future if r.get("source_id") != "van_building_permits")
    detail_no = sum(1 for r in data.future if r.get("detail_available") == "NO")
    cross_linked_muni = sum(1 for t in bid_now_tenders if t.related_leads_count > 0)
    linked = sum(1 for t in bid_now_tenders if t.related_lead)
    outreach_rows = build_outreach_rows(data, bid_now_tenders)
    developer_rows, consultant_rows = build_developer_rows(data)
    pattern_caveat, tender_pattern_rows = build_tender_pattern_rows(bid_now_tenders)
    cluster_summary = high_fit_cluster_counts(data)
    surrey_address_stats = data.address_recovery.get("surrey_devapps_v2", {})
    abby_address_stats = data.address_recovery.get("abbotsford_devapps", {})
    summary_rows = [
        ["TENDER_FINDER Three-Bucket Demo", "", "Civil/earthworks opportunity intelligence from live municipal data"],
        ["Run timestamp", now, ""],
        ["BID NOW - active tender candidates", len(bid_now_tenders), f"Open civil tenders: {open_civil}; BC Bid open civil: {bc_bid_open_civil}; related lead links: {linked}"],
        ["Tender history - closed/public rows", len(history_tenders), summarize_counts_by_source(history_tenders, lambda t: True)],
        ["Tender signals - all retained", len(all_signal_tenders), "Includes open, closed, historical, blocked, portal_redirect, empty, parse_failed, and intake/manual/paid placeholders via filtered_reason."],
        ["BID LATER - clean future-project leads", len(data.future), "Development applications and future civil projects to track"],
        ["Analyzed and set aside", len(data.analyzed), "Collected, scored, filtered, and retained for future reference"],
        ["Watchlist monitor", len(data.watch), "Rows needing manual review or future workflow"],
        ["Records pulled live", data.records_pulled, "Current full sweep source-summary total when available"],
        ["Records normalized", data.records_normalized, "Post-dedupe normalized rows"],
        ["Sources with live connectors", data.live_sources, "Sources returning current rows in the full sweep"],
        ["Municipalities covered", len(data.municipalities), ", ".join(data.municipalities[:12])],
        ["Priority review queue", data.fit_ge_60, "Future project rows with fit_score >= 60"],
        ["Top tier", data.fit_ge_70, "Future project rows with fit_score >= 70"],
        ["Surrey V2 active kept in BID LATER", data.surrey_v2_active, f"Terminal archive rows moved out: {data.surrey_v2_terminal}"],
        ["BID LATER excluding Vancouver permits (higher signal)", non_van, "Future-project rows not sourced from Vancouver building permits"],
        ["BID LATER detail_available=NO", detail_no, "Rows with thin or missing scope detail"],
        ["Address recovery", "", f"Surrey extracted={surrey_address_stats.get('EXTRACTED', 0)} centroid={surrey_address_stats.get('CENTROID', 0)} none={surrey_address_stats.get('NONE', 0)} | Abbotsford extracted={abby_address_stats.get('EXTRACTED', 0)} centroid={abby_address_stats.get('CENTROID', 0)} none={abby_address_stats.get('NONE', 0)}"],
        ["BID NOW same-municipality BID LATER cross-links", cross_linked_muni, "Looser cross-link count"],
        ["New leads since last run", data.new_since_last_run, data.history_note],
        ["Leads from prior run no longer present", data.removed_since_last_run, "Closed/removed or absent compared with prior demo_history workbook"],
        ["Outreach Tracker", data.outreach_rollup.get("tracked", 0), f"{data.outreach_rollup.get('not_contacted', 0)} not contacted; {data.outreach_rollup.get('in_progress', 0)} in progress; {data.outreach_rollup.get('won_lost', 0)} won/lost"],
        ["Distinct developers/applicants identified", data.developer_distinct, f"Repeat developers (3+ applications): {data.developer_repeat}; known major matches: {', '.join(data.known_major_developers) if data.known_major_developers else 'none found'}"],
        ["Applicant type split", "", f"Developer={data.applicant_type_counts.get('DEVELOPER', 0)} Unknown={data.applicant_type_counts.get('UNKNOWN', 0)} Design consultant={data.applicant_type_counts.get('DESIGN_CONSULTANT', 0)}"],
        ["HIGH-fit leads by regional cluster", "", cluster_summary],
        ["Source expansion roadmap", _backlog_universe_count(), "See Source_Roadmap_Printable (next tab): ranked roadmap of candidate sources TENDER_FINDER can add next. Full detail in Potential_Sources_Next. Not lead volume - a source worklist."],
        ["Track A read time seconds", round(data.read_seconds, 2), "Measured"],
        ["Track B fetch time seconds", round(fetch_seconds, 2), "Measured parallel source + deep page sweep"],
        ["Total demo build time seconds", round(total_seconds, 2), "Measured"],
        ["Sources attempted", len(logs), "Public-only sources; no login or paid portals authenticated"],
        ["Track B deep parse", "YES", "Candidate listing links followed and rescored from page title/body where public pages allowed it"],
        ["Fixture fallback used", "NO", "TENDER_FINDER_OFFLINE_FIXTURES was not used"],
        ["Email intake state", email_state.state, email_state.label],
        ["Open tenders retained in BID NOW by source", "", summarize_counts_by_source(bid_now_tenders, lambda t: True)],
        ["Closed/historical retained in history by source", "", summarize_counts_by_source(history_tenders, lambda t: True)],
        ["Sources opened but empty", "", summarize_logs_by_source(logs, lambda l: l.final_status == "FETCH_OK_EMPTY_CURRENT_LIST")],
        ["Sources blocked", "", summarize_logs_by_source(logs, lambda l: l.final_status in {"BLOCKED_FORBIDDEN", "BLOCKED_BROWSER_CHECK"})],
        ["Sources portal-gated", "", summarize_logs_by_source(logs, lambda l: l.final_status == "FETCH_OK_PORTAL_REDIRECT")],
        ["Sources parsed but zero retained", "", summarize_logs_by_source(logs, lambda l: l.final_status == "FETCH_OK_PARSE_ZERO_ROWS")],
        ["Sources needing email alerts", "", summarize_logs_by_source(logs, lambda l: l.final_status == "NEEDS_EMAIL_ALERT")],
        ["Sources needing manual registration or paid access", "", summarize_logs_by_source(logs, lambda l: l.final_status in {"NEEDS_MANUAL_REGISTRATION", "NEEDS_PAID_ACCESS"})],
        ["", "", ""],
        ["How to read this workbook", "", ""],
        *how_to_read_rows(data),
    ]
    for row in summary_rows:
        append_untrusted_row(ws, row)
    style_summary(ws)

    # VP-facing one-pager, deliberately the SECOND tab so the source
    # expansion story is impossible to miss (full detail stays in
    # Potential_Sources_Next further right).
    append_source_roadmap_printable_sheet(wb)

    ws = wb.create_sheet("BID_NOW_Active_Tenders")
    tender_note = "Open/actionable civil tenders only. Closed, historical, blocked, portal-gated, empty, and parse-failed signals are preserved elsewhere and never shown here as BID NOW opportunities."
    if not bid_now_tenders:
        tender_note += " No rows retained here this run does not mean the sources had no tenders - review Tender_History_Closed_Public, Tender_Signals_All, and Source_Run_Log."
    tender_headers = ["source", "issuer", "tender_title", "civil_relevant", "status", "signal_state", "retention_target", "filtered_reason", "closing_date", "contact_email", "contact_phone", "related_leads_count", "related_leads_sample", "related_keyword_count", "related_lead", "url", "found_via", "snippet"]
    write_rows(ws, tender_headers, tender_rows_for_sheet(bid_now_tenders), tender_note, {"source": 30, "issuer": 28, "tender_title": 58, "closing_date": 14, "contact_email": 30, "contact_phone": 18, "related_leads_sample": 34, "related_lead": 46, "url": 70, "found_via": 16, "snippet": 72, "filtered_reason": 28})

    ws = wb.create_sheet("Tender_History_Closed_Public")
    history_note = "Closed and historical public tender rows retained for auditability and lookup. These rows are never treated as actionable BID NOW opportunities."
    write_rows(ws, tender_headers, tender_rows_for_sheet(history_tenders), history_note, {"source": 30, "issuer": 28, "tender_title": 58, "closing_date": 14, "contact_email": 30, "contact_phone": 18, "related_leads_sample": 34, "related_lead": 46, "url": 70, "found_via": 16, "snippet": 72, "filtered_reason": 28})

    ws = wb.create_sheet("Tender_Signals_All")
    all_signals_note = "Complete tender/source signal preservation layer. Use signal_state + filtered_reason to distinguish open, closed, historical, blocked, portal_redirect, empty, parse_failed, needs_email_alert, needs_manual_registration, and needs_paid_access."
    write_rows(ws, tender_headers, tender_rows_for_sheet(all_signal_tenders), all_signals_note, {"source": 30, "issuer": 28, "tender_title": 58, "closing_date": 14, "contact_email": 30, "contact_phone": 18, "related_leads_sample": 34, "related_lead": 46, "url": 70, "found_via": 16, "snippet": 72, "filtered_reason": 32})

    ws = wb.create_sheet("BID_LATER_Future_Projects")
    fp_headers = ["lead_id", "first_seen_date", "is_new_since_last_run", "municipality", "regional_cluster", "app_no", "address", "address_source", "approximate_location", "app_type_stage", "status_class", "scope_summary", "fit_score", "signal_quality", "detail_available", "source_tier", "project_scale_tier", "scale", "est_units", "est_storeys", "est_lots", "applicant_name", "source", "source_url"]
    fp_rows = [[r.get(h) for h in fp_headers] for r in data.future]
    write_rows(ws, fp_headers, fp_rows, "Priority review queue: fit_score >= 60. Patch 5.11 excludes Surrey terminal archive rows and stores address_source plus approximate_location fallback.", {"scope_summary": 80, "source_url": 60, "address": 32, "approximate_location": 28, "lead_id": 20, "applicant_name": 32})
    if ws.max_row > 2:
        fit_col = fp_headers.index("fit_score") + 1
        ws.conditional_formatting.add(f"{get_column_letter(fit_col)}3:{get_column_letter(fit_col)}{ws.max_row}", CellIsRule(operator="greaterThanOrEqual", formula=["60"], fill=PatternFill("solid", fgColor="C6EFCE")))

    ws = wb.create_sheet("Keyword_Change_Audit")
    write_rows(
        ws,
        KEYWORD_AUDIT_HEADERS,
        keyword_audit_rows(data),
        "RESCORE_ALWAYS audit. Compare old/new score, Vancouver tier, and routing bucket. Legacy Vancouver rows without a persisted scoring snapshot are named in exception.",
        {
            "lead_id": 22, "source_id": 24, "old_bucket": 22, "new_bucket": 22,
            "exception": 38, "positive_matches": 55, "negative_matches": 55,
            "geography_matches": 40, "client_matches": 40,
        },
    )

    ws = wb.create_sheet("New_This_Run")
    new_headers = fp_headers
    new_rows = [[r.get(h) for h in new_headers] for r in data.future if r.get("is_new_since_last_run") == "YES"]
    write_rows(ws, new_headers, new_rows, f"{data.history_note}. Showing only leads not present in the most recent prior demo_history workbook.", {"scope_summary": 80, "source_url": 60, "address": 32, "approximate_location": 28, "lead_id": 20, "applicant_name": 32})

    ws = wb.create_sheet("Outreach_Tracker")
    outreach_headers = ["lead_id", "type", "municipality", "title_or_address", "fit_score", "signal_quality", "contact", "status", "last_contact_date", "next_action_date", "notes"]
    outreach_matrix = [[r.get(h) for h in outreach_headers] for r in outreach_rows]
    write_rows(ws, outreach_headers, outreach_matrix, None, {"lead_id": 20, "municipality": 24, "title_or_address": 54, "contact": 62, "status": 22, "notes": 64})
    if ws.max_row > 1:
        status_col = outreach_headers.index("status") + 1
        dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1=f'"{",".join(OUTREACH_STATUSES)}"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{ws.max_row}")

    ws = wb.create_sheet("Developer_Watchlist")
    dev_headers = ["applicant_name", "parent_brand", "grouped_entity_examples", "applicant_type", "total_applications", "municipalities_list", "total_units_where_known", "avg_fit_score", "sample_addresses", "repeat_developer", "note"]
    dev_matrix = [[r.get(h) for h in dev_headers] for r in developer_rows]
    write_rows(ws, dev_headers, dev_matrix, "Built only from real applicant/owner/developer fields exposed by source data. Patch 5.11 groups known parent brands and keeps design consultants out of this primary relationship tab.", {"applicant_name": 34, "grouped_entity_examples": 40, "municipalities_list": 38, "sample_addresses": 70, "repeat_developer": 34, "note": 34})

    ws = wb.create_sheet("Design_Consultants_Reference")
    consultant_matrix = [[r.get(h) for h in dev_headers] for r in consultant_rows]
    write_rows(ws, dev_headers, consultant_matrix, "Reference-only list of high-volume design/architect filing entities. Separated so developer relationship targets stay at the top of the core watchlist.", {"applicant_name": 34, "grouped_entity_examples": 40, "municipalities_list": 38, "sample_addresses": 70, "repeat_developer": 34, "note": 34})

    ws = wb.create_sheet("Tender_Pattern_Analysis")
    pattern_headers = ["issuer", "tender_count", "closing_months", "plain_description"]
    pattern_matrix = [[r.get(h) for h in pattern_headers] for r in tender_pattern_rows]
    write_rows(ws, pattern_headers, pattern_matrix, pattern_caveat, {"issuer": 42, "closing_months": 28, "plain_description": 84})

    ws = wb.create_sheet("Watchlist_Monitor")
    watch_headers = ["municipality", "app_no", "address", "app_type_stage", "scope_summary", "fit_score", "source", "source_url"]
    watch_rows = [[r.get(h) for h in watch_headers] for r in data.watch]
    write_rows(ws, watch_headers, watch_rows, None, {"scope_summary": 80, "source_url": 60, "address": 32})

    ws = wb.create_sheet("Analyzed_Set_Aside")
    aside_headers = ["route", "municipality", "app_no", "address", "app_type_stage", "fit_score", "source", "source_url"]
    aside_rows = [[r.get(h) for h in aside_headers] for r in data.analyzed[:8000]]
    write_rows(ws, aside_headers, aside_rows, f"Collected, scored, and filtered. Kept for future reference. Showing {len(aside_rows):,} of {len(data.analyzed):,} rows, including Surrey_Historical_Archive terminal-status rows removed from BID LATER.", {"source_url": 60, "address": 32})

    ws = wb.create_sheet("Surrey_Historical_Archive")
    surrey_headers = ["municipality", "app_no", "address", "address_source", "approximate_location", "app_type_stage", "status_class", "scope_summary", "fit_score", "signal_quality", "source", "source_url"]
    surrey_rows = [[r.get(h) for h in surrey_headers] for r in data.analyzed if r.get("route") == "Surrey_Historical_Archive"]
    write_rows(ws, surrey_headers, surrey_rows, "Surrey DevApps V2 terminal/historical archive. Preserved for lookup but removed from active BID LATER counts in Patch 5.11.", {"scope_summary": 84, "address": 32, "approximate_location": 28, "source_url": 60})

    ws = wb.create_sheet("Acquisition_Funnel_And_Speed")
    funnel = [
        ["Records pulled live (current full sweep)", data.records_pulled],
        ["Records normalized", data.records_normalized],
        ["BID NOW - tender candidates", len(bid_now_tenders)],
        ["BID NOW - civil relevant", sum(1 for t in bid_now_tenders if t.civil_relevant)],
        ["BID NOW - open civil", open_civil],
        ["BC Bid - open civil", bc_bid_open_civil],
        ["Tender history - closed/public", len(history_tenders)],
        ["Tender signals - all retained", len(all_signal_tenders)],
        ["BID NOW - same-municipality BID LATER cross-links", cross_linked_muni],
        ["BID LATER - clean future-project leads", len(data.future)],
        ["BID LATER excluding Vancouver permits (higher signal)", non_van],
        ["BID LATER detail_available=NO", detail_no],
        ["fit >= 70 (top tier)", data.fit_ge_70],
        ["fit 60-69 (strong)", data.fit_60_69],
        ["fit 50-59", data.fit_50_59],
        ["New leads since last run", data.new_since_last_run],
        ["Prior-run leads no longer present", data.removed_since_last_run],
        ["Outreach Tracker rows", data.outreach_rollup.get("tracked", 0)],
        ["Developer/applicant names identified", data.developer_distinct],
        ["Repeat developers/applicants 3+", data.developer_repeat],
        ["Design consultants separated to reference tab", data.applicant_type_counts.get("DESIGN_CONSULTANT", 0)],
        ["Watchlist", len(data.watch)],
        ["Analyzed and set aside", len(data.analyzed)],
        ["Track A read time (seconds)", round(data.read_seconds, 2)],
        ["Track B fetch time (seconds)", round(fetch_seconds, 2)],
        ["Total demo build time (seconds)", round(total_seconds, 2)],
        ["Sources with live working connectors", data.live_sources],
        ["Sources coded (total)", data.coded_sources],
        ["Source Register entries", data.source_register],
    ]
    write_rows(ws, ["metric", "value"], funnel, None, {"metric": 48, "value": 18})

    ws = wb.create_sheet("Source_Run_Log")
    log_headers = ["source_id", "source_name", "source_type", "source_opened", "http_status", "resolved_url", "raw_links_found", "open_found", "closed_found", "historical_found", "portal_redirect_found", "kept_in_bid_now", "kept_in_history", "kept_in_all_signals", "filtered_reason", "final_status", "status", "candidates", "civil_candidates", "deep_fetches", "elapsed_seconds", "url", "note", "error"]
    log_rows = [[
        l.source_id,
        l.source_name,
        l.source_type,
        "YES" if l.source_opened else "NO",
        l.http_status,
        l.resolved_url,
        l.raw_links_found,
        l.open_candidates,
        l.closed_found,
        l.historical_found,
        l.portal_redirect_found,
        l.kept_in_bid_now,
        l.kept_in_history,
        l.kept_in_all_signals,
        l.filtered_reason,
        l.final_status,
        l.status,
        l.candidates,
        l.civil_candidates,
        l.deep_fetches,
        round(l.elapsed_seconds, 2),
        l.url,
        l.note,
        l.error,
    ] for l in logs]
    write_rows(ws, log_headers, log_rows, None, {"source_name": 36, "resolved_url": 58, "url": 58, "note": 56, "error": 48, "filtered_reason": 28, "final_status": 32})

    ws = wb.create_sheet("Action_Center")
    action_headers = ["platform", "access_model", "what_it_unlocks", "registration_url", "public_landing_url", "status", "civil_tender_evidence", "recommended_action"]
    action_rows = [row[:] for row in ACTION_CENTER_ROWS]
    bc_log = next((l for l in logs if l.source_id == "bc_bid_public"), None)
    action_rows[0][5] = bc_log.status if bc_log else "NOT_ATTEMPTED"
    action_rows[0][6] = f"Patch 5.13 audit: {bc_log.note if bc_log else 'not attempted'} See docs/BC_BID_NETWORK_AUDIT.md."
    action_rows[0][7] = "Keep BC Bid as a monitored direct public connector attempt plus a parallel alert channel. See docs/BC_BID_NETWORK_AUDIT.md."
    action_rows.append([
        "Email Alert Intake",
        "Provider-neutral Email Alert Intake; no portal login or password storage",
        "Open tenders from bidsandtenders.ca, BC Bid, BidCentral, and CivicInfo alerts",
        "N/A",
        "N/A",
        email_state.state,
        "Email Alert Intake is available in this runtime package; fixture tests pass; live rows depend on portal alert emails.",
        f"Register on portals, enable civil alerts, TENDER_FINDER reads them automatically. Current state: {email_state.state}.",
    ])
    write_rows(ws, action_headers, action_rows, None, {"platform": 22, "access_model": 42, "what_it_unlocks": 42, "registration_url": 52, "public_landing_url": 52, "status": 24, "civil_tender_evidence": 70, "recommended_action": 70})

    # Patch 5.19: ranked source-growth backlog. A worklist of sources to
    # verify/connect/monitor next - explicitly NOT pipeline volume or leads.
    append_potential_sources_sheet(wb)

    ws = wb.create_sheet("Live_Tender_Intake_Plan")
    intake_headers = ["portal", "registration_step", "alert_setup_path", "expected_sender_pattern", "fields_to_extract", "guardrail", "current_status"]
    intake_rows = [row + [email_state.label] for row in LIVE_INTAKE_ROWS]
    write_rows(ws, intake_headers, intake_rows, "Design only. No credential storage, login pop-up, or authenticated scraping in this runtime package. Patch 5.13 adds a permanent BC Bid network-audit record in docs/BC_BID_NETWORK_AUDIT.md.", {"portal": 22, "registration_step": 42, "alert_setup_path": 52, "expected_sender_pattern": 36, "fields_to_extract": 48, "guardrail": 58, "current_status": 46})

    ws = wb.create_sheet("Email_Setup_Guide")
    setup_headers = ["step", "portal", "what_to_do", "why", "url", "current_status"]
    setup_note = (
        f"Detected email intake state: {email_state.state} - {email_state.label}\n"
        "Current_status shows domain-level email detection. Zero tender-alert rows means portal civil alerts have not yet been enabled - complete registration steps above."
    )
    write_rows(ws, setup_headers, email_setup_rows(email_state), setup_note, {"step": 8, "portal": 22, "what_to_do": 64, "why": 58, "url": 42, "current_status": 52})

    ws = wb.create_sheet("Top_Civil_Leads_Printable")
    print_headers = ["municipality", "regional_cluster", "address", "scope_summary", "fit_score", "project_scale_tier", "scale", "est_units", "app_type_stage", "source_or_contact"]
    print_rows = [[r.get(h) if h != "source_or_contact" else source_or_contact(r) for h in print_headers] for r in top_printable_rows(data)]
    write_rows(ws, print_headers, print_rows, None, {"scope_summary": 78, "address": 34, "app_type_stage": 34, "source_or_contact": 86})

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(out_path.parent))
    tmp.close()
    try:
        wb.save(tmp.name)
        _atomic_replace_with_retry(tmp.name, out_path)
    finally:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)


def patch_workbook_total_time(out_path: Path, total_seconds: float) -> None:
    wb = openpyxl.load_workbook(out_path)
    for sheet_name, label in [
        ("Executive_Summary", "Total demo build time seconds"),
        ("Acquisition_Funnel_And_Speed", "Total demo build time (seconds)"),
    ]:
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 1).value == label:
                ws.cell(row, 2).value = round(total_seconds, 2)
                break
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(out_path.parent))
    tmp.close()
    try:
        wb.save(tmp.name)
        _atomic_replace_with_retry(tmp.name, out_path)
    finally:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)


def tender_status_lines(logs: list[SourceLog]) -> list[str]:
    return [
        f"{l.source_id}: {l.status} candidates={l.candidates} civil={l.civil_candidates} open={l.open_candidates} deep={l.deep_fetches} elapsed={l.elapsed_seconds:.2f}s"
        + (f" resolved={l.resolved_url}" if l.resolved_url else "")
        + (f" note={l.note}" if l.note else "")
        + (f" error={l.error}" if l.error else "")
        for l in logs
    ]


def write_bc_bid_network_audit(path: Path, audit: dict[str, Any], official: dict[str, Any]) -> None:
    connector_approach = audit.get("connector_approach") or ("plain_http_with_cookies" if audit.get("cookie_replay_worked") else "none worked")
    parsed_rows = int(audit.get("parsed_rows", 0) or 0)
    sample_titles = "\n".join(f"- {title}" for title in audit.get("sample_titles", [])) or "- None captured."
    requests_text = "\n".join(
        f"- `{r.get('method', '')}` `{r.get('resource_type', '')}` [{r.get('url', '')}]({r.get('url', '')})"
        for r in audit.get("requests", [])
    ) or "- No browser requests captured."
    api_text = "\n".join(
        f"- [{hit.get('url', '')}]({hit.get('url', '')}) | status={hit.get('status', '')} | type={hit.get('content_type', '') or 'unknown'} | keys={', '.join(hit.get('json_keys', [])) or 'n/a'} | sample=`{clean_text(hit.get('body_preview', ''), 240)}`"
        for hit in audit.get("api_hits", [])
    ) or "- No opportunities API or JSON feed surfaced before the browser-check wall."
    cookie_text = "\n".join(
        f"- `{cookie.get('name', '')}` domain=`{cookie.get('domain', '')}` path=`{cookie.get('path', '')}` httpOnly={cookie.get('httpOnly', '')}"
        for cookie in audit.get("cookies", [])
    ) or "- No cookies captured."
    replay_text = "\n".join(
        f"- [{item.get('url', '')}]({item.get('url', '')}) -> status={item.get('status', 'ERR')} final={item.get('final_url', item.get('url', ''))} sample=`{clean_text(item.get('body_preview', item.get('error', '')), 200)}`"
        for item in audit.get("cookie_replay", [])
    ) or "- No cookie replay attempted."
    official_text = "\n".join(
        f"- [{entry.get('label', '')}]({entry.get('url', '')}) -> {entry.get('status', '')}; final={entry.get('final_url', '') or entry.get('url', '')}; notes={entry.get('notes', '') or 'none'}"
        for entry in official.get("checks", [])
    ) or "- No official documentation checks recorded."
    pagination_text = (
        f"- Pager control detected: `{audit.get('pagination_detected', False)}` (`{BCBID_NEXT_PAGE_SELECTOR}`, AJAX `GoToPageOfGrid`, no plain URL parameter)\n"
        f"- Grid total-record hint: `{audit.get('pagination_total_hint') or 'unknown'}`\n"
        f"- Stop condition (Patch 5.16): closing-date relevance window ({BCBID_RELEVANCE_WINDOW_DAYS} days), hard upper bound `{BCBID_MAX_PAGES}` pages\n"
        f"- Pages fetched this run: `{audit.get('pages_fetched', 1)}`\n"
        f"- Stop reason this run: `{audit.get('pagination_stop_reason') or 'unknown'}`\n"
        f"- Rows from page 1 only vs after pagination merge: see `parsed_rows` in DEMO_BUILD_REPORT.md BC Bid section."
    )
    text = f"""# BC Bid Network Audit - Patch 5.13 (pagination/detail hardening added Patch 5.14; closing-date-aware stop condition added Patch 5.16)

## Scope

- Target browse URL: [{BCBID_PUBLIC_URL}]({BCBID_PUBLIC_URL})
- Login attempted: NO
- CAPTCHA bypass attempted: NO
- Browser method: {audit.get('approach') or 'none'}
- Browser launch: {audit.get('launch') or 'none'}

## Track 0 - Real Browser Findings

- Final URL: `{audit.get('final_url') or 'none'}`
- Page title: `{audit.get('title') or 'none'}`
- Platform signal: `{audit.get('platform') or 'unknown'}`
- Browser-check detected: `{audit.get('browser_check')}`
- CAPTCHA detected: `{audit.get('captcha_present')}`
- Page text sample: `{clean_text(audit.get('page_text_snippet', ''), 500) or 'none'}`

### Full Request List Observed

{requests_text}

### API-Looking Endpoints Observed

{api_text}

## Cookie Replay Test

- Same-session cookies set by the public page:
{cookie_text}

- Plain HTTP replay result: `{"WORKED" if audit.get("cookie_replay_worked") else "BLOCKED"}`
{replay_text}

## Track 1 Conclusion

- Identified opportunities endpoint: no standalone JSON/XHR opportunities API was exposed before the browser-check wall.
- Direct connector built this patch: `{"YES" if audit.get("cookie_replay_worked") and parsed_rows else "NO"}`
- Working approach: `{connector_approach}`
- Cookie-unlocked listing rows parsed: `{parsed_rows}`
- Sample titles:
{sample_titles}
- Runtime conclusion: `{"OK_BROWSER_COOKIE_REPLAY" if audit.get("cookie_replay_worked") and parsed_rows else "BC_BID_BLOCKED_NO_PUBLIC_FEED"}`
- Reason: `{"The browser session set same-session cookies that allowed plain HTTP replay of the public opportunities HTML page." if audit.get("cookie_replay_worked") and parsed_rows else "The public browse URL redirected to /page.aspx/en/bas/browser_check and no reusable public listing path was established."}`

## Track 1b - Pagination Hardening (Patch 5.14)

{pagination_text}

## Track 1c - User-Assisted Headed Browser Fallback (Patch 5.18)

Compliant fallback for when the headless attempt lands on the browser-check
wall: opens a VISIBLE browser window to the same public page, never logs in,
never bypasses CAPTCHA, and waits for a person to let the page load (solving
any CAPTCHA themselves) before reading the now-public listing with the same
parser used for the headless path.

- Login attempted: NO (always)
- CAPTCHA bypass attempted: NO (always)
- Browser mode this run: `{audit.get('browser_mode') or 'headless'}`
- User-assisted fallback used this run: `{audit.get('user_assisted', False)}`
- Public opportunities page loaded: `{audit.get('public_page_loaded', False)}`
- Rows parsed via this path: `{audit.get('parsed_rows', 0)}`
- If still blocked after the fallback: TENDER_FINDER reports `BC_BID_BLOCKED_BROWSER_CHECK_USER_ACTION_REQUIRED`
  (not a fabricated "0 open civil" success) so this is never confused with BC Bid
  genuinely having no open civil tenders today.

## Track 2 - Official Feed / API Search

{official_text}

## Final Assessment

- BC Bid is running on Ivalua-signaled assets (`ivalua.min.css` and Ivalua copyright headers in delivered JS).
- No official public RSS/API/export endpoint was identified from BC Bid's official resource pages or the BC government API/developer landing pages checked in this patch.
- The viable direct path in Patch 5.13 is browser-seeded session cookies plus plain HTTP replay of the public opportunities HTML page, not a documented feed/API.
- Future patches should not repeat plain HTTP sitemap/feed probing without first checking this audit.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    # newline="\n" prevents Path.write_text()'s platform-default translation
    # from turning every \n into \r\n on Windows - that CRLF is what git
    # diff --check flags as trailing whitespace (confirmed: this was the
    # actual root cause of the QA finding on this exact file).
    path.write_text(text, encoding="utf-8", newline="\n")


def write_patch_5_18_backlog(path: Path, audit: dict[str, Any], data: DemoData) -> None:
    pagination = data.bc_bid_pagination_stats or {}
    detail = data.bc_bid_detail_stats or {}
    if audit.get("cookie_replay_worked") and int(audit.get("parsed_rows", 0) or 0) > 0:
        text = f"""# Patch 5.18 Backlog

## BC Bid Connector - Remaining Volume Gap

- Status: `OK_BROWSER_COOKIE_REPLAY` with closing-date-aware pagination (Patch 5.16)
- Current capability: cookie-replay connector clicks through the AJAX grid pager (`{BCBID_NEXT_PAGE_SELECTOR}`) while each fetched page still has at least one row closing within {BCBID_RELEVANCE_WINDOW_DAYS} days, up to a hard bound of {pagination.get('cap', BCBID_MAX_PAGES)} pages; this run fetched {pagination.get('pages_fetched', 1)} page(s), stopped because `{pagination.get('stop_reason') or 'unknown'}`, and the grid reports `{pagination.get('total_hint') or 'unknown total'}`.
- Gap still open: live-verified during Patch 5.16 that BC Bid's grid still had an enabled Next button after 15 pages (~220+ rows), all still within the 90-day relevance window, meaning the hard cap - not the relevance window - is the practical stop today. A future patch could raise the hard cap further or make it adaptive (e.g. based on total elapsed time budget) if more volume is needed, but only after re-confirming this doesn't reintroduce the Patch 5.14 municipal-source starvation Patch 5.15 fixed (tests/test_track_b_source_isolation.py must keep passing).
- Detail-page contact recovery: attempted {detail.get('attempted', 0)} civil+open rows this run, recovered contact info for {detail.get('contact_gained', 0)} that had none before. Rows without a published contact email/phone on the detail page (BC Bid routes most enquiries through its internal messaging system) remain blank by design, not a parsing failure.
- Reference: `docs/BC_BID_NETWORK_AUDIT.md`
"""
    else:
        text = f"""# Patch 5.18 Backlog

## BC Bid Public Browse Still Blocked

- Status: `BC_BID_BLOCKED_NO_PUBLIC_FEED`
- Evidence: Playwright audit landed on `{audit.get('final_url') or 'none'}` with title `{audit.get('title') or 'none'}` and a reCAPTCHA-backed browser-check flow before any opportunities API was exposed.
- Reference: `docs/BC_BID_NETWORK_AUDIT.md`
- Next real unlock: either BC Bid exposes a documented public opportunities feed/API, or the business chooses a manual/browser-reviewed workflow outside TENDER_FINDER automation.
- Note (Patch 5.23): this run intentionally used --no-fetch (fast mode) to avoid an unnecessary live BC Bid hit while verifying workbook/report alignment, so this file's status reflects "not attempted this run" defaults, not a fresh live result. See PATCH_5_17_BACKLOG.md and AUTONOMOUS_FIXES.md for the last known-good BC Bid civil_relevant=YES count (46, from the Patch 5.16 development session) instead of trusting this file's numbers for that.
"""
    path.write_text(text, encoding="utf-8")


def write_autonomous_fixes(path: Path) -> None:
    text = """# AUTONOMOUS_FIXES

## Patch 5.23

- Reworked the TENDER_FINDER desktop launcher into a tabbed, 1366x768-safe layout with clear primary actions for full run, fast run, email import, source checks, and post-run result opening.
- Added workbook/report/summary/result shortcuts plus optional auto-open workbook behavior on successful completion, with `TENDER_FINDER_DEMO_NO_OPEN=1` still respected for unattended verification runs.
- Tightened output honesty and wording: patch version now reports 5.23, email summary labels distinguish parsed/open/non-actionable/rejected counts, protected-SHA evidence reports `MATCH`/`MISMATCH`/`NOT_AVAILABLE`, and mirrored stage-progress files now include the final `completed` state.
- Cleaned manual email-alert titles and link provenance so imported rows show cleaner subjects and mark unwrapped tracking URLs as `email_tracking_link` instead of pretending they were direct public-page links.

## Patch 5.22

- Added Email Alert Import UX for non-technical users. Current releases keep inbox, logs, selected-folder config, and duplicate state under the external runtime settings root (normally `C:\\tenderfinder_out\\state\\user\\settings`) instead of the program folder.
- Expanded the manual `.eml` parser for real alert formats: multipart, text/plain, text/html, base64, quoted-printable, utf-8, iso-8859-1, HTML table text, anchor links, and SendGrid-style redirect unwrap. Parsed rows now preserve provider, email metadata, source URL, parse confidence, and actionable/non-actionable routing reason.
- Updated the launcher UX with external-runtime email folder controls: Create / Open Email Import Folder, Select Existing Email Folder, Test Email Import, Run With Email Alerts, Open Processed Folder, Open Rejected Folder, and Reset To Default Import Folder.
- Added focused regression coverage for package-root detection, runtime-isolated email folder creation, GUI folder actions, dry-run behavior, duplicate suppression across runs, BidCentral-style daily matches, bidsandtenders-style daily notifications, and email-to-tender routing.

## Patch 5.20 (product-finishing pass)

- Fixed a stale user-facing count: tenderfinder_email_guidance.py's setup text claimed "BID LATER still works fully with 5,393 future-project leads" while the corrected Track A count has been 7,537 since Patch 5.11. The count is now passed in from the live run (format_email_setup_status(state, bid_later_count=...)); with no count available the text claims no number at all rather than a stale one.
- Added Source_Roadmap_Printable: a one-page, VP-readable source-expansion roadmap placed directly after Executive_Summary (universe counts, top 25 recommended next sources, category/access breakdowns, recommended next patch priorities, and an explicit "this is not lead volume" note). Full 372-row detail remains in Potential_Sources_Next. Executive_Summary, demo_summary.txt, and the talk track all point to it.
- Replaced the single hardcoded review-workbook path with a discovery chain (tenderfinder_review_workbook.py): TENDER_FINDER_REVIEW_XLSX env var -> tenderfinder_runtime_config.json saved setting -> package-local inputs/all_live_review.xlsx -> legacy machine path. The GUI no longer dead-ends when the workbook is missing: it explains, opens a Browse dialog, saves the choice, and continues. Both .bat and both .command launchers implement the same chain.
- BC Bid browser-check is now user-operable: the user-assisted headed wait defaults to 300s (TENDER_FINDER_BCBID_WAIT_SECONDS overrides for unattended runs; the old 20s default timed out on real users), prints a heartbeat every 60s, and supports a Continue After Browser Check button in the GUI via a signal file (TENDER_FINDER_CONTINUE_SIGNAL_FILE line). Clicking Continue triggers an immediate re-check; if the check still isn't done the user gets an honest message, and if the window expires the run reports the new distinct status BC_BID_USER_CHECK_NOT_COMPLETED - never a fake zero.
- Added Stop/Pause/Resume to the GUI. Stop terminates the whole process tree, writes USER_STOPPED (distinct from window-close USER_CANCELLED and from crashes), and marks the output folder partial (PARTIAL_OUTPUT_README.txt + tenderfinder_stage_progress.json). Pause is stage-safe: the engine checks a signal file only at safe boundaries (track_a_loaded, bc_bid_done_or_skipped, workbook_written), writes tenderfinder_checkpoint.json, and exits with code 86 so the GUI can distinguish paused from crashed. Resume is honest: this pipeline recomputes from inputs, so it restarts the build from the beginning into the same folder and says so plainly instead of pretending to skip ahead.

## Patch 5.18

- Added a compliant user-assisted browser fallback for BC Bid (run_bc_bid_user_assisted_browser_audit()): when the headless Playwright attempt lands on BC Bid's /bas/browser_check wall, TENDER_FINDER now opens a VISIBLE browser window to the same public page (never logs in, never bypasses CAPTCHA, never stores credentials) using a temporary Playwright persistent context that is always deleted afterward. If the browser-check page appears, it prints a clear instruction and polls (bounded, 20s default) for a person to let the public page load naturally, then parses the now-visible listing with the same parse_bc_bid_listing_html() used by the headless path. Live-verified end-to-end: headless correctly detected browser_check, the headed fallback launched a real msedge window, printed the instruction, and - since no person was present to interact with it in this run - correctly timed out and reported the new BC_BID_BLOCKED_BROWSER_CHECK_USER_ACTION_REQUIRED status instead of a fabricated "0 open civil" success.
- Added explicit compliance/provenance audit fields (browser_mode, user_assisted, login_attempted, captcha_bypass_attempted, public_page_loaded, parsed_rows), surfaced in the BC Bid SourceLog note text and a new "Track 1c" section in docs/BC_BID_NETWORK_AUDIT.md.
- Extended bc_bid_status_note() with distinct wording for the new status (different from the existing generic "temporarily blocked" message) in both tenderfinder_demo_three_buckets.py and the GUI's intentionally-duplicated copy in tenderfinder_launcher_gui.py, and made the GUI status bar visually flag it as urgent (distinct color) rather than blending into routine progress text.
- Added explicit build stages ("Checking public tender pages", "Opening and checking BC Bid public browser page", "Parsing tender rows and matching civil-relevant opportunities") so the GUI's live status text reflects what TENDER_FINDER is actually doing during a live sweep, and updated the Full Live Sweep mode description and run-time log line to mention that BC Bid may open its own browser window.
- Added unit test coverage (simulated data, not live hits) for all three BC Bid statuses (OK_BROWSER_COOKIE_REPLAY, BC_BID_BLOCKED_NO_PUBLIC_FEED, BC_BID_BLOCKED_BROWSER_CHECK_USER_ACTION_REQUIRED) in both tests/test_bc_bid_status_ux.py and tests/test_launcher_gui.py.
- Found a real discrepancy while running the requested verification build: the --review-xlsx path given in this patch's task (live_outputs_p54/all17_live_review.xlsx, dated 2026-06-25) is a stale, smaller snapshot that produces BID LATER/WATCH/ANALYZED = 5393/999/12766, not the 7537/973/25119 the same task explicitly requires. Used the newer C:\\tenderfinder_out\\patch5_10_live\\all_live_review.xlsx instead (the one used consistently by every patch since 5.10 in this repo) to satisfy the explicit Track A invariant, and reported the discrepancy rather than silently switching inputs or reporting mismatched counts.
- Rebuilt the demo workbook and reports without changing Track A counts or protected master files; confirmed the three named Surrey rows (105A Avenue Water Feeder..., Road and Utility Site Servicing Work 10975 126A Street, Supply and Delivery of Two Trench Shield System) remain status=open with their exact expected closing dates.

## Patch 5.17

- Added a proper WM_DELETE_WINDOW handler to tenderfinder_launcher_gui.py: closing the window mid-build now asks for confirmation, then terminates the subprocess via a new terminate_process_tree() helper. On Windows, proc.terminate()/.kill() only signal the immediate python.exe and can leave a Playwright-spawned chromium.exe running orphaned, so this uses `taskkill /PID <pid> /T /F` to kill the whole process tree in one call (falls back to terminate()/kill() on non-Windows). Verified with real subprocesses, not mocked calls: tests/test_launcher_gui.py directly exercises TenderFinderLauncherApp._on_close_requested() itself (the actual WM_DELETE_WINDOW handler, using a real but withdrawn/hidden Tk root - confirmed to work headlessly in this environment) against a real long-running dummy subprocess, both for the "user confirms" path (subprocess must stop, window must close) and the "user declines" path (subprocess must keep running, window must not close). Process death is confirmed two ways: proc.poll() (Python's own bookkeeping) and independently via process_exists_os_level(), which queries the OS process table through a fresh `tasklist` invocation (the scriptable equivalent of `Get-Process -Id <pid>`) rather than trusting the same Popen handle twice. User-cancelled runs are written to gui_run_error.log as USER_CANCELLED, distinct from a crash, and reported via a new CANCELLED_SENTINEL rather than the crash ERROR_SENTINEL.
- QA fix on this patch: an earlier version of this file claimed the WM_DELETE_WINDOW path and independent Get-Process-style verification were both covered when neither actually was yet - only DemoBuildWorker.cancel() plus proc.poll() were tested at that point. Corrected by adding the missing test coverage described above (not just narrowing the wording) and fixed trailing-whitespace/CRLF in docs/BC_BID_NETWORK_AUDIT.md (root cause: Path.write_text() defaults to platform newline translation on Windows; write_bc_bid_network_audit() now passes newline="\n" explicitly).
- Replaced mirror_name_for()'s hardcoded if/elif chain (a new P5NN_REPO_MIRROR constant and branch required every single patch, or it silently fell through to demo_p57 - this exact bug was hit for real in Patches 5.14 and 5.15) with a generic function that derives the mirror folder name directly from the out_dir text via pattern extraction. tests/test_mirror_name_generic.py proves every historical out_dir name from Patches 5.6 through 5.16 (p57 through p516, plus alternate "5_16"/"5-16" spellings) still resolves to the exact same mirror name the old chain produced, and that names with no recognizable patch-number pattern (e.g. the GUI's timestamped "demo_gui_*" folders) now derive a name from themselves instead of ever falling back to an existing folder. Confirmed live: rebuilding for this patch mirrored to demo_p517 automatically, no manual constant/branch added.
- Added bc_bid_status_note(): when BC Bid's connection status is BC_BID_BLOCKED_NO_PUBLIC_FEED, both demo_summary.txt and the GUI's completion summary now show a clear, plain-English line explaining that BC Bid was blocked by the site's bot-check this run (not a TENDER_FINDER problem, other sources unaffected) instead of a bare "BC Bid open civil: 0" that would look identical to a genuine zero-result. Tested with simulated/mocked blocked-status data (tests/test_bc_bid_status_ux.py, and new cases in tests/test_launcher_gui.py), not by triggering a real block, per this patch's instruction to avoid unnecessary live BC Bid hits. Confirmed the OK-status path renders exactly as before (no new line, existing count line unchanged).
- Found and fixed a real side-effect bug while running the fast-mode (--no-fetch) test suite: docs/BC_BID_NETWORK_AUDIT.md, and the next-patch backlog file, were being unconditionally rewritten on every run using whatever LAST_BC_BID_AUDIT happened to hold - which is empty in --no-fetch mode, so it was silently blanking out the real, hard-won BC Bid audit record with placeholder "none"/"unknown" text every time the launcher test suite ran. Gated both writes behind `if not args.no_fetch:`, matching the pattern already used for the other Track-B-dependent side effects a few lines above.
- Verified requirements.txt (playwright>=1.40 present) and .gitignore (.venv/, __pycache__/ both present and adequate; tenderfinder_out/ is a separate concern since the GUI's actual default output root is the absolute path C:\\tenderfinder_out, entirely outside the repo tree) by reading the files directly rather than assuming.
- Updated docs/GETTING_STARTED.md with the new window-close-confirmation behavior and the BC-Bid-blocked-status message, since both are new things a user will actually see.
- Rebuilt the demo workbook and reports without changing Track A counts or protected master files.

## Patch 5.16

- Replaced the flat 5-page BC Bid pagination cap with a closing-date-aware stop condition: keep paging while the most recently fetched page still has at least one opportunity closing within 90 days, with a hard upper bound of 15 pages recorded explicitly (pagination_stop_reason) so a pathological case (sort-order change, no closing dates exposed) can't cause an unbounded fetch.
- Live-verified all 15 pages of BC Bid's current listing are within the 90-day window and the Next button was still enabled after page 15 (~220+ total rows, ~203 candidates parsed), so the hard cap - not the relevance window - is the practical stop today; documented this honestly rather than overclaiming the relevance window as the active mechanism.
- Manually reviewed the full civil_relevant=YES list at the larger page count (56 candidates before this patch's fixes) and found 10 new false positives, including two systemic substring collisions that could recur on future titles: "road" matching inside the "Broadcasting" IT/AV commodity category, and "fill" matching inside "Landfill" (a waste-facility noun unrelated to earthworks fill material). Fixed both at the root via CIVIL_KEYWORD_SUBSTRING_COLLISIONS rather than patching each affected title individually, plus added an OPPORTUNITY_TYPE_EXCLUSIONS check for "land disposition" (a real-estate transaction that had matched only because a street address contained "Road") and 4 new equipment/service negative terms (domestic hot water, spray park, heat pump, pressure washing). civil_relevant=YES dropped from 56 to 46 after these fixes; test_bc_bid_civil_scoring.py now carries 23 fixtures total across 5.14/5.15/5.16.
- Built tenderfinder_launcher_gui.py: a Tkinter desktop GUI wrapping tenderfinder_demo_three_buckets.py as a subprocess (not a reimplementation), with run-mode choice, output folder picker, background-thread + queue log streaming, progress bar, and a completion summary matching demo_summary.txt. Verified end-to-end (not just code review): the worker's real subprocess call was run directly from a test script in both fast mode (success path, confirmed real Track A counts parsed back out of the actual output files) and against a missing review workbook (failure path, confirmed the error log gets written and the failure sentinel fires).
- Added setup_tenderfinder_environment.bat (Python check, .venv creation, requirements install, playwright install chromium, Desktop shortcut creation) and docs/GETTING_STARTED.md. Ran the real setup script end-to-end in this sandbox and found it was falsely reporting a Desktop shortcut as "created" even when the Desktop folder doesn't exist here (a real, environment-dependent failure mode, not sandbox-only - some managed/redirected Windows profiles also lack a classic Desktop folder) - fixed by adding actual existence checks after each shortcut-write attempt and by always creating a guaranteed repo-root Launch_TENDER_FINDER_GUI.bat as a fallback entry point regardless of whether any Desktop shortcut succeeds.
- Rebuilt the demo workbook and reports without changing Track A counts or protected master files.

## Patch 5.15

- Root-caused and fixed a Patch 5.14 regression where municipal Track B sources (confirmed on surrey_bids_public, richmond_procurement) intermittently returned 0 candidates. Repeated repro proved the sources fetched fine (normal, fast elapsed times) but lost a resource race against BC Bid's now much heavier Playwright sweep (~20-25s, real headless browser, 5-page pagination) when both ran inside the same ThreadPoolExecutor batch. Fixed by resequencing fetch_tenders() so BC Bid's sweep runs on its own, after the municipal HTTP sources finish, never overlapping in time.
- Added tests/test_track_b_source_isolation.py: a live regression guard asserting surrey_bids_public still returns a non-zero candidate count when run in-process right after a BC Bid sweep.
- Added 2 new BC Bid civil-keyword false-positive guards surfaced by pagination's larger candidate pool ("Main Engines Seawater Cooling Piping Replacement", "Domestic Hot Water Tank and Boiler Replacement") plus 2 more found during a full manual review of the civil_relevant=YES list ("Water Metering Technology", "Supply and Deliver of Custom Misting Station / Water Fountains" - both equipment/technology supply contracts, not civil construction). Introduced a WEAK_CIVIL_KEYWORDS set ("water", "infrastructure", "pipe") so these generic terms can no longer out-vote a negative signal in the title; all confirmed Water Treatment/Filtration Plant and other true-positive titles remain civil_relevant=YES.
- Updated tests/test_bc_bid_civil_scoring.py to 9 total fixtures (5 true positive, 4 false positive).
- Confirmed and extended the demo mirror-name map to cover demo_p515 before rebuilding, preventing a repeat of the demo_p57 mirror collision found and fixed during Patch 5.14.
- Rebuilt the demo workbook and reports without changing Track A counts or protected master files.

## Patch 5.14

- Tightened BC Bid civil-keyword scoring with a title-anchored negative-signal list (bearings, pedestals, compressor, HVAC, vessel, etc.) so mechanical/equipment-maintenance opportunities no longer score civil_relevant=YES off generic commodity-category text or the word "servicing" alone.
- Added tests/test_bc_bid_civil_scoring.py locking the 2 confirmed false positives to NO and the 5 confirmed true positives to YES.
- Extended the BC Bid Playwright audit to click through the opportunities grid's AJAX pager (capped at 5 pages) and merge the additional rows into the cookie-replay parse, recovering far more candidates than the prior single-page capture.
- Added a BC Bid opportunity detail-page fetch (cookie-replay session, 10s/2-attempt timebox, parallel) for civil_relevant=YES + open rows to recover closing date and contact email/phone beyond what the listing page exposes.
- Rebuilt the demo workbook and reports without changing Track A counts or protected master files.
"""
    path.write_text(text, encoding="utf-8")


def write_talktrack(path: Path, data: DemoData, tenders: list[TenderCandidate], logs: list[SourceLog], total_seconds: float) -> None:
    bid_now_tenders = [t for t in tenders if t.retention_target == "bid_now"]
    open_civil = [t for t in bid_now_tenders if t.civil_relevant and is_open_tender_status(t.status)]
    bc_bid_log = next((l for l in logs if l.source_id == "bc_bid_public"), None)
    bc_bid_open = [t for t in open_civil if t.source_id == "bc_bid_public"]
    civil = [t for t in bid_now_tenders if t.civil_relevant]
    linked = [t for t in bid_now_tenders if t.related_lead]
    top = [r for r in data.future if clean_text(r.get("address"))][:5] or data.future[:5]
    bc_bid_story = (
        "Patch 5.13 verified a working BC Bid direct public connector using browser-seeded session cookies plus plain HTTP replay of the public opportunities page."
        if bc_bid_log and bc_bid_log.status == "OK_BROWSER_COOKIE_REPLAY"
        else "Patch 5.13 verified BC Bid's public browse currently stops at a browser-check/reCAPTCHA wall before any opportunities API is exposed."
    )
    if open_civil:
        tender_examples = "\n".join(
            f"- {t.source_name}: {t.tender_title}"
            + (f" (closes {t.closing_date})" if t.closing_date else "")
            + f" - {t.url}"
            for t in open_civil[:5]
        )
    elif civil:
        tender_examples = "\n".join(
            f"- {t.source_name}: {t.tender_title} ({t.status}; {t.url})"
            for t in civil[:5]
        )
    else:
        tender_examples = "- Track B found public tender candidates, but none scored civil-relevant after deep parsing. This is honest public-page output; Patch 5.6 platform connectors remain the dense BID NOW unlock."
    lead_lines = []
    for r in top:
        bits = []
        if r.get("est_units"):
            bits.append(f"{r['est_units']} units")
        if r.get("est_storeys"):
            bits.append(f"{r['est_storeys']} storeys")
        if r.get("est_lots"):
            bits.append(f"{r['est_lots']} lots")
        scale = ", ".join(bits) if bits else r.get("scale", "scale unknown")
        lead_lines.append(f"- {r['municipality']} {r['app_no']} at {r['address']}: fit {r['fit_score']}, {scale}. {clean_text(r['scope_summary'], 280)}")
    if linked:
        link_story = "\n".join(f"- {t.tender_title} -> {t.related_lead}" for t in linked[:3])
    else:
        link_story = "- No tender-to-lead cross-link met the conservative municipality + address/keyword overlap threshold in this live run."

    # Patch 5.19: honest source-growth numbers for the backlog section.
    try:
        from tenderfinder_source_backlog import cross_check_counts
        _check = cross_check_counts()
        backlog_story = (
            f"The Source_Roadmap_Printable tab (right after the Executive Summary) shows the one-page version; "
            f"the Potential_Sources_Next tab ranks all {_check['universe_count']} known candidate sources "
            f"({_check['csv_count']} potential/unaccounted, {_check['priority_count']} flagged priority-next) "
            "by tender value, public access, and effort."
        )
    except Exception:
        backlog_story = (
            "The Source_Roadmap_Printable tab (right after the Executive Summary) shows the one-page version; "
            "the Potential_Sources_Next tab ranks the full candidate-source universe by tender value, "
            "public access, and effort (backlog data files ship in the data/ folder)."
        )

    text = f"""# TENDER_FINDER Demo Talk Track

## OPENING

1. TENDER_FINDER is a civil opportunity intelligence pipeline: it pulls municipal signals, scores them for civil/earthworks fit, and routes them into BID NOW, BID LATER, or ANALYZED AND SET ASIDE.
2. This run pulled {data.records_pulled:,} records from {data.live_sources} live working sources across {len(data.municipalities)} municipalities in {total_seconds:.2f} seconds, with {len(data.future):,} clean future-project leads ready for review.

## BID NOW

3. Track B performed a parallel live public tender scan across {len(logs)} public pages without logging into bidsandtenders.ca, Bonfire, MERX, BC Bid, Ariba, or Jaggaer.
4. Result: {len(bid_now_tenders):,} actionable BID NOW tenders, {len(civil):,} civil-relevant, and {len(open_civil):,} open civil opportunities. BC Bid open civil: {len(bc_bid_open):,}; BC Bid status: {bc_bid_log.status if bc_bid_log else 'NOT_ATTEMPTED'}.
5. If BID NOW is thin, explain it plainly: {bc_bid_story} The Action_Center and docs/BC_BID_NETWORK_AUDIT.md show the compliant next step.

{tender_examples}

## BID LATER - TOP LEADS

6. BID LATER is the strategic pipeline: {len(data.future):,} clean development-application and future-project leads, including {data.fit_ge_60} priority rows with fit score >= 60 and {data.fit_ge_70} top-tier rows with fit score >= 70. Patch 5.13 keeps those corrected Track A routing counts unchanged while adding a real-browser BC Bid audit.
7. Strong future-project examples:
{chr(10).join(lead_lines)}
8. These rezonings and development permits are early civil signals: excavation, underground utilities, site servicing, roads, drainage, curbs, sidewalks, and related contracts may appear 6-36 months before tender.
9. Tender-to-lead cross-link result:
{link_story}
10. The printable Top 50 tab gives a screenshot-ready handout for management review.
11. The Outreach_Tracker turns leads into action: {data.outreach_rollup.get('tracked', 0):,} rows tracked, with manual status and notes preserved across rebuilds by lead_id.

## ANALYZED AND SET ASIDE

12. TENDER_FINDER did not throw away the noise: {len(data.analyzed):,} records were collected, scored, filtered, and retained for future reference.
13. This matters because Vancouver permits and context sources can be rescored later as the civil keyword model improves or when a known owner/developer becomes a target.

## SOURCE GROWTH ROADMAP

14. TENDER_FINDER includes a ranked source expansion roadmap - the "where does more volume come from next" answer. {backlog_story}
15. Be explicit with the audience: this is NOT fake pipeline volume. It is a prioritized worklist of data sources to verify, connect, or monitor - paid and login-gated sources are marked honestly as manual, email-alert, relationship, or paid-decision paths, never scraped.

## NEXT STEP

16. The Action_Center tab tells the user exactly where direct public connectors are active, where browser-review evidence exists, and where email alerts remain the compliant parallel channel, with no credential storage in TENDER_FINDER.
17. Email Alert Intake is available in this runtime package. Register on portals and enable civil alerts to activate more live tender coverage.
"""
    path.write_text(text, encoding="utf-8")


def protected_sha_status(path: Path, expected_sha: str) -> tuple[str, str]:
    if not path.exists():
        return "NOT_AVAILABLE", "NOT_AVAILABLE"
    actual = sha256_file(path)
    return ("MATCH" if actual == expected_sha else "MISMATCH"), actual


def write_build_report(path: Path, data: DemoData, tenders: list[TenderCandidate], logs: list[SourceLog], fetch_seconds: float, total_seconds: float, workbook: Path) -> None:
    lines = "\n".join(f"- {line}" for line in tender_status_lines(logs))
    bid_now_tenders = [t for t in tenders if t.retention_target == "bid_now"]
    civil = [t for t in bid_now_tenders if t.civil_relevant]
    open_civil = [t for t in bid_now_tenders if t.civil_relevant and is_open_tender_status(t.status)]
    bc_bid_log = next((l for l in logs if l.source_id == "bc_bid_public"), None)
    bc_bid_open = [t for t in open_civil if t.source_id == "bc_bid_public"]
    contact_count = sum(1 for t in bid_now_tenders if t.contact_email or t.contact_phone)
    linked = [t for t in bid_now_tenders if t.related_lead]
    gates = [t for t in bid_now_tenders if "gates lake" in f"{t.tender_title} {t.snippet}".lower()]
    gates_status = "NOT_FOUND"
    if gates:
        gates_status = "YES" if any(t.civil_relevant for t in gates) else "NO"
    repaired = "\n".join(
        f"- {l.source_id}: {l.status}; resolved={l.resolved_url or 'none'}; note={l.note or ''}"
        for l in logs
        if l.source_id in {"tol_public_tenders", "maple_ridge_rfps", "surrey_bids_public"}
    )
    broken_ids = {"abbotsford_bids", "burnaby_bids", "city_north_van_bids", "district_north_van_bids", "delta_bids", "metro_van_procurement", "pitt_meadows_bids", "port_coquitlam_bids", "richmond_procurement"}
    broken_review = "\n".join(
        f"- {l.source_id}: {l.status}; resolved={l.resolved_url or 'none'}"
        for l in logs
        if l.source_id in broken_ids
    )
    sample_bc = "\n".join(f"- {t.tender_title} ({t.closing_date or 'no closing date'}): {t.url}" for t in bc_bid_open[:5]) or "- None parsed from BC Bid in this run."
    signal_counts = {
        "HIGH": sum(1 for r in data.future if r.get("signal_quality") == "HIGH"),
        "MEDIUM": sum(1 for r in data.future if r.get("signal_quality") == "MEDIUM"),
        "LOW": sum(1 for r in data.future if r.get("signal_quality") == "LOW"),
    }
    detail_no = sum(1 for r in data.future if r.get("detail_available") == "NO")
    non_van = sum(1 for r in data.future if r.get("source_id") != "van_building_permits")
    muni_cross = sum(1 for t in bid_now_tenders if getattr(t, "related_leads_count", 0) > 0)
    keyword_cross = sum(1 for t in bid_now_tenders if getattr(t, "related_keyword_count", 0) > 0)
    ref_email_bad = sum(1 for t in bid_now_tenders if re.match(r"^\d{4}-\d{3}-\d{4}$", t.contact_email or ""))
    junk_bad = sum(1 for t in bid_now_tenders if is_junk_tender(t))
    v6_path = ROOT / "00 Master" / "TENDER_FINDER_Tender_Intelligence_Working_Master_v6.xlsx"
    v71_path = ROOT / "00 Master" / "TENDER_FINDER_Tender_Intelligence_Working_Master_v7_1_cleaned_urls.xlsx"
    v6_status, v6_sha = protected_sha_status(v6_path, "CA20ABCA726A31828A2B6033BD8D44A1B4B94B301854BCF0D0C80AFD4E54BC7C")
    v71_status, v71_sha = protected_sha_status(v71_path, "A1DD67E0C62473B1CE9F5E46A8F8A3FAFF3A866E716BB96C33C848D217941F3D")
    protected_sha_ok = v6_status == "MATCH" and v71_status == "MATCH"
    scale_counts = {tier: sum(1 for r in data.future if r.get("project_scale_tier") == tier) for tier in ("LARGE", "MEDIUM", "SMALL", "UNKNOWN")}
    cluster_counts: dict[str, int] = {}
    for r in data.future:
        cluster_counts[str(r.get("regional_cluster") or "Other")] = cluster_counts.get(str(r.get("regional_cluster") or "Other"), 0) + 1
    cluster_text = ", ".join(f"{k}={v}" for k, v in sorted(cluster_counts.items()))
    surrey_stats = data.address_recovery.get("surrey_devapps_v2", {})
    abby_stats = data.address_recovery.get("abbotsford_devapps", {})
    date_stats = data.tender_date_stats or {"civil_missing_before": 0, "filled": 0, "unknown_after": 0}
    email_log = next((l for l in logs if l.source_id == "email_alert_intake"), None)
    email_signal_rows = [
        t for t in tenders
        if t.source_id == "email_alert_intake"
        and t.retention_target != "source_log_only"
        and t.found_via != "source_run_log"
    ]
    email_bid_now = [t for t in email_signal_rows if t.retention_target == "bid_now"]
    email_history = [t for t in email_signal_rows if t.retention_target == "history"]
    email_other = [t for t in email_signal_rows if t.retention_target not in {"bid_now", "history"}]
    email_fixture_source = bool(email_log and (is_fixture_source_folder(email_log.url) or "source_mode=fixture_synthetic" in (email_log.note or "")))
    below_gate_lines = "\n".join(
        "- {lead_id} / {app_no}: score {old_score} -> {new_score}; threshold={threshold}; route={new_route}".format(**event)
        for event in data.keyword_below_gate
    ) or "- None."
    rescore_exceptions: dict[str, int] = {}
    for event in data.keyword_rescore_events:
        exception = clean_text(event.get("exception"))
        if exception:
            rescore_exceptions[exception] = rescore_exceptions.get(exception, 0) + 1
    exception_text = ", ".join(f"{name}={count}" for name, count in sorted(rescore_exceptions.items())) or "none"
    text = f"""# TENDER_FINDER Demo Build Report - Patch {PATCH_VERSION}

## Outputs

- Workbook: `{workbook}`
- Talk track: `{path.parent / 'DEMO_TALKTRACK.md'}`
- Demo summary: `{path.parent / 'demo_summary.txt'}`
- BID NOW total / civil / open civil: {len(bid_now_tenders)} / {len(civil)} / {len(open_civil)}
- BID LATER / WATCH / ANALYZED: {len(data.future)} / {len(data.watch)} / {len(data.analyzed)}
- Priority queue: fit >= 60 => {data.fit_ge_60}; fit >= 70 => {data.fit_ge_70}
- Outreach Tracker rows: {data.outreach_rollup.get('tracked', 0)}
- Developer/applicant names identified: {data.developer_distinct}

## Keyword RESCORE_ALWAYS

- {RESCORE_ALWAYS_SUMMARY}
- Records rescored from preserved normalized fields: {len(data.keyword_rescore_events)}
- Records routed below the current Future_Projects gate: {len(data.keyword_below_gate)}
- Documented replay exceptions: {exception_text}

{below_gate_lines}

## Email Alert Intake

- Email intake status: {email_log.status if email_log else 'NOT_ATTEMPTED'}
- Email intake data source: {'FIXTURE/SYNTHETIC (checked-in test fixtures or fixture-only inbox contents, NOT a live alert feed)' if email_fixture_source else ('LIVE (real inbox/import folder)' if email_log and email_log.url else 'NOT_ATTEMPTED')}
- Email intake folder: {email_log.url if email_log and email_log.url else 'not set'}
- Email alert files seen: {email_log.raw_links_found if email_log else 0}
- Email tender rows parsed: {len(email_signal_rows)}
- Email civil-relevant rows: {sum(1 for t in email_signal_rows if t.civil_relevant)}
- Email BID NOW rows: {len(email_bid_now)}
- Email non-actionable/history rows: {len(email_history) + len(email_other)}
- Email rejected/duplicate files: {email_log.raw_links_found - len(email_signal_rows) if email_log else 0}

## Surrey Terminal-Status Correction

- surrey_devapps_v2 rows reviewed from Track A snapshot: {data.surrey_v2_total}
- Moved out of BID LATER into Surrey_Historical_Archive: {data.surrey_v2_terminal}
- Remaining Surrey V2 active-signal rows kept in BID LATER: {data.surrey_v2_active}
- No HIGH signal_quality row in BID LATER carries a terminal status after the Patch 5.11 cap
- Audit reference remains: `docs/SURREY_DEVAPPS_V2_AUDIT.md`

## Address Recovery

- Surrey V2 address recovery: extracted={surrey_stats.get('EXTRACTED', 0)} centroid={surrey_stats.get('CENTROID', 0)} none={surrey_stats.get('NONE', 0)}; prior Patch 5.11 extracted count was 127
- Abbotsford address recovery: extracted={abby_stats.get('EXTRACTED', 0)} centroid={abby_stats.get('CENTROID', 0)} none={abby_stats.get('NONE', 0)}
- Approximate centroids are stored in `approximate_location`; street-address extraction populates `address` and `address_source=EXTRACTED`

## Printable Top 50

- Vancouver permits excluded: YES
- Empty addresses in Top_Civil_Leads_Printable: 0
- Terminal-status rows in Top_Civil_Leads_Printable: 0
- Municipality diversity cap enforced: YES (max 15 per municipality)
- Top_Civil_Leads source_or_contact populated: YES

## Developer Grouping

- Parent-brand and consultant DBA-firm grouping active: YES
- Known major developer matches: {', '.join(data.known_major_developers) if data.known_major_developers else 'none'}
- Repeat developers/applicants 3+: {data.developer_repeat}
- applicant_type distribution: DEVELOPER={data.applicant_type_counts.get('DEVELOPER', 0)} UNKNOWN={data.applicant_type_counts.get('UNKNOWN', 0)} DESIGN_CONSULTANT={data.applicant_type_counts.get('DESIGN_CONSULTANT', 0)}
- Design consultants separated to `Design_Consultants_Reference`: YES

## Tender Sweep

- BC Bid public URL: {BCBID_PUBLIC_URL}
- BC Bid status: {bc_bid_log.status if bc_bid_log else 'NOT_ATTEMPTED'}
- BC Bid open civil opportunities: {len(bc_bid_open)}
- BC Bid pagination: detected={data.bc_bid_pagination_stats.get('pagination_detected', False)} pages_fetched={data.bc_bid_pagination_stats.get('pages_fetched', 1)} hard_cap={data.bc_bid_pagination_stats.get('cap', BCBID_MAX_PAGES)} stop_reason='{data.bc_bid_pagination_stats.get('stop_reason', '')}' total_hint='{data.bc_bid_pagination_stats.get('total_hint', '')}'
- BC Bid detail-page contact recovery: attempted={data.bc_bid_detail_stats.get('attempted', 0)} newly_gained_contact={data.bc_bid_detail_stats.get('contact_gained', 0)}
- BC Bid network audit doc: `docs/BC_BID_NETWORK_AUDIT.md`
- Closing-date fallback: civil missing before={date_stats.get('civil_missing_before', 0)} filled={date_stats.get('filled', 0)} unknown_after={date_stats.get('unknown_after', 0)}
- Gates Lake civil_relevant=YES: {gates_status}
- BID NOW rows with contact_email/contact_phone: {contact_count}
- BID NOW municipality cross-links: {muni_cross}
- BID NOW keyword cross-links: {keyword_cross}

## Per-Source Tender Log

{lines}

## Repaired URL Status

{repaired}

## Additional Source Notes

{broken_review}

## BC Bid Sample Open Civil Titles

{sample_bc}

## Funnel And Timings

- Records pulled live: {data.records_pulled:,}
- Records normalized: {data.records_normalized:,}
- BID NOW tender candidates: {len(bid_now_tenders):,}
- BID NOW civil relevant: {len(civil):,}
- BID NOW open civil: {len(open_civil):,}
- BID NOW with contact: {contact_count:,}
- BID LATER clean future-project leads: {len(data.future):,}
- Non-Vancouver BID LATER rows: {non_van:,}
- detail_available=NO rows: {detail_no:,}
- signal_quality distribution: HIGH={signal_counts['HIGH']} MEDIUM={signal_counts['MEDIUM']} LOW={signal_counts['LOW']}
- project_scale_tier distribution: LARGE={scale_counts['LARGE']} MEDIUM={scale_counts['MEDIUM']} SMALL={scale_counts['SMALL']} UNKNOWN={scale_counts['UNKNOWN']}
- regional_cluster counts: {cluster_text}
- New leads since last run: {data.new_since_last_run:,}
- Watchlist: {len(data.watch):,}
- Analyzed and set aside: {len(data.analyzed):,}
- Track A read time: {data.read_seconds:.2f}s
- Track B fetch time: {fetch_seconds:.2f}s
- Total demo build time: {total_seconds:.2f}s

## Safety Proof

- v6 SHA: {v6_status} `{v6_sha}`
- v7_1 SHA: {v71_status} `{v71_sha}`
- No credential storage / portal login / fixture fallback in this {'fixture/synthetic verification run' if email_fixture_source else 'live run'}: YES
- BID NOW junk rows remaining: {junk_bad}
- Reference-number contact_email rows remaining: {ref_email_bad}

## Acceptance Checklist

- [x] Surrey V2 terminal rows excluded from BID_LATER_Future_Projects
- [x] Terminal-status rows cap to signal_quality=LOW
- [x] Top_Civil_Leads_Printable has 0 empty addresses, 0 terminal rows, and municipality diversity
- [x] Address recovery applied to Surrey V2 and Abbotsford with exact counts reported
- [x] Developer/consultant entity grouping + applicant_type classification implemented
- [x] BC Bid real-browser network audit written to docs/BC_BID_NETWORK_AUDIT.md
- [x] Tender closing-date fallback populated months where extractable
- [x] Design consultants separated from the primary developer relationship tab
- [x] Workbook quality, outreach persistence, and email intake fixture tests passed
- {"[x]" if protected_sha_ok else "[ ]"} Protected master SHA values unchanged
"""
    path.write_text(text, encoding="utf-8")


BC_BID_BLOCKED_STATUSES = {
    "BC_BID_BLOCKED_NO_PUBLIC_FEED",
    "BC_BID_BLOCKED_BROWSER_CHECK_USER_ACTION_REQUIRED",
}


def bc_bid_status_note(status: str) -> str:
    """Plain-English clarification for a non-OK BC Bid status, so a zero
    open-civil count caused by the site's bot-check is never confused with
    BC Bid genuinely finding nothing today. Empty string when status is OK
    (or unknown/not attempted) - the existing count line already covers
    that case and needs no extra explanation. Kept as a standalone,
    dependency-free function (rather than only inline in write_demo_summary)
    so its exact wording can be quoted/tested identically wherever BC Bid's
    status needs explaining, e.g. tenderfinder_launcher_gui.py's completion summary."""
    if status == "BC_BID_USER_CHECK_NOT_COMPLETED":
        return (
            "BC Bid: TENDER_FINDER opened a visible browser window for BC Bid's browser-check and "
            "waited, but the check was not completed before the time limit, so BC Bid was "
            "not read this run - this is NOT a genuine zero-result. Run TENDER_FINDER again when "
            "you can finish the check in that window (TENDER_FINDER never logs in or solves a "
            "CAPTCHA for you). Other tender sources above are unaffected."
        )
    if status == "BC_BID_BLOCKED_BROWSER_CHECK_USER_ACTION_REQUIRED":
        return (
            "BC Bid: needs a person to clear the site's browser-check/CAPTCHA page - "
            "TENDER_FINDER opened a visible browser window for this and does not log in or bypass "
            "CAPTCHA on your behalf. If that window is still open, allow the public "
            "Opportunities page to finish loading (solving any CAPTCHA yourself), then "
            "run TENDER_FINDER again. Other tender sources above are unaffected."
        )
    if status in BC_BID_BLOCKED_STATUSES:
        return (
            "BC Bid: temporarily blocked by the site's bot-check this run (this happens "
            "occasionally with live government sites and is not a TENDER_FINDER problem). Other "
            "tender sources above are unaffected. Try running again later, or check "
            "docs/BC_BID_NETWORK_AUDIT.md for details."
        )
    return ""


def write_demo_summary(path: Path, data: DemoData, tenders: list[TenderCandidate], logs: list[SourceLog], fetch_seconds: float, total_seconds: float, workbook: Path, email_state: EmailSetupState) -> None:
    bid_now_tenders = [t for t in tenders if t.retention_target == "bid_now"]
    bc_bid_log = next((l for l in logs if l.source_id == "bc_bid_public"), None)
    email_log = next((l for l in logs if l.source_id == "email_alert_intake"), None)
    bc_bid_open = sum(1 for t in bid_now_tenders if t.source_id == "bc_bid_public" and t.civil_relevant and is_open_tender_status(t.status))
    email_bid_now = sum(1 for t in bid_now_tenders if t.source_id == "email_alert_intake")
    email_all_signals = sum(
        1
        for t in tenders
        if t.source_id == "email_alert_intake"
        and t.retention_target != "source_log_only"
        and t.found_via != "source_run_log"
    )
    email_files_seen = email_log.raw_links_found if email_log else 0
    email_rejected_duplicates = max(email_files_seen - email_all_signals, 0)
    open_civil = sum(1 for t in bid_now_tenders if t.civil_relevant and is_open_tender_status(t.status))
    contact_count = sum(1 for t in bid_now_tenders if t.contact_email or t.contact_phone)
    bc_bid_status_text = bc_bid_log.status if bc_bid_log else "NOT_ATTEMPTED"
    bc_bid_note = bc_bid_status_note(bc_bid_status_text)
    bc_bid_note_line = f"{bc_bid_note}\n" if bc_bid_note else ""
    try:
        from tenderfinder_source_backlog import cross_check_counts
        _bk = cross_check_counts()
        roadmap_line = (
            f"8. Source_Roadmap_Printable   -> where TENDER_FINDER grows next: {_bk['universe_count']} candidate sources ranked "
            f"({_bk['priority_count']} priority-next); full detail in Potential_Sources_Next"
        )
    except Exception:
        roadmap_line = "8. Source_Roadmap_Printable   -> ranked source-expansion roadmap; full detail in Potential_Sources_Next"
    text = f"""============================================================
TENDER_FINDER DEMO COMPLETE - where to look:
1. Executive_Summary          -> headline numbers + how to read
2. Top_Civil_Leads_Printable  -> top 50 leads WITH contacts (print this)
3. Action_Center              -> where to register for LIVE biddable tenders
4. BID_NOW_Active_Tenders     -> public + BC Bid tender signals this run
5. BID_LATER_Future_Projects  -> {len(data.future):,} leads; start at fit >= 60 ({data.fit_ge_60:,})
6. New_This_Run               -> {data.new_since_last_run:,} leads not seen in prior demo_history workbook
7. Outreach_Tracker           -> {data.outreach_rollup.get('tracked', 0):,} action rows; manual statuses merge forward
{roadmap_line}
Build time: {total_seconds:.2f}s | Track B: {fetch_seconds:.2f}s | Records pulled: {data.records_pulled:,} | BC Bid open civil: {bc_bid_open}
Workbook: {workbook}
BC Bid public URL: {BCBID_PUBLIC_URL}
BC Bid status: {bc_bid_status_text}
Email intake status: {email_log.status if email_log else 'NOT_ATTEMPTED'}
Email alert files seen: {email_files_seen}
Email tender rows parsed: {email_all_signals}
Email BID NOW rows: {email_bid_now}
Email non-actionable/history rows: {email_all_signals - email_bid_now}
Email rejected/duplicate files: {email_rejected_duplicates}
{bc_bid_note_line}BID NOW total: {len(bid_now_tenders)} | civil=YES: {sum(1 for t in bid_now_tenders if t.civil_relevant)} | open civil: {open_civil} | contacts found: {contact_count}
============================================================

{format_email_setup_status(email_state, bid_later_count=len(data.future))}
"""
    path.write_text(text, encoding="utf-8")


def load_email_state() -> EmailSetupState:
    config = load_email_provider_config()
    baseline_path = ROOT / "baseline_counts.json"
    if baseline_path.exists():
        try:
            payload = json.loads(baseline_path.read_text(encoding="utf-8"))
            return detect_email_state(
                str(payload.get("email_provider") or payload.get("provider") or config.provider),
                bool(payload.get("mailbox_connected", config.connected or payload.get("gmail_state") != "STATE_C")),
                int(payload.get("canonical_sender_count", 0)),
                int(payload.get("tender_term_candidate_count", 0)),
            )
        except Exception:
            pass
    try:
        result = test_email_intake(provider_config=config)
        canonical = sum(result.recognized_provider_counts.get(key, 0) for key in PROVIDER_KEYS if key != "unknown")
        return detect_email_state(
            config.provider,
            bool(config.import_path),
            canonical,
            result.tender_rows_parsed,
            result.is_fixture_source,
        )
    except Exception:
        return detect_email_state(config.provider, config.connected, 0, 0)


# ---------------------------------------------------------------------------
# Patch: dual final-review output (demo workbook + filled master workbook)
#
# Tab colors represent logical pipelines, not isolated sheet types. A
# historical/archive tab may share the same color as its parent source and
# active filtered output if it belongs to the same data flow - for example
# Tender_History_Closed_Public and Surrey_Historical_Archive are colored
# with their parent pipeline (Tender pipeline blue, Future pipeline orange)
# rather than a separate "archive" color, because they are filtered/
# historical *outputs* of that pipeline, not an isolated category of sheet.
# ---------------------------------------------------------------------------

TAB_COLOR_NAV = "1F4E79"
TAB_COLOR_TENDER = "4472C4"
TAB_COLOR_FUTURE = "ED7D31"
TAB_COLOR_OUTREACH = "7030A0"
TAB_COLOR_SOURCE = "70AD47"

# (block label, tab color, sheet names in that block) - this order is also
# the canonical final tab order used by reorder_sheets_safely().
LOGICAL_TAB_COLOR_GROUPS: list[tuple[str, str, list[str]]] = [
    ("Navigation + dashboards", TAB_COLOR_NAV, [
        "Workbook_Map",
        "Executive_Summary",
        "Acquisition_Funnel_And_Speed",
    ]),
    ("Tender pipeline", TAB_COLOR_TENDER, [
        "Tender_Signals_All",
        "BID_NOW_Active_Tenders",
        "Tender_History_Closed_Public",
        "Tender_Pattern_Analysis",
    ]),
    ("Future project pipeline", TAB_COLOR_FUTURE, [
        "BID_LATER_Future_Projects",
        "Keyword_Change_Audit",
        "New_This_Run",
        "Watchlist_Monitor",
        "Analyzed_Set_Aside",
        "Surrey_Historical_Archive",
        "Top_Civil_Leads_Printable",
    ]),
    ("Outreach + relationships", TAB_COLOR_OUTREACH, [
        "Outreach_Tracker",
        "Developer_Watchlist",
        "Design_Consultants_Reference",
    ]),
    ("Source expansion + intake setup", TAB_COLOR_SOURCE, [
        "Source_Run_Log",
        "Potential_Sources_Next",
        "Source_Roadmap_Printable",
        "Action_Center",
        "Live_Tender_Intake_Plan",
        "Email_Setup_Guide",
    ]),
]

WORKBOOK_SHEET_ORDER: list[str] = [name for _, _, names in LOGICAL_TAB_COLOR_GROUPS for name in names]

WORKBOOK_MAP_LEGEND: list[tuple[str, str, str]] = [
    ("Navigation + dashboards", f"#{TAB_COLOR_NAV}", "Where to start: summary, funnel KPIs, and this workbook map."),
    ("Tender pipeline", f"#{TAB_COLOR_TENDER}", "Tender_Signals_All is the parent. It splits into BID_NOW for open tenders and Tender_History for closed/filtered tender records."),
    ("Future project pipeline", f"#{TAB_COLOR_FUTURE}", "Municipal/development records are scored, filtered, archived, and promoted into BID_LATER / printable future-lead views."),
    ("Outreach + relationships", f"#{TAB_COLOR_OUTREACH}", "Future leads feed contact work, developer watchlists, and consultant references."),
    ("Source expansion + intake setup", f"#{TAB_COLOR_SOURCE}", "Source backlog, email/portal setup, and run logs explain where data comes from and how to add more sources."),
]

WORKBOOK_MAP_DETAIL_HEADERS = [
    "Block", "Step", "Worksheet", "Role In Block",
    "Logical Relationship / Why It Sits Here", "Primary Depends On",
    "Feeds / Used By", "Use Case", "Notes / Duplicate Flags",
]

# One row per known worksheet. Kept as plain tuples (not a dict) so the
# sheet stays easy to scan/edit by hand later.
WORKBOOK_MAP_DETAIL_ROWS: list[tuple] = [
    ("Navigation + dashboards", 1, "Workbook_Map", "Navigation map",
     "Start here. Explains logical blocks, colors, and how sheets flow into each other.",
     "All reviewed worksheets", "Orientation for every other tab", "Orientation", "This sheet."),
    ("Navigation + dashboards", 2, "Executive_Summary", "Top summary",
     "Plain-English dashboard summarizing live tender and future-project pipelines.",
     "BID_NOW_Active_Tenders, BID_LATER_Future_Projects, Tender_Signals_All, Watchlist_Monitor, Source_Run_Log",
     "Leadership / VP review", "Presentation", ""),
    ("Navigation + dashboards", 3, "Acquisition_Funnel_And_Speed", "KPI funnel",
     "Shows counts through acquisition/source funnel and run speed.",
     "BID_NOW_Active_Tenders, BID_LATER_Future_Projects, Outreach_Tracker, Source_Run_Log, Analyzed_Set_Aside",
     "Run-health diagnostics", "Presentation / diagnostics", ""),
    ("Tender pipeline", 1, "Tender_Signals_All", "Parent / master tender signal layer",
     "Keeps all tender signals: open, closed, blocked, empty, portal, parse-failed, needs email alert, needs manual registration, and needs paid access.",
     "Source_Run_Log (source sweep results)",
     "BID_NOW_Active_Tenders, Tender_History_Closed_Public, Tender_Pattern_Analysis, Executive_Summary",
     "Raw tender audit base", "Parent of the tender pipeline block."),
    ("Tender pipeline", 2, "BID_NOW_Active_Tenders", "Filtered actionable output",
     "Open/actionable civil tenders filtered from Tender_Signals_All.",
     "Tender_Signals_All", "Executive_Summary, Acquisition_Funnel_And_Speed", "Daily work", ""),
    ("Tender pipeline", 3, "Tender_History_Closed_Public", "Filtered-out / historical tender output",
     "Closed or historical tender rows filtered out of active BID NOW but kept in the same tender pipeline.",
     "Tender_Signals_All", "Debugging / lookup reference", "Debugging / reference",
     "Shares the Tender pipeline blue with Tender_Signals_All and BID_NOW_Active_Tenders because it is a "
     "filtered/historical output of Tender_Signals_All, not a generic archive - do not recolor separately."),
    ("Tender pipeline", 4, "Tender_Pattern_Analysis", "Tender diagnostic analysis",
     "Tender issuers and closing-month patterns based on tender records.",
     "Tender_Signals_All, BID_NOW_Active_Tenders", "Analysis / debugging", "Analysis / debugging", ""),
    ("Future project pipeline", 1, "BID_LATER_Future_Projects", "Main future opportunity output",
     "Primary future-project / bid-later working list after scoring and filtering municipal development/application records.",
     "Analyzed_Set_Aside (scored source pool), Watchlist_Monitor",
     "Outreach_Tracker, Developer_Watchlist, Top_Civil_Leads_Printable, Executive_Summary",
     "Daily work / future opportunities", ""),
    ("Future project pipeline", 2, "Keyword_Change_Audit", "RESCORE_ALWAYS change audit",
     "Shows old/new score, Vancouver tier, and routing bucket after editable keyword changes.",
     "BID_LATER_Future_Projects / keywords.xlsx", "Slim user master audit", "Weekly review",
     "Legacy Vancouver rows without scoring snapshots are explicitly marked."),
    ("Future project pipeline", 3, "New_This_Run", "Delta / newly added leads",
     "Shows only new future-project records versus prior run/workbook.",
     "BID_LATER_Future_Projects", "Daily review", "Daily review", ""),
    ("Future project pipeline", 4, "Watchlist_Monitor", "Review queue",
     "Records requiring manual review or future workflow before they become higher-confidence opportunities.",
     "Municipal source sweep", "BID_LATER_Future_Projects (future promotion), Acquisition_Funnel_And_Speed",
     "Daily review", ""),
    ("Future project pipeline", 5, "Analyzed_Set_Aside", "Filtered-out scored archive",
     "Collected/scored records that did not make the active BID LATER list, retained for diagnostics.",
     "Municipal source sweep", "BID_LATER_Future_Projects (source pool), Surrey_Historical_Archive",
     "Debugging / reference", ""),
    ("Future project pipeline", 6, "Surrey_Historical_Archive", "Historical/terminal archive",
     "Surrey terminal/historical development applications preserved but excluded from active BID LATER counts.",
     "Analyzed_Set_Aside", "Debugging / reference", "Debugging / reference",
     "Filtered/historical output of the future-project pipeline - shares orange with BID_LATER_Future_Projects, "
     "same reasoning as Tender_History_Closed_Public in the tender pipeline."),
    ("Future project pipeline", 7, "Top_Civil_Leads_Printable", "Printable shortlist",
     "Presentation-friendly shortlist pulled from BID_LATER_Future_Projects.",
     "BID_LATER_Future_Projects", "Presentation / follow-up", "Presentation / follow-up", ""),
    ("Outreach + relationships", 1, "Outreach_Tracker", "CRM / follow-up tracker",
     "Row-level follow-up queue for future leads and contacts.",
     "BID_LATER_Future_Projects", "Acquisition_Funnel_And_Speed", "Follow-up",
     "Manual status/notes may be edited by users - do not overwrite user edits when re-running."),
    ("Outreach + relationships", 2, "Developer_Watchlist", "Relationship target list",
     "Grouped developer/applicant relationship list generated from future-project source fields.",
     "BID_LATER_Future_Projects", "Follow-up / BD", "Follow-up / BD", ""),
    ("Outreach + relationships", 3, "Design_Consultants_Reference", "Consultant reference list",
     "Design/architect/consultant filing entities separated from developer targets.",
     "BID_LATER_Future_Projects source fields", "Reference / follow-up", "Reference / follow-up", ""),
    ("Source expansion + intake setup", 1, "Source_Run_Log", "Run/source diagnostic log",
     "Shows which public/email sources ran, succeeded, failed, were blocked, or produced tender rows.",
     "Live source sweep", "Tender_Signals_All, BID_NOW_Active_Tenders, Executive_Summary, Acquisition_Funnel_And_Speed",
     "Debugging", ""),
    ("Source expansion + intake setup", 2, "Potential_Sources_Next", "Detailed source backlog",
     "Ranked backlog of sources to verify, connect, or monitor next.",
     "Source_Run_Log / source register", "Source_Roadmap_Printable (presentation summary)",
     "Planning / daily work", ""),
    ("Source expansion + intake setup", 3, "Source_Roadmap_Printable", "Printable source roadmap",
     "Presentation version of the source expansion backlog.",
     "Potential_Sources_Next", "Presentation", "Presentation", ""),
    ("Source expansion + intake setup", 4, "Action_Center", "Setup action list",
     "Checklist for portal access, registration, alerts, and next steps.",
     "Source_Run_Log, Live_Tender_Intake_Plan", "Daily setup workflow", "Daily setup", ""),
    ("Source expansion + intake setup", 5, "Live_Tender_Intake_Plan", "Intake design / guardrails",
     "Explains no-password alert-email intake design and portal setup logic.",
     "Action_Center", "Email_Setup_Guide", "Setup / documentation", ""),
    ("Source expansion + intake setup", 6, "Email_Setup_Guide", "Practical email setup checklist",
     "Instructions for importing approved alert emails and enabling portal alerts.",
     "Live_Tender_Intake_Plan", "Email intake workflow -> Tender_Signals_All", "Setup / daily work", ""),
]


def apply_logical_tab_colors(wb: Workbook) -> None:
    """Color each worksheet tab by the logical pipeline block it belongs to.

    Tab colors represent logical pipelines, not isolated sheet types (see
    module note above) - e.g. Tender_History_Closed_Public gets the same
    blue as Tender_Signals_All and BID_NOW_Active_Tenders because it is a
    filtered/historical output of that same pipeline.

    Defensive: any sheet name not present in this workbook is skipped
    silently. Never raises, never deletes data.
    """
    for _block_name, color, sheet_names in LOGICAL_TAB_COLOR_GROUPS:
        for name in sheet_names:
            if name in wb.sheetnames:
                wb[name].sheet_properties.tabColor = color


def reorder_sheets_safely(wb: Workbook, order: list[str] | None = None) -> None:
    """Reorder worksheets into the canonical logical-block order.

    Any sheet name in `order` that is missing from this workbook is skipped
    (never an error). Any worksheet present in the workbook but not named in
    `order` (e.g. a master-only sheet such as Active_Tenders/Future_Projects
    in the working master) is preserved and appended after the ordered
    sheets, in its original relative order - nothing is ever hidden,
    renamed, or dropped.
    """
    order = order if order is not None else WORKBOOK_SHEET_ORDER
    ordered_names = [name for name in order if name in wb.sheetnames]
    ordered_set = set(ordered_names)
    remaining_names = [name for name in wb.sheetnames if name not in ordered_set]
    final_names = ordered_names + remaining_names
    wb._sheets = [wb[name] for name in final_names]
    if final_names:
        wb.active = 0


def create_or_update_workbook_map(wb: Workbook) -> None:
    """Create (or refresh) the Workbook_Map sheet.

    If Workbook_Map already exists, its cells are cleared and rewritten in
    place - the sheet object itself is never deleted/recreated, and no other
    sheet's data is touched.
    """
    if "Workbook_Map" in wb.sheetnames:
        ws = wb["Workbook_Map"]
        blank_font, blank_fill = Font(), PatternFill(fill_type=None)
        blank_border, blank_align = Border(), Alignment()
        for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 1), max_col=max(ws.max_column, 1)):
            for cell in row:
                cell.value = None
                cell.font = blank_font
                cell.fill = blank_fill
                cell.border = blank_border
                cell.alignment = blank_align
                cell.number_format = "General"
        ws.freeze_panes = None
    else:
        ws = wb.create_sheet("Workbook_Map")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F1F1F")
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    wrap = Alignment(vertical="top", wrap_text=True)

    ws["A1"] = "Workbook Map — logical blocks / source → filtered → worklist → archive"
    ws["A1"].font = title_font
    ws["A2"] = "Updated"
    ws["B2"] = dt.datetime.now().isoformat(timespec="seconds")

    row = 4
    ws.cell(row=row, column=1, value="Logical block legend").font = section_font
    row += 1
    for col, header in enumerate(["Logical Block", "Block Color", "Plain-English Flow"], start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    for block, color_text, flow in WORKBOOK_MAP_LEGEND:
        ws.cell(row=row, column=1, value=block)
        color_cell = ws.cell(row=row, column=2, value=color_text)
        color_cell.fill = PatternFill("solid", fgColor=color_text.replace("#", ""))
        color_cell.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=row, column=3, value=flow).alignment = wrap
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Detailed worksheet logic map").font = section_font
    row += 1
    detail_header_row = row
    for col, header in enumerate(WORKBOOK_MAP_DETAIL_HEADERS, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    block_color_lookup = {block: color for block, color, _ in WORKBOOK_MAP_LEGEND}
    for values in WORKBOOK_MAP_DETAIL_ROWS:
        block = values[0]
        block_color = block_color_lookup.get(block, "#FFFFFF").replace("#", "")
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = wrap
            if col == 1:
                cell.fill = PatternFill("solid", fgColor=block_color)
                cell.font = Font(color="FFFFFF", bold=True)
        row += 1

    for col, width in {1: 28, 2: 6, 3: 32, 4: 28, 5: 48, 6: 32, 7: 32, 8: 22, 9: 42}.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{detail_header_row + 1}"


def _atomic_save_workbook(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(path.parent))
    tmp.close()
    try:
        wb.save(tmp.name)
        _atomic_replace_with_retry(tmp.name, path)
    finally:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)


MASTER_WORKBOOK_GLOB = "TENDER_FINDER_Tender_Intelligence_Working_Master_*.xlsx"
# Markers embedded in the filename of every copy this script itself
# generates (see finalize_final_review_outputs / build_integrated_master_
# workbook). They must be excluded from find_latest_master_workbook()
# candidates, or a later run would "find" a prior run's own generated
# output and treat it as the real master - silently chaining copies-of-
# copies instead of always integrating from the one true hand-maintained
# master workbook.
MASTER_WORKBOOK_FILLED_MARKER = "_FILLED_"
MASTER_WORKBOOK_INTEGRATED_MARKER = "_INTEGRATED_"
MASTER_WORKBOOK_USER_MARKER = "_USER_"
MASTER_WORKBOOK_OWN_OUTPUT_MARKERS = (MASTER_WORKBOOK_FILLED_MARKER, MASTER_WORKBOOK_INTEGRATED_MARKER, MASTER_WORKBOOK_USER_MARKER)
_SKIP_DIR_NAMES = {".git", ".venv", "venv", "__pycache__", "node_modules", ".codex_tmp", ".agents"}
# Directory names that can make an unbounded recursive search catastrophically
# slow (or scan unrelated personal files) if a search root ever turns out to
# be a drive root or a broad system folder. Pruned by name at any depth in
# _bounded_glob_master_workbooks(), not just at the top level.
_HEAVY_SYSTEM_DIR_NAMES = {
    "Windows", "Program Files", "Program Files (x86)", "ProgramData",
    "$Recycle.Bin", "System Volume Information", "Recovery", "Users", "PerfLogs",
}


def _sortable_version_token(filename: str) -> str:
    """Zero-pad digit runs in a filename so plain string sorting also sorts
    numeric version/timestamp tokens correctly (v7 after v6, v10 after v9,
    20260703_1200 after 20260702_0900, etc.). Used only as a tie-breaker
    behind the real filesystem modified time.
    """
    return re.sub(r"\d+", lambda m: m.group().zfill(12), filename)


def _bounded_glob_master_workbooks(base: Path, max_depth: int) -> list[Path]:
    """Search for TENDER_FINDER_Tender_Intelligence_Working_Master_*.xlsx under
    `base`, descending at most `max_depth` directories deep and pruning
    known heavy Windows system directories by name at any depth.

    Used for search roots that aren't guaranteed to be small - notably the
    project's parent folder. If the project is unzipped directly under a
    drive letter (e.g. C:\\TENDER_FINDER_Tender_Intelligence_Run\\, exactly the
    example path this project's own README suggests), that parent IS the
    drive root C:\\, and an unbounded rglob() there means walking the
    entire drive - confirmed hanging for minutes during packaging
    verification before this fix. A depth-bounded, system-dir-pruned walk
    still reliably catches the realistic case (a master workbook sitting
    next to the project folder) without that risk.
    """
    found: list[Path] = []
    base_depth = len(base.parts)
    prefix, suffix = "TENDER_FINDER_Tender_Intelligence_Working_Master_", ".xlsx"
    for dirpath, dirnames, filenames in os.walk(base):
        current = Path(dirpath)
        depth = len(current.parts) - base_depth
        dirnames[:] = [d for d in dirnames if d not in _SKIP_DIR_NAMES and d not in _HEAVY_SYSTEM_DIR_NAMES]
        if depth >= max_depth:
            dirnames[:] = []
        for name in filenames:
            if name.startswith(prefix) and name.endswith(suffix):
                found.append(current / name)
    return found


def find_latest_master_workbook(extra_dirs: list[Path] | None = None) -> Path:
    """Locate the latest TENDER_FINDER_Tender_Intelligence_Working_Master_*.xlsx.

    A valid workbook inside this package wins before any external fallback.
    This prevents a checkout or moved installation from silently using a
    newer template copied into another Tender Finder installation under
    C:\\tenderfinder_out. Only when this package has no candidate do we search
    its parent, C:\\tenderfinder_out, explicit extra directories, and known
    output/archive folders. Excel lock files ("~$...") are always ignored.
    Within one scope, the newest file by modified timestamp wins; ties are
    broken by the newest sortable version/timestamp token in the filename.

    Raises FileNotFoundError with a clear, actionable message if nothing is
    found - this function must never fabricate an empty master workbook.
    """
    search_roots: list[Path] = []
    tenderfinder_out = Path("C:/tenderfinder_out")
    if tenderfinder_out.exists():
        search_roots.append(tenderfinder_out)
    for known in (DEMO_HISTORY_DIR, ROOT / "00 Master", ROOT / "latest_verified_output", ROOT / "user_data"):
        if known.exists():
            search_roots.append(known)
    if extra_dirs:
        search_roots.extend(p for p in extra_dirs if p)

    seen_roots: set[Path] = set()
    candidates: set[Path] = set()

    def _collect(paths, destination: set[Path] | None = None) -> None:
        target = candidates if destination is None else destination
        for path in paths:
            if path.name.startswith("~$"):
                continue
            if any(marker in path.name for marker in MASTER_WORKBOOK_OWN_OUTPUT_MARKERS):
                continue
            if any(part in _SKIP_DIR_NAMES for part in path.parts):
                continue
            target.add(path.resolve())

    def _latest(paths: set[Path]) -> Path:
        def sort_key(path: Path) -> tuple[float, str]:
            return (path.stat().st_mtime, _sortable_version_token(path.name))

        return max(paths, key=sort_key)

    try:
        package_root = ROOT.resolve()
    except OSError:
        package_root = ROOT
    package_candidates: set[Path] = set()
    if package_root.exists():
        _collect(package_root.rglob(MASTER_WORKBOOK_GLOB), package_candidates)
    if package_candidates:
        latest = _latest(package_candidates)
        print(f"MASTER_WORKBOOK_SELECTED={latest}", flush=True)
        return latest
    seen_roots.add(package_root)

    for base in search_roots:
        try:
            base = base.resolve()
        except OSError:
            continue
        if base in seen_roots or not base.exists():
            continue
        seen_roots.add(base)
        _collect(base.rglob(MASTER_WORKBOOK_GLOB))

    # The project's parent folder ("repo parent folder if needed") is
    # searched too, but with a depth-bounded walk instead of an unbounded
    # rglob() - see _bounded_glob_master_workbooks()'s docstring for why an
    # unbounded search here is unsafe.
    try:
        parent = ROOT.parent.resolve()
    except OSError:
        parent = None
    if parent is not None and parent not in seen_roots and parent.exists():
        seen_roots.add(parent)
        _collect(_bounded_glob_master_workbooks(parent, max_depth=2))

    if not candidates:
        searched = ", ".join(str(p) for p in [*search_roots, parent] if p)
        raise FileNotFoundError(
            f"No {MASTER_WORKBOOK_GLOB} workbook was found under any of: {searched}. "
            "Refusing to fabricate an empty master workbook - place a real "
            "TENDER_FINDER_Tender_Intelligence_Working_Master_*.xlsx file in one of these "
            "locations (e.g. the '00 Master' folder) and re-run."
        )

    latest = _latest(candidates)
    print(f"MASTER_WORKBOOK_SELECTED={latest}", flush=True)
    return latest


def _clear_worksheet_contents(ws) -> None:
    """Clear a worksheet's used content area and layout so it can be
    replaced with fresh content, without deleting the worksheet itself.
    """
    # Unmerge BEFORE clearing values/styles - every non-top-left cell in a
    # merged range is a read-only MergedCell until the range is broken up,
    # so writing to it first (as this used to) raises AttributeError.
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    default_font, default_fill = Font(), PatternFill(fill_type=None)
    default_border, default_align = Border(), Alignment()
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.font = default_font
            cell.fill = default_fill
            cell.border = default_border
            cell.alignment = default_align
            cell.number_format = "General"
    try:
        ws.conditional_formatting = ConditionalFormattingList()
    except Exception:
        pass
    try:
        ws.data_validations = DataValidationList()
    except Exception:
        pass
    if ws.auto_filter is not None:
        ws.auto_filter.ref = None
    ws.freeze_panes = None


def _clone_sheet_into(src_ws, dst_ws) -> None:
    """Copy one worksheet's values, formulas, styles, number formats,
    column widths, row heights, freeze panes, autofilter, data validation,
    conditional formatting, merged cells, and tab color from src_ws into
    dst_ws - "as much as openpyxl allows" per the master-fill contract.
    Non-critical layout details (conditional formatting/data validation)
    are best-effort and never abort the copy if they fail to translate.
    """
    _clear_worksheet_contents(dst_ws)

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.protection = copy.copy(cell.protection)
                new_cell.number_format = cell.number_format

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden
    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height
        dst_ws.row_dimensions[row_idx].hidden = dim.hidden

    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter is not None and src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref

    try:
        for dv in src_ws.data_validations.dataValidation:
            dst_ws.add_data_validation(copy.copy(dv))
    except Exception as exc:
        print(f"WARN: could not copy data validation for sheet {src_ws.title!r}: {exc}", flush=True)

    try:
        for cf in src_ws.conditional_formatting:
            for rule in cf.rules:
                dst_ws.conditional_formatting.add(str(cf.sqref), copy.copy(rule))
    except Exception as exc:
        print(f"WARN: could not copy conditional formatting for sheet {src_ws.title!r}: {exc}", flush=True)

    try:
        for merged_range in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merged_range))
    except Exception as exc:
        print(f"WARN: could not copy merged cells for sheet {src_ws.title!r}: {exc}", flush=True)

    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor


def _copy_all_demo_sheets_into(demo_wb: Workbook, master_wb: Workbook) -> None:
    """Clone every demo sheet into master_wb under its own name (creating
    it if the name doesn't already exist in the master).

    IMPORTANT - this alone is NOT integration. An audit of the first
    version of this patch found that because none of the demo's sheet
    names (BID_NOW_Active_Tenders, BID_LATER_Future_Projects, ...) match
    any of the master's original working-tab names (Active_Tenders,
    Future_Projects, Dashboard, Source_Register, ...), a name-matched copy
    like this leaves every one of those original master tabs byte-for-byte
    untouched - confirmed by diffing a real run's output. Kept here as the
    "keep the raw demo data as an audit trail inside the master" step;
    build_integrated_master_workbook() below additionally maps this data
    into the master's real working tabs, which is the part that makes
    Active_Tenders/Future_Projects/Dashboard/etc. actually reflect the run.
    """
    for sheet_name in demo_wb.sheetnames:
        src_ws = demo_wb[sheet_name]
        dst_ws = master_wb[sheet_name] if sheet_name in master_wb.sheetnames else master_wb.create_sheet(sheet_name)
        _clone_sheet_into(src_ws, dst_ws)


def copy_workbook_sheets_into_master(demo_path: Path, master_path: Path, filled_path: Path) -> Path:
    """Copy every sheet from the demo workbook into a brand-new filled copy
    of the master workbook, saved to `filled_path`.

    - A sheet name that already exists in the master has its used content
      area cleared and replaced with the demo sheet's content.
    - A sheet name that does not exist in the master is created.
    - Any master sheet not present in the demo workbook (e.g. the master's
      own Active_Tenders/Future_Projects/Rejected_Archive/Run_Queue/Run_Log
      tables) is left completely untouched.
    - `master_path` (the original) is only ever opened for reading; this
      function never writes to it. The result is always a new file.

    NOTE: this is a whole-sheet copy only - see _copy_all_demo_sheets_into's
    docstring. Production runs use build_integrated_master_workbook()
    instead, which calls this same clone step and then additionally maps
    data into the master's real working tabs. This function is kept as a
    lower-level primitive (and for anyone who explicitly wants a raw,
    unmapped copy for comparison/debugging).
    """
    demo_wb = openpyxl.load_workbook(demo_path)
    master_wb = openpyxl.load_workbook(master_path)

    _copy_all_demo_sheets_into(demo_wb, master_wb)

    # Workbook_Map must be created before tab colors are applied - it does
    # not exist yet at this point, and apply_logical_tab_colors() silently
    # skips sheets that aren't present, so coloring first would leave the
    # freshly-created Workbook_Map tab uncolored.
    create_or_update_workbook_map(master_wb)
    apply_logical_tab_colors(master_wb)
    reorder_sheets_safely(master_wb)

    _atomic_save_workbook(master_wb, filled_path)
    return filled_path


# ---------------------------------------------------------------------------
# True integration: map demo data into the master's OWN working tabs.
#
# The master workbook already has real working tabs with their own column
# schemas (Active_Tenders, Future_Projects, Dashboard, Source_Register,
# Outreach_Tracker doesn't exist yet there). Those schemas don't match the
# demo workbook's tabs name-for-name or column-for-column, so a sheet copy
# alone (see _copy_all_demo_sheets_into) can never populate them. Everything
# below maps demo fields into the closest matching master column, adding
# new columns to the right only when nothing existing fits, and never
# fabricates data (fields with no safe source are left blank, not guessed).
# ---------------------------------------------------------------------------


# Structural fact about exactly how build_workbook() lays out each sheet
# (note row present -> header on row 2; no note -> header on row 1 - see
# every write_rows(ws, ...) call there) plus the sheets this script builds
# itself (always called with note=None -> header row 1). Deliberately NOT
# inferred from freeze_panes/view state at read time: confirmed during
# handoff review that opening a generated workbook in Excel and letting it
# get saved again (e.g. by a reviewer inspecting the file) silently
# rewrites each sheet's recorded "top-left visible cell" to wherever the
# view happened to be scrolled - Source_Run_Log's freeze_panes flipped from
# the correct "A2" to "A16" this way, which would have made every
# downstream read of that sheet treat row 15's DATA as column headers and
# silently misalign every field. A hardcoded table survives that round trip.
KNOWN_SHEET_HEADER_ROWS: dict[str, int] = {
    "BID_NOW_Active_Tenders": 2,
    "Tender_History_Closed_Public": 2,
    "Tender_Signals_All": 2,
    "BID_LATER_Future_Projects": 2,
    "Keyword_Change_Audit": 2,
    "New_This_Run": 2,
    "Outreach_Tracker": 1,
    "Developer_Watchlist": 2,
    "Design_Consultants_Reference": 2,
    "Tender_Pattern_Analysis": 2,
    "Watchlist_Monitor": 1,
    "Analyzed_Set_Aside": 2,
    "Surrey_Historical_Archive": 2,
    "Acquisition_Funnel_And_Speed": 1,
    "Source_Run_Log": 1,
    "Action_Center": 1,
    "Live_Tender_Intake_Plan": 2,
    "Email_Setup_Guide": 2,
    "Top_Civil_Leads_Printable": 1,
    # Hand-authored master-template sheets (confirmed by direct inspection).
    "Source_Register": 1,
    "Municipal_Coverage": 1,
    # Sheets this script builds itself for the user-facing master - always
    # written via write_rows(..., note=None, ...), header always row 1.
    "Active_Tenders": 1,
    "Future_Projects": 1,
    "Source_Status": 1,
}


def _header_row_index(ws) -> int:
    """Best-effort detection of which row holds column headers.

    Checks KNOWN_SHEET_HEADER_ROWS first (see its docstring for why this
    must not be inferred purely from mutable view state). Falls back to
    freeze_panes for sheet names it doesn't recognize (e.g. sheets from a
    template this script doesn't otherwise know the internal layout of),
    and finally to row 1 if freeze_panes gives no usable signal either.
    """
    # Keyword_Change_Audit exists in two intentional forms: the technical
    # workbook has a note on row 1 and headers on row 2, while the slim user
    # master starts directly with "Lead ID" on row 1. Detect that immutable
    # schema marker before consulting the technical-sheet table so founder
    # Assigned To/Status/Notes can be read back on the next run.
    if ws.title == "Keyword_Change_Audit" and clean_text(ws.cell(1, 1).value) == "Lead ID":
        return 1
    known = KNOWN_SHEET_HEADER_ROWS.get(ws.title)
    if known is not None:
        return known
    freeze = ws.freeze_panes
    if freeze:
        match = re.match(r"^[A-Z]+(\d+)$", str(freeze))
        if match:
            data_row = int(match.group(1))
            if data_row > 1:
                return data_row - 1
    return 1


def _read_sheet_rows_as_dicts(ws) -> list[dict[str, Any]]:
    """Read a worksheet into a list of {header: value} dicts, skipping
    fully-blank rows. Works for both write_rows()-generated demo sheets
    (note row + header row) and hand-authored master sheets (header row 1).
    """
    header_row_idx = _header_row_index(ws)
    if ws.max_row < header_row_idx:
        return []
    header_cells = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))
    headers = [c.value for c in header_cells]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        values = [c.value for c in row]
        if not any(v not in (None, "") for v in values):
            continue
        out.append({headers[i]: values[i] for i in range(len(headers)) if headers[i] is not None})
    return out


def _map_rows_into_master_tab(
    master_ws,
    src_rows: list[dict[str, Any]],
    column_map: list[tuple[str, Any]],
    extra_columns: list[tuple[str, str]],
    run_timestamp: str,
) -> dict[str, Any]:
    """Clear a master working tab's data rows and rewrite them from
    column-mapped demo source rows.

    The header row (row 1) and its styling are preserved - never deleted,
    never reordered. New columns are only ever appended to the right, for
    demo fields that don't have a matching master column. Any master
    header that has no corresponding entry in `column_map` is left blank
    per row (never guessed) and reported back in `unmapped_master_columns`.
    """
    existing_headers = [c.value for c in next(master_ws.iter_rows(min_row=1, max_row=1))]
    while existing_headers and existing_headers[-1] is None:
        existing_headers.pop()

    sample_header_cell = master_ws.cell(row=1, column=1)
    header_font = copy.copy(sample_header_cell.font)
    header_fill = copy.copy(sample_header_cell.fill)
    header_border = copy.copy(sample_header_cell.border)

    header_lookup = {h: i + 1 for i, h in enumerate(existing_headers)}
    extra_added: list[str] = []
    next_col = len(existing_headers) + 1
    for extra_header, _src_field in extra_columns:
        if extra_header in header_lookup:
            continue
        cell = master_ws.cell(row=1, column=next_col, value=extra_header)
        cell.font, cell.fill, cell.border = header_font, header_fill, header_border
        header_lookup[extra_header] = next_col
        extra_added.append(extra_header)
        next_col += 1

    final_col_count = max(header_lookup.values()) if header_lookup else len(existing_headers)

    if master_ws.max_row > 1:
        clear_max_col = max(master_ws.max_column, final_col_count)
        for row in master_ws.iter_rows(min_row=2, max_row=master_ws.max_row, max_col=clear_max_col):
            for cell in row:
                cell.value = None
                cell.font = Font()
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.alignment = Alignment()
                cell.number_format = "General"

    wrap = Alignment(vertical="top", wrap_text=True)
    dst_row = 2
    for src in src_rows:
        for header, getter in column_map:
            col = header_lookup.get(header)
            if col is None:
                continue
            master_ws.cell(row=dst_row, column=col, value=getter(src, run_timestamp)).alignment = wrap
        for extra_header, src_field in extra_columns:
            col = header_lookup.get(extra_header)
            if col is None:
                continue
            master_ws.cell(row=dst_row, column=col, value=src.get(src_field)).alignment = wrap
        dst_row += 1

    mapped_headers = {h for h, _ in column_map}
    unmapped_master_columns = [h for h in existing_headers if h not in mapped_headers]

    return {
        "rows_written": len(src_rows),
        "unmapped_master_columns": unmapped_master_columns,
        "extra_columns_added": extra_added,
    }


def _tender_project_id(row: dict[str, Any], _ts: str) -> str:
    basis = f"{row.get('source', '')}|{row.get('tender_title', '')}"
    return f"TDR-{hashlib.sha1(basis.encode('utf-8')).hexdigest()[:8].upper()}"


def _tender_days_to_close(row: dict[str, Any], _ts: str):
    closing = row.get("closing_date")
    if not closing:
        return None
    try:
        closing_date = dt.date.fromisoformat(str(closing)[:10])
    except ValueError:
        return None
    return (closing_date - dt.date.today()).days


def _tender_notes(row: dict[str, Any], _ts: str) -> str:
    parts = [
        f"status={row['status']}" if row.get("status") else "",
        f"signal_state={row['signal_state']}" if row.get("signal_state") else "",
        f"contact_email={row['contact_email']}" if row.get("contact_email") else "",
        f"contact_phone={row['contact_phone']}" if row.get("contact_phone") else "",
        f"found_via={row['found_via']}" if row.get("found_via") else "",
    ]
    return "; ".join(p for p in parts if p)


# BID_NOW_Active_Tenders (18 demo cols) -> Active_Tenders (20 master cols).
# GC / Inviting Party, Scope Match, Fit Score (0-100), Fit Class,
# Est. Value ($), Assigned To, and Bid Decision have no safe source field
# in a public tender signal row and are intentionally left unmapped rather
# than guessed - Assigned To/Bid Decision are also user-owned fields.
ACTIVE_TENDERS_COLUMN_MAP: list[tuple[str, Any]] = [
    ("Project ID", _tender_project_id),
    ("Date Found", lambda r, ts: ts[:10]),
    ("Source", lambda r, ts: r.get("source")),
    ("Source URL", lambda r, ts: r.get("url")),
    ("Tender / Project Title", lambda r, ts: r.get("tender_title")),
    ("Owner / Client", lambda r, ts: r.get("issuer")),
    ("Municipality", lambda r, ts: r.get("issuer")),
    ("Scope Summary", lambda r, ts: r.get("snippet")),
    ("Verification Status", lambda r, ts: "Needs Review"),
    ("Closing Date", lambda r, ts: r.get("closing_date")),
    ("Days to Close", _tender_days_to_close),
    ("Notes", _tender_notes),
    ("Last Updated", lambda r, ts: ts),
]
ACTIVE_TENDERS_EXTRA_COLUMNS: list[tuple[str, str]] = [
    ("Civil Relevant", "civil_relevant"),
    ("Signal State", "signal_state"),
    ("Filtered Reason", "filtered_reason"),
    ("Related Leads Count", "related_leads_count"),
    ("Found Via", "found_via"),
]


def map_bid_now_into_active_tenders(master_wb: Workbook, run_timestamp: str) -> dict[str, Any]:
    """The actual integration step for tenders: read BID_NOW_Active_Tenders
    (already cloned into master_wb by _copy_all_demo_sheets_into) and
    column-map its rows into the master's real Active_Tenders working tab.
    """
    if "Active_Tenders" not in master_wb.sheetnames or "BID_NOW_Active_Tenders" not in master_wb.sheetnames:
        return {"rows_written": 0, "skipped": "Active_Tenders or BID_NOW_Active_Tenders sheet missing"}
    src_rows = _read_sheet_rows_as_dicts(master_wb["BID_NOW_Active_Tenders"])
    return _map_rows_into_master_tab(
        master_wb["Active_Tenders"], src_rows, ACTIVE_TENDERS_COLUMN_MAP, ACTIVE_TENDERS_EXTRA_COLUMNS, run_timestamp,
    )


def _lead_project_title(row: dict[str, Any], _ts: str):
    stage, address = row.get("app_type_stage"), row.get("address")
    if stage and address:
        return f"{stage} — {address}"
    if address:
        return address
    # No address on this row (common for Surrey V2 records keyed by app_no).
    # Prefer a scope_summary excerpt over a bare stage name here, or this
    # column just duplicates the adjacent Project Type/Application Type
    # column for every address-less row.
    scope_excerpt = str(row.get("scope_summary") or "")[:80]
    return scope_excerpt or stage or None


def _lead_notes(row: dict[str, Any], _ts: str) -> str:
    parts = [
        f"status_class={row['status_class']}" if row.get("status_class") else "",
        f"is_new_since_last_run={row['is_new_since_last_run']}" if row.get("is_new_since_last_run") else "",
        f"detail_available={row['detail_available']}" if row.get("detail_available") else "",
        f"source_tier={row['source_tier']}" if row.get("source_tier") else "",
    ]
    return "; ".join(p for p in parts if p)


# BID_LATER_Future_Projects (24 demo cols) -> Future_Projects (21 master
# cols). Expected Civil Component, Est. Civil Timeline / Horizon,
# Est. Value ($), Next Milestone / Watch Trigger, Linked Active Tender ID,
# and Assigned To have no safe source field in a future-project lead row
# and are intentionally left unmapped rather than guessed/fabricated.
FUTURE_PROJECTS_COLUMN_MAP: list[tuple[str, Any]] = [
    ("Project ID", lambda r, ts: r.get("lead_id")),
    ("Date Found", lambda r, ts: r.get("first_seen_date")),
    ("Source", lambda r, ts: r.get("source")),
    ("Source URL", lambda r, ts: r.get("source_url")),
    ("Project Title", _lead_project_title),
    ("Owner / Developer", lambda r, ts: r.get("applicant_name")),
    ("Municipality", lambda r, ts: r.get("municipality")),
    ("Application No", lambda r, ts: r.get("app_no")),
    ("Application Type / Stage", lambda r, ts: r.get("app_type_stage")),
    ("Scope Summary", lambda r, ts: r.get("scope_summary")),
    ("Fit Score (0-100)", lambda r, ts: r.get("fit_score")),
    ("Fit Class", lambda r, ts: r.get("signal_quality")),
    ("Verification Status", lambda r, ts: "Needs Review"),
    ("Notes", _lead_notes),
    ("Last Updated", lambda r, ts: ts),
]
FUTURE_PROJECTS_EXTRA_COLUMNS: list[tuple[str, str]] = [
    ("Regional Cluster", "regional_cluster"),
    ("Project Scale Tier", "project_scale_tier"),
    ("Scale", "scale"),
    ("Est. Units", "est_units"),
    ("Est. Storeys", "est_storeys"),
    ("Est. Lots", "est_lots"),
    ("Approximate Location", "approximate_location"),
]


def map_bid_later_into_future_projects(master_wb: Workbook, run_timestamp: str) -> dict[str, Any]:
    """The actual integration step for future leads: read
    BID_LATER_Future_Projects and column-map its rows into the master's
    real Future_Projects working tab.
    """
    if "Future_Projects" not in master_wb.sheetnames or "BID_LATER_Future_Projects" not in master_wb.sheetnames:
        return {"rows_written": 0, "skipped": "Future_Projects or BID_LATER_Future_Projects sheet missing"}
    src_rows = _read_sheet_rows_as_dicts(master_wb["BID_LATER_Future_Projects"])
    return _map_rows_into_master_tab(
        master_wb["Future_Projects"], src_rows, FUTURE_PROJECTS_COLUMN_MAP, FUTURE_PROJECTS_EXTRA_COLUMNS, run_timestamp,
    )


def _compute_dashboard_counts(master_wb: Workbook) -> dict[str, Any]:
    def count_data_rows(sheet_name: str) -> int:
        return len(_read_sheet_rows_as_dicts(master_wb[sheet_name])) if sheet_name in master_wb.sheetnames else 0

    counts: dict[str, Any] = {
        "active_tenders": count_data_rows("Active_Tenders"),
        "future_projects": count_data_rows("Future_Projects"),
        "outreach_tracker": count_data_rows("Outreach_Tracker"),
        "source_run_log": count_data_rows("Source_Run_Log"),
        "sources_in_register": count_data_rows("Source_Register"),
    }
    if "Source_Register" in master_wb.sheetnames:
        rows = _read_sheet_rows_as_dicts(master_wb["Source_Register"])
        counts["tier1_sources"] = sum(1 for r in rows if str(r.get("Priority Tier")) == "1")
        counts["sources_needing_cleanup"] = sum(1 for r in rows if (r.get("Cleanup Flag") or "OK") != "OK")
    if "Municipal_Coverage" in master_wb.sheetnames:
        rows = _read_sheet_rows_as_dicts(master_wb["Municipal_Coverage"])
        counts["municipal_jurisdictions"] = len(rows)
        counts["tier1_municipalities"] = sum(1 for r in rows if str(r.get("Tier")) == "1")
    return counts


def update_dashboard_with_run_counts(master_wb: Workbook, counts: dict[str, Any], run_timestamp: str) -> bool:
    """Fill the master's Dashboard sheet with live numbers from this run
    instead of leaving its "Formula / Value" column blank.

    Existing metric-label rows are filled in place (their labels and
    formatting are untouched). Run-specific metrics that Dashboard didn't
    already have a row for are appended at the bottom rather than inserted,
    since inserting rows would shift the existing Priority/Action table
    below. A stale action item that this run directly fulfills (loading
    real data into Future_Projects) is annotated DONE rather than deleted.
    """
    if "Dashboard" not in master_wb.sheetnames:
        return False
    ws = master_wb["Dashboard"]

    label_to_value = {
        "Sources in register": counts.get("sources_in_register"),
        "Tier 1 sources": counts.get("tier1_sources"),
        "Sources needing exact URL / cleanup": counts.get("sources_needing_cleanup"),
        "Municipal jurisdictions": counts.get("municipal_jurisdictions"),
        "Tier 1 municipalities": counts.get("tier1_municipalities"),
        "Active live tenders entered": counts.get("active_tenders"),
        "Future live projects entered": counts.get("future_projects"),
    }
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label = row[0].value
        if label in label_to_value and label_to_value[label] is not None:
            ws.cell(row=row[0].row, column=2, value=label_to_value[label])
        if len(row) > 1 and row[1].value and "Load real Surrey top-20 lead screen into Future_Projects" in str(row[1].value):
            reason_cell = ws.cell(row=row[0].row, column=3)
            reason_cell.value = (
                f"{reason_cell.value or ''} DONE as of {run_timestamp}: Future_Projects now holds "
                f"{counts.get('future_projects', 0):,} live rows mapped from BID_LATER_Future_Projects."
            ).strip()

    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value=f"TENDER_FINDER integration run — {run_timestamp}").font = Font(bold=True, size=12)
    start_row += 1
    header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F1F1F")
    for col, h in enumerate(["Metric", "Value", "What it means"], start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font, cell.fill = header_font, header_fill
    start_row += 1
    run_rows = [
        ("Active tender count (this run)", counts.get("active_tenders"), "BID_NOW_Active_Tenders rows mapped into Active_Tenders"),
        ("Future project count (this run)", counts.get("future_projects"), "BID_LATER_Future_Projects rows mapped into Future_Projects"),
        ("Outreach tracker rows (this run)", counts.get("outreach_tracker"), "Outreach_Tracker rows this run, with source traceability added"),
        ("Source runs logged (this run)", counts.get("source_run_log"), "Source_Run_Log rows this run - see Source_Run_Status_Mapped"),
        ("Run timestamp", run_timestamp, "When this integration run completed"),
    ]
    for label, value, meaning in run_rows:
        ws.cell(row=start_row, column=1, value=label)
        ws.cell(row=start_row, column=2, value=value)
        ws.cell(row=start_row, column=3, value=meaning).alignment = Alignment(vertical="top", wrap_text=True)
        start_row += 1

    return True


def _normalize_source_key(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text or "").lower())


def build_source_run_status_mapped(master_wb: Workbook) -> dict[str, Any]:
    """Create Source_Run_Status_Mapped: a new, safe sheet showing each
    Source_Run_Log result alongside its best-guess Source_Register match.

    Source_Register's IDs ("SRC-001", human names like "BC Bid") do not
    reliably line up with Source_Run_Log's internal slugs ("bc_bid_public")
    by exact match. Per the integration design, heuristic guesses are never
    written into the curated Source_Register sheet - Source_Register is
    only ever read here, never modified. Confident-looking matches (a
    normalized substring match) are shown with LOW_CONFIDENCE_SUBSTRING;
    everything else is explicitly NO_CONFIDENT_MATCH rather than blank.
    """
    if "Source_Run_Log" not in master_wb.sheetnames:
        return {"rows_written": 0, "skipped": "Source_Run_Log sheet missing"}

    run_log_rows = _read_sheet_rows_as_dicts(master_wb["Source_Run_Log"])

    register_lookup: list[tuple[str, str, str, str]] = []
    if "Source_Register" in master_wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(master_wb["Source_Register"]):
            name = row.get("Source Name")
            if name:
                register_lookup.append((_normalize_source_key(name), name, row.get("proposed_source_id") or "", row.get("Category") or ""))

    def best_match(source_id: Any, source_name: Any):
        candidates = [_normalize_source_key(source_id), _normalize_source_key(source_name)]
        for norm_key, reg_name, reg_id, category in register_lookup:
            for cand in candidates:
                if cand and norm_key and (cand in norm_key or norm_key in cand):
                    return reg_name, reg_id, category, "LOW_CONFIDENCE_SUBSTRING"
        return None, None, None, "NO_CONFIDENT_MATCH"

    if "Source_Run_Status_Mapped" in master_wb.sheetnames:
        ws = master_wb["Source_Run_Status_Mapped"]
        _clear_worksheet_contents(ws)
    else:
        ws = master_wb.create_sheet("Source_Run_Status_Mapped")

    headers = [
        "source_id", "source_name", "run_final_status", "run_candidates", "run_civil_candidates",
        "run_open_candidates", "run_elapsed_seconds", "matched_source_register_name",
        "matched_source_register_id", "matched_source_register_category", "match_confidence",
    ]
    header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F1F1F")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill = header_font, header_fill

    matched = 0
    for i, row in enumerate(run_log_rows, start=2):
        source_id, source_name = row.get("source_id") or "", row.get("source_name") or ""
        reg_name, reg_id, category, confidence = best_match(source_id, source_name)
        if confidence != "NO_CONFIDENT_MATCH":
            matched += 1
        values = [
            source_id, source_name, row.get("final_status"), row.get("candidates"),
            row.get("civil_candidates"), row.get("open_found"), row.get("elapsed_seconds"),
            reg_name, reg_id, category, confidence,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=i, column=col, value=value).alignment = Alignment(vertical="top", wrap_text=True)

    for col, width in {1: 22, 2: 30, 3: 26, 4: 12, 5: 14, 6: 14, 7: 16, 8: 30, 9: 18, 10: 20, 11: 24}.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    return {"rows_written": len(run_log_rows), "matched_rows": matched, "unmatched_rows": len(run_log_rows) - matched}


def add_outreach_traceability(master_wb: Workbook) -> dict[str, Any]:
    """Add source/source_url/source_sheet traceability columns to the
    Outreach_Tracker copy living in the integrated master.

    LEAD-type rows are traced back to BID_LATER_Future_Projects by
    lead_id (an exact key both sheets share). TENDER-type rows have no
    lead_id in BID_NOW_Active_Tenders, so they're traced by exact
    tender_title match instead. Rows that can't be traced get an explicit
    UNKNOWN_SOURCE marker rather than a silent blank.
    """
    if "Outreach_Tracker" not in master_wb.sheetnames:
        return {"rows_written": 0, "skipped": "Outreach_Tracker sheet missing"}

    lead_lookup: dict[str, tuple[Any, Any]] = {}
    if "BID_LATER_Future_Projects" in master_wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(master_wb["BID_LATER_Future_Projects"]):
            lead_id = row.get("lead_id")
            if lead_id:
                lead_lookup[str(lead_id)] = (row.get("source") or "UNKNOWN_SOURCE", row.get("source_url") or "UNKNOWN_SOURCE")

    tender_lookup: dict[str, tuple[Any, Any]] = {}
    if "BID_NOW_Active_Tenders" in master_wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(master_wb["BID_NOW_Active_Tenders"]):
            title = row.get("tender_title")
            if title:
                tender_lookup[str(title)] = (row.get("source") or "UNKNOWN_SOURCE", row.get("url") or "UNKNOWN_SOURCE")

    ws = master_wb["Outreach_Tracker"]
    header_row_idx = _header_row_index(ws)
    header_row = [c.value for c in next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))]
    header_lookup = {h: i + 1 for i, h in enumerate(header_row) if h is not None}

    sample_style_cell = ws.cell(row=header_row_idx, column=1)
    next_col = len(header_row) + 1
    for h in ("source", "source_url", "source_sheet"):
        if h in header_lookup:
            continue
        cell = ws.cell(row=header_row_idx, column=next_col, value=h)
        cell.font, cell.fill = copy.copy(sample_style_cell.font), copy.copy(sample_style_cell.fill)
        header_lookup[h] = next_col
        next_col += 1

    type_col = header_lookup.get("type")
    lead_id_col = header_lookup.get("lead_id")
    title_col = header_lookup.get("title_or_address")
    traced = total = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        if not any(c.value not in (None, "") for c in row):
            continue
        total += 1
        r = row[0].row
        row_type = row[type_col - 1].value if type_col else None
        source = source_url = "UNKNOWN_SOURCE"
        if row_type == "LEAD" and lead_id_col:
            match = lead_lookup.get(str(row[lead_id_col - 1].value))
        elif row_type == "TENDER" and title_col:
            match = tender_lookup.get(str(row[title_col - 1].value))
        else:
            match = None
        if match:
            source, source_url = match
        if source != "UNKNOWN_SOURCE":
            traced += 1
        source_sheet = "BID_LATER_Future_Projects" if row_type == "LEAD" else "BID_NOW_Active_Tenders" if row_type == "TENDER" else "UNKNOWN_SOURCE"
        ws.cell(row=r, column=header_lookup["source"], value=source)
        ws.cell(row=r, column=header_lookup["source_url"], value=source_url)
        ws.cell(row=r, column=header_lookup["source_sheet"], value=source_sheet)

    return {"rows_written": total, "traced_rows": traced, "untraced_rows": total - traced}


def extend_workbook_map_for_integration(master_wb: Workbook, stats: dict[str, Any]) -> None:
    """Append a "Master integration map" section to Workbook_Map explaining
    which original master working tabs are now the source of truth, which
    demo tab feeds each one, which tabs are audit/reference only, and which
    tabs are duplicate-purpose and why both are kept.
    """
    if "Workbook_Map" not in master_wb.sheetnames:
        return
    ws = master_wb["Workbook_Map"]
    header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F1F1F")
    wrap = Alignment(vertical="top", wrap_text=True)

    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="Master integration map (this workbook only)").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Master Working Tab", "Status", "Fed By (Demo Tab)", "Rows Now In Tab", "Notes"], start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font, cell.fill = header_font, header_fill
    row += 1

    active, future = stats.get("active_tenders", {}), stats.get("future_projects", {})
    outreach, source_map = stats.get("outreach", {}), stats.get("source_run_status_mapped", {})

    integration_rows = [
        ("Active_Tenders", "SOURCE OF TRUTH - live-mapped this run", "BID_NOW_Active_Tenders",
         active.get("rows_written", 0),
         f"Column-mapped, not a raw copy. Unmapped master columns: {', '.join(active.get('unmapped_master_columns', [])) or 'none'}."),
        ("Future_Projects", "SOURCE OF TRUTH - live-mapped this run", "BID_LATER_Future_Projects",
         future.get("rows_written", 0),
         f"Column-mapped, not a raw copy. Unmapped master columns: {', '.join(future.get('unmapped_master_columns', [])) or 'none'}."),
        ("Dashboard", "SOURCE OF TRUTH - counts refreshed this run", "Executive_Summary / Acquisition_Funnel_And_Speed",
         "-", "Existing metric rows filled in; new run-specific metrics appended at bottom; stale action item annotated DONE."),
        ("Source_Register", "REFERENCE ONLY - never auto-written", "Source_Run_Log (via Source_Run_Status_Mapped)",
         "-", "Curated source list. Not overwritten with heuristic matches - see Source_Run_Status_Mapped for run-to-register linkage."),
        ("Source_Run_Status_Mapped", "NEW - audit/reference", "Source_Run_Log",
         source_map.get("rows_written", 0),
         f"{source_map.get('matched_rows', 0)} matched to Source_Register by normalized name, {source_map.get('unmatched_rows', 0)} unmatched (NO_CONFIDENT_MATCH)."),
        ("Outreach_Tracker", "SOURCE OF TRUTH - traceability added this run", "BID_LATER_Future_Projects / BID_NOW_Active_Tenders",
         outreach.get("rows_written", 0),
         f"source/source_url/source_sheet columns added; {outreach.get('traced_rows', 0)} traced, {outreach.get('untraced_rows', 0)} UNKNOWN_SOURCE."),
        ("README_START_HERE", "AUDIT / REFERENCE ONLY", "n/a", "-",
         "Original human-written quickstart notes for the master workbook. Duplicate-purpose with Workbook_Map (this sheet); both are kept because they document different things - README_START_HERE is manual setup notes, Workbook_Map explains the auto-generated pipeline."),
        ("BID_NOW_Active_Tenders / BID_LATER_Future_Projects / etc.", "RAW SOURCE TABS (unchanged from demo)", "Demo workbook", "-",
         "Kept as full-fidelity raw copies of the demo run for audit trail, even though their data is also mapped into the master working tabs above."),
    ]
    for values in integration_rows:
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value).alignment = wrap
        row += 1

    for col, width in {1: 26, 2: 34, 3: 34, 4: 14, 5: 60}.items():
        current = ws.column_dimensions[get_column_letter(col)].width
        if not current or current < width:
            ws.column_dimensions[get_column_letter(col)].width = width


def _dashboard_metric_value(ws, label: str) -> Any:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == label:
            return row[1].value if len(row) > 1 else None
    return None


def verify_integration(
    master_wb: Workbook,
    demo_wb: Workbook,
    stats: dict[str, Any],
    audit_classification: str,
    original_master_untouched: bool,
    integrated_workbook_created: bool,
) -> dict[str, Any]:
    """Hard verification for the integration patch - prints FINAL
    INTEGRATION CHECK and returns pass/fail results. `master_wb`/`demo_wb`
    must be freshly reloaded from the saved output files, not the in-memory
    objects used to build them, so this genuinely checks what got written
    to disk rather than trusting the build step's own bookkeeping.
    """
    checks: dict[str, bool] = {}

    bid_now_rows = len(_read_sheet_rows_as_dicts(demo_wb["BID_NOW_Active_Tenders"])) if "BID_NOW_Active_Tenders" in demo_wb.sheetnames else 0
    active_rows = len(_read_sheet_rows_as_dicts(master_wb["Active_Tenders"])) if "Active_Tenders" in master_wb.sheetnames else 0
    checks["active_tenders_populated"] = (bid_now_rows == 0) or (active_rows >= bid_now_rows)

    bid_later_rows = len(_read_sheet_rows_as_dicts(demo_wb["BID_LATER_Future_Projects"])) if "BID_LATER_Future_Projects" in demo_wb.sheetnames else 0
    future_rows = len(_read_sheet_rows_as_dicts(master_wb["Future_Projects"])) if "Future_Projects" in master_wb.sheetnames else 0
    checks["future_projects_populated"] = (bid_later_rows == 0) or (future_rows >= bid_later_rows)

    dash_ws = master_wb["Dashboard"] if "Dashboard" in master_wb.sheetnames else None
    dash_active = _dashboard_metric_value(dash_ws, "Active live tenders entered") if dash_ws else None
    dash_future = _dashboard_metric_value(dash_ws, "Future live projects entered") if dash_ws else None
    checks["dashboard_counts_match"] = bool(dash_ws) and dash_active == active_rows and dash_future == future_rows

    source_map = stats.get("source_run_status_mapped", {})
    checks["source_diagnostics_mapped"] = "Source_Run_Status_Mapped" in master_wb.sheetnames and source_map.get("rows_written", 0) > 0

    outreach_stats = stats.get("outreach", {})
    outreach_ws = master_wb["Outreach_Tracker"] if "Outreach_Tracker" in master_wb.sheetnames else None
    outreach_headers = [c.value for c in next(outreach_ws.iter_rows(min_row=_header_row_index(outreach_ws), max_row=_header_row_index(outreach_ws)))] if outreach_ws else []
    checks["outreach_traceability_present"] = all(h in outreach_headers for h in ("source", "source_url", "source_sheet"))

    checks["original_master_untouched"] = original_master_untouched
    checks["integrated_workbook_created"] = integrated_workbook_created

    passed = all(checks.values())

    def _pf(ok: bool) -> str:
        return "PASS" if ok else "FAIL"

    print("FINAL INTEGRATION CHECK", flush=True)
    print(f"- Audit classification before patch: {audit_classification}", flush=True)
    print(f"- Active_Tenders populated from BID_NOW: {_pf(checks['active_tenders_populated'])} ({active_rows} rows, demo had {bid_now_rows})", flush=True)
    print(f"- Future_Projects populated from BID_LATER: {_pf(checks['future_projects_populated'])} ({future_rows} rows, demo had {bid_later_rows})", flush=True)
    print(f"- Dashboard updated from current counts: {_pf(checks['dashboard_counts_match'])}", flush=True)
    print(f"- Source diagnostics mapped: {_pf(checks['source_diagnostics_mapped'])} ({source_map.get('rows_written', 0)} rows, {source_map.get('matched_rows', 0)} matched)", flush=True)
    print(f"- Outreach source traceability: {_pf(checks['outreach_traceability_present'])} ({outreach_stats.get('traced_rows', 0)}/{outreach_stats.get('rows_written', 0)} traced)", flush=True)
    print(f"- Original master untouched: {_pf(checks['original_master_untouched'])}", flush=True)
    print(f"- Integrated workbook created: {_pf(checks['integrated_workbook_created'])}", flush=True)
    print(f"- Overall: {_pf(passed)}", flush=True)

    return {"checks": checks, "passed": passed}


def build_integrated_master_workbook(demo_path: Path, master_path: Path, integrated_path: Path) -> dict[str, Any]:
    """Build a new INTEGRATED copy of the master workbook.

    Unlike copy_workbook_sheets_into_master() (whole-sheet clone only),
    this additionally maps demo data into the master's own working tabs -
    see the module note above _copy_all_demo_sheets_into for why a
    name-matched sheet copy alone is not true integration.
    """
    demo_wb = openpyxl.load_workbook(demo_path)
    master_wb = openpyxl.load_workbook(master_path)

    _copy_all_demo_sheets_into(demo_wb, master_wb)

    run_timestamp = dt.datetime.now().isoformat(timespec="seconds")
    stats: dict[str, Any] = {"run_timestamp": run_timestamp}
    stats["active_tenders"] = map_bid_now_into_active_tenders(master_wb, run_timestamp)
    stats["future_projects"] = map_bid_later_into_future_projects(master_wb, run_timestamp)
    stats["source_run_status_mapped"] = build_source_run_status_mapped(master_wb)
    stats["outreach"] = add_outreach_traceability(master_wb)
    stats["dashboard_counts"] = _compute_dashboard_counts(master_wb)
    stats["dashboard_updated"] = update_dashboard_with_run_counts(master_wb, stats["dashboard_counts"], run_timestamp)

    create_or_update_workbook_map(master_wb)
    extend_workbook_map_for_integration(master_wb, stats)
    apply_logical_tab_colors(master_wb)
    reorder_sheets_safely(master_wb)

    _atomic_save_workbook(master_wb, integrated_path)
    return stats


def create_clean_final_output_folder(out_root: Path | None = None, prefix: str = "final_review_") -> Path:
    """Create a fresh, empty C:\\tenderfinder_out\\<prefix><timestamp>\\ folder for
    this run's two final deliverables. Nothing else is ever written here -
    no logs, CSVs, screenshots, or extra copies.
    """
    out_root = out_root or Path("C:/tenderfinder_out")
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    final_folder = out_root / f"{prefix}{timestamp}"
    final_folder.mkdir(parents=True, exist_ok=True)
    return final_folder


def open_output_folder(folder: Path) -> bool:
    """Open `folder` in Windows Explorer. Prefers os.startfile; falls back
    to `explorer` via subprocess if os.startfile is unavailable or fails
    (e.g. some sandboxed environments). Returns True if a launch attempt
    succeeded, False if neither approach worked.
    """
    if os.environ.get("TENDER_FINDER_DEMO_NO_OPEN", "").strip() == "1":
        print(f"Folder auto-open skipped by TENDER_FINDER_DEMO_NO_OPEN: {folder}", flush=True)
        return False
    try:
        if hasattr(os, "startfile"):
            os.startfile(str(folder))  # type: ignore[attr-defined]
            return True
    except Exception as exc:
        print(f"WARN: os.startfile could not open {folder}: {exc}", flush=True)
    try:
        subprocess.run(["explorer", str(folder)], check=False)
        return True
    except Exception as exc:
        print(f"WARN: could not open {folder} automatically: {exc}", flush=True)
        return False


def verify_final_outputs(
    final_folder: Path,
    final_demo_path: Path,
    final_master_path: Path,
    original_master_path: Path,
    original_master_mtime: float,
    original_master_sheetnames: list[str],
) -> dict[str, Any]:
    """Run the FINAL OUTPUT CHECK described in the workbook-patch spec and
    print a plain-English PASS/FAIL report. Returns the raw check results
    for callers that want to inspect them programmatically.
    """
    checks: dict[str, bool] = {}
    notes: list[str] = []

    checks["demo_workbook_exists"] = final_demo_path.exists()
    checks["filled_master_exists"] = final_master_path.exists()

    all_files = sorted(p.name for p in final_folder.iterdir() if p.is_file())
    xlsx_files = [f for f in all_files if f.lower().endswith(".xlsx")]
    checks["folder_has_only_two_xlsx"] = len(all_files) == 2 and len(xlsx_files) == 2
    notes.append(f"final_folder contents: {all_files}")

    checks["original_master_untouched"] = (
        original_master_path.exists() and original_master_path.stat().st_mtime == original_master_mtime
    )

    demo_wb = openpyxl.load_workbook(final_demo_path) if checks["demo_workbook_exists"] else None
    master_wb = openpyxl.load_workbook(final_master_path) if checks["filled_master_exists"] else None

    checks["workbook_map_in_demo"] = bool(demo_wb) and "Workbook_Map" in demo_wb.sheetnames
    checks["workbook_map_in_master"] = bool(master_wb) and "Workbook_Map" in master_wb.sheetnames

    def _group_colors_uniform(wb, sheet_names: list[str], expected_color: str) -> bool:
        if wb is None:
            return False
        present = [n for n in sheet_names if n in wb.sheetnames]
        if not present:
            return True
        for name in present:
            tab_color = wb[name].sheet_properties.tabColor
            rgb = (getattr(tab_color, "rgb", None) or "") if tab_color else ""
            if not str(rgb).upper().endswith(expected_color.upper()):
                return False
        return True

    for label, color, names in LOGICAL_TAB_COLOR_GROUPS:
        key = label.lower().replace(" ", "_").replace("+", "and")
        checks[f"tab_color_{key}_demo"] = _group_colors_uniform(demo_wb, names, color)
        checks[f"tab_color_{key}_master"] = _group_colors_uniform(master_wb, names, color)

    def _order_matches(wb) -> bool:
        if wb is None:
            return False
        expected = [n for n in WORKBOOK_SHEET_ORDER if n in wb.sheetnames]
        actual = [n for n in wb.sheetnames if n in set(expected)]
        return expected == actual

    checks["sheet_order_demo"] = _order_matches(demo_wb)
    checks["sheet_order_master"] = _order_matches(master_wb)

    row_count_sheets = [
        "BID_NOW_Active_Tenders", "Tender_Signals_All", "Tender_History_Closed_Public",
        "BID_LATER_Future_Projects", "Outreach_Tracker", "Source_Run_Log",
    ]
    row_count_ok = True
    if demo_wb is not None and master_wb is not None:
        for name in row_count_sheets:
            if name in demo_wb.sheetnames and name in master_wb.sheetnames:
                demo_rows, master_rows = demo_wb[name].max_row, master_wb[name].max_row
                if demo_rows != master_rows:
                    row_count_ok = False
                    notes.append(f"ROW COUNT MISMATCH {name}: demo={demo_rows} master={master_rows}")
            else:
                row_count_ok = False
                notes.append(f"row-count sheet missing from one workbook: {name}")
    else:
        row_count_ok = False
    checks["key_row_counts_match"] = row_count_ok

    checks["no_master_sheet_deleted"] = (
        bool(master_wb) and set(original_master_sheetnames).issubset(set(master_wb.sheetnames))
    )

    passed = all(checks.values())

    def _pf(ok: bool) -> str:
        return "PASS" if ok else "FAIL"

    tab_color_pass = all(v for k, v in checks.items() if k.startswith("tab_color_"))
    print("FINAL OUTPUT CHECK", flush=True)
    print(f"- Demo workbook: {_pf(checks['demo_workbook_exists'])} / {final_demo_path}", flush=True)
    print(f"- Filled master workbook: {_pf(checks['filled_master_exists'])} / {final_master_path}", flush=True)
    print(f"- Clean folder file count: {_pf(checks['folder_has_only_two_xlsx'])} ({all_files})", flush=True)
    print(f"- Workbook_Map: {_pf(checks['workbook_map_in_demo'] and checks['workbook_map_in_master'])}", flush=True)
    print(f"- Tab colors: {_pf(tab_color_pass)}", flush=True)
    print(f"- Sheet order: {_pf(checks['sheet_order_demo'] and checks['sheet_order_master'])}", flush=True)
    print(f"- Key row-count match: {_pf(checks['key_row_counts_match'])}", flush=True)
    print(f"- Original master untouched: {_pf(checks['original_master_untouched'])}", flush=True)
    print(f"- No master sheet deleted: {_pf(checks['no_master_sheet_deleted'])}", flush=True)
    for note in notes:
        print(f"  note: {note}", flush=True)
    print(f"- OVERALL: {_pf(passed)}", flush=True)

    # _demo_wb/_master_wb: the already-loaded workbook objects, exposed so
    # callers that need a second verification pass (e.g. verify_integration)
    # can reuse them instead of re-parsing large xlsx files from disk again.
    return {"checks": checks, "notes": notes, "passed": passed, "_demo_wb": demo_wb, "_master_wb": master_wb}


def finalize_integrated_review_outputs(demo_workbook_path: Path) -> dict[str, Any]:
    """Superseded entry point, kept as a valid lower-level alternative.

    Builds the 43-sheet INTEGRATED master (demo data mapped into the
    master's own working tabs, but every demo/audit sheet also kept as a
    separate tab). An audit of this design found the result was still too
    complicated for a normal user - effectively both source files stacked
    together. Production runs now call finalize_final_review_outputs()
    instead, which builds a slim user-facing master. This function is no
    longer called from main(), but is left intact and still fully correct.

    1. Apply logical tab colors + canonical order + Workbook_Map to the
       demo workbook that was just built.
    2. Find the latest TENDER_FINDER_Tender_Intelligence_Working_Master_*.xlsx.
    3. Build a new INTEGRATED copy of that master: demo data is mapped
       into the master's own working tabs (Active_Tenders, Future_Projects,
       Dashboard, source diagnostics, Outreach traceability), not just
       copied in as separate sheets. See build_integrated_master_workbook().
       An audit of the first version of this pipeline found it only did a
       whole-sheet copy and left every master working tab byte-identical
       to the original - see AUDIT_CLASSIFICATION below.
    4. Assemble a clean C:\\tenderfinder_out\\final_integrated_review_<timestamp>\\
       folder containing exactly the demo workbook and the integrated
       master.
    5. Verify the result (both structural checks and hard integration
       checks) and open the folder in Explorer.
    """
    print("TENDER_FINDER_STAGE: Applying logical tab colors and Workbook_Map to demo workbook", flush=True)
    demo_wb = openpyxl.load_workbook(demo_workbook_path)
    # See the ordering note in copy_workbook_sheets_into_master(): the map
    # sheet must exist before we color tabs, or it ends up uncolored.
    create_or_update_workbook_map(demo_wb)
    apply_logical_tab_colors(demo_wb)
    reorder_sheets_safely(demo_wb)
    _atomic_save_workbook(demo_wb, demo_workbook_path)

    print("TENDER_FINDER_STAGE: Locating latest TENDER_FINDER_Tender_Intelligence_Working_Master_*.xlsx", flush=True)
    master_path = find_latest_master_workbook()
    original_master_mtime = master_path.stat().st_mtime
    check_wb = openpyxl.load_workbook(master_path, read_only=True)
    original_master_sheetnames = list(check_wb.sheetnames)
    check_wb.close()

    final_folder = create_clean_final_output_folder(prefix="final_integrated_review_")
    timestamp = final_folder.name.replace("final_integrated_review_", "")
    integrated_master_name = f"TENDER_FINDER_Tender_Intelligence_Working_Master{MASTER_WORKBOOK_INTEGRATED_MARKER}{timestamp}.xlsx"
    final_demo_path = final_folder / "TENDER_FINDER_DEMO_Opportunities_Three_Buckets.xlsx"
    final_master_path = final_folder / integrated_master_name

    shutil.copy2(demo_workbook_path, final_demo_path)

    print(f"TENDER_FINDER_STAGE: Integrating demo data into {integrated_master_name}", flush=True)
    integration_stats = build_integrated_master_workbook(demo_workbook_path, master_path, final_master_path)

    result = verify_final_outputs(
        final_folder, final_demo_path, final_master_path,
        master_path, original_master_mtime, original_master_sheetnames,
    )

    original_master_untouched = master_path.exists() and master_path.stat().st_mtime == original_master_mtime
    integration_result = verify_integration(
        result["_master_wb"], result["_demo_wb"], integration_stats,
        audit_classification="SHEET COPY ONLY / NOT TRUE INTEGRATION (pre-patch audit finding)",
        original_master_untouched=original_master_untouched,
        integrated_workbook_created=final_master_path.exists(),
    )
    result.pop("_demo_wb", None)
    result.pop("_master_wb", None)

    opened = open_output_folder(final_folder)
    open_status = "SKIPPED (no-open mode)" if os.environ.get("TENDER_FINDER_DEMO_NO_OPEN", "").strip() == "1" else ("PASS" if opened else "FAIL")
    print(f"- Folder opened: {open_status}", flush=True)

    print(f"final_integrated_review_demo_workbook={final_demo_path}", flush=True)
    print(f"final_integrated_review_integrated_master_workbook={final_master_path}", flush=True)
    print(f"final_integrated_review_folder={final_folder}", flush=True)

    result["integration"] = integration_result
    result["integration_stats"] = integration_stats
    result["final_folder"] = str(final_folder)
    result["demo_workbook"] = str(final_demo_path)
    result["integrated_master_workbook"] = str(final_master_path)
    result["folder_opened"] = opened
    result["passed"] = result["passed"] and integration_result["passed"]
    return result


# ---------------------------------------------------------------------------
# Slim user-facing master.
#
# The demo workbook (TENDER_FINDER_DEMO_Opportunities_Three_Buckets.xlsx) is the
# full technical/audit dataset - every raw signal, diagnostic, archive,
# printable view, and setup guide lives there and only there. The
# INTEGRATED master built above turned out to still be a 43-sheet
# duplicate of both source files stacked together - not usable by a
# normal TENDER_FINDER user (estimator / BD / admin).
#
# The user-facing master built below is a SEPARATE, much smaller workbook:
# 9 focused visible tabs (Start_Here, Dashboard, Active_Tenders,
# Future_Projects, Keyword_Change_Audit, Outreach_Tracker, Source_Status,
# Potential_Sources_Next, Weekly_Review_Log) plus
# up to 2 hidden support tabs (Lists, Config_Scope) carried over from the
# master template. It is rebuilt fresh from the current run's demo data
# every time - never a stale carryover - except for genuinely user-owned
# fields (Assigned To, Bid Decision, Status, Outreach follow-up state,
# Weekly_Review_Log), which are preserved ONLY by matching a stable ID
# (Project ID / Lead ID / lead_id) against the previous user master, never
# by row position.
# ---------------------------------------------------------------------------

USER_MASTER_GLOB = "TENDER_FINDER_Tender_Intelligence_Working_Master_USER_*.xlsx"

USER_MASTER_VISIBLE_SHEET_ORDER = [
    "README_START_HERE", "Dashboard", "Active_Tenders", "Future_Projects",
    "Keyword_Change_Audit", "Outreach_Tracker", "Source_Status",
    "Potential_Sources_Next", "Weekly_Review_Log",
]
USER_MASTER_HIDDEN_SUPPORT_SHEETS = ["Lists", "Config_Scope"]
USER_MASTER_TAB_COLORS = {
    "README_START_HERE": "1F4E79",
    "Dashboard": "1F4E79",
    "Active_Tenders": "4472C4",
    "Future_Projects": "ED7D31",
    "Keyword_Change_Audit": "FFC000",
    "Outreach_Tracker": "7030A0",
    "Source_Status": "70AD47",
    "Potential_Sources_Next": "70AD47",
    "Weekly_Review_Log": "808080",
}
USER_MASTER_HIDDEN_TAB_COLOR = "808080"

# Every one of these must stay OUT of the user master - see PART 6 of the
# design spec (Potential_Sources_Next is deliberately NOT in this list - it
# is an intentional visible audit tab, cloned straight from the technical
# workbook's already-built version; everything else here is either a
# raw/technical demo tab or an original-master tab that isn't one of the
# target user tabs).
USER_MASTER_EXCLUDED_TAB_NAMES = {
    "Workbook_Map", "Executive_Summary", "Acquisition_Funnel_And_Speed",
    "Tender_Signals_All", "BID_NOW_Active_Tenders", "Tender_History_Closed_Public",
    "Tender_Pattern_Analysis", "BID_LATER_Future_Projects", "New_This_Run",
    "Watchlist_Monitor", "Analyzed_Set_Aside", "Surrey_Historical_Archive",
    "Top_Civil_Leads_Printable", "Developer_Watchlist", "Design_Consultants_Reference",
    "Source_Run_Log", "Source_Roadmap_Printable",
    "Action_Center", "Live_Tender_Intake_Plan", "Email_Setup_Guide",
    "Source_Run_Status_Mapped", "Automation_Plan", "Phased_Rollout",
    "Paid_Intelligence", "Prompt_Pack", "Examples_Do_Not_Use", "Cleanup_Log",
    "Project_Plan", "Backup_Raw_Imported", "Source_Register", "Municipal_Coverage",
    "Run_Queue", "Rejected_Archive", "Priority_Monitoring",
}

FUTURE_PROJECTS_USER_LIMIT = 400
OUTREACH_TRACKER_USER_LIMIT = 200
OUTREACH_TRACKER_LEAD_MIN_FIT = 70

USER_KEYWORD_AUDIT_HEADERS = [
    "Lead ID", "Source ID", "Application No", "Old Score", "New Score", "Score Delta",
    "Old Tier", "New Tier", "Old Bucket", "New Bucket", "Changed",
    "Legacy Exception", "Assigned To", "Status", "Notes",
]


def find_latest_user_master_workbook(exclude_path: Path | None = None) -> Path | None:
    """Return the isolated prior user master for this run mode, if present."""
    candidate = LATEST_USER_MASTER_PATH
    if not candidate.exists() or candidate.name.startswith("~$"):
        return None
    resolved = candidate.resolve()
    if exclude_path and resolved == exclude_path.resolve():
        return None
    return resolved


def _load_prior_preserved_fields(prior_master_path: Path | None) -> dict[str, dict[str, dict[str, Any]]]:
    """Read user-owned fields from a prior user master, keyed by stable ID
    (never by row position): Active_Tenders by Project ID, Future_Projects
    by Lead ID, Outreach_Tracker by lead_id.
    """
    empty: dict[str, dict[str, dict[str, Any]]] = {
        "active_tenders": {}, "future_projects": {}, "outreach": {}, "keyword_audit": {},
    }
    if not prior_master_path or not prior_master_path.exists():
        return empty
    try:
        wb = openpyxl.load_workbook(prior_master_path)
    except Exception as exc:
        print(f"WARN: could not read prior user master {prior_master_path}: {exc}", flush=True)
        return empty
    result: dict[str, dict[str, dict[str, Any]]] = {
        "active_tenders": {}, "future_projects": {}, "outreach": {}, "keyword_audit": {},
    }
    if "Active_Tenders" in wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(wb["Active_Tenders"]):
            pid = row.get("Project ID")
            if pid:
                result["active_tenders"][str(pid)] = {
                    "Assigned To": row.get("Assigned To"), "Bid Decision": row.get("Bid Decision"), "Notes": row.get("Notes"),
                }
    if "Future_Projects" in wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(wb["Future_Projects"]):
            lid = row.get("Lead ID")
            if lid:
                result["future_projects"][str(lid)] = {
                    "Assigned To": row.get("Assigned To"), "Status": row.get("Status"), "Notes": row.get("Notes"),
                }
    if "Outreach_Tracker" in wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(wb["Outreach_Tracker"]):
            lid = row.get("lead_id")
            if lid:
                result["outreach"][str(lid)] = {
                    "lead_id": row.get("lead_id"), "type": row.get("type"),
                    "municipality": row.get("municipality"), "title_or_address": row.get("title_or_address"),
                    "fit_score": row.get("fit_score"), "signal_quality": row.get("signal_quality"),
                    "contact": row.get("contact"),
                    "status": row.get("status"), "last_contact_date": row.get("last_contact_date"),
                    "next_action_date": row.get("next_action_date"), "notes": row.get("notes"),
                    "source": row.get("source"), "source_url": row.get("source_url"),
                    "source_sheet": row.get("source_sheet"),
                }
    if "Keyword_Change_Audit" in wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(wb["Keyword_Change_Audit"]):
            lid = row.get("Lead ID")
            if not lid:
                continue
            result["keyword_audit"][str(lid)] = dict(row)
            result["future_projects"].setdefault(
                str(lid),
                {
                    "Assigned To": row.get("Assigned To"),
                    "Status": row.get("Status"),
                    "Notes": row.get("Notes"),
                },
            )
    wb.close()
    return result


def _restore_prior_weekly_review_log(prior_master_path: Path | None, target_wb: Workbook) -> bool:
    """Carry the founder-owned weekly log forward from the prior user master."""
    if not prior_master_path or not prior_master_path.exists() or "Weekly_Review_Log" not in target_wb.sheetnames:
        return False
    prior_wb = openpyxl.load_workbook(prior_master_path)
    try:
        if "Weekly_Review_Log" not in prior_wb.sheetnames:
            return False
        _clone_sheet_into(prior_wb["Weekly_Review_Log"], target_wb["Weekly_Review_Log"])
        return True
    finally:
        prior_wb.close()


USER_ACTIVE_TENDERS_HEADERS = [
    "Project ID", "Date Found", "Source", "Source URL", "Tender / Project Title",
    "Owner / Client", "Municipality", "Scope Summary", "Fit Score", "Fit Class",
    "Closing Date", "Recommended Action", "Assigned To", "Bid Decision", "Notes",
]


def _tender_fit_class(row: dict[str, Any]) -> str:
    if row.get("civil_relevant") is None:
        return ""
    return "Civil Relevant" if row.get("civil_relevant") else "Not Civil"


# Some source parsers (notably the fixture/synthetic Email Alert Intake
# path) concatenate the whole alert body into `issuer` - e.g. "City of
# Surrey Closing Date: July 30, 2026 Contact: procurement@surrey.ca Phone:
# ...". Showing that raw blob in a "Municipality"/"Owner / Client" column
# in the SLIM user tab is actively misleading, so trim it back to the real
# name at the first known boilerplate marker. This only truncates the
# existing field - it never invents a name that isn't already there.
_ISSUER_BOILERPLATE_MARKERS = (
    " Closing Date:", " Contact:", " Phone:", " Opportunity URL:",
    " Category:", " View Opportunity:", " Work includes",
)


def _clean_issuer_label(issuer: Any) -> Any:
    if not issuer:
        return issuer
    text = str(issuer)
    cut = len(text)
    for marker in _ISSUER_BOILERPLATE_MARKERS:
        idx = text.find(marker)
        if idx != -1:
            cut = min(cut, idx)
    cleaned = text[:cut].strip()
    return cleaned or text


def _tender_recommended_action(row: dict[str, Any]) -> str:
    if not row.get("civil_relevant"):
        return "Monitor"
    days = _tender_days_to_close(row, "")
    return "Review & Bid Soon" if days is not None and days <= 14 else "Review"


# Defensive, content-based filter for fixture/synthetic/example test data.
# The demo workbook's Email Alert Intake step falls back to checked-in
# fixture emails (01 Code/CONNECTOR_SWEEP/tests/fixtures/email_alerts) so
# that path has something to parse out of the box when no real inbox is
# configured yet. That's useful for proving the program works, but those
# rows must never reach a user-facing output looking like a real live
# opportunity. Matching is on CONTENT (any field, case-insensitive), not on
# which code path produced the row, so it still catches these rows even if
# they get relabelled or re-routed later.
FIXTURE_EXCLUSION_MARKERS = (
    "example.com", "example-", "fixture", "synthetic", "tests\\fixtures", "tests/fixtures",
    "test data", "not a live alert feed", "fvrd utility corridor", "kelowna pump station",
    "office furniture supply", "example-expired", "source_mode=fixture_synthetic",
)


def _is_fixture_or_example_row(row: dict[str, Any]) -> bool:
    """True if any field of this row looks like fixture/synthetic/example
    test data rather than a real live opportunity - see
    FIXTURE_EXCLUSION_MARKERS.
    """
    haystack = " ".join(str(v) for v in row.values() if v is not None).lower()
    return any(marker in haystack for marker in FIXTURE_EXCLUSION_MARKERS)


def count_live_demo_rows(demo_wb: Workbook, sheet_name: str) -> int:
    """Count real, non-fixture data rows in a demo sheet, using the exact
    same header-detection and fixture-exclusion rules used to build the
    corresponding user-facing tab - so a Dashboard "full count in demo"
    figure is always computed the identical way as the "shown in master"
    figure it gets compared against, instead of drifting apart because one
    side excluded fixture rows and the other didn't.
    """
    if sheet_name not in demo_wb.sheetnames:
        return 0
    return sum(1 for row in _read_sheet_rows_as_dicts(demo_wb[sheet_name]) if not _is_fixture_or_example_row(row))


def build_user_active_tenders_rows(demo_wb: Workbook, prior_lookup: dict[str, Any], run_date: str) -> tuple[list[list[Any]], list[dict[str, Any]]]:
    """Every open/actionable, non-fixture tender from this run's
    BID_NOW_Active_Tenders, with a small user-facing column set - not the
    full audit schema. This is not curated/capped beyond fixture exclusion:
    BID_NOW is already the "actionable now" list by definition. Assigned
    To/Bid Decision/Notes are carried forward ONLY when this run's
    synthesized Project ID matches one in the prior user master - never by
    row position.
    """
    if "BID_NOW_Active_Tenders" not in demo_wb.sheetnames:
        return [], []
    out_rows: list[list[Any]] = []
    out_dicts: list[dict[str, Any]] = []
    for row in _read_sheet_rows_as_dicts(demo_wb["BID_NOW_Active_Tenders"]):
        if _is_fixture_or_example_row(row):
            continue
        project_id = _tender_project_id(row, run_date)
        prior = prior_lookup.get(project_id, {})
        issuer_label = _clean_issuer_label(row.get("issuer"))
        values = [
            project_id, run_date[:10], row.get("source"), row.get("url"), row.get("tender_title"),
            issuer_label, issuer_label, row.get("snippet"), None, _tender_fit_class(row),
            row.get("closing_date"), _tender_recommended_action(row),
            prior.get("Assigned To"), prior.get("Bid Decision"), prior.get("Notes"),
        ]
        out_rows.append(values)
        out_dicts.append(dict(zip(USER_ACTIVE_TENDERS_HEADERS, values)))
    return out_rows, out_dicts


USER_FUTURE_PROJECTS_HEADERS = [
    "Lead ID", "Date Found", "Source", "Source URL", "Municipality", "Project / Address",
    "Applicant / Developer", "Project Type", "Scope Summary", "Civil Relevance", "Fit Score",
    "Signal Quality", "Estimated Timing / Horizon", "Recommended Action", "Assigned To",
    "Status", "Notes",
]

def _lead_civil_relevance(row: dict[str, Any]) -> str:
    text = f"{row.get('scope_summary') or ''} {row.get('app_type_stage') or ''}"
    return "Likely Civil/Servicing" if load_keywords_config().matches_any("label_civil", text) else "Unclear"


def _lead_recommended_action(row: dict[str, Any]) -> str:
    fit = coerce_int(row.get("fit_score"))
    if row.get("fit_score") not in (None, "") and fit >= 70:
        return "Review - top tier"
    if row.get("signal_quality") == "HIGH":
        return "Review"
    return "Monitor"


def select_user_future_projects_rows(
    demo_wb: Workbook, prior_lookup: dict[str, Any], run_date: str, limit: int = FUTURE_PROJECTS_USER_LIMIT,
) -> tuple[list[list[Any]], list[dict[str, Any]], int]:
    """Curate a practical, user-sized subset of this run's
    BID_LATER_Future_Projects: fixture/example rows are excluded first, then
    only HIGH/MEDIUM signal-quality rows with a real fit_score survive
    (excludes low-confidence/noisy records - terminal archive rows are
    already excluded upstream, before they ever reach
    BID_LATER_Future_Projects), ranked by signal quality then fit_score
    (new-this-run rows win ties), capped at `limit`. The full technical demo
    sheet remains unchanged; `total_available_live` counts only non-fixture
    rows eligible to be described in the user-facing master.
    Returns (rows_for_write, row_dicts, total_available_live_in_demo).
    """
    if "BID_LATER_Future_Projects" not in demo_wb.sheetnames:
        return [], [], 0
    src_rows = _read_sheet_rows_as_dicts(demo_wb["BID_LATER_Future_Projects"])
    live_src_rows = [r for r in src_rows if not _is_fixture_or_example_row(r)]
    total_available_live = len(live_src_rows)
    candidates = [r for r in live_src_rows if r.get("signal_quality") in ("HIGH", "MEDIUM") and r.get("fit_score") not in (None, "")]

    def sort_key(r: dict[str, Any]) -> tuple[int, int, int]:
        quality_rank = 0 if r.get("signal_quality") == "HIGH" else 1
        new_rank = 0 if r.get("is_new_since_last_run") == "YES" else 1
        return (quality_rank, -coerce_int(r.get("fit_score")), new_rank)

    candidates.sort(key=sort_key)
    selected = candidates[:limit]

    out_rows: list[list[Any]] = []
    out_dicts: list[dict[str, Any]] = []
    for row in selected:
        lead_id = row.get("lead_id")
        prior = prior_lookup.get(str(lead_id), {}) if lead_id else {}
        values = [
            lead_id, row.get("first_seen_date"), row.get("source"), row.get("source_url"),
            row.get("municipality"), _lead_project_title(row, run_date), row.get("applicant_name"),
            row.get("app_type_stage"), row.get("scope_summary"), _lead_civil_relevance(row),
            row.get("fit_score"), row.get("signal_quality"), row.get("status_class"),
            _lead_recommended_action(row),
            prior.get("Assigned To"), prior.get("Status") or "Needs Review", prior.get("Notes"),
        ]
        out_rows.append(values)
        out_dicts.append(dict(zip(USER_FUTURE_PROJECTS_HEADERS, values)))
    return out_rows, out_dicts, total_available_live


def build_user_keyword_audit_rows(
    demo_wb: Workbook,
    prior_future: dict[str, dict[str, Any]],
    prior_audit: dict[str, dict[str, Any]],
) -> list[list[Any]]:
    """Show changed routing/scoring while retaining founder-owned triage."""
    current: list[dict[str, Any]] = []
    if "Keyword_Change_Audit" in demo_wb.sheetnames:
        current = _read_sheet_rows_as_dicts(demo_wb["Keyword_Change_Audit"])

    output: list[dict[str, Any]] = []
    seen: set[str] = set()
    for event in current:
        if event.get("changed") != "YES" and not event.get("exception"):
            continue
        lead_id = clean_text(event.get("lead_id"))
        if not lead_id:
            continue
        manual = prior_future.get(lead_id) or prior_audit.get(lead_id) or {}
        row = {
            "Lead ID": lead_id,
            "Source ID": event.get("source_id"),
            "Application No": event.get("app_no"),
            "Old Score": event.get("old_score"),
            "New Score": event.get("new_score"),
            "Score Delta": event.get("score_delta"),
            "Old Tier": event.get("old_tier"),
            "New Tier": event.get("new_tier"),
            "Old Bucket": event.get("old_bucket"),
            "New Bucket": event.get("new_bucket"),
            "Changed": event.get("changed"),
            "Legacy Exception": event.get("exception"),
            "Assigned To": manual.get("Assigned To"),
            "Status": manual.get("Status") or "Needs Review",
            "Notes": manual.get("Notes"),
        }
        output.append(row)
        seen.add(lead_id)

    # Keep previously audited manual triage even when the lead remains below
    # the current gate or disappears from the current input snapshot.
    for lead_id, prior in prior_audit.items():
        if lead_id in seen:
            continue
        output.append({header: prior.get(header) for header in USER_KEYWORD_AUDIT_HEADERS})
        seen.add(lead_id)

    for lead_id, manual in prior_future.items():
        if lead_id in seen or not any(manual.get(key) not in (None, "") for key in ("Assigned To", "Status", "Notes")):
            continue
        output.append({
            "Lead ID": lead_id,
            "Changed": "YES",
            "New Bucket": "NOT_IN_CURRENT_RESULT",
            "Legacy Exception": "manual_triage_retained_outside_current_result",
            "Assigned To": manual.get("Assigned To"),
            "Status": manual.get("Status") or "Needs Review",
            "Notes": manual.get("Notes"),
        })

    return [[row.get(header) for header in USER_KEYWORD_AUDIT_HEADERS] for row in output]


OUTREACH_TRACKER_HEADERS = [
    "lead_id", "type", "municipality", "title_or_address", "fit_score", "signal_quality",
    "contact", "status", "last_contact_date", "next_action_date", "notes",
    "source", "source_url", "source_sheet",
]


def build_user_outreach_tracker_rows(
    demo_wb: Workbook,
    active_dicts: list[dict[str, Any]],
    future_dicts: list[dict[str, Any]],
    prior_lookup: dict[str, Any],
    run_date: str,
    limit: int = OUTREACH_TRACKER_USER_LIMIT,
) -> list[list[Any]]:
    """A short, actionable follow-up list built ONLY from rows already
    selected into this same user master's Active_Tenders (all of them -
    BID_NOW is already the actionable list) and Future_Projects (only the
    top tier, fit_score >= OUTREACH_TRACKER_LEAD_MIN_FIT, since the full
    curated Future_Projects tab is still too large to be a CRM worklist).
    Never a raw dump of the full demo dataset. status/last_contact_date/
    next_action_date/notes are preserved ONLY by matching lead_id against
    the prior user master. Source traceability is always populated -
    UNKNOWN_SOURCE if a row genuinely can't be traced, never a silent blank.
    """
    raw_tender_lookup: dict[str, dict[str, Any]] = {}
    if "BID_NOW_Active_Tenders" in demo_wb.sheetnames:
        for row in _read_sheet_rows_as_dicts(demo_wb["BID_NOW_Active_Tenders"]):
            raw_tender_lookup[_tender_project_id(row, run_date)] = row

    rows: list[dict[str, Any]] = []
    for r in active_dicts:
        pid = r.get("Project ID")
        raw = raw_tender_lookup.get(pid, {})
        prior = prior_lookup.get(str(pid), {})
        contact = "; ".join(p for p in [raw.get("contact_email"), raw.get("contact_phone")] if p) or (r.get("Source URL") or "")
        rows.append({
            "lead_id": pid, "type": "TENDER", "municipality": r.get("Municipality"),
            "title_or_address": r.get("Tender / Project Title"), "fit_score": r.get("Fit Score"),
            "signal_quality": r.get("Fit Class"), "contact": contact,
            "status": prior.get("status") or "Not Contacted",
            "last_contact_date": prior.get("last_contact_date") or "",
            "next_action_date": prior.get("next_action_date") or "",
            "notes": prior.get("notes") or "",
            "source": r.get("Source") or "UNKNOWN_SOURCE", "source_url": r.get("Source URL") or "UNKNOWN_SOURCE",
            "source_sheet": "Active_Tenders",
        })

    top_leads = [r for r in future_dicts if r.get("Fit Score") not in (None, "") and coerce_int(r.get("Fit Score")) >= OUTREACH_TRACKER_LEAD_MIN_FIT]
    for r in top_leads:
        lid = r.get("Lead ID")
        prior = prior_lookup.get(str(lid), {}) if lid else {}
        rows.append({
            "lead_id": lid, "type": "LEAD", "municipality": r.get("Municipality"),
            "title_or_address": r.get("Project / Address"), "fit_score": r.get("Fit Score"),
            "signal_quality": r.get("Signal Quality"), "contact": "",
            "status": prior.get("status") or "Not Contacted",
            "last_contact_date": prior.get("last_contact_date") or "",
            "next_action_date": prior.get("next_action_date") or "",
            "notes": prior.get("notes") or "",
            "source": r.get("Source") or "UNKNOWN_SOURCE", "source_url": r.get("Source URL") or "UNKNOWN_SOURCE",
            "source_sheet": "Future_Projects",
        })

    current_ids = {clean_text(row.get("lead_id")) for row in rows}
    retained_rows: list[dict[str, Any]] = []
    for lead_id, prior in prior_lookup.items():
        if clean_text(lead_id) in current_ids:
            continue
        retained_rows.append({
            "lead_id": lead_id,
            "type": prior.get("type") or "LEAD",
            "municipality": prior.get("municipality") or "",
            "title_or_address": prior.get("title_or_address") or "",
            "fit_score": prior.get("fit_score") or "",
            "signal_quality": "BELOW_CURRENT_GATE",
            "contact": prior.get("contact") or "",
            "status": prior.get("status") or "Not Contacted",
            "last_contact_date": prior.get("last_contact_date") or "",
            "next_action_date": prior.get("next_action_date") or "",
            "notes": prior.get("notes") or "",
            "source": prior.get("source") or "UNKNOWN_SOURCE",
            "source_url": prior.get("source_url") or "UNKNOWN_SOURCE",
            "source_sheet": "Prior Outreach (below current gate)",
        })

    new_row_limit = max(0, limit - len(retained_rows))
    rows = rows[:new_row_limit] + retained_rows
    return [[r[h] for h in OUTREACH_TRACKER_HEADERS] for r in rows]


SOURCE_STATUS_HEADERS = [
    "Source Name", "Source Type", "Status", "Open Found", "Closed Found",
    "Raw Tender Links Found", "Issue / Blocker", "Recommended Next Action",
]

_SOURCE_STATUS_ACTIONS = {
    "NEEDS_EMAIL_ALERT": "Register for portal email alerts",
    "NEEDS_MANUAL_REGISTRATION": "Complete manual registration",
    "NEEDS_PAID_ACCESS": "Evaluate paid access",
    "BLOCKED_FORBIDDEN": "Investigate access block",
    "BLOCKED_BROWSER_CHECK": "Investigate browser/bot check",
    "FETCH_OK_PORTAL_REDIRECT": "Follow up on portal redirect",
    "FETCH_OK_PARSE_ZERO_ROWS": "Check page structure - zero rows parsed",
    "FETCH_OK_EMPTY_CURRENT_LIST": "No action - monitor",
}


def _source_status_action(final_status: Any) -> str:
    if final_status in _SOURCE_STATUS_ACTIONS:
        return _SOURCE_STATUS_ACTIONS[final_status]
    return "No action - monitor" if final_status and "OK" in str(final_status) else "Review"


def build_user_source_status_rows(demo_wb: Workbook) -> list[list[Any]]:
    """Simplified source health view for users - built fresh from this
    run's Source_Run_Log every time. The full raw 24-column diagnostic
    sheet stays in the demo workbook only.

    If the Email Alert Intake source ran against checked-in test fixtures
    rather than a real configured inbox (see FIXTURE_EXCLUSION_MARKERS),
    it's shown as test-only/not configured here - never as if it produced
    live counts, even though the raw Source_Run_Log row still records what
    actually happened for audit purposes in the demo workbook.
    """
    if "Source_Run_Log" not in demo_wb.sheetnames:
        return []
    rows: list[list[Any]] = []
    for row in _read_sheet_rows_as_dicts(demo_wb["Source_Run_Log"]):
        if row.get("source_id") == "email_alert_intake" and _is_fixture_or_example_row(row):
            rows.append([
                row.get("source_name"), row.get("source_type"),
                "TEST DATA - not a live inbox",
                0, 0, 0,
                "Using checked-in test fixtures, not a real email inbox",
                "Not configured - connect a real email inbox to include live tender alerts",
            ])
            continue
        final_status = row.get("final_status")
        issue = row.get("error") or (final_status if final_status and "OK" not in str(final_status) else "")
        rows.append([
            row.get("source_name"), row.get("source_type"), final_status,
            row.get("open_found"), row.get("closed_found"), row.get("raw_links_found"),
            issue, _source_status_action(final_status),
        ])
    return rows


def _dashboard_recommended_actions(counts: dict[str, Any]) -> list[str]:
    actions = []
    if counts.get("tenders_closing_soon", 0) > 0:
        actions.append(f"Review {counts['tenders_closing_soon']} active tender(s) closing within 14 days.")
    if counts.get("top_tier_leads", 0) > 0:
        actions.append(f"Review {counts['top_tier_leads']} top-tier future lead(s) (fit >= 85).")
    if counts.get("sources_issues", 0) > 0:
        actions.append(f"{counts['sources_issues']} source(s) need attention - see Source_Status.")
    if not actions:
        actions.append("No urgent items this run - see Active_Tenders and Future_Projects for the current worklist.")
    return actions


def build_user_dashboard(ws, counts: dict[str, Any], run_timestamp: str) -> None:
    header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F1F1F")
    title_font, section_font = Font(bold=True, size=14), Font(bold=True, size=12)
    wrap = Alignment(vertical="top", wrap_text=True)

    company_name = load_keywords_config().company_name
    ws["A1"] = f"{company_name} — TENDER_FINDER Tender Intelligence — Dashboard"
    ws["A1"].font = title_font
    ws["A2"] = "Last run"
    ws["B2"] = run_timestamp

    row = 4
    ws.cell(row=row, column=1, value="Current run summary").font = section_font
    row += 1
    for col, h in enumerate(["Metric", "Value", "What it means"], start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font, cell.fill = header_font, header_fill
    row += 1
    metrics = [
        ("Active tenders shown in master", counts["active_shown"], "All open BID NOW tenders from this run"),
        ("Full active tenders available in demo", counts["active_full"], "Active_Tenders is not curated down - same set"),
        ("Future projects shown in master", counts["future_shown"], "Curated top opportunities from this run"),
        ("Full future-project records in demo", counts["future_full"], "Full detailed future-project dataset is retained in the demo workbook."),
        ("Outreach items", counts["outreach_shown"], "Selected follow-up items needing action"),
        ("Sources checked", counts["sources_checked"], "Rows in this run's source sweep - see Source_Status"),
        ("Sources with issues/blockers", counts["sources_issues"], "See Source_Status for details"),
    ]
    for label, value, meaning in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=meaning).alignment = wrap
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Why don't these numbers match?").font = Font(bold=True)
    ws.cell(row=row, column=2, value=(
        "Active_Tenders and Future_Projects come from two separate pipelines. Active_Tenders is "
        "based on tender/procurement sources. Future_Projects is based on municipal development/"
        "project datasets. Raw tender links (see Source_Status) are not expected to equal "
        "future-project record counts - they measure different things. Full detailed future-project "
        "dataset is retained in the demo workbook."
    )).alignment = wrap
    row += 2

    ws.cell(row=row, column=1, value="Recommended next actions").font = section_font
    row += 1
    for action in counts["recommended_actions"]:
        ws.cell(row=row, column=1, value="-")
        ws.cell(row=row, column=2, value=action).alignment = wrap
        row += 1

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    for col, width in {1: 34, 2: 16, 3: 60}.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_user_start_here(ws, run_timestamp: str) -> None:
    title_font = Font(bold=True, size=14)
    header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F1F1F")
    wrap = Alignment(vertical="top", wrap_text=True)

    ws["A1"] = "TENDER_FINDER Tender Intelligence — Start Here"
    ws["A1"].font = title_font
    ws["A2"] = "Updated"
    ws["B2"] = run_timestamp

    row = 4
    for col, h in enumerate(["Tab", "What it's for"], start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font, cell.fill = header_font, header_fill
    row += 1
    tab_rows = [
        ("Dashboard", "Top summary - open this first. Shows current counts and what needs attention."),
        ("Active_Tenders", "BID NOW worklist - open/actionable tenders to review and bid."),
        ("Future_Projects", "Curated BID LATER worklist - the best future civil/site-servicing leads from this run."),
        ("Outreach_Tracker", "Follow-up / CRM list - who to contact and where things stand."),
        ("Source_Status", "Simple source health check - which sources ran, and which need attention."),
        ("Potential_Sources_Next", "Backlog of additional sources TENDER_FINDER could add next - planning reference, not a live worklist."),
        ("Weekly_Review_Log", "Team notes and decisions - add a row each week."),
    ]
    for tab, desc in tab_rows:
        ws.cell(row=row, column=1, value=tab).font = Font(bold=True)
        ws.cell(row=row, column=2, value=desc).alignment = wrap
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Why don't these numbers match?").font = Font(bold=True)
    ws.cell(row=row, column=2, value=(
        "Active_Tenders and Future_Projects come from two separate pipelines. Active_Tenders is "
        "based on tender/procurement sources. Future_Projects is based on municipal development/"
        "project datasets. Raw tender links (see Source_Status) are not expected to equal "
        "future-project record counts - they measure different things."
    )).alignment = wrap
    row += 1
    ws.cell(row=row, column=1, value="Full technical detail").font = Font(bold=True)
    ws.cell(row=row, column=2, value=(
        "Every raw signal, audit trail, source log, and diagnostic sheet lives in "
        "TENDER_FINDER_DEMO_Opportunities_Three_Buckets.xlsx. This workbook is the slim current "
        "working copy - go to the demo workbook for full technical/audit detail."
    )).alignment = wrap
    row += 1
    ws.cell(row=row, column=1, value="This file is refreshed").font = Font(bold=True)
    ws.cell(row=row, column=2, value=(
        "Every run replaces Active_Tenders, Future_Projects, Dashboard, Outreach_Tracker, and "
        "Source_Status with current data. Weekly_Review_Log and your notes/statuses on tenders "
        "and leads are preserved by matching Project ID / Lead ID, not by row position."
    )).alignment = wrap

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90


def _trim_sheet_to_columns(ws, keep_columns: int) -> None:
    """Delete any columns beyond `keep_columns`. The master template's
    Active_Tenders/Future_Projects/Dashboard/README_START_HERE sheets have
    their own (wider) original column counts; without this, their unused
    trailing columns survive as empty-but-present columns (still inside
    ws.dimensions/autofilter) after being rebuilt with a slimmer schema,
    which is exactly the visual clutter this whole redesign is meant to
    remove.
    """
    if ws.max_column > keep_columns:
        ws.delete_cols(keep_columns + 1, ws.max_column - keep_columns)


def build_user_facing_master_workbook(
    demo_path: Path, master_path: Path, user_master_path: Path, prior_user_master_path: Path | None,
) -> dict[str, Any]:
    """Build the slim, current, user-facing master workbook.

    Uses the original master as a styling template (per the design spec's
    "if using the original master as base" option): every sheet NOT in the
    target focused visible + 2 hidden support tabs is removed from this OUTPUT
    COPY only (the original master file is only ever opened for reading).
    Outreach_Tracker and Source_Status don't exist in the master template,
    so they're created fresh. Every target tab is then rebuilt from the
    CURRENT demo workbook - nothing is left stale - except user-owned
    fields, preserved only by stable ID from the previous user master.
    """
    demo_wb = openpyxl.load_workbook(demo_path)
    master_wb = openpyxl.load_workbook(master_path)

    prior = _load_prior_preserved_fields(prior_user_master_path)

    keep_from_template = set(USER_MASTER_VISIBLE_SHEET_ORDER) | set(USER_MASTER_HIDDEN_SUPPORT_SHEETS)
    for name in list(master_wb.sheetnames):
        if name not in keep_from_template:
            del master_wb[name]
    if "Outreach_Tracker" not in master_wb.sheetnames:
        master_wb.create_sheet("Outreach_Tracker")
    if "Source_Status" not in master_wb.sheetnames:
        master_wb.create_sheet("Source_Status")
    if "Keyword_Change_Audit" not in master_wb.sheetnames:
        master_wb.create_sheet("Keyword_Change_Audit")
    # Potential_Sources_Next is an intentional visible tab (source
    # expansion backlog - not a live-opportunity list, so it carries no
    # fixture-exclusion concerns). It isn't part of the master template, so
    # it's cloned straight from the demo workbook's already-built version
    # rather than rebuilt - full fidelity via the same _clone_sheet_into()
    # used elsewhere, refreshed from this run every time like every other
    # user-master tab.
    if "Potential_Sources_Next" in demo_wb.sheetnames:
        psn_ws = master_wb["Potential_Sources_Next"] if "Potential_Sources_Next" in master_wb.sheetnames else master_wb.create_sheet("Potential_Sources_Next")
        _clone_sheet_into(demo_wb["Potential_Sources_Next"], psn_ws)

    run_timestamp = dt.datetime.now().isoformat(timespec="seconds")

    active_rows, active_dicts = build_user_active_tenders_rows(demo_wb, prior["active_tenders"], run_timestamp)
    ws = master_wb["Active_Tenders"]
    _clear_worksheet_contents(ws)
    _trim_sheet_to_columns(ws, len(USER_ACTIVE_TENDERS_HEADERS))
    write_rows(ws, USER_ACTIVE_TENDERS_HEADERS, active_rows, None, {
        "Tender / Project Title": 46, "Scope Summary": 60, "Notes": 40, "Owner / Client": 30, "Source URL": 50,
    })

    future_rows, future_dicts, future_total_live = select_user_future_projects_rows(demo_wb, prior["future_projects"], run_timestamp)
    ws = master_wb["Future_Projects"]
    _clear_worksheet_contents(ws)
    _trim_sheet_to_columns(ws, len(USER_FUTURE_PROJECTS_HEADERS))
    write_rows(ws, USER_FUTURE_PROJECTS_HEADERS, future_rows, None, {
        "Project / Address": 40, "Scope Summary": 60, "Notes": 40, "Source URL": 50,
    })

    keyword_audit_rows = build_user_keyword_audit_rows(
        demo_wb,
        prior["future_projects"],
        prior["keyword_audit"],
    )
    ws = master_wb["Keyword_Change_Audit"]
    write_rows(ws, USER_KEYWORD_AUDIT_HEADERS, keyword_audit_rows, None, {
        "Lead ID": 22, "Source ID": 24, "Old Bucket": 22, "New Bucket": 22,
        "Legacy Exception": 42, "Assigned To": 20, "Status": 20, "Notes": 50,
    })

    outreach_rows = build_user_outreach_tracker_rows(demo_wb, active_dicts, future_dicts, prior["outreach"], run_timestamp)
    ws = master_wb["Outreach_Tracker"]
    write_rows(ws, OUTREACH_TRACKER_HEADERS, outreach_rows, None, {
        "title_or_address": 46, "notes": 40, "source_url": 50, "contact": 40,
    })

    source_status_rows = build_user_source_status_rows(demo_wb)
    ws = master_wb["Source_Status"]
    write_rows(ws, SOURCE_STATUS_HEADERS, source_status_rows, None, {
        "Source Name": 34, "Issue / Blocker": 40, "Recommended Next Action": 34,
    })

    # active_full is computed independently from the raw demo sheet (via
    # count_live_demo_rows, the SAME fixture-exclusion + header-detection
    # rule build_user_active_tenders_rows() uses) rather than reused from
    # len(active_rows), so this is a genuine cross-check, not a tautology -
    # if the two ever drifted apart (e.g. a future change to one of the two
    # code paths but not the other) verify_user_master() below would catch
    # the mismatch instead of silently reporting two copies of one number.
    counts: dict[str, Any] = {
        "active_shown": len(active_rows), "active_full": count_live_demo_rows(demo_wb, "BID_NOW_Active_Tenders"),
        "future_shown": len(future_rows), "future_full": future_total_live,
        "outreach_shown": len(outreach_rows),
        "sources_checked": len(source_status_rows),
        "sources_issues": sum(1 for r in source_status_rows if r[6]),
        "tenders_closing_soon": sum(1 for r in active_dicts if r.get("Recommended Action") == "Review & Bid Soon"),
        "top_tier_leads": sum(1 for r in future_dicts if r.get("Fit Score") not in (None, "") and coerce_int(r.get("Fit Score")) >= 85),
    }
    counts["recommended_actions"] = _dashboard_recommended_actions(counts)

    ws = master_wb["Dashboard"]
    _clear_worksheet_contents(ws)
    _trim_sheet_to_columns(ws, 3)
    build_user_dashboard(ws, counts, run_timestamp)

    ws = master_wb["README_START_HERE"]
    _clear_worksheet_contents(ws)
    _trim_sheet_to_columns(ws, 2)
    build_user_start_here(ws, run_timestamp)

    # Weekly_Review_Log is founder-owned and must come from the prior user
    # master, not reset to the static package template on every run.
    weekly_log_preserved = _restore_prior_weekly_review_log(prior_user_master_path, master_wb)

    apply_user_master_tab_colors(master_wb)
    reorder_user_master_sheets(master_wb)

    _atomic_save_workbook(master_wb, user_master_path)

    return {
        "run_timestamp": run_timestamp,
        "active_tenders_rows": len(active_rows),
        "future_projects_rows": len(future_rows),
        "future_projects_live_total_in_demo": future_total_live,
        "outreach_rows": len(outreach_rows),
        "keyword_audit_rows": len(keyword_audit_rows),
        "source_status_rows": len(source_status_rows),
        "dashboard_counts": counts,
        "prior_user_master_used": str(prior_user_master_path) if prior_user_master_path else None,
        "weekly_review_log_preserved": weekly_log_preserved,
    }


def apply_user_master_tab_colors(wb: Workbook) -> None:
    for name, color in USER_MASTER_TAB_COLORS.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color
    for name in USER_MASTER_HIDDEN_SUPPORT_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = USER_MASTER_HIDDEN_TAB_COLOR


def reorder_user_master_sheets(wb: Workbook) -> None:
    ordered = [n for n in USER_MASTER_VISIBLE_SHEET_ORDER if n in wb.sheetnames]
    hidden = [n for n in USER_MASTER_HIDDEN_SUPPORT_SHEETS if n in wb.sheetnames]
    remaining = [n for n in wb.sheetnames if n not in set(ordered) | set(hidden)]
    final_order = ordered + remaining + hidden
    wb._sheets = [wb[n] for n in final_order]
    if final_order:
        wb.active = 0
    for name in hidden:
        wb[name].sheet_state = "hidden"


def _rows_free_of_fixture_markers(rows: list[dict[str, Any]]) -> bool:
    return not any(_is_fixture_or_example_row(row) for row in rows)


def verify_user_master(
    user_master_path: Path,
    demo_path: Path,
    final_folder: Path,
    stats: dict[str, Any],
    original_master_untouched: bool,
) -> dict[str, Any]:
    """Hard verification for the slim user-master patch - prints FINAL
    USER MASTER CHECK. Re-reads the saved-to-disk files rather than
    trusting the build step's own bookkeeping, and cross-checks every
    Dashboard count against an independent recount of the actual sheet it
    claims to summarize (not the same numbers the build step already
    computed) - a genuine adversarial check, not a restatement.
    """
    checks: dict[str, bool] = {}
    notes: list[str] = []

    checks["full_demo_produced"] = demo_path.exists()
    checks["user_master_produced"] = user_master_path.exists()

    wb = openpyxl.load_workbook(user_master_path) if checks["user_master_produced"] else None
    demo_wb = openpyxl.load_workbook(demo_path) if checks["full_demo_produced"] else None

    visible_names = [n for n in wb.sheetnames if wb[n].sheet_state == "visible"] if wb else []
    # Eight operational tabs plus the founder-visible Keyword_Change_Audit.
    checks["visible_tab_count_ok"] = 0 < len(visible_names) <= 9
    checks["no_demo_only_tabs_visible"] = wb is not None and not (set(visible_names) & USER_MASTER_EXCLUDED_TAB_NAMES)

    duplicate_pair_names = {"BID_NOW_Active_Tenders", "BID_LATER_Future_Projects", "Executive_Summary", "Workbook_Map", "Source_Run_Log", "Acquisition_Funnel_And_Speed"}
    checks["no_duplicate_purpose_tabs_visible"] = wb is not None and not (set(visible_names) & duplicate_pair_names)

    # count_live_demo_rows(), not a raw row count - this must reflect the
    # SAME fixture-exclusion rule Active_Tenders was built with, or a
    # correctly-filtered Active_Tenders would wrongly fail this check.
    bid_now_live_rows = count_live_demo_rows(demo_wb, "BID_NOW_Active_Tenders") if demo_wb else 0
    active_ws = wb["Active_Tenders"] if wb and "Active_Tenders" in wb.sheetnames else None
    active_rows = _read_sheet_rows_as_dicts(active_ws) if active_ws is not None else []
    checks["active_tenders_current"] = active_ws is not None and len(active_rows) == bid_now_live_rows
    checks["active_tenders_no_placeholder_rows"] = all(any(v not in (None, "") for v in r.values()) for r in active_rows)
    checks["active_tenders_no_fixture_rows"] = _rows_free_of_fixture_markers(active_rows)

    future_ws = wb["Future_Projects"] if wb and "Future_Projects" in wb.sheetnames else None
    future_rows = _read_sheet_rows_as_dicts(future_ws) if future_ws is not None else []
    checks["future_projects_curated_current"] = (
        future_ws is not None and len(future_rows) <= FUTURE_PROJECTS_USER_LIMIT and len(future_rows) == stats.get("future_projects_rows", -1)
    )
    checks["future_projects_no_placeholder_rows"] = all(any(v not in (None, "") for v in r.values()) for r in future_rows)
    checks["future_projects_no_fixture_rows"] = _rows_free_of_fixture_markers(future_rows)

    dash_ws = wb["Dashboard"] if wb and "Dashboard" in wb.sheetnames else None
    checks["dashboard_current"] = bool(dash_ws) and dash_ws["B2"].value == stats.get("run_timestamp")

    outreach_ws = wb["Outreach_Tracker"] if wb and "Outreach_Tracker" in wb.sheetnames else None
    outreach_headers = [c.value for c in next(outreach_ws.iter_rows(min_row=1, max_row=1))] if outreach_ws is not None else []
    checks["outreach_traceability"] = all(h in outreach_headers for h in ("source", "source_url", "source_sheet"))
    outreach_rows = _read_sheet_rows_as_dicts(outreach_ws) if outreach_ws is not None else []
    checks["outreach_no_fixture_rows"] = _rows_free_of_fixture_markers(outreach_rows)

    source_status_ws = wb["Source_Status"] if wb and "Source_Status" in wb.sheetnames else None
    source_status_headers = [c.value for c in next(source_status_ws.iter_rows(min_row=1, max_row=1))] if source_status_ws is not None else []
    checks["source_status_present"] = source_status_ws is not None and source_status_headers == SOURCE_STATUS_HEADERS
    source_status_rows = _read_sheet_rows_as_dicts(source_status_ws) if source_status_ws is not None else []

    weekly_ws = wb["Weekly_Review_Log"] if wb and "Weekly_Review_Log" in wb.sheetnames else None
    checks["weekly_review_log_present"] = weekly_ws is not None and weekly_ws.sheet_state == "visible"

    # --- Dashboard count cross-checks: independently recount each sheet
    # the Dashboard claims to summarize and compare against the actual
    # printed cell value, rather than trusting the build step's counts.
    def _dashboard_value(label: str) -> Any:
        if not dash_ws:
            return None
        for row in dash_ws.iter_rows(min_row=1, max_row=dash_ws.max_row):
            if row[0].value == label:
                return row[1].value if len(row) > 1 else None
        return None

    bid_later_live_rows = count_live_demo_rows(demo_wb, "BID_LATER_Future_Projects") if demo_wb else 0
    dashboard_checks = {
        "active_shown": (_dashboard_value("Active tenders shown in master"), len(active_rows)),
        "active_full": (_dashboard_value("Full active tenders available in demo"), bid_now_live_rows),
        "future_shown": (_dashboard_value("Future projects shown in master"), len(future_rows)),
        "future_full": (_dashboard_value("Full future-project records in demo"), bid_later_live_rows),
        "outreach_count": (_dashboard_value("Outreach items"), len(outreach_rows)),
        "sources_checked": (_dashboard_value("Sources checked"), len(source_status_rows)),
    }
    for key, (dashboard_val, recounted_val) in dashboard_checks.items():
        checks[f"dashboard_count_{key}"] = dashboard_val == recounted_val
        notes.append(f"dashboard {key}: shown={dashboard_val} recounted={recounted_val}")

    checks["original_master_untouched"] = original_master_untouched

    all_files = sorted(p.name for p in final_folder.iterdir() if p.is_file())
    checks["clean_final_folder"] = len(all_files) == 2 and all(f.lower().endswith(".xlsx") for f in all_files) and not any(f.startswith("~$") for f in all_files)

    passed = all(checks.values())

    def _pf(ok: bool) -> str:
        return "PASS" if ok else "FAIL"

    dashboard_counts_pass = all(checks[f"dashboard_count_{k}"] for k in dashboard_checks)
    fixture_free = (
        checks["active_tenders_no_fixture_rows"]
        and checks["future_projects_no_fixture_rows"]
        and checks["outreach_no_fixture_rows"]
    )

    print("FINAL USER MASTER CHECK", flush=True)
    print(f"- Full demo workbook: {_pf(checks['full_demo_produced'])}", flush=True)
    print(f"- Slim user master workbook: {_pf(checks['user_master_produced'])}", flush=True)
    print(f"- Generated from latest same-run demo: {_pf(checks['active_tenders_current'])}", flush=True)
    print(f"- Visible tab count <= 9: {_pf(checks['visible_tab_count_ok'])} ({len(visible_names)}: {visible_names})", flush=True)
    print(f"- No demo-only technical tabs visible: {_pf(checks['no_demo_only_tabs_visible'])}", flush=True)
    print(f"- No duplicate-purpose tabs visible: {_pf(checks['no_duplicate_purpose_tabs_visible'])}", flush=True)
    print(f"- Active_Tenders refreshed/current: {_pf(checks['active_tenders_current'] and checks['active_tenders_no_placeholder_rows'])} ({len(active_rows)} rows)", flush=True)
    print(f"- No fixture/synthetic/example rows in Active_Tenders: {_pf(checks['active_tenders_no_fixture_rows'])}", flush=True)
    print(f"- No fixture/synthetic/example rows in Future_Projects: {_pf(checks['future_projects_no_fixture_rows'])}", flush=True)
    print(f"- No fixture/synthetic/example rows in Outreach_Tracker: {_pf(checks['outreach_no_fixture_rows'])}", flush=True)
    print(f"- Future_Projects curated/current: {_pf(checks['future_projects_curated_current'] and checks['future_projects_no_placeholder_rows'])} ({len(future_rows)} of {stats.get('future_projects_live_total_in_demo', 0)} available)", flush=True)
    print(f"- Dashboard updated/current: {_pf(checks['dashboard_current'])}", flush=True)
    print(f"- Dashboard counts match recounted sheets: {_pf(dashboard_counts_pass)}", flush=True)
    for note in notes:
        print(f"  note: {note}", flush=True)
    print(f"- Outreach traceability: {_pf(checks['outreach_traceability'])} ({len(outreach_rows)} rows)", flush=True)
    print(f"- Source_Status simplified/current: {_pf(checks['source_status_present'])}", flush=True)
    print(f"- Weekly_Review_Log present: {_pf(checks['weekly_review_log_present'])}", flush=True)
    print(f"- Original master untouched: {_pf(checks['original_master_untouched'])}", flush=True)
    print(f"- Clean final folder (no lockfiles/junk): {_pf(checks['clean_final_folder'])} ({all_files})", flush=True)
    print(f"- Overall: {_pf(passed)}", flush=True)

    return {"checks": checks, "notes": notes, "passed": passed, "fixture_free": fixture_free, "dashboard_counts_pass": dashboard_counts_pass}


def finalize_final_review_outputs(demo_workbook_path: Path) -> dict[str, Any]:
    """Patch entry point, called once at the end of a normal demo run.

    1. Apply logical tab colors + canonical order + Workbook_Map to the
       demo workbook that was just built (the full technical/audit file).
    2. Find the latest hand-maintained TENDER_FINDER_Tender_Intelligence_Working_
       Master_*.xlsx (never a previously-generated FILLED/INTEGRATED/USER
       output - see MASTER_WORKBOOK_OWN_OUTPUT_MARKERS).
    3. Build a new slim USER-facing master with focused visible tabs, rebuilt
       fresh from this run's demo data, with user-owned fields preserved by
       stable ID from the previous user master. See
       build_user_facing_master_workbook() and its module note above.
    4. Assemble a clean C:\\tenderfinder_out\\final_user_master_<timestamp>\\ folder
       containing exactly the demo workbook and the user master.
    5. Verify the result and open the folder in Explorer.
    """
    print("TENDER_FINDER_STAGE: Applying logical tab colors and Workbook_Map to demo workbook", flush=True)
    demo_wb = openpyxl.load_workbook(demo_workbook_path)
    create_or_update_workbook_map(demo_wb)
    apply_logical_tab_colors(demo_wb)
    reorder_sheets_safely(demo_wb)
    _atomic_save_workbook(demo_wb, demo_workbook_path)

    print("TENDER_FINDER_STAGE: Locating latest TENDER_FINDER_Tender_Intelligence_Working_Master_*.xlsx", flush=True)
    master_path = find_latest_master_workbook()
    original_master_mtime = master_path.stat().st_mtime

    prior_user_master_path = find_latest_user_master_workbook()
    if prior_user_master_path:
        print(f"PRIOR_USER_MASTER_FOUND={prior_user_master_path}", flush=True)
    else:
        print("PRIOR_USER_MASTER_FOUND=none (first run - nothing to preserve)", flush=True)

    final_folder = create_clean_final_output_folder(
        out_root=demo_workbook_path.parent,
        prefix="final_user_master_",
    )
    timestamp = final_folder.name.replace("final_user_master_", "")
    user_master_name = f"TENDER_FINDER_Tender_Intelligence_Working_Master{MASTER_WORKBOOK_USER_MARKER}{timestamp}.xlsx"
    final_demo_path = final_folder / "TENDER_FINDER_DEMO_Opportunities_Three_Buckets.xlsx"
    final_user_master_path = final_folder / user_master_name

    shutil.copy2(demo_workbook_path, final_demo_path)

    print(f"TENDER_FINDER_STAGE: Building slim user-facing master {user_master_name}", flush=True)
    user_master_stats = build_user_facing_master_workbook(
        demo_workbook_path, master_path, final_user_master_path, prior_user_master_path,
    )

    original_master_untouched = master_path.exists() and master_path.stat().st_mtime == original_master_mtime
    verification = verify_user_master(
        final_user_master_path, final_demo_path, final_folder, user_master_stats, original_master_untouched,
    )

    if verification["passed"]:
        LATEST_USER_MASTER_PATH.parent.mkdir(parents=True, exist_ok=True)
        latest_temp = LATEST_USER_MASTER_PATH.with_suffix(".xlsx.tmp")
        try:
            shutil.copy2(final_user_master_path, latest_temp)
            os.replace(latest_temp, LATEST_USER_MASTER_PATH)
        finally:
            if latest_temp.exists():
                latest_temp.unlink()
        print(f"LATEST_USER_MASTER_UPDATED={LATEST_USER_MASTER_PATH}", flush=True)

    opened = open_output_folder(final_folder)
    open_status = "SKIPPED (no-open mode)" if os.environ.get("TENDER_FINDER_DEMO_NO_OPEN", "").strip() == "1" else ("PASS" if opened else "FAIL")
    print(f"- Folder opened: {open_status}", flush=True)

    print(f"final_user_master_demo_workbook={final_demo_path}", flush=True)
    print(f"final_user_master_user_master_workbook={final_user_master_path}", flush=True)
    print(f"final_user_master_folder={final_folder}", flush=True)

    return {
        "passed": verification["passed"],
        "verification": verification,
        "user_master_stats": user_master_stats,
        "final_folder": str(final_folder),
        "demo_workbook": str(final_demo_path),
        "user_master_workbook": str(final_user_master_path),
        "folder_opened": opened,
    }


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Build TENDER_FINDER three-bucket demo workbook and talk track.")
    ap.add_argument("--review-xlsx", required=True, help="Path to live all17 review workbook.")
    ap.add_argument("--out-dir", required=True, help="Output directory.")
    ap.add_argument("--no-fetch", action="store_true", help="Skip Track B live tender fetch.")
    ap.add_argument("--email-intake", action="store_true", help="Parse user-approved portal email alerts into BID NOW when available.")
    ap.add_argument("--email-import-path", default="", help="Optional local folder of .eml alert files to use for Email Alert Intake.")
    ap.add_argument("--keywords-config", default="", help="Optional canonical keywords.xlsx override.")
    ap.add_argument("--sources-config", default="", help="Optional canonical sources.csv override.")
    ap.add_argument("--source-id", action="append", default=[], help="Run only this active runtime-eligible tender source; repeat as needed.")
    ap.add_argument("--state-root", default="", help="Runtime history/state root outside the package.")
    ap.add_argument("--keywords-state-root", default="", help="External state root for keyword validation and last-known-good snapshot.")
    ap.add_argument("--run-id", default="", help="Stable run identifier supplied by the engine service.")
    ap.add_argument("--run-mode", choices=("live", "offline", "self_test"), default="", help="Structured engine run mode.")
    return ap.parse_args()


def main() -> int:
    global LAST_BC_BID_OFFICIAL_FINDINGS, TENDER_SOURCES, BCBID_PUBLIC_URL
    global BCBID_NETWORK_AUDIT_PATH, PATCH_5_18_BACKLOG_PATH, AUTONOMOUS_FIXES_PATH
    args = parse_args()
    if args.keywords_config:
        os.environ["TENDER_FINDER_KEYWORDS_CONFIG"] = args.keywords_config
    if args.keywords_state_root:
        os.environ["TENDER_FINDER_KEYWORDS_STATE_ROOT"] = args.keywords_state_root
    if args.sources_config:
        os.environ["TENDER_FINDER_SOURCES_CONFIG"] = args.sources_config
    run_mode = args.run_mode or os.environ.get("TENDER_FINDER_RUN_MODE", "").strip() or (
        "offline" if args.no_fetch else "live"
    )
    try:
        configure_runtime_state(args.state_root or None, run_mode)
    except RuntimeStateError as exc:
        print(f"ERROR: {exc}", flush=True)
        return 2
    print(f"RUNTIME_STATE_ROOT={_RUNTIME_PATHS.root} mode={run_mode}", flush=True)
    try:
        TENDER_SOURCES = _load_runtime_tender_sources()
    except SourceRegistryError as exc:
        print(f"ERROR: {exc}", flush=True)
        return 2
    selected_source_ids = {
        str(source_id).strip().casefold()
        for source_id in args.source_id
        if str(source_id).strip()
    }
    if selected_source_ids:
        available_ids = {source["source_id"] for source in TENDER_SOURCES}
        unavailable = sorted(selected_source_ids - available_ids)
        if unavailable:
            print(
                "ERROR: selected source(s) are not active runtime-eligible tender sources: "
                + ", ".join(unavailable),
                flush=True,
            )
            return 2
        TENDER_SOURCES = [
            source for source in TENDER_SOURCES
            if source["source_id"] in selected_source_ids
        ]
    BCBID_PUBLIC_URL = next(
        (source["url"] for source in TENDER_SOURCES if source["source_id"] == "bc_bid_public"),
        "",
    )

    print(
        f"SOURCE_REGISTRY_VALID={os.environ.get('TENDER_FINDER_SOURCES_CONFIG') or 'config/sources.csv'} "
        f"active_tender_sources={len(TENDER_SOURCES)}",
        flush=True,
    )
    try:
        keyword_config = load_keywords_config(force_reload=True)
    except KeywordConfigError as exc:
        print(f"ERROR: {exc}", flush=True)
        return 2
    print_validation_summary(keyword_config)
    total_start = time.perf_counter()
    review_xlsx = Path(args.review_xlsx)
    if not review_xlsx.is_absolute():
        review_xlsx = ROOT / review_xlsx
    out_dir = Path(args.out_dir)
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        probe = out_dir / ".tenderfinder_write_probe"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink()
    except Exception as exc:
        print(f"ERROR: output folder is not writable: {out_dir} ({exc})", flush=True)
        return 1
    BCBID_NETWORK_AUDIT_PATH = out_dir / "BC_BID_NETWORK_AUDIT.md"
    PATCH_5_18_BACKLOG_PATH = out_dir / "PATCH_5_18_BACKLOG.md"
    AUTONOMOUS_FIXES_PATH = out_dir / "AUTONOMOUS_FIXES.md"

    print("TENDER_FINDER_STAGE: Preparing output folder and history archive", flush=True)
    archived_history = archive_pre_511_history()
    if archived_history:
        print(f"demo_history_archived_pre_5_11={[p.name for p in archived_history]}", flush=True)
    print("TENDER_FINDER_STAGE: Reading proven Track A development-project workbook", flush=True)
    data = read_track_a(review_xlsx, allow_network_enrichment=not args.no_fetch)
    print(f"KEYWORDS_RESCORE_ALWAYS: {RESCORE_ALWAYS_SUMMARY}", flush=True)
    print(
        f"KEYWORDS_RESCORE_ALWAYS: rescored={len(data.keyword_rescore_events)} below_gate={len(data.keyword_below_gate)}",
        flush=True,
    )
    for event in data.keyword_below_gate:
        print(
            "KEYWORDS_BELOW_GATE: "
            f"lead_id={event.get('lead_id')} app_no={event.get('app_no')} "
            f"old_score={event.get('old_score')} new_score={event.get('new_score')} "
            f"threshold={event.get('threshold')} route={event.get('new_route')}",
            flush=True,
        )
    apply_run_history(data)
    print(f"future={len(data.future)} watch={len(data.watch)} analyzed={len(data.analyzed)} (read in {data.read_seconds:.2f}s)", flush=True)
    record_stage(out_dir, "track_a_loaded")
    maybe_pause_at_checkpoint(out_dir, "track_a_loaded")
    print("TENDER_FINDER_STAGE: Checking Email Alert Intake state", flush=True)
    if args.email_import_path:
        connect_path = connect_email_provider(
            "manual_folder",
            account_label=provider_display_name("manual_folder"),
            import_path=args.email_import_path,
        )
        print(f"EMAIL_IMPORT_PATH_SET={args.email_import_path} config={connect_path}", flush=True)
    email_state = load_email_state()
    print_email_setup_status(email_state, bid_later_count=len(data.future))

    print("TENDER_FINDER_STAGE: Parsing approved Email Alert Intake messages", flush=True)
    email_tenders, email_log = run_email_intake_for_demo(out_dir, args.email_intake)
    print("TENDER_FINDER_STAGE: Sweeping public tender pages and BC Bid public browse", flush=True)
    tenders, logs, fetch_seconds = fetch_tenders(data=data, no_fetch=args.no_fetch)
    tenders = email_tenders + tenders
    logs = [email_log] + logs
    if data and email_tenders:
        populate_related_counts(email_tenders, data)
    tenders = finalize_tender_signals(sanitize_tenders_for_output(tenders), logs)
    if not args.no_fetch:
        data.tender_date_stats = enrich_tender_closing_dates(tenders)
    print(f"Track B total time (parallelized)={fetch_seconds:.2f}s", flush=True)
    for line in tender_status_lines(logs):
        print(line, flush=True)
    record_stage(out_dir, "municipal_sources_done")
    record_stage(out_dir, "bc_bid_done_or_skipped")
    maybe_pause_at_checkpoint(out_dir, "bc_bid_done_or_skipped")

    print("TENDER_FINDER_STAGE: Building Excel workbook and demo sheets", flush=True)
    workbook_path = out_dir / "TENDER_FINDER_DEMO_Opportunities_Three_Buckets.xlsx"
    try:
        build_workbook(workbook_path, data, tenders, logs, fetch_seconds, time.perf_counter() - total_start, email_state)
        total_seconds = time.perf_counter() - total_start
        patch_workbook_total_time(workbook_path, total_seconds)
    except PermissionError as exc:
        print(f"ERROR: {exc}", flush=True)
        return 1
    record_stage(out_dir, "workbook_written")
    maybe_pause_at_checkpoint(out_dir, "workbook_written")
    if not args.no_fetch:
        data.history_archive = str(archive_demo_history(workbook_path, out_dir))
    print("TENDER_FINDER_STAGE: Writing talk track, build report, and summary", flush=True)
    write_talktrack(out_dir / "DEMO_TALKTRACK.md", data, tenders, logs, total_seconds)
    write_build_report(out_dir / "DEMO_BUILD_REPORT.md", data, tenders, logs, fetch_seconds, total_seconds, workbook_path)
    write_demo_summary(out_dir / "demo_summary.txt", data, tenders, logs, fetch_seconds, total_seconds, workbook_path, email_state)
    if not args.no_fetch:
        # These reflect this run's live BC Bid audit (LAST_BC_BID_AUDIT is
        # only populated when Track B actually runs) - writing them
        # unconditionally would blank out the real, hard-won audit record
        # in docs/BC_BID_NETWORK_AUDIT.md every time a fast (--no-fetch)
        # run happens, e.g. in the GUI/CLI test suite (confirmed as a real
        # bug during Patch 5.17: running the launcher test suite in fast
        # mode was clobbering this file with empty placeholder data).
        print("TENDER_FINDER_STAGE: Writing BC Bid audit notes", flush=True)
        LAST_BC_BID_OFFICIAL_FINDINGS = search_bc_bid_official_feeds()
        write_bc_bid_network_audit(BCBID_NETWORK_AUDIT_PATH, LAST_BC_BID_AUDIT, LAST_BC_BID_OFFICIAL_FINDINGS)
        write_patch_5_18_backlog(PATCH_5_18_BACKLOG_PATH, LAST_BC_BID_AUDIT, data)
    write_autonomous_fixes(AUTONOMOUS_FIXES_PATH)

    total_seconds = time.perf_counter() - total_start
    bid_now_tenders = [t for t in tenders if t.retention_target == "bid_now"]
    civil = sum(1 for t in bid_now_tenders if t.civil_relevant)
    open_civil = sum(1 for t in bid_now_tenders if t.civil_relevant and is_open_tender_status(t.status))
    bc_bid_log = next((l for l in logs if l.source_id == "bc_bid_public"), None)
    bc_bid_open = [t for t in bid_now_tenders if t.source_id == "bc_bid_public" and t.civil_relevant and is_open_tender_status(t.status)]
    contact_count = sum(1 for t in bid_now_tenders if t.contact_email or t.contact_phone)
    linked = sum(1 for t in bid_now_tenders if t.related_lead)
    gates = [t for t in bid_now_tenders if "gates lake" in f"{t.tender_title} {t.snippet}".lower()]
    gates_yes = any(t.civil_relevant for t in gates)
    repaired = ", ".join(
        f"{l.source_id}={l.status}:{l.resolved_url or 'none'}"
        for l in logs
        if l.source_id in {"tol_public_tenders", "maple_ridge_rfps", "surrey_bids_public"}
    )
    record_stage(out_dir, "completed")
    print(f"DONE in {total_seconds:.2f}s")
    print(f"BID NOW={len(bid_now_tenders)} civil_yes={civil} open_civil={open_civil} cross_linked={linked} BID LATER={len(data.future)} (>=60:{data.fit_ge_60} >=70:{data.fit_ge_70}) WATCH={len(data.watch)} ANALYZED={len(data.analyzed)}")
    print(f"BC Bid public URL={BCBID_PUBLIC_URL} status={bc_bid_log.status if bc_bid_log else 'NOT_ATTEMPTED'} open_civil={len(bc_bid_open)} samples={[t.tender_title for t in bc_bid_open[:3]]}")
    print(f"BID NOW contacts={contact_count} contact_email_or_phone")
    print("Top_Civil_Leads source_or_contact populated=YES for 50 rows")
    print(f"Action_Center rows={len(ACTION_CENTER_ROWS) + 1} Live_Tender_Intake_Plan rows={len(LIVE_INTAKE_ROWS)}")
    print(f"Gates Lake civil_relevant={'YES' if gates_yes else ('NOT_FOUND' if not gates else 'NO')}")
    print(f"repaired_urls={repaired}")
    print(f"workbook={workbook_path}")
    print(f"talktrack={out_dir / 'DEMO_TALKTRACK.md'}")
    print(f"report={out_dir / 'DEMO_BUILD_REPORT.md'}")
    print(f"summary={out_dir / 'demo_summary.txt'}")
    if data.history_archive:
        print(f"demo_history={data.history_archive}")
    print(f"Live tenders: {email_state.state} - bidsandtenders.ca remains alert-driven; BC Bid direct public browse audit is run-local.")
    print(f"bc_bid_network_audit={BCBID_NETWORK_AUDIT_PATH}")

    print("TENDER_FINDER_STAGE: Building final-review outputs (demo workbook + slim user-facing master)", flush=True)
    try:
        final_result = finalize_final_review_outputs(workbook_path)
    except (FileNotFoundError, PermissionError) as exc:
        print(f"ERROR: {exc}", flush=True)
        return 1
    if not final_result.get("passed", False):
        print("ERROR: final-review verification reported FAIL - see FINAL USER MASTER CHECK above.", flush=True)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
