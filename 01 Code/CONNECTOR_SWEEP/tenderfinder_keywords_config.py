#!/usr/bin/env python3
"""Shared, strict loader for the founder-editable ``config/keywords.xlsx``.

All active tender scoring and keyword gates load through this module.  Invalid
or missing configuration is a hard stop: callers must never fall back to old
hardcoded business rules because that would make spreadsheet edits misleading.
"""
from __future__ import annotations

import os
import re
import json
import shutil
import uuid
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

try:
    import regex as bounded_regex
except ImportError:  # Reported as a configuration/dependency error below.
    bounded_regex = None

from tenderfinder_package_paths import (
    detect_package_root,
    keywords_config_path,
    keywords_template_path,
)
from tenderfinder_runtime import (
    atomic_write_json,
    runtime_paths,
    sha256_file,
    timestamp_now,
    user_settings_root,
)


CONFIG_ENV_VAR = "TENDER_FINDER_KEYWORDS_CONFIG"
KEYWORDS_STATE_ROOT_ENV_VAR = "TENDER_FINDER_KEYWORDS_STATE_ROOT"
VALIDATION_REPORT_FILENAME = "keywords_validation_last.txt"
LKG_WORKBOOK_FILENAME = "keywords_last_known_good.xlsx"
LKG_METADATA_FILENAME = "keywords_last_known_good.json"
PROFILE_SHEET = "Company_Profile"
KEYWORDS_SHEET = "Keywords"
INSTRUCTIONS_SHEET = "Instructions"
REQUIRED_SHEETS = (PROFILE_SHEET, KEYWORDS_SHEET, INSTRUCTIONS_SHEET)
PROFILE_COLUMNS = (
    "company_name",
    "company_description",
    "regions",
    "work_types",
    "known_clients",
)
KEYWORD_COLUMNS = (
    "keyword",
    "match_type",
    "weight",
    "category",
    "explanation",
    "active",
    "param",
)
MATCH_TYPES = frozenset({"contains", "exact", "regex"})
CATEGORIES = frozenset(
    {
        "positive",
        "negative",
        "geography",
        "client",
        "gate_include",
        "gate_exclude",
        "gate_weak",
        "gate_collision",
        "label_civil",
        "van_signal_primary",
        "van_signal_secondary",
        "tender_match",
    }
)
SCORING_CATEGORIES = frozenset({"positive", "negative", "geography", "client"})
PROFILE_RULE_DEFAULTS = {
    "regions": ("geography", 8, "Company Profile region"),
    "work_types": ("positive", 9, "Company Profile work type"),
    "known_clients": ("client", 6, "Company Profile known client"),
}

MAX_REGEX_PATTERN_LENGTH = 256
MAX_REGEX_TEXT_LENGTH = 100_000
REGEX_TIMEOUT_SECONDS = 0.02
_REGEX_STRESS_TEXTS = (
    "a" * 4096 + "!",
    "0" * 4096 + "X",
    "word " * 2048 + "!",
)
_UNSAFE_REGEX_CHECKS = (
    (re.compile(r"\\[1-9]|\(\?P="), "backreferences are not allowed"),
    (re.compile(r"\(\?(?:[=!]|<[=!])"), "lookaround assertions are not allowed"),
    (re.compile(r"\(\?\("), "conditional groups are not allowed"),
    (
        re.compile(r"\((?:\\.|[^()])*(?:\*|\+)(?:\\.|[^()])*\)(?:\*|\+|\{\d*,\})"),
        "nested unbounded quantifiers are not allowed",
    ),
)


class KeywordConfigError(RuntimeError):
    """Raised when the live workbook cannot safely control scoring."""

    def __init__(self, path: Path, errors: Iterable[str]):
        self.path = Path(path)
        self.errors = tuple(str(error) for error in errors)
        template = keywords_template_path(detect_package_root(self.path.parent))
        details = "\n".join(f"  - {error}" for error in self.errors)
        super().__init__(
            "Keyword configuration is invalid; TENDER_FINDER stopped before scoring.\n"
            f"Workbook: {self.path}\n"
            "Fix the listed issue(s). If the live file is missing or unusable, copy\n"
            f"{template}\n"
            f"to {self.path} as keywords.xlsx, then fill in Company_Profile.\n"
            f"{details}"
        )


class KeywordRegexRuntimeError(RuntimeError):
    """Raised when a validated workbook regex exceeds its runtime budget."""


@dataclass(frozen=True)
class KeywordRule:
    keyword: str
    match_type: str
    weight: int
    category: str
    explanation: str = ""
    active: bool = True
    param: str = ""
    source: str = KEYWORDS_SHEET
    row_number: int = 0

    @property
    def identity(self) -> tuple[str, str]:
        return (self.keyword.strip().casefold(), self.category)

    def matches(self, text: Any) -> bool:
        if not self.active:
            return False
        value = "" if text is None else str(text)
        if self.match_type == "exact":
            return value.strip().lower() == self.keyword.strip().lower()
        if self.match_type == "regex":
            # Historical tender regexes searched already-lowercased text.
            if bounded_regex is None:
                raise KeywordRegexRuntimeError(
                    "The required 'regex' package is unavailable. Run setup_venv.bat, "
                    "then validate keywords.xlsx again."
                )
            bounded_value = value.lower()[:MAX_REGEX_TEXT_LENGTH]
            try:
                return bounded_regex.search(
                    self.keyword,
                    bounded_value,
                    timeout=REGEX_TIMEOUT_SECONDS,
                ) is not None
            except TimeoutError as exc:
                location = f" row {self.row_number}" if self.row_number else ""
                raise KeywordRegexRuntimeError(
                    f"Keyword regex{location} exceeded the {REGEX_TIMEOUT_SECONDS:.2f}s "
                    "safety limit. Simplify or disable the rule before rerunning."
                ) from exc
        return self.keyword.lower() in value.lower()


@dataclass(frozen=True)
class KeywordConfig:
    path: Path
    company_name: str
    company_description: str
    regions: tuple[str, ...]
    work_types: tuple[str, ...]
    known_clients: tuple[str, ...]
    rules: tuple[KeywordRule, ...]
    requested_path: Path | None = None
    source_kind: str = "canonical"
    loaded_at: str = ""
    last_validation_at: str = ""
    validation_errors: tuple[str, ...] = ()
    last_known_good_path: Path | None = None
    last_known_good_status: str = "not_managed"
    last_known_good_saved_at: str = ""

    @property
    def active_keyword_count(self) -> int:
        return sum(1 for rule in self.rules if rule.active)

    @property
    def inactive_keyword_count(self) -> int:
        return sum(1 for rule in self.rules if not rule.active)

    @property
    def category_counts(self) -> dict[str, int]:
        counts: dict[str, int] = {}
        for rule in self.rules:
            if rule.active:
                counts[rule.category] = counts.get(rule.category, 0) + 1
        return dict(sorted(counts.items()))

    def rules_for(self, category: str) -> tuple[KeywordRule, ...]:
        return tuple(rule for rule in self.rules if rule.active and rule.category == category)

    def matching_rules(self, category: str, text: Any) -> tuple[KeywordRule, ...]:
        return tuple(rule for rule in self.rules_for(category) if rule.matches(text))

    def matches_any(self, category: str, text: Any) -> bool:
        return any(rule.matches(text) for rule in self.rules_for(category))

    def match_count(self, category: str, text: Any) -> int:
        return sum(1 for rule in self.rules_for(category) if rule.matches(text))

    def collision_terms(self) -> tuple[str, ...]:
        terms: list[str] = []
        for rule in self.rules_for("gate_collision"):
            terms.extend(part.strip().casefold() for part in rule.param.split("|") if part.strip())
        return tuple(dict.fromkeys(terms))

    def exclusion_rules(self, mode: str) -> tuple[KeywordRule, ...]:
        return tuple(rule for rule in self.rules_for("gate_exclude") if rule.param == mode)


@dataclass(frozen=True)
class KeywordStatePaths:
    settings: Path
    validation_report: Path
    last_known_good_workbook: Path
    last_known_good_metadata: Path


_CACHE: dict[tuple[Path, Path, bool], KeywordConfig] = {}


def clear_keywords_cache() -> None:
    _CACHE.clear()


def resolve_keywords_path(path: str | Path | None = None, root: str | Path | None = None) -> Path:
    if path is not None:
        return Path(path).expanduser().resolve()
    override = os.environ.get(CONFIG_ENV_VAR, "").strip()
    if override:
        return Path(override).expanduser().resolve()
    package_root = detect_package_root(Path(root).resolve() if root is not None else None)
    return keywords_config_path(package_root).resolve()


def keyword_state_paths(
    *,
    root: str | Path | None = None,
    state_root: str | Path | None = None,
    create: bool = False,
) -> KeywordStatePaths:
    """Return external keyword validation/LKG paths for one state context."""
    package_root = detect_package_root(Path(root).resolve() if root is not None else None)
    selected = str(
        state_root
        or os.environ.get(KEYWORDS_STATE_ROOT_ENV_VAR, "").strip()
    ).strip()
    if selected:
        settings = runtime_paths(
            selected,
            mode="keywords",
            package_root=package_root,
            create=create,
        ).settings
    else:
        settings = user_settings_root(package_root=package_root, create=create)
    return KeywordStatePaths(
        settings=settings,
        validation_report=settings / VALIDATION_REPORT_FILENAME,
        last_known_good_workbook=settings / LKG_WORKBOOK_FILENAME,
        last_known_good_metadata=settings / LKG_METADATA_FILENAME,
    )


def keyword_validation_report_path(
    *,
    root: str | Path | None = None,
    state_root: str | Path | None = None,
) -> Path:
    return keyword_state_paths(root=root, state_root=state_root).validation_report


def _normalized_header(value: Any) -> str:
    return str(value or "").strip().casefold()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _split_profile_values(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;\r\n]+", text) if part.strip()]


def _sheet_records(ws: Any, required_columns: tuple[str, ...], errors: list[str]) -> list[tuple[int, dict[str, Any]]]:
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_normalized_header(value) for value in first_row]
    missing = [column for column in required_columns if column not in headers]
    if missing:
        errors.append(f"Sheet {ws.title}: missing column(s): {', '.join(missing)}")
        return []
    positions = {column: headers.index(column) for column in required_columns}
    records: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {
            column: values[index] if index < len(values) else None
            for column, index in positions.items()
        }
        if any(_text(value) for value in record.values()):
            records.append((row_number, record))
    return records


def _parse_weight(value: Any, category: str, row_number: int, errors: list[str]) -> int:
    if value is None or _text(value) == "":
        if category in SCORING_CATEGORIES:
            errors.append(f"Sheet {KEYWORDS_SHEET} row {row_number}: weight is required for category '{category}'")
        return 0
    if isinstance(value, bool):
        errors.append(f"Sheet {KEYWORDS_SHEET} row {row_number}: weight must be a whole number")
        return 0
    try:
        number = float(value)
    except (TypeError, ValueError):
        errors.append(f"Sheet {KEYWORDS_SHEET} row {row_number}: weight '{value}' is not a number")
        return 0
    if not number.is_integer():
        errors.append(f"Sheet {KEYWORDS_SHEET} row {row_number}: weight must be a whole number")
        return 0
    return int(number)


def _validate_regex(keyword: str, row_number: int, errors: list[str]) -> None:
    prefix = f"Sheet {KEYWORDS_SHEET} row {row_number}:"
    if len(keyword) > MAX_REGEX_PATTERN_LENGTH:
        errors.append(
            f"{prefix} regex is {len(keyword)} characters; maximum is "
            f"{MAX_REGEX_PATTERN_LENGTH}"
        )
        return
    for check, message in _UNSAFE_REGEX_CHECKS:
        if check.search(keyword):
            errors.append(f"{prefix} unsafe regex: {message}")
            return
    if bounded_regex is None:
        errors.append(
            f"{prefix} regex rules require the 'regex' package. Run setup_venv.bat."
        )
        return
    try:
        compiled = bounded_regex.compile(keyword)
    except Exception as exc:
        errors.append(f"{prefix} invalid regex '{keyword}': {exc}")
        return
    try:
        for sample in _REGEX_STRESS_TEXTS:
            compiled.search(sample, timeout=REGEX_TIMEOUT_SECONDS)
    except TimeoutError:
        errors.append(
            f"{prefix} unsafe regex exceeded the {REGEX_TIMEOUT_SECONDS:.2f}s "
            "validation safety limit"
        )


def _write_report(
    path: Path,
    status: str,
    lines: Iterable[str],
    *,
    package_root: Path,
    state_root: str | Path | None,
) -> None:
    report = keyword_state_paths(
        root=package_root,
        state_root=state_root,
        create=True,
    ).validation_report
    try:
        report.parent.mkdir(parents=True, exist_ok=True)
        content = [
            "TENDER_FINDER keyword configuration validation",
            f"Checked: {datetime.now().astimezone().isoformat(timespec='seconds')}",
            f"Workbook: {path}",
            f"Status: {status}",
            "",
            *lines,
            "",
        ]
        report.write_text("\n".join(content), encoding="utf-8")
    except OSError:
        # The configuration error remains authoritative even if a read-only
        # folder prevents the convenience report from being written.
        pass


def _raise_invalid(
    path: Path,
    errors: list[str],
    *,
    package_root: Path,
    state_root: str | Path | None,
) -> None:
    _write_report(
        path,
        "INVALID",
        (f"- {error}" for error in errors),
        package_root=package_root,
        state_root=state_root,
    )
    raise KeywordConfigError(path, errors)


def _load_uncached(
    path: Path,
    *,
    package_root: Path,
    state_root: str | Path | None,
    allow_empty_profile: bool = False,
) -> KeywordConfig:
    if not path.exists():
        _raise_invalid(
            path,
            ["Workbook is missing."],
            package_root=package_root,
            state_root=state_root,
        )
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        _raise_invalid(
            path,
            [f"Workbook could not be opened: {exc}"],
            package_root=package_root,
            state_root=state_root,
        )

    errors: list[str] = []
    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in wb.sheetnames]
    if missing_sheets:
        errors.append(f"Missing required sheet(s): {', '.join(missing_sheets)}")

    profile_records: list[tuple[int, dict[str, Any]]] = []
    keyword_records: list[tuple[int, dict[str, Any]]] = []
    if PROFILE_SHEET in wb.sheetnames:
        profile_records = _sheet_records(wb[PROFILE_SHEET], PROFILE_COLUMNS, errors)
    if KEYWORDS_SHEET in wb.sheetnames:
        keyword_records = _sheet_records(wb[KEYWORDS_SHEET], KEYWORD_COLUMNS, errors)

    company_names: list[str] = []
    descriptions: list[str] = []
    regions: list[str] = []
    work_types: list[str] = []
    known_clients: list[str] = []
    for _, record in profile_records:
        if _text(record["company_name"]):
            company_names.append(_text(record["company_name"]))
        if _text(record["company_description"]):
            descriptions.append(_text(record["company_description"]))
        regions.extend(_split_profile_values(record["regions"]))
        work_types.extend(_split_profile_values(record["work_types"]))
        known_clients.extend(_split_profile_values(record["known_clients"]))
    company_name = company_names[0] if company_names else ""
    company_description = descriptions[0] if descriptions else ""
    if len({name.casefold() for name in company_names}) > 1:
        errors.append(f"Sheet {PROFILE_SHEET}: company_name contains conflicting values")
    if len({value.casefold() for value in descriptions}) > 1:
        errors.append(f"Sheet {PROFILE_SHEET}: company_description contains conflicting values")
    if not company_name and not allow_empty_profile:
        errors.append(f"Sheet {PROFILE_SHEET}: company_name is required")

    profile_rules: dict[tuple[str, str], KeywordRule] = {}
    profile_lists = {
        "regions": regions,
        "work_types": work_types,
        "known_clients": known_clients,
    }
    for field_name, values in profile_lists.items():
        category, weight, explanation = PROFILE_RULE_DEFAULTS[field_name]
        for value in values:
            rule = KeywordRule(
                keyword=value,
                match_type="contains",
                weight=weight,
                category=category,
                explanation=explanation,
                source=PROFILE_SHEET,
            )
            if rule.identity in profile_rules:
                errors.append(f"Sheet {PROFILE_SHEET}: duplicate '{value}' in {field_name}")
            else:
                profile_rules[rule.identity] = rule

    keyword_rules: dict[tuple[str, str], KeywordRule] = {}
    for row_number, record in keyword_records:
        keyword = _text(record["keyword"])
        match_type = _text(record["match_type"]).casefold()
        category = _text(record["category"]).casefold()
        active_text = _text(record["active"]).upper()
        param = _text(record["param"]).casefold()
        if not keyword:
            errors.append(f"Sheet {KEYWORDS_SHEET} row {row_number}: keyword is required")
        if match_type not in MATCH_TYPES:
            errors.append(
                f"Sheet {KEYWORDS_SHEET} row {row_number}: match_type must be one of "
                f"{', '.join(sorted(MATCH_TYPES))}"
            )
        if category not in CATEGORIES:
            errors.append(
                f"Sheet {KEYWORDS_SHEET} row {row_number}: category must be one of "
                f"{', '.join(sorted(CATEGORIES))}"
            )
        if active_text not in {"Y", "N"}:
            errors.append(f"Sheet {KEYWORDS_SHEET} row {row_number}: active must be Y or N")
        weight = _parse_weight(record["weight"], category, row_number, errors)
        if match_type == "regex" and keyword:
            _validate_regex(keyword, row_number, errors)
        if category == "gate_collision" and not param:
            errors.append(f"Sheet {KEYWORDS_SHEET} row {row_number}: gate_collision requires collision terms in param")
        if category == "gate_exclude" and param not in {"absolute_title", "negative_title"}:
            errors.append(
                f"Sheet {KEYWORDS_SHEET} row {row_number}: gate_exclude param must be "
                "absolute_title or negative_title"
            )
        rule = KeywordRule(
            keyword=keyword,
            match_type=match_type,
            weight=weight,
            category=category,
            explanation=_text(record["explanation"]),
            active=active_text == "Y",
            param=param,
            source=KEYWORDS_SHEET,
            row_number=row_number,
        )
        if keyword and category:
            if rule.identity in keyword_rules:
                errors.append(
                    f"Sheet {KEYWORDS_SHEET} row {row_number}: duplicate (keyword, category) "
                    f"pair '{keyword}', '{category}'"
                )
            else:
                keyword_rules[rule.identity] = rule

    if errors:
        wb.close()
        _raise_invalid(
            path,
            errors,
            package_root=package_root,
            state_root=state_root,
        )

    # Explicit Keywords rows win over Profile-generated defaults with the
    # same normalized (keyword, category) key.
    merged_rules = dict(profile_rules)
    merged_rules.update(keyword_rules)
    rules = tuple(
        sorted(
            merged_rules.values(),
            key=lambda rule: (rule.category, rule.keyword.casefold(), rule.param, rule.source),
        )
    )
    config = KeywordConfig(
        path=path,
        company_name=company_name,
        company_description=company_description,
        regions=tuple(dict.fromkeys(regions)),
        work_types=tuple(dict.fromkeys(work_types)),
        known_clients=tuple(dict.fromkeys(known_clients)),
        rules=rules,
    )
    wb.close()
    _write_report(
        path,
        "VALID",
        (
            f"Company: {config.company_name or '(blank template profile)'}",
            f"Active keyword rules: {config.active_keyword_count}",
            f"Regions: {len(config.regions)}",
            f"Work types: {len(config.work_types)}",
            f"Known clients: {len(config.known_clients)}",
        ),
        package_root=package_root,
        state_root=state_root,
    )
    return config


def _atomic_copy(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.name}.{uuid.uuid4().hex}.tmp")
    try:
        with source.open("rb") as source_handle, temporary.open("wb") as target_handle:
            shutil.copyfileobj(source_handle, target_handle)
            target_handle.flush()
            os.fsync(target_handle.fileno())
        os.replace(temporary, destination)
    finally:
        if temporary.exists():
            temporary.unlink()


def _save_last_known_good(
    canonical_path: Path,
    config: KeywordConfig,
    state_paths: KeywordStatePaths,
) -> str:
    saved_at = timestamp_now()
    _atomic_copy(canonical_path, state_paths.last_known_good_workbook)
    snapshot_hash = sha256_file(state_paths.last_known_good_workbook)
    atomic_write_json(
        state_paths.last_known_good_metadata,
        {
            "schema_version": 1,
            "canonical_path": str(canonical_path),
            "canonical_sha256": sha256_file(canonical_path),
            "snapshot_path": str(state_paths.last_known_good_workbook),
            "snapshot_sha256": snapshot_hash,
            "saved_at": saved_at,
            "company_name": config.company_name,
            "active_keyword_count": config.active_keyword_count,
            "inactive_keyword_count": config.inactive_keyword_count,
            "category_counts": config.category_counts,
        },
    )
    return saved_at


def _validated_last_known_good(
    requested_path: Path,
    state_paths: KeywordStatePaths,
) -> tuple[Path | None, dict[str, Any], str]:
    workbook = state_paths.last_known_good_workbook
    metadata_path = state_paths.last_known_good_metadata
    if not workbook.exists() or not metadata_path.exists():
        return None, {}, "no validated snapshot has been saved"
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        if not isinstance(metadata, dict):
            raise ValueError("metadata is not an object")
        recorded_canonical = Path(str(metadata.get("canonical_path", ""))).resolve()
        if recorded_canonical != requested_path.resolve():
            raise ValueError("snapshot belongs to a different canonical workbook")
        recorded_snapshot = Path(str(metadata.get("snapshot_path", ""))).resolve()
        if recorded_snapshot != workbook.resolve():
            raise ValueError("snapshot path does not match keyword state")
        recorded_hash = str(metadata.get("snapshot_sha256", "")).strip().lower()
        actual_hash = sha256_file(workbook).lower()
        if not recorded_hash or actual_hash != recorded_hash:
            raise ValueError("snapshot SHA-256 does not match metadata")
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
        return None, {}, f"snapshot metadata/checksum is invalid: {exc}"
    return workbook, metadata, "ready"


def inspect_last_known_good(
    path: str | Path | None = None,
    *,
    root: str | Path | None = None,
    state_root: str | Path | None = None,
) -> dict[str, Any]:
    package_root = detect_package_root(Path(root).resolve() if root is not None else None)
    requested = resolve_keywords_path(path, package_root)
    paths = keyword_state_paths(root=package_root, state_root=state_root)
    workbook, metadata, detail = _validated_last_known_good(requested, paths)
    return {
        "status": "ready" if workbook else "unavailable",
        "detail": detail,
        "path": str(workbook or paths.last_known_good_workbook),
        "saved_at": str(metadata.get("saved_at", "")),
        "sha256": str(metadata.get("snapshot_sha256", "")),
    }


def load_keywords_config(
    path: str | Path | None = None,
    *,
    root: str | Path | None = None,
    force_reload: bool = False,
    allow_empty_profile: bool = False,
    allow_last_known_good: bool | None = None,
    state_root: str | Path | None = None,
) -> KeywordConfig:
    package_root = detect_package_root(Path(root).resolve() if root is not None else None)
    resolved = resolve_keywords_path(path, package_root)
    managed_canonical = resolved == keywords_config_path(package_root).resolve()
    lkg_enabled = (
        managed_canonical
        if allow_last_known_good is None
        else bool(allow_last_known_good)
    ) and not allow_empty_profile
    state_paths = keyword_state_paths(
        root=package_root,
        state_root=state_root,
        create=lkg_enabled,
    )
    cache_key = (resolved, state_paths.settings.resolve(), lkg_enabled)
    if force_reload:
        _CACHE.pop(cache_key, None)
    cached = _CACHE.get(cache_key)
    if not allow_empty_profile and not force_reload and cached:
        return cached

    checked_at = timestamp_now()
    try:
        config = _load_uncached(
            resolved,
            package_root=package_root,
            state_root=state_root,
            allow_empty_profile=allow_empty_profile,
        )
    except KeywordConfigError as canonical_error:
        if not lkg_enabled:
            raise
        snapshot, metadata, lkg_detail = _validated_last_known_good(
            resolved,
            state_paths,
        )
        if snapshot is None:
            combined = [
                *canonical_error.errors,
                f"Last-known-good fallback unavailable: {lkg_detail}.",
            ]
            _write_report(
                resolved,
                "INVALID_NO_FALLBACK",
                (f"- {error}" for error in combined),
                package_root=package_root,
                state_root=state_root,
            )
            raise KeywordConfigError(resolved, combined) from canonical_error
        try:
            fallback = _load_uncached(
                snapshot,
                package_root=package_root,
                state_root=state_root,
                allow_empty_profile=False,
            )
        except KeywordConfigError as fallback_error:
            combined = [
                *canonical_error.errors,
                *(
                    f"Last-known-good snapshot: {error}"
                    for error in fallback_error.errors
                ),
            ]
            _write_report(
                resolved,
                "INVALID_FALLBACK_REJECTED",
                (f"- {error}" for error in combined),
                package_root=package_root,
                state_root=state_root,
            )
            raise KeywordConfigError(resolved, combined) from fallback_error
        config = replace(
            fallback,
            requested_path=resolved,
            source_kind="last_known_good",
            loaded_at=checked_at,
            last_validation_at=checked_at,
            validation_errors=canonical_error.errors,
            last_known_good_path=snapshot,
            last_known_good_status="in_use",
            last_known_good_saved_at=str(metadata.get("saved_at", "")),
        )
        _write_report(
            resolved,
            "FALLBACK_LAST_KNOWN_GOOD",
            (
                f"Effective workbook: {snapshot}",
                f"Snapshot saved: {config.last_known_good_saved_at or '(unknown)'}",
                *(f"Canonical error: {error}" for error in canonical_error.errors),
            ),
            package_root=package_root,
            state_root=state_root,
        )
        _CACHE[cache_key] = config
        return config

    config = replace(
        config,
        requested_path=resolved,
        source_kind="canonical",
        loaded_at=checked_at,
        last_validation_at=checked_at,
    )
    if lkg_enabled:
        try:
            saved_at = _save_last_known_good(resolved, config, state_paths)
            config = replace(
                config,
                last_known_good_path=state_paths.last_known_good_workbook,
                last_known_good_status="ready",
                last_known_good_saved_at=saved_at,
            )
        except OSError as exc:
            config = replace(
                config,
                last_known_good_path=state_paths.last_known_good_workbook,
                last_known_good_status=f"snapshot_write_failed: {exc}",
            )
        _write_report(
            resolved,
            "VALID_CANONICAL",
            (
                f"Effective workbook: {resolved}",
                f"Active keyword rules: {config.active_keyword_count}",
                f"Inactive keyword rules: {config.inactive_keyword_count}",
                f"Last-known-good status: {config.last_known_good_status}",
                f"Last-known-good saved: {config.last_known_good_saved_at or '(not saved)'}",
            ),
            package_root=package_root,
            state_root=state_root,
        )
    if not allow_empty_profile:
        _CACHE[cache_key] = config
    return config


def validation_summary(config: KeywordConfig) -> str:
    return (
        f"Company: {config.company_name} | Active keywords: {config.active_keyword_count} | "
        f"Rules source: {config.source_kind}"
    )


def print_validation_summary(config: KeywordConfig) -> None:
    print(f"KEYWORDS_CONFIG_VALID={config.path}", flush=True)
    print(f"KEYWORDS_CONFIG_REQUESTED={config.requested_path or config.path}", flush=True)
    print(f"KEYWORDS_CONFIG_SOURCE={config.source_kind}", flush=True)
    print(f"KEYWORDS_PROFILE={validation_summary(config)}", flush=True)
