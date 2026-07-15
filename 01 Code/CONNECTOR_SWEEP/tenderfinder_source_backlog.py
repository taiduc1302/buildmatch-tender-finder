"""Patch 5.19: source growth backlog loader + TENDER_FINDER ranking.

Reads the source-universe backlog shipped in data/ (XLSX preferred, CSV as
cross-check/fallback) and produces the ranked "Potential_Sources_Next" sheet
for the demo workbook.

This is NOT pipeline volume and must never be presented as leads or tenders.
It is a prioritized worklist of sources TENDER_FINDER could verify, connect, or
monitor next. Paid/login/relationship-gated sources are kept visible and
marked honestly - they are never routed to an automatic live connector,
consistent with the standing invariants (no portal login, no password
storage, no paid-source scraping).
"""
from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path
from typing import Any

from tenderfinder_excel_safety import append_untrusted_row, set_untrusted_cell

DATA_DIR = Path(__file__).resolve().parent / "data"
BACKLOG_XLSX = DATA_DIR / "TENDER_FINDER_Source_Universe_Backlog_v2_EXPANDED.xlsx"
BACKLOG_CSV = DATA_DIR / "TENDER_FINDER_Potential_Unaccounted_Sources_v2_EXPANDED.csv"
UNIVERSE_SHEET = "All_Source_Universe_v2"
PRIORITY_SHEET = "Priority_Next_60"

CORE_COLUMNS = [
    "source_id", "source_name", "category", "owner_org", "geography",
    "source_type", "signal_type", "suggested_output_route", "priority_tier",
    "automation_feasibility", "access_status", "final_package_status",
    "duplicate_status", "next_action", "notes", "evidence_source",
]

RANK_COLUMNS = [
    "tenderfinder_rank", "priority_score", "recommended_phase",
    "recommended_connector_type", "demo_value", "implementation_effort",
    "risk_note", "next_step_plain_english",
]

# Sheet layout: rank fields the reader sorts by first, then the preserved
# core columns, then the remaining TENDER_FINDER annotations.
SHEET_COLUMNS = (
    ["tenderfinder_rank", "priority_score", "recommended_phase"]
    + CORE_COLUMNS
    + ["recommended_connector_type", "demo_value", "implementation_effort",
       "risk_note", "next_step_plain_english"]
)

# Access statuses that mean a human/business decision stands between TENDER_FINDER
# and the data. These must NEVER be treated as automatic live connectors.
GATED_ACCESS = {"login_or_relationship_required", "paid_or_login_required"}

# Connector types allowed for gated sources - all human-mediated.
NON_AUTOMATIC_CONNECTOR_TYPES = {"email_alert_intake", "manual_or_paid_decision"}

CATEGORY_BASE_SCORE = {
    "A_active_tender": 40,
    "B_dev_applications": 26,
    "D_capital_future_infrastructure": 24,
    "C_council_agendas": 22,
    "F_gc_developer_invites": 20,
    "I_first_nations_indigenous_infrastructure": 18,
    "K_private_developer_pipeline": 16,
    "G_news_early_signal": 12,
    "E_paid_intelligence": 8,
}

TIER_BONUS = {"1": 30, "2": 15, "3": 0}

ACCESS_BONUS = {
    "public_or_public_with_bot_checks": 20,
    "public_to_verify": 14,
    "to_verify": 8,
    "login_or_relationship_required": 2,
    "paid_or_login_required": 0,
}

ROUTE_BONUS = {
    "Active_Tenders": 10,
    "Future_Projects": 5,
    "Run_Queue": 4,
    "varies": 2,
    "Paid_Intelligence": 0,
}

PRIORITY_60_BONUS = 12
STRATEGIC_GATED_TIER1_BONUS = 6  # gated but marked tier 1 = strategically important

DUPLICATE_ADJUSTMENT = {
    "accounted_final": -25,
    "possible_duplicate_or_parent_source": -8,
    "baseline_existing_concept_preserve": -5,
    "new_candidate_to_dedupe": 0,
}

FEASIBILITY_CONNECTOR_TYPE = {
    "partial_automated_public_page": "live_html_connector",
    "manual_or_rss_html": "rss_or_html_watcher",
    "manual_or_html_watch": "html_watch_with_manual_review",
    "manual_or_pdf_p3": "pdf_manual_review",
    "manual_or_email_forwarded": "email_alert_intake",
    "paid_login_or_manual": "manual_or_paid_decision",
    "to_assess": "to_assess_likely_html_watch",
}

CATEGORY_PLAIN_ENGLISH = {
    "A_active_tender": "Active tender portals - open bid opportunities you could respond to now",
    "B_dev_applications": "Development applications - rezoning/subdivision/servicing filings that become civil work later",
    "C_council_agendas": "Council agendas and minutes - early approval and capital-award signals",
    "D_capital_future_infrastructure": "Capital plans and future infrastructure - budgeted projects before they go to tender",
    "E_paid_intelligence": "Paid intelligence services - subscription products; business decision, never scraped",
    "F_gc_developer_invites": "General contractor / developer invite networks - sub-trade opportunities via relationships",
    "G_news_early_signal": "News and announcements - earliest but least structured project signals",
    "I_first_nations_indigenous_infrastructure": "First Nations and Indigenous infrastructure - community and partnership-driven projects",
    "K_private_developer_pipeline": "Private developer pipelines - land assembly and servicing signals before public filings",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rows_from_sheet(ws) -> list[dict[str, str]]:
    """Parse a backlog sheet that has a 2-3 row banner above the real header
    row. The header row is found by looking for 'source_id' in column 1."""
    raw = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(raw[:8]):
        if row and _clean(row[0]) == "source_id":
            header_idx = i
            break
    if header_idx is None:
        return []
    headers = [_clean(c) for c in raw[header_idx]]
    rows = []
    for r in raw[header_idx + 1:]:
        if not any(v not in (None, "") for v in r):
            continue
        rows.append({h: _clean(v) for h, v in zip(headers, r) if h})
    return rows


def load_universe_rows(xlsx_path: Path = BACKLOG_XLSX, csv_path: Path = BACKLOG_CSV) -> tuple[list[dict[str, str]], str]:
    """Load the full source universe. Prefers the XLSX All_Source_Universe_v2
    sheet; falls back to the CSV (potential/unaccounted subset) if the XLSX
    is missing or unreadable. Returns (rows, source_label)."""
    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            try:
                if UNIVERSE_SHEET in wb.sheetnames:
                    rows = _rows_from_sheet(wb[UNIVERSE_SHEET])
                    if rows:
                        return rows, f"xlsx:{UNIVERSE_SHEET}"
            finally:
                wb.close()
        except Exception:
            pass
    rows = load_csv_rows(csv_path)
    return rows, "csv_fallback"


def load_csv_rows(csv_path: Path = BACKLOG_CSV) -> list[dict[str, str]]:
    if not csv_path.exists():
        return []
    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        return [{k: _clean(v) for k, v in row.items() if k} for row in csv.DictReader(f)]


def load_priority_ids(xlsx_path: Path = BACKLOG_XLSX) -> set[str]:
    if not xlsx_path.exists():
        return set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            if PRIORITY_SHEET not in wb.sheetnames:
                return set()
            return {r["source_id"] for r in _rows_from_sheet(wb[PRIORITY_SHEET]) if r.get("source_id")}
        finally:
            wb.close()
    except Exception:
        return set()


def cross_check_counts(xlsx_path: Path = BACKLOG_XLSX, csv_path: Path = BACKLOG_CSV) -> dict[str, Any]:
    """Cross-check the CSV (potential/unaccounted) against the XLSX universe.
    Reported in the sheet summary; a mismatch is surfaced, not hidden."""
    universe, origin = load_universe_rows(xlsx_path, csv_path)
    csv_rows = load_csv_rows(csv_path)
    universe_ids = {r.get("source_id", "") for r in universe}
    missing = sorted({r.get("source_id", "") for r in csv_rows} - universe_ids)
    return {
        "universe_count": len(universe),
        "universe_origin": origin,
        "csv_count": len(csv_rows),
        "priority_count": len(load_priority_ids(xlsx_path)),
        "csv_ids_missing_from_universe": missing,
    }


def is_gated(row: dict[str, str]) -> bool:
    return row.get("access_status", "") in GATED_ACCESS


def priority_score(row: dict[str, str], priority_ids: set[str]) -> int:
    score = CATEGORY_BASE_SCORE.get(row.get("category", ""), 10)
    score += TIER_BONUS.get(row.get("priority_tier", ""), 0)
    score += ACCESS_BONUS.get(row.get("access_status", ""), 5)
    score += ROUTE_BONUS.get(row.get("suggested_output_route", ""), 2)
    if row.get("source_id", "") in priority_ids:
        score += PRIORITY_60_BONUS
    score += DUPLICATE_ADJUSTMENT.get(row.get("duplicate_status", ""), 0)
    if is_gated(row) and row.get("priority_tier") == "1":
        score += STRATEGIC_GATED_TIER1_BONUS
    return score


def recommended_connector_type(row: dict[str, str]) -> str:
    connector = FEASIBILITY_CONNECTOR_TYPE.get(
        row.get("automation_feasibility", ""), "to_assess_likely_html_watch")
    if is_gated(row):
        # Hard rule: gated sources never get an automatic live connector.
        if connector != "email_alert_intake":
            connector = "manual_or_paid_decision"
    return connector


def recommended_phase(row: dict[str, str], score: int) -> str:
    dup = row.get("duplicate_status", "")
    if dup == "accounted_final":
        return "Phase 0 - already coded in package"
    if dup == "possible_duplicate_or_parent_source":
        return "Phase 4 - dedupe before any build"
    if is_gated(row):
        return "Phase 3 - relationship / email-alert / paid decision"
    if score >= 65:
        return "Phase 1 - verify and connect next"
    if score >= 45:
        return "Phase 2 - monitor or lightweight watch"
    return "Phase 4 - hold for later review"


def demo_value(row: dict[str, str]) -> str:
    if is_gated(row):
        # Can't be shown live in a demo without an account/subscription.
        return "medium" if row.get("category") == "A_active_tender" or row.get("priority_tier") == "1" else "low"
    if row.get("category") == "A_active_tender" or row.get("suggested_output_route") == "Active_Tenders":
        return "high"
    if row.get("category") in {"B_dev_applications", "C_council_agendas",
                               "D_capital_future_infrastructure",
                               "I_first_nations_indigenous_infrastructure"}:
        return "medium"
    if row.get("suggested_output_route") == "Future_Projects":
        return "medium"
    return "low"


def implementation_effort(row: dict[str, str]) -> str:
    if is_gated(row) or row.get("automation_feasibility") == "paid_login_or_manual":
        return "high"
    if row.get("automation_feasibility") in {"partial_automated_public_page", "manual_or_rss_html"}:
        return "low"
    return "medium"


def risk_note(row: dict[str, str]) -> str:
    notes = []
    dup = row.get("duplicate_status", "")
    if dup == "possible_duplicate_or_parent_source":
        notes.append("Possible duplicate/parent of an already-coded source - dedupe before building anything.")
    elif dup == "accounted_final":
        notes.append("Already coded in the current package - listed for completeness, no new work needed.")
    elif dup == "baseline_existing_concept_preserve":
        notes.append("Preserved from the original 68-source register - re-verify before re-coding.")
    access = row.get("access_status", "")
    if access == "paid_or_login_required":
        notes.append("Paid or login-gated - do not scrape; needs a subscription/account decision.")
    elif access == "login_or_relationship_required":
        notes.append("Needs a login or relationship contact - route via email alerts or manual checks, never scraping.")
    elif access == "public_or_public_with_bot_checks":
        notes.append("Public but may show a bot-check - handle honestly, same as BC Bid.")
    elif access in {"public_to_verify", "to_verify"}:
        notes.append("URL and public access not yet verified.")
    return " ".join(notes) or "No specific risk flagged - verify the URL before building."


def next_step_plain_english(row: dict[str, str], phase: str, connector: str) -> str:
    if phase.startswith("Phase 0"):
        return "Nothing to do - this source is already coded and running in the package."
    if phase.startswith("Phase 4 - dedupe"):
        return "Check whether this is the same as an already-coded source before doing anything."
    if phase.startswith("Phase 3"):
        if connector == "email_alert_intake":
            return "Register for this portal's email alerts and let TENDER_FINDER's email intake read them - no login automation."
        return "Business decision needed: paid subscription or a relationship contact. Do not scrape."
    if phase.startswith("Phase 1"):
        if connector == "live_html_connector":
            return "Verify the public page loads, then build or enable an HTML connector for it."
        if connector == "rss_or_html_watcher":
            return "Verify the feed or listing page, then add an RSS/HTML watcher."
        if connector == "email_alert_intake":
            return "Sign up for the public email alerts and route them through TENDER_FINDER's email intake."
        return "Verify the public page, then pick the lightest connector that fits it."
    if phase.startswith("Phase 2"):
        return "Set up a lightweight periodic check (e.g. monthly) before investing in a connector."
    return "Revisit after higher-priority sources are connected."


def rank_universe(rows: list[dict[str, str]], priority_ids: set[str]) -> list[dict[str, Any]]:
    """Score, annotate, and deterministically sort every universe row. No
    rows are dropped - duplicates stay visible with their status marked."""
    annotated = []
    for row in rows:
        score = priority_score(row, priority_ids)
        phase = recommended_phase(row, score)
        connector = recommended_connector_type(row)
        out = {col: row.get(col, "") for col in CORE_COLUMNS}
        out["priority_score"] = score
        out["recommended_phase"] = phase
        out["recommended_connector_type"] = connector
        out["demo_value"] = demo_value(row)
        out["implementation_effort"] = implementation_effort(row)
        out["risk_note"] = risk_note(row)
        out["next_step_plain_english"] = next_step_plain_english(row, phase, connector)
        annotated.append(out)
    annotated.sort(key=lambda r: (
        -r["priority_score"],
        r.get("priority_tier") or "9",
        r.get("category", ""),
        r.get("source_id", ""),
    ))
    for i, row in enumerate(annotated, 1):
        row["tenderfinder_rank"] = i
    return annotated


def build_summary_rows(ranked: list[dict[str, Any]], check: dict[str, Any]) -> list[list[Any]]:
    """The one-page business-readable summary block for the top of the sheet."""
    coded = sum(1 for r in ranked if r.get("final_package_status") == "in_final_coded_sources")
    by_category = Counter(r.get("category", "") for r in ranked)
    by_access = Counter(r.get("access_status", "") for r in ranked)
    rows: list[list[Any]] = [
        ["SOURCE GROWTH BACKLOG - ranked worklist of sources to verify, connect, or monitor next", "", ""],
        ["This is NOT pipeline volume or leads. It is a to-do list of data sources, honestly marked where access is gated.", "", ""],
        ["", "", ""],
        ["Total source universe", check["universe_count"], f"loaded from {check['universe_origin']}"],
        ["Sources already coded in package", coded, "duplicate_status=accounted_final rows stay visible below"],
        ["Potential / unaccounted sources", check["csv_count"], "cross-checked against the CSV backlog file"],
        ["Priority next 60", check["priority_count"], "flagged in the Priority_Next_60 backlog sheet; scored higher below"],
    ]
    missing = check.get("csv_ids_missing_from_universe") or []
    if missing:
        rows.append(["CSV cross-check WARNING", len(missing),
                     "CSV sources missing from the universe sheet: " + ", ".join(missing[:10])])
    else:
        rows.append(["CSV cross-check", "OK", "every CSV source id is present in the universe sheet"])
    rows.append(["", "", ""])
    rows.append(["COUNT BY CATEGORY", "", ""])
    for cat, n in sorted(by_category.items(), key=lambda kv: (-kv[1], kv[0])):
        rows.append([cat, n, CATEGORY_PLAIN_ENGLISH.get(cat, "")])
    rows.append(["", "", ""])
    rows.append(["COUNT BY ACCESS STATUS", "", ""])
    for status, n in sorted(by_access.items(), key=lambda kv: (-kv[1], kv[0])):
        rows.append([status, n, "gated - manual/email/paid path only" if status in GATED_ACCESS else ""])
    rows.append(["", "", ""])
    rows.append(["TOP 15 RECOMMENDED NEXT SOURCES", "", "already-coded and dedupe-first rows excluded"])
    shown = 0
    for r in ranked:
        if r.get("duplicate_status") in {"accounted_final", "possible_duplicate_or_parent_source"}:
            continue
        rows.append([f"#{r['tenderfinder_rank']} {r.get('source_name', '')}",
                     r["priority_score"],
                     f"{r.get('category', '')} | {r['next_step_plain_english']}"])
        shown += 1
        if shown >= 15:
            break
    return rows


def build_potential_sources_content(
    xlsx_path: Path = BACKLOG_XLSX, csv_path: Path = BACKLOG_CSV,
) -> tuple[list[list[Any]], list[str], list[list[Any]]]:
    """Everything the workbook writer needs: (summary_rows, headers, data_rows)."""
    universe, _origin = load_universe_rows(xlsx_path, csv_path)
    priority_ids = load_priority_ids(xlsx_path)
    ranked = rank_universe(universe, priority_ids)
    check = cross_check_counts(xlsx_path, csv_path)
    summary = build_summary_rows(ranked, check)
    data_rows = [[r.get(col, "") for col in SHEET_COLUMNS] for r in ranked]
    return summary, list(SHEET_COLUMNS), data_rows


def build_roadmap_priorities(ranked: list[dict[str, Any]]) -> list[list[Any]]:
    """3-5 honest 'what to build next' bullets derived from the ranking
    itself - no invented volume, just counts of what each move unlocks."""
    p1 = [r for r in ranked if r["recommended_phase"].startswith("Phase 1")]
    p1_active = [r for r in p1 if r.get("category") == "A_active_tender"]
    p2 = [r for r in ranked if r["recommended_phase"].startswith("Phase 2")]
    email_route = [r for r in ranked if r["recommended_connector_type"] == "email_alert_intake"]
    paid = [r for r in ranked if r["recommended_phase"].startswith("Phase 3")
            and r["recommended_connector_type"] == "manual_or_paid_decision"]
    return [
        ["1. Connect public active-tender pages", len(p1_active),
         "Phase 1 active-tender sources (public, tier-ranked) - the direct BID NOW volume unlock"],
        ["2. Verify remaining Phase 1 sources", len(p1) - len(p1_active),
         "Public dev-application / capital-plan / council sources ready to verify and connect"],
        ["3. Stand up lightweight watchers", len(p2),
         "Phase 2 sources worth a monthly check before investing in connectors"],
        ["4. Enable portal email alerts", len(email_route),
         "Sources best served by TENDER_FINDER's existing email intake - register once, no login scraping"],
        ["5. Make paid/relationship decisions", len(paid),
         "Gated sources needing a subscription or contact decision - flagged honestly, never scraped"],
    ]


def append_source_roadmap_printable_sheet(wb) -> None:
    """One-page, VP-readable source-expansion roadmap. Placed right after
    Executive_Summary by build_workbook so nobody has to hunt for it. The
    full 372-row detail lives in Potential_Sources_Next."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Source_Roadmap_Printable")
    if not BACKLOG_XLSX.exists() and not BACKLOG_CSV.exists():
        ws.cell(1, 1, "Source backlog data files were not found in the data/ folder next to the script - no roadmap could be generated for this build.")
        ws.cell(1, 1).font = Font(name="Arial", bold=True, color="A43A1E")
        return

    universe, _ = load_universe_rows()
    ranked = rank_universe(universe, load_priority_ids())
    check = cross_check_counts()
    coded = sum(1 for r in ranked if r.get("final_package_status") == "in_final_coded_sources")
    by_category = Counter(r.get("category", "") for r in ranked)
    by_access = Counter(r.get("access_status", "") for r in ranked)

    rows: list[list[Any]] = [
        ["WHERE TENDER_FINDER CAN GROW NEXT", "", ""],
        ["This is not lead volume. This is a ranked source expansion roadmap: the data sources TENDER_FINDER can verify, connect, or monitor next, scored by tender value, public access, and effort.", "", ""],
        ["", "", ""],
        ["THE NUMBERS", "", ""],
        ["Total source universe", check["universe_count"], "every candidate source known to TENDER_FINDER today"],
        ["Already coded and running", coded, "sources live in the current package"],
        ["Potential / unaccounted", check["csv_count"], "candidate sources not yet connected"],
        ["Priority Next 60", check["priority_count"], "pre-flagged shortlist, boosted in the ranking below"],
        ["", "", ""],
        ["RECOMMENDED NEXT PATCH PRIORITIES", "count", "what it unlocks"],
    ]
    rows.extend(build_roadmap_priorities(ranked))
    rows.append(["", "", ""])
    rows.append(["TOP 25 RECOMMENDED NEXT SOURCES", "score", "category | plain-English next step"])
    shown = 0
    for r in ranked:
        if r.get("duplicate_status") in {"accounted_final", "possible_duplicate_or_parent_source"}:
            continue
        rows.append([f"#{r['tenderfinder_rank']} {r.get('source_name', '')}",
                     r["priority_score"],
                     f"{r.get('category', '')} | {r['next_step_plain_english']}"])
        shown += 1
        if shown >= 25:
            break
    rows.append(["", "", ""])
    rows.append(["BY CATEGORY", "count", "what the category means"])
    for cat, n in sorted(by_category.items(), key=lambda kv: (-kv[1], kv[0])):
        rows.append([cat, n, CATEGORY_PLAIN_ENGLISH.get(cat, "")])
    rows.append(["", "", ""])
    rows.append(["BY ACCESS STATUS", "count", ""])
    for status, n in sorted(by_access.items(), key=lambda kv: (-kv[1], kv[0])):
        rows.append([status, n, "gated - manual/email/paid path only, never scraped" if status in GATED_ACCESS else ""])
    rows.append(["", "", ""])
    rows.append(["Full detail: see the Potential_Sources_Next tab (all rows, filterable, with rank/score/phase/risk fields).", "", ""])

    for row in rows:
        append_untrusted_row(ws, row)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    section_headers = {"THE NUMBERS", "RECOMMENDED NEXT PATCH PRIORITIES",
                       "TOP 25 RECOMMENDED NEXT SOURCES", "BY CATEGORY", "BY ACCESS STATUS"}
    for r in range(2, ws.max_row + 1):
        label = _clean(ws.cell(r, 1).value)
        is_section = label in section_headers
        for c in range(1, 4):
            cell = ws.cell(r, c)
            cell.font = Font(name="Arial", size=10, bold=is_section or r == 2)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_section:
                cell.fill = PatternFill("solid", fgColor="D9E2F3")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 96


def append_potential_sources_sheet(wb) -> None:
    """Create and fill the Potential_Sources_Next sheet on an open workbook.
    If the backlog data files are missing, the sheet still exists with an
    honest explanation instead of fabricated content."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Potential_Sources_Next")
    if not BACKLOG_XLSX.exists() and not BACKLOG_CSV.exists():
        ws.cell(1, 1, "Source growth backlog data files were not found in the data/ folder next to the script - no ranked source list could be generated for this build.")
        ws.cell(1, 1).font = Font(name="Arial", bold=True, color="A43A1E")
        return

    summary, headers, data_rows = build_potential_sources_content()
    for row in summary:
        append_untrusted_row(ws, row)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    for r in range(2, ws.max_row + 1):
        label = _clean(ws.cell(r, 1).value)
        bold = label.startswith(("COUNT BY", "TOP 15")) or r == 2
        for c in range(1, 4):
            ws.cell(r, c).font = Font(name="Arial", size=10, bold=bold)
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    header_row = ws.max_row + 2
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for offset, row in enumerate(data_rows, 1):
        for col, value in enumerate(row, 1):
            cell = set_untrusted_cell(ws.cell(header_row + offset, col), value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = f"A{header_row + 1}"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(data_rows)}"
    widths = {
        "tenderfinder_rank": 30, "priority_score": 12, "recommended_phase": 34,
        "source_id": 30, "source_name": 42, "category": 30, "owner_org": 30,
        "geography": 24, "source_type": 24, "signal_type": 34,
        "suggested_output_route": 20, "priority_tier": 11,
        "automation_feasibility": 26, "access_status": 26,
        "final_package_status": 30, "duplicate_status": 30,
        "next_action": 40, "notes": 44, "evidence_source": 26,
        "recommended_connector_type": 28, "demo_value": 12,
        "implementation_effort": 18, "risk_note": 54,
        "next_step_plain_english": 62,
    }
    for col, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(header, 20)
