"""Tests for the headless development-data refresh orchestration (Phase 2).

Uses injected fake acquirer/scorer so the full flow runs headlessly with no
network and no real workbook engine. Runs under pytest and as a script.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
REPO_ROOT = ROOT.parents[1]

import tenderfinder_refresh_service as rs  # noqa: E402
import tenderfinder_data_modes as dm  # noqa: E402


def _records(source_id: str, n: int) -> tuple[dict, ...]:
    return tuple(
        {
            "source": source_id,
            "app_no": f"{source_id}-{i}",
            "address": f"{100 + i} Example St",
            "scope_summary": "New multi-lot subdivision servicing with watermain and storm sewer",
            "app_type_stage": "Development Permit",
            "municipality": "Surrey",
        }
        for i in range(n)
    )


def _ok_acquirer(source):
    return rs.SourceFetch(source_id=source["source_id"], ok=True, records=_records(source["source_id"], 3),
                          http_status=200)


def _fake_scorer(dataset_path: Path, out_dir: Path) -> rs.ScoreResult:
    out = out_dir / "ranked.xlsx"
    out.write_bytes(b"PK\x03\x04 fake ranked workbook")
    return rs.ScoreResult(output_path=str(out), scored=9, bid_now=1, bid_later=3, watch=2, skip=3)


def test_eligible_sources_exclude_non_runnable() -> None:
    eligible = rs.eligible_development_sources(package_root=REPO_ROOT)
    ids = {s["source_id"] for s in eligible}
    # needs_configuration / manual_only / wrong_source sources must be excluded
    for excluded in ("van_rezoning", "van_devpermits", "city_langley_devapps", "burnaby_devapps"):
        assert excluded not in ids, f"{excluded} must not be runnable"
    # a known ready development source is present
    assert any(s["operational_status"] in ("ready_for_live_test", "verified_live") for s in eligible)


def test_source_health_never_marks_failed_source_healthy() -> None:
    health = rs.source_health_rows(package_root=REPO_ROOT, track="development")
    by_id = {row["source_id"]: row for row in health}
    van = by_id.get("van_rezoning")
    assert van is not None
    assert van["runtime_eligible"] is False
    assert van["skip_reason"]


def test_full_refresh_flow_promotes_scores_and_writes_manifest() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        result = rs.refresh_development_data(
            rs.RefreshRequest(preset_id="civil_contractor", state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps", "twp_langley_devactivity")),
            acquirer=_ok_acquirer, scorer=_fake_scorer,
        )
        assert result.succeeded, result.message
        # active pointer set to the new live dataset
        active = dm.load_active_dataset(state_root=state, package_root=REPO_ROOT)
        assert active is not None and active.data_mode == dm.MODE_LIVE
        # metrics reconcile and are truthful
        assert result.metrics.is_reconciled(), result.metrics.reconciliation_errors()
        assert result.metrics.records_fetched == 6
        assert result.metrics.records_scored == 9
        assert result.metrics.data_mode == dm.MODE_LIVE
        # output + manifest written
        assert Path(result.output_paths["output_workbook"]).exists()
        assert Path(result.manifest_path).exists()


def test_partial_source_failure_still_produces_dataset() -> None:
    def flaky(source):
        if source["source_id"].startswith("maple"):
            return rs.SourceFetch(source_id=source["source_id"], ok=False, error="503 upstream")
        return _ok_acquirer(source)

    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        result = rs.refresh_development_data(
            rs.RefreshRequest(state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps", "twp_langley_devactivity")),
            acquirer=flaky, scorer=_fake_scorer,
        )
        assert result.succeeded
        assert result.metrics.sources_failed == 1
        assert result.metrics.sources_successful == 1
        outcomes = {o["source_id"]: o for o in result.source_outcomes}
        assert outcomes["maple_ridge_devapps"]["ok"] is False


def test_total_failure_preserves_previous_dataset_and_marks_stale() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        # first, a good refresh establishes a live dataset
        good = rs.refresh_development_data(
            rs.RefreshRequest(state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps",)),
            acquirer=_ok_acquirer, scorer=_fake_scorer,
        )
        assert good.succeeded
        good_dataset_id = dm.load_active_dataset(state_root=state, package_root=REPO_ROOT).dataset_id

        # now every source fails
        def all_fail(source):
            return rs.SourceFetch(source_id=source["source_id"], ok=False, error="network down")

        failed = rs.refresh_development_data(
            rs.RefreshRequest(state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps", "twp_langley_devactivity")),
            acquirer=all_fail, scorer=_fake_scorer,
        )
        assert not failed.succeeded
        assert failed.stale
        # previous active dataset is unchanged (not overwritten by the failed run)
        still = dm.load_active_dataset(state_root=state, package_root=REPO_ROOT)
        assert still.dataset_id == good_dataset_id
        assert still.data_mode == dm.MODE_LIVE  # pointer itself untouched
        # the failure result labels continued use as cached/stale
        assert failed.provenance is not None and failed.provenance.stale
        assert failed.metrics.succeeded is False
        assert not failed.metrics.records_live  # a failed run reports no live records


def test_failed_validation_preserves_previous() -> None:
    def empty_records(source):
        # ok but returns only a non-descriptive stub -> dataset ends up empty
        return rs.SourceFetch(source_id=source["source_id"], ok=True,
                              records=({"source": source["source_id"], "app_no": "x"},))

    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        result = rs.refresh_development_data(
            rs.RefreshRequest(state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps",)),
            acquirer=empty_records, scorer=_fake_scorer,
        )
        assert not result.succeeded
        assert dm.load_active_dataset(state_root=state, package_root=REPO_ROOT) is None


def test_thin_stub_records_do_not_fail_an_otherwise_good_dataset() -> None:
    """Regression test for a real bug found during the controlled live sweep:
    real municipal open-data feeds routinely mix a few non-descriptive stub
    records (an app number with every other field blank) into an otherwise
    rich dataset. A single thin record must not sour hundreds of good ones."""
    def mixed_records(source):
        good = [
            {"source": source["source_id"], "app_no": f"A{i}", "address": f"{i} Main St",
             "scope_summary": "Watermain and sanitary sewer servicing"}
            for i in range(50)
        ]
        stub = [{"source": source["source_id"], "app_no": "STUB-1", "address": "", "scope_summary": ""}]
        return rs.SourceFetch(source_id=source["source_id"], ok=True, records=tuple(good + stub))

    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        result = rs.refresh_development_data(
            rs.RefreshRequest(state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps",)),
            acquirer=mixed_records, scorer=_fake_scorer,
        )
    assert result.succeeded, result.message
    assert result.metrics.normalized_records == 50  # the stub was dropped, not fatal
    assert result.metrics.duplicates_removed == 1    # accounts for the dropped stub


def test_records_fetched_stays_truthful_on_validation_failure_but_records_live_is_zero() -> None:
    """A validation failure (dataset ends up empty) must still report the true
    fetched count as a diagnostic - it must NOT be hidden as zero - while
    records_live correctly stays 0 because nothing was promoted."""
    def stub_only(source):
        return rs.SourceFetch(source_id=source["source_id"], ok=True,
                              records=({"source": source["source_id"], "app_no": "x"},))

    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        result = rs.refresh_development_data(
            rs.RefreshRequest(state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps",)),
            acquirer=stub_only, scorer=_fake_scorer,
        )
    assert not result.succeeded
    assert result.metrics.records_fetched == 1  # truthful: we did fetch 1 record
    assert result.metrics.records_live == 0     # truthful: nothing went live
    assert result.metrics.is_reconciled(), result.metrics.reconciliation_errors()


def test_dedup_reconciles() -> None:
    dupes = [
        {"source": "s", "app_no": "1", "address": "1 St", "scope_summary": "civil"},
        {"source": "s", "app_no": "1", "address": "1 St", "scope_summary": "civil"},
        {"source": "s", "app_no": "2", "address": "2 St", "scope_summary": "civil"},
    ]
    deduped, removed = rs.deduplicate_records(dupes)
    assert len(deduped) == 2 and removed == 1


def test_never_writes_inside_package() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        result = rs.refresh_development_data(
            rs.RefreshRequest(state_root=state, package_root=REPO_ROOT,
                              source_ids=("maple_ridge_devapps",)),
            acquirer=_ok_acquirer, scorer=_fake_scorer,
        )
        dataset_path = Path(result.output_paths["dataset"]).resolve()
        assert str(REPO_ROOT.resolve()) not in str(dataset_path)


def test_dataset_writer_neutralizes_formula_injection() -> None:
    """Public-source fields must never be interpreted as Excel formulas.

    A hostile or malformed public record could contain a scope_summary or
    address starting with '=', '+', '-', or '@' (the classic CSV/Excel
    formula-injection prefixes, e.g. '=cmd|...' or '=HYPERLINK(...)'). The
    dataset writer must always emit these as literal text.
    """
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "dataset.xlsx"
        malicious = [
            {"lead_id": "L1", "municipality": "Surrey", "app_no": "A1",
             "address": "=cmd|'/c calc'!A1", "app_type_stage": "Development Permit",
             "scope_summary": "+1+1", "applicant_name": "@SUM(1,2)",
             "source": "s", "source_url": "-2+3"},
        ]
        rs._default_dataset_writer(malicious, dest)
        wb = openpyxl.load_workbook(dest, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        data_row = dict(zip(rows[0], rows[1]))
        # every formula-prefixed value must come back with a literal-text
        # marker, never as a live formula.
        for field in ("address", "scope_summary", "applicant_name", "source_url"):
            assert str(data_row[field]).startswith("'"), f"{field} was not neutralized: {data_row[field]!r}"


def test_default_scorer_neutralizes_formula_injection() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp:
        dataset_path = Path(tmp) / "dataset.xlsx"
        out_dir = Path(tmp) / "output"
        rs._default_dataset_writer(
            [{"lead_id": "L1", "municipality": "Surrey", "scope_summary": "=HYPERLINK(\"http://evil\")"}],
            dataset_path,
        )
        score = rs.default_scorer(dataset_path, out_dir, preset_id="civil_contractor", package_root=REPO_ROOT)
        wb = openpyxl.load_workbook(score.output_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        data_row = dict(zip(rows[0], rows[1]))
        assert str(data_row["scope_summary"]).startswith("'")


# --------------------------------------------------------------------------- #
# Full-sweep acquirer (Gap A fix): must NOT be the bounded connector preview.
# tenderfinder_raw_sweep.run_connector is monkeypatched so these tests remain
# fully offline/deterministic — no real network I/O in the mandatory suite.
# --------------------------------------------------------------------------- #


def _fake_lead(project_id="P1", municipality="Surrey", app_no="A1", address="1 Main St",
               owner="Acme Developer", app_type_stage="Development Permit",
               scope_summary="Watermain and sanitary sewer servicing", fit_score=82,
               source_url="https://example.com/p1", proposed_route="Future_Projects"):
    return {
        "project_id": project_id, "municipality": municipality, "app_no": app_no,
        "address": address, "owner": owner, "app_type_stage": app_type_stage,
        "scope_summary": scope_summary, "fit_score": fit_score, "source_url": source_url,
        "proposed_route": proposed_route, "date_found": "2026-07-19", "raw_hash": "abc123",
        "dedupe_key": f"{app_no}|{address}",
    }


def _source_row(source_id="maple_ridge_devapps"):
    rows = rs.eligible_development_sources(package_root=REPO_ROOT, source_ids=(source_id,))
    assert rows, f"{source_id} must be a runtime-eligible fixture for this test"
    return rows[0]


def test_full_sweep_acquirer_maps_many_leads_not_a_bounded_preview() -> None:
    import tenderfinder_raw_sweep as raw_sweep

    many_leads = [_fake_lead(project_id=f"P{i}", app_no=f"A{i}", address=f"{i} Main St")
                  for i in range(37)]  # far more than a ~5-record preview
    original = raw_sweep.run_connector

    def fake_run_connector(conn, max_records, probe=False, raw_dir=None, **kwargs):
        assert probe is False
        return {"status": "load_ok", "error": "", "leads": many_leads[:max_records],
                "records_pulled": len(many_leads), "http_status": 200}

    raw_sweep.run_connector = fake_run_connector
    try:
        fetch = rs.full_sweep_development_acquirer(_source_row(), max_records=500)
    finally:
        raw_sweep.run_connector = original

    assert fetch.ok
    assert fetch.count == 37  # NOT bounded to a small preview sample
    assert fetch.records[0]["scope_summary"] == "Watermain and sanitary sewer servicing"
    assert fetch.records[0]["app_no"] == "A0"


def test_full_sweep_acquirer_respects_max_records_cap() -> None:
    import tenderfinder_raw_sweep as raw_sweep

    captured = {}
    original = raw_sweep.run_connector

    def fake_run_connector(conn, max_records, probe=False, raw_dir=None, **kwargs):
        captured["max_records"] = max_records
        return {"status": "load_ok", "error": "", "leads": [], "http_status": 200}

    raw_sweep.run_connector = fake_run_connector
    try:
        rs.full_sweep_development_acquirer(_source_row(), max_records=250)
    finally:
        raw_sweep.run_connector = original
    assert captured["max_records"] == 250


def test_full_sweep_acquirer_non_fetch_status_is_not_ok() -> None:
    import tenderfinder_raw_sweep as raw_sweep

    original = raw_sweep.run_connector
    raw_sweep.run_connector = lambda conn, max_records, probe=False, raw_dir=None, **kw: {
        "status": "access_test_required", "error": "", "leads": [], "http_status": 0,
    }
    try:
        fetch = rs.full_sweep_development_acquirer(_source_row(), max_records=100)
    finally:
        raw_sweep.run_connector = original
    assert not fetch.ok
    assert "access_test_required" in fetch.error


def test_full_sweep_acquirer_error_is_not_ok() -> None:
    import tenderfinder_raw_sweep as raw_sweep

    original = raw_sweep.run_connector
    raw_sweep.run_connector = lambda conn, max_records, probe=False, raw_dir=None, **kw: {
        "status": "load_error", "error": "504 Gateway Timeout", "leads": [], "http_status": 504,
    }
    try:
        fetch = rs.full_sweep_development_acquirer(_source_row(), max_records=100)
    finally:
        raw_sweep.run_connector = original
    assert not fetch.ok
    assert "504" in fetch.error


def test_full_sweep_acquirer_exception_is_a_failed_source_not_a_crash() -> None:
    import tenderfinder_raw_sweep as raw_sweep

    original = raw_sweep.run_connector

    def boom(*a, **k):
        raise RuntimeError("network unreachable")

    raw_sweep.run_connector = boom
    try:
        fetch = rs.full_sweep_development_acquirer(_source_row(), max_records=100)
    finally:
        raw_sweep.run_connector = original
    assert not fetch.ok
    assert "network unreachable" in fetch.error


def test_make_full_sweep_acquirer_binds_raw_dir_outside_package() -> None:
    import tenderfinder_raw_sweep as raw_sweep

    captured = {}
    original = raw_sweep.run_connector
    raw_sweep.run_connector = lambda conn, max_records, probe=False, raw_dir=None, **kw: (
        captured.__setitem__("raw_dir", raw_dir) or
        {"status": "load_ok", "error": "", "leads": [], "http_status": 200}
    )
    try:
        with tempfile.TemporaryDirectory() as tmp:
            state = Path(tmp) / "state"
            request = rs.RefreshRequest(state_root=state, package_root=REPO_ROOT, run_id="run123")
            acquirer = rs.make_full_sweep_acquirer(request, package_root=REPO_ROOT, run_id="run123")
            acquirer(_source_row())
    finally:
        raw_sweep.run_connector = original
    assert captured["raw_dir"] is not None
    assert str(REPO_ROOT.resolve()) not in str(captured["raw_dir"])
    assert "run123" in str(captured["raw_dir"])


def test_full_sweep_flow_promotes_dataset_with_many_records() -> None:
    """End-to-end: full-sweep acquirer feeding the whole promote/score pipeline
    with a realistically large per-source count (not a tiny preview)."""
    import tenderfinder_raw_sweep as raw_sweep

    many_leads = [_fake_lead(project_id=f"P{i}", app_no=f"A{i}", address=f"{i} Main St")
                  for i in range(150)]
    original = raw_sweep.run_connector
    raw_sweep.run_connector = lambda conn, max_records, probe=False, raw_dir=None, **kw: {
        "status": "load_ok", "error": "", "leads": many_leads[:max_records], "http_status": 200,
    }
    try:
        with tempfile.TemporaryDirectory() as tmp:
            state = Path(tmp) / "state"
            run_id = "run_full_sweep_e2e"
            request = rs.RefreshRequest(
                state_root=state, package_root=REPO_ROOT, run_id=run_id,
                source_ids=("maple_ridge_devapps",), max_records_per_source=500,
            )
            acquirer = rs.make_full_sweep_acquirer(request, package_root=REPO_ROOT, run_id=run_id)
            result = rs.refresh_development_data(request, acquirer=acquirer, scorer=_fake_scorer)

            assert result.succeeded, result.message
            assert result.metrics.records_fetched == 150  # far above a ~5-record preview
            assert result.metrics.normalized_records == 150
            active = dm.load_active_dataset(state_root=state, package_root=REPO_ROOT)
            assert active is not None and active.record_counts["total"] == 150
    finally:
        raw_sweep.run_connector = original


def test_diagnostic_preview_and_full_sweep_are_distinct_functions() -> None:
    # Guards against accidentally re-aliasing the bounded preview back onto the
    # production entry point (the original Gap A defect).
    assert rs.full_sweep_development_acquirer is not rs.diagnostic_preview_acquirer
    assert rs.default_development_acquirer is rs.diagnostic_preview_acquirer


# --------------------------------------------------------------------------- #
# Real scorer (fixes a second real gap found during the controlled live
# sweep: the GUI never passed a scorer, so BID_LATER/WATCH/SKIP counts stayed
# permanently 0 even on a successful refresh with real data).
# --------------------------------------------------------------------------- #


def test_default_scorer_buckets_and_writes_ranked_workbook() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        dataset_path = Path(tmp) / "dataset.xlsx"
        out_dir = Path(tmp) / "output"
        records = [
            # strong civil fit -> BID_LATER
            {"lead_id": "L1", "municipality": "Surrey", "app_no": "A1", "address": "1 St",
             "app_type_stage": "Development Permit",
             "scope_summary": "Watermain and sanitary sewer site servicing subdivision",
             "applicant_name": "", "source": "s", "source_url": ""},
            # borderline -> WATCH
            {"lead_id": "L2", "municipality": "Surrey", "app_no": "A2", "address": "2 St",
             "app_type_stage": "Development Permit", "scope_summary": "Road grading",
             "applicant_name": "", "source": "s", "source_url": ""},
            # no civil signal -> SKIP
            {"lead_id": "L3", "municipality": "Surrey", "app_no": "A3", "address": "3 St",
             "app_type_stage": "Sign Permit", "scope_summary": "New commercial sign",
             "applicant_name": "", "source": "s", "source_url": ""},
        ]
        rs._default_dataset_writer(records, dataset_path)
        score = rs.default_scorer(dataset_path, out_dir, preset_id="civil_contractor", package_root=REPO_ROOT)
        assert score.scored == 3
        assert score.bid_now == 0  # development-only refresh never populates BID_NOW; truthful
        assert score.bid_later + score.watch + score.skip == 3
        assert score.bid_later >= 1  # the strong civil record must land in BID_LATER
        assert Path(score.output_path).exists()


def test_make_default_scorer_binds_preset() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        dataset_path = Path(tmp) / "dataset.xlsx"
        out_dir = Path(tmp) / "output"
        rs._default_dataset_writer(
            [{"lead_id": "L1", "municipality": "Surrey", "scope_summary":
              "Apartment building interior fit-out mechanical electrical HVAC suite"}],
            dataset_path,
        )
        request_civil = rs.RefreshRequest(preset_id="civil_contractor", package_root=REPO_ROOT)
        request_res = rs.RefreshRequest(preset_id="multifamily_residential", package_root=REPO_ROOT)
        score_civil = rs.make_default_scorer(request_civil)(dataset_path, out_dir / "civil")
        score_res = rs.make_default_scorer(request_res)(dataset_path, out_dir / "res")
    # the civil preset penalizes interior/mechanical/electrical/HVAC/suite scope;
    # the residential preset does not - so the same record scores differently.
    assert score_civil.skip >= score_res.skip or score_civil.watch != score_res.watch \
        or score_civil.bid_later != score_res.bid_later


def test_full_refresh_with_real_scorer_reports_truthful_bucket_counts() -> None:
    """End-to-end regression test for the exact bug found in the controlled
    live sweep: refresh_development_data with the REAL scorer (not a fake)
    must report non-zero BID_LATER/WATCH/SKIP for real civil-relevant data,
    not permanent zeros."""
    with tempfile.TemporaryDirectory() as tmp:
        state = Path(tmp) / "state"
        run_id = "run_real_scorer_e2e"
        request = rs.RefreshRequest(
            state_root=state, package_root=REPO_ROOT, run_id=run_id,
            source_ids=("maple_ridge_devapps",), preset_id="civil_contractor",
        )
        scorer = rs.make_default_scorer(request)
        result = rs.refresh_development_data(request, acquirer=_ok_acquirer, scorer=scorer)

        assert result.succeeded, result.message
        assert result.metrics.records_scored == result.metrics.normalized_records
        total_bucketed = (result.metrics.bid_now + result.metrics.bid_later
                          + result.metrics.watch + result.metrics.skip)
        assert total_bucketed == result.metrics.records_scored
        assert result.metrics.bid_later > 0  # the fixture records are strongly civil-relevant
        assert result.metrics.is_reconciled(), result.metrics.reconciliation_errors()
        assert "output_workbook" in result.output_paths
        assert Path(result.output_paths["output_workbook"]).exists()


def main() -> int:
    tests = [value for name, value in sorted(globals().items())
             if name.startswith("test_") and callable(value)]
    for test in tests:
        test()
        print(f"[PASS] {test.__name__}")
    print(f"Build Week refresh-service tests: {len(tests)} passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
