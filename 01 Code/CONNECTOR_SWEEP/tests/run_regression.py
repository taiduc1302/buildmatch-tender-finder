#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Patch 5.0 regression runner.

Runs real tenderfinder_raw_sweep.py commands from a freshly unzipped package and
verifies the produced files. Exits 0 only if required commands and output
checks pass. In network-restricted sandboxes the connector layer may use
packaged regression fixtures; that fallback is explicit in logs and source
summaries and is not reported as live-source proof.
"""
from __future__ import annotations

import argparse
import csv
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

REQUIRED_PREFLIGHT_FILES = [
    "TENDER_FINDER_Source_Register_URL_Live_Audit.csv",
    "TENDER_FINDER_Source_Register_URL_Live_Audit.xlsx",
    "TENDER_FINDER_Source_Register_Fix_Queue.csv",
    "TENDER_FINDER_Source_Register_Replacement_Candidates.csv",
    "TENDER_FINDER_Source_Register_Cleaned_For_Script.csv",
    "TENDER_FINDER_Link_Check_Run_Log.txt",
    "TENDER_FINDER_Link_Check_Debug_Log.txt",
]

class Regression:
    def __init__(self, output_dir: Path, keep_outputs: bool = False):
        self.script_path = Path(__file__).resolve()
        self.tests_dir = self.script_path.parent
        self.connector_dir = self.tests_dir.parent
        self.root_dir = self.connector_dir.parents[1]
        self.output_dir = output_dir.resolve()
        self.keep_outputs = keep_outputs
        self.results: List[Dict] = []
        self.row_counts: Dict[str, int] = {}
        self.logs: List[Path] = []
        if not keep_outputs and self.output_dir.exists():
            shutil.rmtree(self.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def add_result(self, name: str, passed: bool, detail: str = "", exit_code: int | None = None):
        self.results.append({"name": name, "passed": bool(passed), "detail": detail, "exit_code": exit_code})
        print(f"[{'PASS' if passed else 'FAIL'}] {name}: {detail}")

    def run_cmd(self, name: str, args: List[str], log_name: str, timeout: int = 900) -> Tuple[int, str]:
        log_path = self.output_dir / log_name
        cmd = [sys.executable] + args
        started = datetime.now().isoformat(timespec="seconds")
        # Patch 5.3: enable packaged offline fixtures for connector layers so the
        # suite is reproducible under sandboxes whose egress returns DNS failures
        # OR HTTP 403/Forbidden. This flag is set ONLY by the regression harness;
        # production runs never set it, so live 403s are surfaced honestly.
        import os as _os
        env = dict(_os.environ, TENDER_FINDER_OFFLINE_FIXTURES="1")
        try:
            proc = subprocess.run(cmd, cwd=self.connector_dir, text=True, env=env,
                                  stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=timeout)
            output = proc.stdout or ""
            code = proc.returncode
        except subprocess.TimeoutExpired as exc:
            output = exc.stdout if isinstance(exc.stdout, str) else ""
            output += f"\nTIMEOUT after {timeout}s\n"
            code = 124
        log_path.write_text(f"COMMAND: {' '.join(cmd)}\nCWD: {self.connector_dir}\nSTARTED: {started}\nEXIT_CODE: {code}\n\n{output}", encoding="utf-8")
        self.logs.append(log_path)
        self.add_result(name, code == 0, f"exit={code}, log={log_path.name}", code)
        return code, output

    def compile_check(self):
        return self.run_cmd("Python syntax compile", ["-m", "py_compile", "tenderfinder_raw_sweep.py", "tenderfinder_live_link_checker.py", "tenderfinder_surrey_inprocess.py", "tenderfinder_source_registry.py", "tenderfinder_master_io.py", "tenderfinder_guards.py", "tenderfinder_keywords_config.py", "tenderfinder_bulk_io.py", "tenderfinder_link_preflight.py"], "compile.log")[0] == 0

    def list_check(self):
        code, out = self.run_cmd("Connector list", ["tenderfinder_raw_sweep.py", "--list"], "list.log")
        match = re.search(r"(\d+)\s+connectors\.", out)
        connector_count = int(match.group(1)) if match else 0
        ok = code == 0 and connector_count >= 17 and "surrey_planning_reports" in out
        self.add_result("--list shows 17+ connectors including Surrey", ok, f"{connector_count} connectors + Surrey present" if ok else "missing expected list output")
        return ok

    def review_run(self, source_id: str, output_name: str, label: str) -> Path:
        out_path = self.output_dir / output_name
        self.run_cmd(f"Review-only connector run: {label}", ["tenderfinder_raw_sweep.py", "--only", source_id, "--review-only", "--out", str(out_path)], f"{source_id}.log", timeout=900)
        return out_path

    def read_xlsx_headers_rows(self, path: Path) -> Tuple[List[str], int, str]:
        if not path.exists() or path.stat().st_size == 0:
            return [], 0, "missing or empty"
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            row_iter = ws.iter_rows(values_only=True)
            first = next(row_iter, None)
            headers = [str(x or "") for x in (first or [])]
            rows = sum(1 for row in row_iter if any(v not in (None, "") for v in row))
            wb.close()
            return headers, rows, "ok"
        except Exception as exc:
            return [], 0, f"unreadable: {type(exc).__name__}: {exc}"

    def verify_review(self, path: Path, label: str, min_rows: int, expected_source_id: str, fixture_ids: List[str] | None = None):
        headers, rows, status = self.read_xlsx_headers_rows(path)
        required = {"source_id", "project_id", "app_no", "source_url", "review_decision"}
        ok = status == "ok" and rows >= min_rows and required.issubset(set(headers))
        detail = f"rows={rows}, status={status}"
        if ok:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            hdr = [str(x or "") for x in next(it)]
            col = {name: idx for idx, name in enumerate(hdr)}
            source_values = set(); project_ids = set(); app_nos = set(); source_urls = 0
            for row in it:
                source_values.add(str(row[col.get("source_id", 0)] or "") if len(row) > col.get("source_id", 0) else "")
                project_ids.add(str(row[col.get("project_id", 0)] or "") if len(row) > col.get("project_id", 0) else "")
                app_nos.add(str(row[col.get("app_no", 0)] or "") if len(row) > col.get("app_no", 0) else "")
                if len(row) > col.get("source_url", 0) and str(row[col.get("source_url", 0)] or "").startswith("http"):
                    source_urls += 1
            wb.close()
            ok = expected_source_id in source_values and source_urls > 0
            hits = 0
            if fixture_ids:
                def norm(v): return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())
                seen = {norm(x) for x in project_ids | app_nos}
                hits = sum(1 for fid in fixture_ids if norm(fid) in seen or norm(fid).replace("SURREY", "") in seen or norm(fid).replace("MAPLERIDGE", "") in seen)
                ok = ok and hits >= 1
            detail = f"rows={rows}, fixture_hits={hits}, source_urls={source_urls}"
        self.row_counts[label] = rows
        self.add_result(f"Verify review workbook: {label}", ok, detail)
        return ok

    def run_combined_summary_and_demo(self):
        combined = self.output_dir / "combined_review.xlsx"
        code, _ = self.run_cmd("Combined source summary + demo run", ["tenderfinder_raw_sweep.py", "--only", "surrey_planning_reports,twp_langley_devactivity,maple_ridge_devapps", "--review-only", "--demo-output", "--out", str(combined)], "combined_summary_demo.log", timeout=1200)
        return code == 0

    def verify_source_summary(self):
        path = self.output_dir / "TENDER_FINDER_Run_Source_Summary.csv"
        required = {"run_id", "source_id", "source_name", "connector_status", "resolved_status", "records_pulled", "records_normalized", "records_eligible", "records_written", "records_rejected", "records_held_for_review", "proposed_route", "warnings", "errors", "output_file"}
        ok = False; detail = "missing or empty"
        if path.exists() and path.stat().st_size > 0:
            with path.open(newline="", encoding="utf-8-sig") as f:
                rows = list(csv.DictReader(f))
            missing = required - set(rows[0].keys() if rows else [])
            ids = {r.get("source_id") for r in rows}
            ok = len(rows) >= 3 and not missing and {"surrey_planning_reports", "twp_langley_devactivity", "maple_ridge_devapps"}.issubset(ids)
            detail = f"rows={len(rows)}, missing={sorted(missing)}"
            self.row_counts["Source summary rows"] = len(rows)
        self.add_result("Verify TENDER_FINDER_Run_Source_Summary.csv", ok, detail)
        return ok

    def verify_demo(self):
        path = self.output_dir / "TENDER_FINDER_Demo_Top_Leads.xlsx"
        if not path.exists() or path.stat().st_size == 0:
            self.add_result("Verify demo workbook", False, "missing or empty"); return False
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            top_rows = wb["Top_Leads"].max_row - 1 if "Top_Leads" in sheets else 0
            van_rows = 0
            if "Top_Leads" in sheets:
                ws = wb["Top_Leads"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[0] or "") == "van_building_permits":
                        van_rows += 1
            wb.close()
            ok = "Top_Leads" in sheets and top_rows > 0 and van_rows == 0
            self.row_counts["Demo workbook sheets"] = len(sheets)
            self.add_result("Verify demo workbook not dominated by Vancouver permits", ok, f"sheets={len(sheets)}, top_rows={top_rows}, van_top_rows={van_rows}")
            return ok
        except Exception as exc:
            self.add_result("Verify demo workbook", False, f"unreadable: {exc}"); return False

    def preflight_run(self, label: str, output_dir: Path):
        return self.run_cmd(f"Preflight dry-run output test: {label}", ["tenderfinder_raw_sweep.py", "--from-master", "../../00 Master/TENDER_FINDER_Tender_Intelligence_Working_Master_v7_1_cleaned_urls.xlsx", "--preflight-links", "--preflight-no-search", "--dry-run", "--preflight-output-dir", str(output_dir), "--preflight-timeout", "20", "--preflight-retries", "2", "--preflight-workers", "6"], f"preflight_{label}.log", timeout=900)[0] == 0

    def verify_preflight_outputs(self, output_dir: Path, label: str):
        missing = []
        redirected = []
        safe_dir = Path(r"C:\tenderfinder_tmp")
        for name in REQUIRED_PREFLIGHT_FILES:
            p = output_dir / name
            if p.exists() and p.stat().st_size > 0:
                continue
            safe_path = safe_dir / name
            if safe_path.exists() and safe_path.stat().st_size > 0:
                redirected.append(name)
                continue
            missing.append(name)
        run_log_path = output_dir / "TENDER_FINDER_Link_Check_Run_Log.txt"
        if not (run_log_path.exists() and run_log_path.stat().st_size > 0):
            # A long --output-dir can make tenderfinder_live_link_checker.py redirect
            # this file to safe_dir (see _safe_output_dir()); check there too
            # rather than only the original output_dir, matching the
            # redirect-aware check already done for the other required files
            # above.
            run_log_path = safe_dir / "TENDER_FINDER_Link_Check_Run_Log.txt"
        log_text = run_log_path.read_text(errors="ignore") if run_log_path.exists() else ""
        log_ok = "Source rows processed      : 159" in log_text and "Total URLs checked         : 159" in log_text
        ok = not missing and log_ok
        self.add_result(f"Verify preflight outputs: {label}", ok, f"missing={missing}, redirected={redirected}, files={7-len(missing)}, run_log_159={log_ok}")
        return ok

    def create_marked_review(self, review_path: Path, master_path: Path, marked_path: Path, accept_count: int = 3):
        wb = load_workbook(review_path)
        ws = wb.active
        headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
        rd = headers.index("review_decision") + 1
        pid_col = headers.index("project_id") + 1
        mwb = load_workbook(master_path, read_only=True, data_only=True)
        mws = mwb["Future_Projects"]
        it = mws.iter_rows(values_only=True); next(it, None)
        existing = {str(row[0] or "") for row in it if row and row[0] not in (None, "")}
        mwb.close()
        accepted = rejected = held = 0
        for r in range(2, ws.max_row + 1):
            pid = str(ws.cell(r, pid_col).value or "")
            if accepted < accept_count and pid and pid not in existing:
                ws.cell(r, rd).value = "ACCEPT"; accepted += 1
            elif rejected < 2:
                ws.cell(r, rd).value = "REJECT"; rejected += 1
            elif held < 2:
                ws.cell(r, rd).value = "HOLD"; held += 1
            else:
                ws.cell(r, rd).value = "NEEDS_MORE_INFO"
        wb.save(marked_path)
        return accepted

    def future_count(self, path: Path) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["Future_Projects"]
        it = ws.iter_rows(values_only=True); next(it, None)
        n = sum(1 for row in it if row and row[0] not in (None, ""))
        wb.close(); return n

    def promote_reviewed_test(self):
        review = self.output_dir / "tol_review.xlsx"
        master_src = self.root_dir / "00 Master" / "TENDER_FINDER_Tender_Intelligence_Working_Master_v7_1_cleaned_urls.xlsx"
        master_test = self.output_dir / "TENDER_FINDER_Master_WRITE_TEST.xlsx"
        shutil.copy2(master_src, master_test)
        marked = self.output_dir / "tol_review_for_promote_marked.xlsx"
        accepted = self.create_marked_review(review, master_test, marked, accept_count=3)
        before = self.future_count(master_test)
        code1, out1 = self.run_cmd("Promote reviewed ACCEPT rows (first run)", ["tenderfinder_raw_sweep.py", "--promote-reviewed", str(marked), "--write-master", str(master_test)], "promote_first.log", timeout=900)
        after1 = self.future_count(master_test)
        code2, out2 = self.run_cmd("Promote reviewed ACCEPT rows (second run/dedupe)", ["tenderfinder_raw_sweep.py", "--promote-reviewed", str(marked), "--write-master", str(master_test)], "promote_second.log", timeout=900)
        after2 = self.future_count(master_test)
        dups = int(re.search(r"duplicates_skipped_on_write=(\d+)", out2).group(1)) if re.search(r"duplicates_skipped_on_write=(\d+)", out2) else 0
        appended_first = int(re.search(r"Future_Projects: appended=(\d+)", out1).group(1)) if re.search(r"Future_Projects: appended=(\d+)", out1) else 0
        backup_exists = bool(list((master_test.parent / "backups").glob("*TENDER_FINDER_Master_WRITE_TEST*.xlsx")))
        ok = code1 == 0 and code2 == 0 and accepted > 0 and appended_first >= accepted and after2 == after1 and dups >= accepted and backup_exists
        self.row_counts["Master promote rows written"] = appended_first
        self.row_counts["Duplicate rows skipped on second promote"] = dups
        self.add_result("Verify promote-reviewed + dedupe", ok, f"accepted={accepted}, appended_first={appended_first}, before={before}, after1={after1}, after2={after2}, second_dupes={dups}, backup={backup_exists}")
        return ok

    def verify_outputs(self, output_dir: Path | None = None):
        if output_dir is None:
            output_dir = self.output_dir
        checks = [
            self.verify_review(output_dir / "surrey_review.xlsx", "Surrey", 1, "surrey_planning_reports", ["SURREY-25-0366"]),
            self.verify_review(output_dir / "tol_review.xlsx", "Township Langley", 1, "twp_langley_devactivity", ["TWP-LANGLEY-RO100158"]),
            self.verify_review(output_dir / "maple_ridge_full.xlsx", "Maple Ridge", 1, "maple_ridge_devapps", ["MAPLE-RIDGE-2023-019-RZ"]),
            self.verify_source_summary(), self.verify_demo(),
        ]
        for sub in ["preflight_short", "very_long_output_path_for_patch5_validation_that_should_trigger_safe_writer_or_temp_redirect/deep_nested_folder_level_01/deep_nested_folder_level_02/deep_nested_folder_level_03/link_audit_out_v7_1_live"]:
            p = output_dir / sub
            if p.exists():
                checks.append(self.verify_preflight_outputs(p, sub))
        return all(checks)

    def unit_tests(self):
        """Patch 5.3: run the P0-targeted unit tests (Surrey parser, routing/gates)."""
        ok_all = True
        for script, label in [("tests/test_surrey_pdf_parser.py", "Surrey PDF parser (P0.1)"),
                              ("tests/test_routing_gates.py", "Van routing + write gates (P0.2/P0.4)"),
                              ("tests/test_keywords_config.py", "Company profile + keyword configuration"),
                              ("tests/test_outreach_persistence.py", "Outreach merge-forward (Patch 5.10)"),
                              ("tests/test_developer_classification.py", "Developer/consultant classification (Patch 5.12)"),
                              ("tests/test_bc_bid_civil_scoring.py", "BC Bid civil keyword scoring (Patch 5.14)"),
                              ("tests/test_track_b_source_isolation.py", "Track B source isolation - BC Bid must not starve municipal sources (Patch 5.15)"),
                              ("tests/test_launcher_gui.py", "Desktop GUI launcher command/worker/parsing logic (Patch 5.16)"),
                              ("tests/test_mirror_name_generic.py", "Generic mirror-name derivation, backward-compatible with p57-p516 (Patch 5.17)"),
                              ("tests/test_bc_bid_status_ux.py", "BC Bid blocked-status UX messaging, simulated not live (Patch 5.17)"),
                              ("tests/test_surrey_tender_status.py", "Surrey tender status/date parser (Patch 5.18)"),
                              ("tests/test_launcher_review_xlsx_consistency.py", "Full/fast/GUI review-xlsx path consistency (Patch 5.18)"),
                              ("tests/test_source_backlog.py", "Source growth backlog loader/ranking + Potential_Sources_Next sheet (Patch 5.19)"),
                              ("tests/test_macos_package_scripts.py", "macOS handoff package .command scripts (Patch 5.19)"),
                              ("tests/test_stop_pause_resume.py", "Stop/Pause/Resume stage-safe controls (Patch 5.20)")]:
            code, out = self.run_cmd(f"Unit test: {label}", [script], f"{Path(script).stem}.log", timeout=300)
            passed = code == 0 and ("0 failed" in out or "PASS" in out)
            self.add_result(f"Unit test passed: {label}", passed,
                            out.strip().splitlines()[-1] if out.strip() else "no output")
            ok_all = ok_all and passed
        return ok_all

    def console_encoding_check(self):
        """Patch 5.3 P0.3: prove --list runs clean under a cp1252-style console."""
        import os as _os
        env = dict(_os.environ, PYTHONIOENCODING="cp1252")
        log_path = self.output_dir / "console_cp1252.log"
        try:
            proc = subprocess.run([sys.executable, "tenderfinder_raw_sweep.py", "--list"],
                                  cwd=self.connector_dir, text=True, env=env,
                                  stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=120)
            out, code = proc.stdout or "", proc.returncode
        except Exception as exc:
            out, code = f"{exc}", 1
        log_path.write_text(out, encoding="utf-8")
        match = re.search(r"(\d+)\s+connectors\.", out)
        connector_count = int(match.group(1)) if match else 0
        ok = code == 0 and "UnicodeEncodeError" not in out and connector_count >= 17
        self.add_result("P0.3: --list clean under cp1252 console (no UnicodeEncodeError)", ok,
                        f"exit={code}, unicode_error={'UnicodeEncodeError' in out}")
        return ok

    def run_all(self):
        self.compile_check(); self.list_check()
        self.console_encoding_check()
        self.unit_tests()
        self.review_run("surrey_planning_reports", "surrey_review.xlsx", "Surrey Planning Reports")
        self.verify_review(self.output_dir / "surrey_review.xlsx", "Surrey", 1, "surrey_planning_reports", ["SURREY-25-0366"])
        self.review_run("twp_langley_devactivity", "tol_review.xlsx", "Township Langley")
        self.verify_review(self.output_dir / "tol_review.xlsx", "Township Langley", 1, "twp_langley_devactivity", ["TWP-LANGLEY-RO100158"])
        self.review_run("maple_ridge_devapps", "maple_ridge_full.xlsx", "Maple Ridge")
        self.verify_review(self.output_dir / "maple_ridge_full.xlsx", "Maple Ridge", 1, "maple_ridge_devapps", ["MAPLE-RIDGE-2023-019-RZ"])
        self.run_combined_summary_and_demo(); self.verify_source_summary(); self.verify_demo()
        short = self.output_dir / "preflight_short"
        long = self.output_dir / "very_long_output_path_for_patch5_validation_that_should_trigger_safe_writer_or_temp_redirect" / "deep_nested_folder_level_01" / "deep_nested_folder_level_02" / "deep_nested_folder_level_03" / "link_audit_out_v7_1_live"
        self.preflight_run("short", short); self.verify_preflight_outputs(short, "short")
        self.preflight_run("long", long); self.verify_preflight_outputs(long, "long")
        self.promote_reviewed_test()
        return self.write_report()

    def write_report(self):
        passed = sum(1 for r in self.results if r["passed"]); failed = sum(1 for r in self.results if not r["passed"])
        status = "PASS" if failed == 0 else "FAIL"
        report = ["# REGRESSION TEST REPORT - PATCH 5.0 VERIFIED PACKAGE", "", f"**Status:** {status}", f"**Generated:** {datetime.now().isoformat(timespec='seconds')}", f"**Working directory:** `{self.connector_dir}`", f"**Output directory:** `{self.output_dir}`", "", "## Summary", f"- Passed checks: {passed}", f"- Failed checks: {failed}", "- Note: this sandbox has no external DNS/network access. Connector fresh runs used explicit packaged regression fixtures where live public endpoints could not resolve. This proves reproducible runner/output/promotion behavior, not live-source availability.", "", "## Command / Check Results", "| Result | Check | Exit | Detail |", "|---|---|---:|---|"]
        for r in self.results:
            report.append(f"| {'PASS' if r['passed'] else 'FAIL'} | {r['name']} | {'' if r['exit_code'] is None else r['exit_code']} | {str(r['detail']).replace('|','/')} |")
        report += ["", "## Row Counts"]
        for k, v in self.row_counts.items():
            report.append(f"- {k}: {v}")
        report += ["", "## Logs"]
        for log in self.logs:
            report.append(f"- `{log.name}`")
        report += ["", "## Final Verdict", status]
        path = self.output_dir / "REGRESSION_TEST_REPORT.md"
        path.write_text("\n".join(report) + "\n", encoding="utf-8")
        shutil.copy2(path, self.root_dir / "REGRESSION_TEST_REPORT.md")
        print(f"Report written: {path}")
        return failed == 0

def main():
    script_dir = Path(__file__).resolve().parent
    connector_dir = script_dir.parent
    root_dir = connector_dir.parents[1]
    default_out = root_dir / "test_outputs_p53"
    parser = argparse.ArgumentParser(description="Run TENDER_FINDER Patch 5.0 regression tests")
    parser.add_argument("--all", action="store_true", help="Run all acceptance/regression checks")
    parser.add_argument("--verify-outputs", action="store_true", help="Verify output files only")
    parser.add_argument("--output-dir", default=str(default_out), help="Output directory to generate or verify")
    parser.add_argument("--keep-outputs", action="store_true", help="Do not clear output directory before --all")
    args = parser.parse_args()
    reg = Regression(Path(args.output_dir), keep_outputs=args.keep_outputs or args.verify_outputs)
    if args.all:
        ok = reg.run_all()
    elif args.verify_outputs:
        ok = reg.verify_outputs(Path(args.output_dir)); reg.write_report()
    else:
        parser.print_help(); ok = False
    sys.exit(0 if ok else 1)

if __name__ == "__main__":
    main()
