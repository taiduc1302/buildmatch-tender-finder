from __future__ import annotations

import hashlib
import os
import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

CONNECTOR_DIR = Path(__file__).resolve().parents[1]
PACKAGE_ROOT = CONNECTOR_DIR.parents[1]
sys.path.insert(0, str(CONNECTOR_DIR))

import tenderfinder_guards as guards  # noqa: E402
import tenderfinder_launcher_gui as gui  # noqa: E402
from tenderfinder_keywords_config import (  # noqa: E402
    CONFIG_ENV_VAR,
    KEYWORDS_SHEET,
    KeywordConfigError,
    KeywordRule,
    clear_keywords_cache,
    inspect_last_known_good,
    keyword_state_paths,
    load_keywords_config,
)


LIVE = PACKAGE_ROOT / "config" / "keywords.xlsx"
TEMPLATE = PACKAGE_ROOT / "config" / "keywords_template.xlsx"


def _minimal_workbook(path: Path, keyword_rows: list[list[object]]) -> None:
    wb = Workbook()
    profile = wb.active
    profile.title = "Company_Profile"
    profile.append(["company_name", "company_description", "regions", "work_types", "known_clients"])
    profile.append(["Example Co", "Civil contractor", "Surrey", "earthwork", "Example Owner"])
    keywords = wb.create_sheet("Keywords")
    keywords.append(["keyword", "match_type", "weight", "category", "explanation", "active", "param"])
    for row in keyword_rows:
        keywords.append(row)
    instructions = wb.create_sheet("Instructions")
    instructions.append(["Read me"])
    wb.save(path)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_live_and_template_validate() -> None:
    live = load_keywords_config(LIVE, force_reload=True)
    template = load_keywords_config(TEMPLATE, force_reload=True, allow_empty_profile=True)
    assert live.company_name == "Meridian Civil Works"
    assert live.active_keyword_count == 227
    assert template.active_keyword_count == 0
    assert {"Company_Profile", "Keywords", "Instructions"}.issubset(load_workbook(LIVE, read_only=True).sheetnames)
    for path in (LIVE, TEMPLATE):
        wb = load_workbook(path)
        assert wb["Company_Profile"].freeze_panes == "A2"
        assert wb["Keywords"].freeze_panes == "A2"
        assert len(wb["Keywords"].data_validations.dataValidation) == 3
        assert all(wb["Company_Profile"].cell(1, column).comment for column in range(1, 6))
        wb.close()


def test_match_types() -> None:
    assert KeywordRule("road", "contains", 0, "gate_include").matches("Broad Roadwork")
    assert not KeywordRule("road", "exact", 0, "gate_include").matches("roadwork")
    assert KeywordRule(" road ", "exact", 0, "gate_include").matches("ROAD")
    assert KeywordRule(r"\brfps?\b", "regex", 0, "tender_match").matches("New RFP posted")


def test_exact_scoring_matches_preserved_fields() -> None:
    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "exact.xlsx"
        _minimal_workbook(
            path,
            [["Precision Field Title", "exact", 17, "positive", "field-aware exact", "Y", ""]],
        )
        old_config = os.environ.get(CONFIG_ENV_VAR)
        os.environ[CONFIG_ENV_VAR] = str(path)
        clear_keywords_cache()
        try:
            assert guards.score_civil_fit(
                "Prefix Precision Field Title suffix",
                "Cedar Harbour",
                match_fields=["Prefix", "Precision Field Title", "suffix"],
            ) == 52
            assert guards.score_civil_fit(
                "Prefix Precision Field Title suffix",
                "Cedar Harbour",
            ) == 35
        finally:
            if old_config is None:
                os.environ.pop(CONFIG_ENV_VAR, None)
            else:
                os.environ[CONFIG_ENV_VAR] = old_config
            clear_keywords_cache()


def test_inactive_rule_is_ignored() -> None:
    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "inactive.xlsx"
        _minimal_workbook(
            path,
            [["trenching", "contains", 9, "positive", "disabled", "N", ""]],
        )
        config = load_keywords_config(path, force_reload=True)
        assert not config.matches_any("positive", "trenching and excavation")
        assert all(rule.keyword != "trenching" for rule in config.rules_for("positive"))


def test_tybo_scoring_compatibility() -> None:
    assert guards.score_civil_fit("subdivision servicing water main storm sanitary", "Surrey") == 88
    assert guards.score_civil_fit("interior tenant improvement kitchen renovation", "Burnaby") == 0
    assert guards.score_civil_fit("earthwork for Polygon", "Pitt Meadows") == 58
    assert guards.score_civil_fit("civil roadwork grading earthwork subdivision servicing", "Surrey") == 100


def test_keywords_row_overrides_profile_rule() -> None:
    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "override.xlsx"
        _minimal_workbook(
            path,
            [["earthwork", "contains", 13, "positive", "Explicit override", "Y", ""]],
        )
        config = load_keywords_config(path, force_reload=True)
        rule = next(rule for rule in config.rules_for("positive") if rule.keyword == "earthwork")
        assert rule.weight == 13
        assert rule.source == KEYWORDS_SHEET


def test_duplicate_and_bad_regex_are_hard_errors() -> None:
    with tempfile.TemporaryDirectory() as temp:
        folder = Path(temp)
        state_root = folder / "state"
        duplicate = folder / "duplicate.xlsx"
        _minimal_workbook(
            duplicate,
            [
                ["trench", "contains", 9, "positive", "one", "Y", ""],
                ["TRENCH", "contains", 9, "positive", "two", "Y", ""],
            ],
        )
        try:
            load_keywords_config(duplicate, force_reload=True, state_root=state_root)
        except KeywordConfigError as exc:
            assert "duplicate (keyword, category)" in str(exc)
        else:
            raise AssertionError("duplicate keyword/category pair did not stop validation")

        bad_regex = folder / "bad_regex.xlsx"
        _minimal_workbook(
            bad_regex,
            [["(", "regex", 0, "tender_match", "broken", "Y", ""]],
        )
        try:
            load_keywords_config(bad_regex, force_reload=True, state_root=state_root)
        except KeywordConfigError as exc:
            assert "invalid regex" in str(exc)
            report = keyword_state_paths(
                root=PACKAGE_ROOT,
                state_root=state_root,
            ).validation_report
            assert report.exists()
            assert PACKAGE_ROOT not in report.parents
        else:
            raise AssertionError("invalid regex did not stop validation")


def test_malformed_weight_and_active_report_sheet_row() -> None:
    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "malformed.xlsx"
        _minimal_workbook(
            path,
            [["trench", "contains", 9.5, "positive", "broken", "maybe", ""]],
        )
        try:
            load_keywords_config(path, force_reload=True)
        except KeywordConfigError as exc:
            message = str(exc)
            assert "Sheet Keywords row 2: weight must be a whole number" in message
            assert "Sheet Keywords row 2: active must be Y or N" in message
        else:
            raise AssertionError("malformed weight/active values did not stop validation")


def test_missing_custom_workbook_has_no_implicit_fallback() -> None:
    with tempfile.TemporaryDirectory() as temp:
        missing = Path(temp) / "keywords.xlsx"
        try:
            load_keywords_config(missing, force_reload=True)
        except KeywordConfigError as exc:
            message = str(exc)
            assert "Workbook is missing" in message
            assert "keywords_template.xlsx" in message
            assert "as keywords.xlsx" in message
        else:
            raise AssertionError("missing workbook unexpectedly loaded")


def test_last_known_good_fallback_is_external_verified_and_visible() -> None:
    with tempfile.TemporaryDirectory() as temp:
        folder = Path(temp)
        canonical = folder / "keywords.xlsx"
        state_root = folder / "state"
        shutil.copy2(LIVE, canonical)
        clear_keywords_cache()

        first = load_keywords_config(
            canonical,
            root=PACKAGE_ROOT,
            state_root=state_root,
            force_reload=True,
            allow_last_known_good=True,
        )
        assert first.source_kind == "canonical"
        assert first.last_known_good_status == "ready"
        assert first.last_known_good_path is not None
        assert first.last_known_good_path.exists()
        assert PACKAGE_ROOT not in first.last_known_good_path.parents
        assert first.active_keyword_count == 227

        canonical.unlink()
        fallback = load_keywords_config(
            canonical,
            root=PACKAGE_ROOT,
            state_root=state_root,
            force_reload=True,
            allow_last_known_good=True,
        )
        assert fallback.source_kind == "last_known_good"
        assert fallback.requested_path == canonical.resolve()
        assert fallback.path == first.last_known_good_path
        assert fallback.last_known_good_status == "in_use"
        assert fallback.active_keyword_count == 227
        assert any("Workbook is missing" in error for error in fallback.validation_errors)

        status = inspect_last_known_good(
            canonical,
            root=PACKAGE_ROOT,
            state_root=state_root,
        )
        assert status["status"] == "ready"
        assert status["saved_at"]


def test_cache_is_stable_until_explicit_reload() -> None:
    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "keywords.xlsx"
        shutil.copy2(LIVE, path)
        clear_keywords_cache()
        first = load_keywords_config(path)
        second = load_keywords_config(path)
        assert first is second
        wb = load_workbook(path)
        wb["Company_Profile"]["A2"] = "Changed Company"
        wb.save(path)
        os.utime(path, None)
        third = load_keywords_config(path)
        assert third is first
        refreshed = load_keywords_config(path, force_reload=True)
        assert refreshed is not first
        assert refreshed.company_name == "Changed Company"


def test_gui_helpers_are_headless_and_read_live_config() -> None:
    result = gui.validate_keywords_for_gui(PACKAGE_ROOT, force_reload=True)
    assert result["company_name"] == "Meridian Civil Works"
    assert result["active_keyword_count"] == 227
    assert result["inactive_keyword_count"] == 0
    assert len(result["category_counts"]) == 12
    assert result["source_kind"] == "canonical"
    assert result["last_known_good_status"] == "ready"
    assert result["validation_errors"] == []
    assert result["rescore_semantics"].startswith("Scores, gates, labels, and bucket routing always reflect current keywords.xlsx")
    assert "legacy_vancouver_scoring_text_unavailable" in result["rescore_exceptions"]
    assert "tenderfinder_agent2.py remains isolated/static" in result["rescore_exceptions"]
    assert gui.keywords_folder(PACKAGE_ROOT) == PACKAGE_ROOT / "config"


def test_validation_does_not_modify_live_workbook() -> None:
    before = _sha256(LIVE)
    load_keywords_config(LIVE, force_reload=True)
    after = _sha256(LIVE)
    assert after == before


def main() -> int:
    tests = [
        test_live_and_template_validate,
        test_match_types,
        test_exact_scoring_matches_preserved_fields,
        test_inactive_rule_is_ignored,
        test_tybo_scoring_compatibility,
        test_keywords_row_overrides_profile_rule,
        test_duplicate_and_bad_regex_are_hard_errors,
        test_malformed_weight_and_active_report_sheet_row,
        test_missing_custom_workbook_has_no_implicit_fallback,
        test_last_known_good_fallback_is_external_verified_and_visible,
        test_cache_is_stable_until_explicit_reload,
        test_gui_helpers_are_headless_and_read_live_config,
        test_validation_does_not_modify_live_workbook,
    ]
    for test in tests:
        test()
        print(f"[PASS] {test.__name__}")
    print(f"Keyword configuration tests: {len(tests)} passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
