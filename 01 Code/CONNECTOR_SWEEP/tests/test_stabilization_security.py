from __future__ import annotations

import datetime as dt
import csv
import sys
import tempfile
from email.message import Message
from pathlib import Path

import openpyxl


CODE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = CODE_DIR.parents[1]
sys.path.insert(0, str(CODE_DIR))

import tenderfinder_demo_three_buckets as demo  # noqa: E402
import tenderfinder_raw_sweep as raw_sweep  # noqa: E402
from tenderfinder_email_intake import EmailTender, write_outputs as write_email_outputs  # noqa: E402
from tenderfinder_excel_safety import (  # noqa: E402
    append_untrusted_row,
    safe_untrusted_excel_value,
)
from tenderfinder_url_safety import (  # noqa: E402
    URLSafetyError,
    install_playwright_public_route,
    safe_requests_request,
    safe_urllib_fetch,
    validate_public_url,
    validate_public_url_syntax,
)


PUBLIC_IPV4 = "93.184.216.34"


def public_resolver(host: str, port: int, **_kwargs):
    return [(2, 1, 6, "", (PUBLIC_IPV4, port))]


def mixed_resolver(host: str, port: int, **_kwargs):
    return [
        (2, 1, 6, "", (PUBLIC_IPV4, port)),
        (2, 1, 6, "", ("127.0.0.1", port)),
    ]


def test_excel_formula_injection_is_neutralized() -> None:
    hostile = [
        '=HYPERLINK("https://example.com","click")',
        "+SUM(1,2)",
        "-1+2",
        "@SUM(1,2)",
        "   =SUM(1,2)",
        "\t@SUM(1,2)",
    ]
    ordinary = ["ordinary text", "https://example.com"]
    real_date = dt.date(2026, 7, 15)

    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "formula_safety.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        append_untrusted_row(ws, [*hostile, *ordinary, 42, real_date])
        # Intentional application formulas bypass the untrusted writer.
        ws.cell(2, 1).value = "=SUM(20,22)"
        wb.save(path)
        wb.close()

        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb.active
        for column, original in enumerate(hostile, start=1):
            cell = ws.cell(1, column)
            assert cell.data_type != "f", (original, cell.value, cell.data_type)
            assert str(cell.value).startswith("'"), (original, cell.value)
        assert ws.cell(1, 7).value == ordinary[0]
        assert ws.cell(1, 8).value == ordinary[1]
        assert ws.cell(1, 9).value == 42 and ws.cell(1, 9).data_type == "n"
        assert ws.cell(1, 10).data_type == "d"
        assert ws.cell(2, 1).data_type == "f"
        wb.close()


def test_runtime_workbook_writer_uses_central_guard() -> None:
    with tempfile.TemporaryDirectory() as temp:
        path = Path(temp) / "runtime_writer.xlsx"
        wb = openpyxl.Workbook()
        demo.write_rows(
            wb.active,
            ["title", "score", "found"],
            [["=HYPERLINK(\"https://example.com\",\"click\")", 17, dt.date(2026, 7, 15)]],
        )
        wb.save(path)
        wb.close()
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb.active
        assert ws["A2"].data_type != "f"
        assert ws["A2"].value.startswith("'")
        assert ws["B2"].data_type == "n"
        assert ws["C2"].data_type == "d"
        wb.close()


def test_csv_export_uses_central_guard() -> None:
    hostile = '=HYPERLINK("https://example.com","click")'
    with tempfile.TemporaryDirectory() as temp:
        output_dir = Path(temp)
        row = EmailTender(
            provider="fixture",
            source="fixture",
            issuer="Example issuer",
            tender_title=hostile,
            civil_relevant=True,
            status="open",
            closing_date="2026-08-01",
            contact_email="",
            contact_phone="",
            url="https://example.com/tender",
            found_via="fixture",
            snippet="ordinary text",
            message_id="fixture-1",
            received_date="2026-07-15",
        )
        write_email_outputs([row], output_dir)
        csv_path = output_dir / "BID_NOW_from_email_alerts.csv"

        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            exported = next(csv.DictReader(handle))
        assert exported["tender_title"].startswith("'")
        assert exported["civil_relevant"] == "True"


def test_private_and_malformed_urls_are_rejected_without_network() -> None:
    blocked = [
        "http://127.0.0.1",
        "http://localhost",
        "http://10.0.0.1",
        "http://172.16.0.1",
        "http://192.168.1.1",
        "http://169.254.169.254/latest/meta-data",
        "http://[::1]",
        "http://0.0.0.0",
        "file:///etc/passwd",
        "ftp://example.com/file",
        r"\\server\share\file",
        "//example.com/path",
        "not a URL",
        "https://user:password@example.com/path",
        "https://example.com\\@127.0.0.1/",
    ]
    for value in blocked:
        try:
            validate_public_url_syntax(value)
        except URLSafetyError:
            pass
        else:
            raise AssertionError(f"unsafe URL was accepted: {value}")


def test_dns_resolution_is_fail_closed() -> None:
    valid = validate_public_url(
        "https://public.example.org/opportunities",
        resolver=public_resolver,
    )
    assert valid.addresses == (PUBLIC_IPV4,)
    assert valid.url == "https://public.example.org/opportunities"

    try:
        validate_public_url("https://mixed.example.org/", resolver=mixed_resolver)
    except URLSafetyError as exc:
        assert "non-public" in str(exc)
    else:
        raise AssertionError("mixed public/private DNS response was accepted")

    def failing_resolver(_host: str, _port: int, **_kwargs):
        raise OSError("offline resolution failure")

    try:
        validate_public_url("https://unresolved.example.org/", resolver=failing_resolver)
    except URLSafetyError as exc:
        assert "resolution failed" in str(exc)
    else:
        raise AssertionError("unresolved hostname was accepted")


def test_no_fetch_address_enrichment_never_calls_network_adapter() -> None:
    original = demo.arcgis_query_paged
    calls: list[str] = []

    def forbidden(layer_url: str, *_args, **_kwargs):
        calls.append(layer_url)
        raise AssertionError("network adapter was called in no-fetch enrichment")

    demo.arcgis_query_paged = forbidden
    try:
        result = demo.load_address_enrichment(allow_network=False)
    finally:
        demo.arcgis_query_paged = original

    assert calls == []
    assert result == {"surrey_devapps_v2": {}, "abbotsford_devapps": {}}


def test_safe_http_metadata_is_preserved_for_live_evidence_without_extra_request() -> None:
    original = raw_sweep.safe_requests_request
    calls: list[str] = []

    class Response:
        status_code = 200
        text = '{"features": []}'

        def raise_for_status(self):
            return None

    def fake_safe_request(_method, url, **_kwargs):
        from tenderfinder_url_safety import SafeRequestsResult

        calls.append(url)
        return SafeRequestsResult(
            Response(),
            "https://public.example.org/final?f=json",
            (url, "https://public.example.org/final?f=json"),
        )

    raw_sweep.safe_requests_request = fake_safe_request
    try:
        audit = {}
        body = raw_sweep.http_get("https://public.example.org/start", audit=audit)
    finally:
        raw_sweep.safe_requests_request = original

    assert body == '{"features": []}'
    assert calls == ["https://public.example.org/start"]
    assert audit == {
        "http_status": 200,
        "final_validated_url": "https://public.example.org/final?f=json",
        "redirect_chain": [
            "https://public.example.org/start",
            "https://public.example.org/final?f=json",
        ],
    }


def test_arcgis_query_honors_small_controlled_limit_at_request_time() -> None:
    original = raw_sweep.http_get_json
    requested_urls: list[str] = []

    def fake_get_json(url, timeout=90, audit=None):
        requested_urls.append(url)
        return {"features": [{"attributes": {"OBJECTID": index}} for index in range(5)]}

    raw_sweep.http_get_json = fake_get_json
    try:
        rows = raw_sweep.query_arcgis_layer(
            "https://public.example.org/FeatureServer/0",
            max_records=5,
            page=250,
            pause=0,
            max_retries=1,
            where="STATUS <> 'Concluded'",
            order_by="OBJECTID DESC",
        )
    finally:
        raw_sweep.http_get_json = original

    assert len(rows) == 5
    assert len(requested_urls) == 1
    assert "resultRecordCount=5" in requested_urls[0]
    assert "where=STATUS+%3C%3E+%27Concluded%27" in requested_urls[0]
    assert "orderByFields=OBJECTID+DESC" in requested_urls[0]


class FakeRequestsResponse:
    def __init__(self, status: int, url: str, location: str = ""):
        self.status_code = status
        self.url = url
        self.headers = {"Location": location} if location else {}
        self.closed = False

    def close(self) -> None:
        self.closed = True


def test_requests_redirect_to_private_is_rejected_before_second_request() -> None:
    calls: list[str] = []

    def request(_method: str, url: str, **_kwargs):
        calls.append(url)
        return FakeRequestsResponse(302, url, "http://127.0.0.1/private")

    try:
        safe_requests_request(
            "GET",
            "https://public.example.org/start",
            request_callable=request,
            resolver=public_resolver,
        )
    except URLSafetyError as exc:
        assert "non-public" in str(exc)
    else:
        raise AssertionError("private redirect was followed")
    assert calls == ["https://public.example.org/start"]


class FakeUrllibResponse:
    status = 302

    def __init__(self):
        self.headers = Message()
        self.headers["Location"] = "http://10.0.0.1/private"

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def geturl(self) -> str:
        return "https://public.example.org/start"

    def read(self, _limit: int) -> bytes:
        return b""


class FakeOpener:
    def __init__(self):
        self.calls = 0

    def open(self, _request, timeout: float):
        self.calls += 1
        return FakeUrllibResponse()


def test_urllib_redirect_to_private_is_rejected_before_second_request() -> None:
    opener = FakeOpener()
    try:
        safe_urllib_fetch(
            "https://public.example.org/start",
            opener=opener,
            resolver=public_resolver,
        )
    except URLSafetyError as exc:
        assert "non-public" in str(exc)
    else:
        raise AssertionError("private redirect was followed")
    assert opener.calls == 1


def test_playwright_route_blocks_private_subresources() -> None:
    class Context:
        handler = None

        def route(self, pattern, handler):
            assert pattern == "**/*"
            self.handler = handler

    class Route:
        continued = False
        aborted = False

        def continue_(self):
            self.continued = True

        def abort(self, reason):
            assert reason == "blockedbyclient"
            self.aborted = True

    class Request:
        def __init__(self, url):
            self.url = url

    blocked = []
    context = Context()
    install_playwright_public_route(
        context,
        resolver=public_resolver,
        on_blocked=lambda url, reason: blocked.append((url, reason)),
    )
    public_route = Route()
    context.handler(public_route, Request("https://public.example.org/app.js"))
    assert public_route.continued and not public_route.aborted

    private_route = Route()
    context.handler(private_route, Request("http://169.254.169.254/latest/meta-data"))
    assert private_route.aborted and not private_route.continued
    assert blocked and "non-public" in blocked[0][1]

    data_route = Route()
    context.handler(data_route, Request("data:text/plain,hello"))
    assert data_route.continued and not data_route.aborted


def main() -> int:
    tests = [
        test_excel_formula_injection_is_neutralized,
        test_runtime_workbook_writer_uses_central_guard,
        test_csv_export_uses_central_guard,
        test_private_and_malformed_urls_are_rejected_without_network,
        test_dns_resolution_is_fail_closed,
        test_no_fetch_address_enrichment_never_calls_network_adapter,
        test_safe_http_metadata_is_preserved_for_live_evidence_without_extra_request,
        test_arcgis_query_honors_small_controlled_limit_at_request_time,
        test_requests_redirect_to_private_is_rejected_before_second_request,
        test_urllib_redirect_to_private_is_rejected_before_second_request,
        test_playwright_route_blocks_private_subresources,
    ]
    for test in tests:
        test()
        print(f"[PASS] {test.__name__}")
    print(f"Stabilization security tests: {len(tests)} passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
