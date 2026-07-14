from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from tenderfinder_demo_three_buckets import (  # noqa: E402
    OUTREACH_TRACKER_HEADERS,
    _is_fixture_or_example_row,
    build_user_facing_master_workbook,
    build_user_outreach_tracker_rows,
    count_live_demo_rows,
    protected_sha_status,
    select_user_future_projects_rows,
)
from tenderfinder_email_intake import parse_eml_dir  # noqa: E402


FUTURE_HEADERS = [
    "lead_id", "first_seen_date", "is_new_since_last_run", "municipality",
    "regional_cluster", "app_no", "address", "address_source",
    "approximate_location", "app_type_stage", "status_class", "scope_summary",
    "fit_score", "signal_quality", "detail_available", "source_tier",
    "project_scale_tier", "scale", "est_units", "est_storeys", "est_lots",
    "applicant_name", "source", "source_url",
]


def _future_row(lead_id: str, address: str, scope: str, source_url: str, fit_score: int) -> list[object]:
    values = {
        "lead_id": lead_id,
        "first_seen_date": "2026-07-14",
        "is_new_since_last_run": "YES",
        "municipality": "Surrey",
        "regional_cluster": "Fraser Valley",
        "app_no": lead_id,
        "address": address,
        "address_source": "application",
        "approximate_location": address,
        "app_type_stage": "Development Permit",
        "status_class": "OPEN",
        "scope_summary": scope,
        "fit_score": fit_score,
        "signal_quality": "HIGH",
        "detail_available": "YES",
        "source_tier": "1",
        "project_scale_tier": "MEDIUM",
        "scale": "MEDIUM",
        "applicant_name": "Clean Civil Developer",
        "source": "Municipal Development Feed",
        "source_url": source_url,
    }
    return [values.get(header) for header in FUTURE_HEADERS]


def _mixed_demo_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BID_LATER_Future_Projects"
    ws.append(["Mixed clean/fixture test rows"])
    ws.append(FUTURE_HEADERS)
    ws.append(_future_row(
        "CLEAN-001", "100 Main Street", "Watermain and drainage servicing",
        "https://surrey.ca/development/clean-001", 90,
    ))
    ws.append(_future_row(
        "FIXTURE-001", "200 Example Avenue", "Synthetic fixture civil project",
        "https://example.com/development/fixture-001", 95,
    ))
    return wb


def _minimal_master(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "README_START_HERE"
    for sheet_name in ("Dashboard", "Active_Tenders", "Future_Projects", "Weekly_Review_Log"):
        wb.create_sheet(sheet_name)
    wb.save(path)


def test_protected_sha_status_missing() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        status, sha = protected_sha_status(Path(tmp) / "missing.xlsx", "abc123")
        assert status == "NOT_AVAILABLE"
        assert sha == "NOT_AVAILABLE"


def test_email_fixture_titles_are_clean() -> None:
    fixtures = Path(__file__).resolve().parent / "fixtures" / "email_alerts"
    rows = parse_eml_dir(fixtures)
    titles = {row.tender_title for row in rows}
    assert "Fraser Valley Utility Corridor Upgrade" in titles
    assert "Kelowna Pump Station Upgrade" in titles
    assert "Drainage Culvert Replacement" in titles
    assert all("Daily Matches" not in title for title in titles)
    assert all("Daily Notification" not in title for title in titles)
    assert all("BC Bid Alert" not in title for title in titles)


def test_tracking_links_are_labeled_when_unwrapped() -> None:
    fixtures = Path(__file__).resolve().parent / "fixtures" / "email_alerts"
    rows = parse_eml_dir(fixtures)
    tracking_row = next(row for row in rows if row.tender_title == "Fraser Valley Utility Corridor Upgrade")
    assert tracking_row.url == "https://example.com/tenders/fvrd-utility-corridor"
    assert tracking_row.found_via == "email_tracking_link"


def test_source_text_has_no_stale_patch_57_talktrack_or_old_email_summary() -> None:
    source = (ROOT / "tenderfinder_demo_three_buckets.py").read_text(encoding="utf-8")
    assert "Patch 5.7 built and deployed the email alert intake module." not in source
    assert "Email intake kept in BID NOW:" not in source
    assert '["x] Protected master SHA values unchanged' not in source


def test_future_selector_and_outreach_exclude_fixture_rows() -> None:
    demo_wb = _mixed_demo_workbook()

    future_rows, future_dicts, future_total_live = select_user_future_projects_rows(
        demo_wb, {}, "2026-07-14T12:00:00",
    )

    assert len(future_rows) == 1
    assert len(future_dicts) == 1
    assert future_total_live == 1
    assert future_dicts[0]["Lead ID"] == "CLEAN-001"
    assert not any(_is_fixture_or_example_row(row) for row in future_dicts)

    outreach_rows = build_user_outreach_tracker_rows(
        demo_wb, [], future_dicts, {}, "2026-07-14T12:00:00",
    )
    outreach_dicts = [dict(zip(OUTREACH_TRACKER_HEADERS, row)) for row in outreach_rows]
    assert len(outreach_dicts) == 1
    assert outreach_dicts[0]["lead_id"] == "CLEAN-001"
    assert not any(_is_fixture_or_example_row(row) for row in outreach_dicts)


def test_dashboard_future_full_matches_non_fixture_validator_recount() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        demo_path = tmp_path / "demo.xlsx"
        master_path = tmp_path / "master.xlsx"
        user_master_path = tmp_path / "user-master.xlsx"

        demo_wb = _mixed_demo_workbook()
        demo_wb.save(demo_path)
        _minimal_master(master_path)

        stats = build_user_facing_master_workbook(
            demo_path, master_path, user_master_path, None,
        )

        validator_demo_wb = openpyxl.load_workbook(demo_path)
        validator_recount = count_live_demo_rows(
            validator_demo_wb, "BID_LATER_Future_Projects",
        )
        validator_demo_wb.close()

        result_wb = openpyxl.load_workbook(user_master_path)
        dashboard_ws = result_wb["Dashboard"]
        dashboard_future_full = next(
            row[1].value
            for row in dashboard_ws.iter_rows(min_row=1, max_row=dashboard_ws.max_row)
            if row[0].value == "Full future-project records in demo"
        )
        future_rows = [
            row for row in result_wb["Future_Projects"].iter_rows(min_row=2, values_only=True)
            if any(value not in (None, "") for value in row)
        ]
        result_wb.close()

        assert validator_recount == 1
        assert stats["dashboard_counts"]["future_full"] == validator_recount
        assert dashboard_future_full == validator_recount
        assert len(future_rows) == 1


def main() -> int:
    tests = [
        test_protected_sha_status_missing,
        test_email_fixture_titles_are_clean,
        test_tracking_links_are_labeled_when_unwrapped,
        test_source_text_has_no_stale_patch_57_talktrack_or_old_email_summary,
        test_future_selector_and_outreach_exclude_fixture_rows,
        test_dashboard_future_full_matches_non_fixture_validator_recount,
    ]
    for test in tests:
        test()
        print(f"[PASS] {test.__name__}")
    print("Patch 5.23 output consistency tests: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
