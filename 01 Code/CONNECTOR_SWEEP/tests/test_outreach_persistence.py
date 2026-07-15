from __future__ import annotations

import tempfile
from pathlib import Path
import sys

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from tenderfinder_demo_three_buckets import (  # noqa: E402
    DemoData,
    apply_run_history,
    build_outreach_rows,
    build_user_outreach_tracker_rows,
    OUTREACH_TRACKER_HEADERS,
    stable_lead_id,
)


def make_row(municipality: str, app_no: str, address: str, fit: int = 70) -> dict:
    row = {
        "municipality": municipality,
        "app_no": app_no,
        "address": address,
        "source": "Unit Test Source",
        "source_id": "unit_test_source",
        "fit_score": fit,
        "signal_quality": "HIGH",
        "detail_available": "YES",
        "scope_summary": "Unit test civil lead",
    }
    row["lead_id"] = stable_lead_id(row)
    return row


def write_prior(path: Path, carried_row: dict, dropped_row: dict) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BID_LATER_Future_Projects"
    ws.append(["lead_id", "first_seen_date", "municipality", "app_no", "address", "source", "source_id"])
    ws.append([
        carried_row["lead_id"],
        "2026-06-01",
        carried_row["municipality"],
        carried_row["app_no"],
        carried_row["address"],
        carried_row["source"],
        carried_row["source_id"],
    ])
    ws = wb.create_sheet("Outreach_Tracker")
    ws.append(["lead_id", "type", "municipality", "title_or_address", "fit_score", "signal_quality", "contact", "status", "last_contact_date", "next_action_date", "notes"])
    ws.append([
        carried_row["lead_id"],
        "LEAD",
        carried_row["municipality"],
        carried_row["address"],
        carried_row["fit_score"],
        carried_row["signal_quality"],
        "Planning desk",
        "Contacted",
        "2026-06-15",
        "2026-07-01",
        "Left voicemail with estimator.",
    ])
    ws.append([
        dropped_row["lead_id"],
        "LEAD",
        dropped_row["municipality"],
        dropped_row["address"],
        dropped_row["fit_score"],
        dropped_row["signal_quality"],
        "Project manager",
        "Meeting Scheduled",
        "2026-06-20",
        "2026-07-10",
        "Keep this tracked even if keywords move it below gate.",
    ])
    wb.save(path)


def main() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        history_dir = Path(tmp)
        carried = make_row("Coquitlam", "TEST-1", "100 Test St")
        new = make_row("Coquitlam", "TEST-2", "200 Test St")
        dropped = make_row("Coquitlam", "TEST-3", "300 Test St", fit=45)
        write_prior(history_dir / "demo_p510_20260601_000000.xlsx", carried, dropped)

        data = DemoData()
        data.future = [carried, new]
        apply_run_history(data, history_dir)
        rows = build_outreach_rows(data, [])
        by_id = {row["lead_id"]: row for row in rows}

        assert by_id[carried["lead_id"]]["status"] == "Contacted"
        assert by_id[carried["lead_id"]]["notes"] == "Left voicemail with estimator."
        assert by_id[carried["lead_id"]]["last_contact_date"] == "2026-06-15"
        assert by_id[new["lead_id"]]["status"] == "Not Contacted"
        assert by_id[dropped["lead_id"]]["status"] == "Meeting Scheduled"
        assert by_id[dropped["lead_id"]]["signal_quality"] == "BELOW_CURRENT_GATE"
        assert by_id[dropped["lead_id"]]["contact"] == "Project manager"
        assert carried["is_new_since_last_run"] == "NO"
        assert new["is_new_since_last_run"] == "YES"

        prior_user_lookup = {
            dropped["lead_id"]: {
                "type": "LEAD",
                "municipality": dropped["municipality"],
                "title_or_address": dropped["address"],
                "fit_score": dropped["fit_score"],
                "signal_quality": "HIGH",
                "contact": "Project manager",
                "status": "Meeting Scheduled",
                "last_contact_date": "2026-06-20",
                "next_action_date": "2026-07-10",
                "notes": "Keep this tracked even if keywords move it below gate.",
                "source": "Municipal planning feed",
                "source_url": "https://planning.coquitlam.ca/records/TEST-3",
            }
        }
        slim_values = build_user_outreach_tracker_rows(
            openpyxl.Workbook(), [], [], prior_user_lookup, "2026-07-14"
        )
        slim_row = dict(zip(OUTREACH_TRACKER_HEADERS, slim_values[0]))
        assert slim_row["lead_id"] == dropped["lead_id"]
        assert slim_row["status"] == "Meeting Scheduled"
        assert slim_row["signal_quality"] == "BELOW_CURRENT_GATE"
        assert slim_row["source_sheet"] == "Prior Outreach (below current gate)"

    print("Outreach persistence test: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
