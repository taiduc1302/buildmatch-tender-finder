from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parents[1]
sys.path.insert(0, str(ROOT))

from tenderfinder_source_backlog import (  # noqa: E402
    BACKLOG_CSV,
    BACKLOG_XLSX,
    GATED_ACCESS,
    NON_AUTOMATIC_CONNECTOR_TYPES,
    SHEET_COLUMNS,
    build_potential_sources_content,
    cross_check_counts,
    load_csv_rows,
    load_priority_ids,
    load_universe_rows,
    rank_universe,
)

# Patch 5.19: the source growth backlog is a ranked WORKLIST of sources to
# verify/connect/monitor - never presented as pipeline volume. These tests
# pin the loader counts to the shipped backlog files, prove no rows are
# silently dropped, prove ranking is deterministic, and prove gated
# (paid/login/relationship) sources are never routed to an automatic live
# connector.

EXPECTED_UNIVERSE_ROWS = 372
EXPECTED_CSV_ROWS = 351
EXPECTED_PRIORITY_ROWS = 60


def test_backlog_data_files_ship_with_the_code() -> None:
    assert BACKLOG_XLSX.exists(), f"missing {BACKLOG_XLSX}"
    assert BACKLOG_CSV.exists(), f"missing {BACKLOG_CSV}"


def test_universe_loads_from_xlsx_with_expected_count() -> None:
    rows, origin = load_universe_rows()
    assert origin == "xlsx:All_Source_Universe_v2", f"expected XLSX universe, got {origin}"
    assert len(rows) == EXPECTED_UNIVERSE_ROWS, f"expected {EXPECTED_UNIVERSE_ROWS} rows, got {len(rows)}"
    assert all(r.get("source_id") for r in rows), "every universe row must have a source_id"


def test_csv_loads_with_expected_count() -> None:
    rows = load_csv_rows()
    assert len(rows) == EXPECTED_CSV_ROWS, f"expected {EXPECTED_CSV_ROWS} CSV rows, got {len(rows)}"


def test_priority_60_loads() -> None:
    ids = load_priority_ids()
    assert len(ids) == EXPECTED_PRIORITY_ROWS, f"expected {EXPECTED_PRIORITY_ROWS} priority ids, got {len(ids)}"


def test_csv_cross_check_finds_no_missing_ids() -> None:
    check = cross_check_counts()
    assert check["csv_ids_missing_from_universe"] == [], (
        "CSV sources missing from the XLSX universe: "
        + ", ".join(check["csv_ids_missing_from_universe"][:10])
    )


def test_universe_falls_back_to_csv_when_xlsx_missing() -> None:
    missing = Path(tempfile.gettempdir()) / "tenderfinder_no_such_backlog.xlsx"
    rows, origin = load_universe_rows(xlsx_path=missing)
    assert origin == "csv_fallback"
    assert len(rows) == EXPECTED_CSV_ROWS


def test_ranking_drops_no_rows() -> None:
    rows, _ = load_universe_rows()
    ranked = rank_universe(rows, load_priority_ids())
    assert len(ranked) == len(rows), "ranking must keep every universe row visible"
    assert {r["source_id"] for r in ranked} == {r["source_id"] for r in rows}
    # Duplicates/parent sources stay visible, marked - not removed.
    dupes = [r for r in ranked if r["duplicate_status"] == "possible_duplicate_or_parent_source"]
    assert dupes, "expected the known parent-source rows to remain in the output"
    for r in dupes:
        assert "duplicate" in r["risk_note"].lower() or "dedupe" in r["risk_note"].lower()


def test_ranking_is_deterministic() -> None:
    rows, _ = load_universe_rows()
    pids = load_priority_ids()
    first = [(r["tenderfinder_rank"], r["source_id"], r["priority_score"]) for r in rank_universe(rows, pids)]
    second = [(r["tenderfinder_rank"], r["source_id"], r["priority_score"]) for r in rank_universe(list(reversed(rows)), pids)]
    assert first == second, "ranking must not depend on input row order"


def test_ranking_sorted_by_score_then_tier() -> None:
    rows, _ = load_universe_rows()
    ranked = rank_universe(rows, load_priority_ids())
    scores = [r["priority_score"] for r in ranked]
    assert scores == sorted(scores, reverse=True), "rows must be sorted by priority_score descending"
    assert [r["tenderfinder_rank"] for r in ranked] == list(range(1, len(ranked) + 1))


def test_gated_sources_never_get_automatic_connectors() -> None:
    """Paid/login/relationship sources must be marked honestly - email-alert,
    manual, or paid-decision paths only. Never an automatic live connector."""
    rows, _ = load_universe_rows()
    ranked = rank_universe(rows, load_priority_ids())
    gated = [r for r in ranked if r["access_status"] in GATED_ACCESS]
    assert gated, "test data should contain gated sources"
    for r in gated:
        assert r["recommended_connector_type"] in NON_AUTOMATIC_CONNECTOR_TYPES, (
            f"{r['source_id']} is gated ({r['access_status']}) but got connector "
            f"type {r['recommended_connector_type']!r}"
        )
        assert "scrape" in r["risk_note"].lower() or "paid" in r["risk_note"].lower() or "login" in r["risk_note"].lower()


def test_sheet_content_has_all_columns_and_summary() -> None:
    summary, headers, data_rows = build_potential_sources_content()
    assert headers == SHEET_COLUMNS
    for required in ["tenderfinder_rank", "priority_score", "recommended_phase",
                     "recommended_connector_type", "demo_value",
                     "implementation_effort", "risk_note",
                     "next_step_plain_english", "source_id", "duplicate_status"]:
        assert required in headers, f"missing column {required}"
    assert len(data_rows) == EXPECTED_UNIVERSE_ROWS
    summary_text = "\n".join(str(cell) for row in summary for cell in row)
    assert "NOT pipeline volume" in summary_text
    assert "Total source universe" in summary_text
    assert "TOP 15 RECOMMENDED NEXT SOURCES" in summary_text
    assert "COUNT BY CATEGORY" in summary_text
    assert "COUNT BY ACCESS STATUS" in summary_text


def test_workbook_gets_potential_sources_next_sheet() -> None:
    """Prove the sheet lands in a real openpyxl workbook via the same
    function build_workbook calls."""
    from openpyxl import Workbook, load_workbook
    from tenderfinder_source_backlog import append_potential_sources_sheet

    wb = Workbook()
    append_potential_sources_sheet(wb)
    assert "Potential_Sources_Next" in wb.sheetnames
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sheet_check.xlsx"
        wb.save(path)
        rb = load_workbook(path, read_only=True)
        ws = rb["Potential_Sources_Next"]
        rows = list(ws.iter_rows(values_only=True))
        rb.close()
    text = "\n".join(str(c) for row in rows for c in row if c is not None)
    assert "SOURCE GROWTH BACKLOG" in text
    header_row = next(r for r in rows if r and r[0] == "tenderfinder_rank")
    data_after_header = rows[rows.index(header_row) + 1:]
    populated = [r for r in data_after_header if any(v not in (None, "") for v in r)]
    assert len(populated) == EXPECTED_UNIVERSE_ROWS


def test_source_roadmap_printable_sheet() -> None:
    """Product QA task B: the VP-facing one-pager must exist, sit right
    after Executive_Summary, and carry the required blocks including the
    'not lead volume' honesty note and the top-25 list."""
    from openpyxl import Workbook
    from tenderfinder_source_backlog import append_source_roadmap_printable_sheet

    wb = Workbook()
    wb.active.title = "Executive_Summary"
    append_source_roadmap_printable_sheet(wb)
    assert wb.sheetnames.index("Source_Roadmap_Printable") == 1
    ws = wb["Source_Roadmap_Printable"]
    text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    for required in ["WHERE TENDER_FINDER CAN GROW NEXT", "not lead volume",
                     "Total source universe", "Already coded and running",
                     "Potential / unaccounted", "Priority Next 60",
                     "RECOMMENDED NEXT PATCH PRIORITIES",
                     "TOP 25 RECOMMENDED NEXT SOURCES",
                     "BY CATEGORY", "BY ACCESS STATUS",
                     "Potential_Sources_Next"]:
        assert required in text, f"roadmap sheet missing: {required}"
    top25 = sum(1 for row in ws.iter_rows(values_only=True)
                if row[0] and str(row[0]).startswith("#"))
    assert top25 == 25, f"expected 25 top-source rows, got {top25}"


def main() -> int:
    tests = [
        test_source_roadmap_printable_sheet,
        test_backlog_data_files_ship_with_the_code,
        test_universe_loads_from_xlsx_with_expected_count,
        test_csv_loads_with_expected_count,
        test_priority_60_loads,
        test_csv_cross_check_finds_no_missing_ids,
        test_universe_falls_back_to_csv_when_xlsx_missing,
        test_ranking_drops_no_rows,
        test_ranking_is_deterministic,
        test_ranking_sorted_by_score_then_tier,
        test_gated_sources_never_get_automatic_connectors,
        test_sheet_content_has_all_columns_and_summary,
        test_workbook_gets_potential_sources_next_sheet,
    ]
    for test in tests:
        test()
        print(f"[PASS] {test.__name__}")
    print("Source backlog ranking test: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
