from __future__ import annotations

import hashlib
import os
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl


CODE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = CODE_DIR.parents[1]
sys.path.insert(0, str(CODE_DIR))

import tenderfinder_demo_three_buckets as demo  # noqa: E402
from tenderfinder_engine import RunRequest, _base_manifest, prepare_run, test_source_definition  # noqa: E402
from tenderfinder_keywords_config import (  # noqa: E402
    KeywordRegexRuntimeError,
    KeywordRule,
    MAX_REGEX_PATTERN_LENGTH,
    _validate_regex,
)
from tenderfinder_package_paths import (  # noqa: E402
    ensure_email_alert_dirs,
    load_user_config,
    save_user_config,
    tenderfinder_user_config_path,
)
from tenderfinder_runtime import (  # noqa: E402
    STATE_ROOT_ENV_VAR,
    RuntimeStateError,
    resolve_state_root,
    runtime_paths,
)
from tenderfinder_source_registry import (  # noqa: E402
    SOURCE_COLUMNS,
    SourceRegistryError,
    load_source_rows,
    registry_summary,
    set_source_active,
    upsert_source,
)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def source_row(**values: str) -> dict[str, str]:
    row = {column: "" for column in SOURCE_COLUMNS}
    row.update(
        {
            "source_id": "fixture_rss_source",
            "name": "Fixture RSS source",
            "track": "tender",
            "active": "N",
            "adapter": "public_listing",
            "municipality": "Surrey",
            "url": "https://public.example.org/opportunities",
            "rss": "https://public.example.org/opportunities/feed.xml",
            "no_retry": "Y",
            "access_status": "public_no_login",
            "notes": "Offline source-extension proof; never fetched.",
        }
    )
    row.update(values)
    return row


def test_editable_regex_is_bounded() -> None:
    errors: list[str] = []
    _validate_regex("a" * (MAX_REGEX_PATTERN_LENGTH + 1), 2, errors)
    assert errors and "maximum" in errors[0]

    errors = []
    _validate_regex(r"(a+)+$", 3, errors)
    assert errors and "nested unbounded quantifiers" in errors[0]

    errors = []
    _validate_regex(r"(road)\1", 4, errors)
    assert errors and "backreferences" in errors[0]

    rule = KeywordRule(
        keyword=r"(a|aa)+$",
        match_type="regex",
        weight=0,
        category="tender_match",
        row_number=999,
    )
    try:
        rule.matches("a" * 20_000 + "!")
    except KeywordRegexRuntimeError as exc:
        assert "0.02s safety limit" in str(exc)
    else:
        raise AssertionError("pathological runtime regex was not stopped")


def test_source_registry_crud_fixture_preview_and_restore() -> None:
    canonical = REPO_ROOT / "config" / "sources.csv"
    canonical_before = sha256(canonical)
    with tempfile.TemporaryDirectory() as temp:
        temp_dir = Path(temp)
        registry = temp_dir / "sources.csv"
        pristine = temp_dir / "sources.pristine.csv"
        shutil.copyfile(canonical, registry)
        shutil.copyfile(canonical, pristine)

        upsert_source(source_row(), registry)
        rows = load_source_rows(registry)
        added = next(row for row in rows if row["source_id"] == "fixture_rss_source")
        assert added["active"] == "N"

        fixture = temp_dir / "source_fixture.xml"
        fixture.write_text(
            """<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0"><channel><title>Public procurement</title><item>
<title>Invitation to Tender - Watermain Replacement and Road Restoration</title>
<link>/opportunities/ITT-2026-01</link>
<description>Open opportunity. Closing date: July 31, 2026.</description>
</item></channel></rss>
""",
            encoding="utf-8",
        )
        preview = test_source_definition(
            "fixture_rss_source",
            root=REPO_ROOT,
            sources_path=registry,
            fixture_path=fixture,
            allow_network=False,
        )
        assert preview["passed"] is True
        assert preview["network_used"] is False
        assert preview["status_code"] == "PASS_ADAPTER_FIXTURE"
        assert preview["candidate_count"] == 1
        assert preview["normalized_preview"][0]["source_id"] == "fixture_rss_source"
        assert "Watermain Replacement" in preview["normalized_preview"][0]["tender_title"]

        set_source_active("fixture_rss_source", True, registry)
        edited = dict(next(row for row in load_source_rows(registry) if row["source_id"] == "fixture_rss_source"))
        edited["name"] = "Edited fixture RSS source"
        upsert_source(edited, registry)
        set_source_active("fixture_rss_source", False, registry)
        final_row = next(row for row in load_source_rows(registry) if row["source_id"] == "fixture_rss_source")
        assert final_row["name"] == "Edited fixture RSS source"
        assert final_row["active"] == "N"
        assert not registry.with_suffix(".csv.tmp").exists()

        # Disabled drafts can be incomplete, but cannot be enabled until ready.
        incomplete = source_row(
            source_id="incomplete_source_draft",
            name="",
            adapter="",
            url="",
            rss="",
        )
        upsert_source(incomplete, registry)
        draft_result = test_source_definition(
            "incomplete_source_draft", root=REPO_ROOT, sources_path=registry
        )
        assert draft_result["status_code"] == "DRAFT_INCOMPLETE"
        assert draft_result["passed"] is False
        try:
            set_source_active("incomplete_source_draft", True, registry)
        except SourceRegistryError as exc:
            assert "name is required" in str(exc)
        else:
            raise AssertionError("incomplete source draft was enabled")

        # Unsupported portals remain storable only as disabled drafts and
        # report the honest adapter requirement.
        custom = source_row(source_id="custom_portal_draft", adapter="custom_javascript_portal")
        upsert_source(custom, registry)
        custom_result = test_source_definition(
            "custom_portal_draft", root=REPO_ROOT, sources_path=registry
        )
        assert custom_result["status_code"] == "CUSTOM_ADAPTER_REQUIRED"
        try:
            set_source_active("custom_portal_draft", True, registry)
        except SourceRegistryError as exc:
            assert "custom adapter required" in str(exc)
        else:
            raise AssertionError("unsupported custom portal was enabled")

        try:
            upsert_source(
                source_row(source_id="private_url_draft", url="http://127.0.0.1/private", rss=""),
                registry,
            )
        except SourceRegistryError as exc:
            assert "non-public address" in str(exc)
        else:
            raise AssertionError("private-network source URL was accepted")

        shutil.copyfile(pristine, registry)
        assert sha256(registry) == sha256(pristine)

    assert sha256(canonical) == canonical_before


def test_runtime_isolation_and_manifest_identity() -> None:
    try:
        resolve_state_root(REPO_ROOT / "runtime_state", package_root=REPO_ROOT)
    except RuntimeStateError as exc:
        assert "outside the package repository" in str(exc)
    else:
        raise AssertionError("repository-local runtime state was allowed")

    with tempfile.TemporaryDirectory() as temp:
        base = Path(temp)
        state = base / "state"
        paths = runtime_paths(state, mode="offline", package_root=REPO_ROOT)
        assert paths.root == state.resolve()
        assert paths.history.is_dir() and paths.settings.is_dir()

        plan = prepare_run(
            RunRequest(
                review_xlsx=REPO_ROOT / "inputs" / "all_live_review.xlsx",
                out_dir=base / "out",
                mode="offline",
                python_exe=sys.executable,
                state_root=state,
                run_id="focused_manifest_test",
            ),
            root=REPO_ROOT,
        )
        assert "--no-fetch" in plan.command
        assert plan.state_root == state.resolve()
        manifest = _base_manifest(plan)
        assert manifest["offline"] is True
        assert len(manifest["keywords_sha256"]) == 64
        assert len(manifest["sources_sha256"]) == 64
        assert manifest["source_registry"]["total"] == 39

        previous_state = os.environ.get(STATE_ROOT_ENV_VAR)
        os.environ[STATE_ROOT_ENV_VAR] = str(state)
        try:
            email_dirs = ensure_email_alert_dirs(REPO_ROOT)
            assert email_dirs["user_data"] == (state / "settings").resolve()
            for key in ("inbox", "processed", "rejected", "logs"):
                email_dirs[key].resolve().relative_to(state.resolve())
                try:
                    email_dirs[key].resolve().relative_to(REPO_ROOT.resolve())
                except ValueError:
                    pass
                else:
                    raise AssertionError(f"email runtime directory leaked into repository: {email_dirs[key]}")
            saved = save_user_config({"email_import_path": str(email_dirs["inbox"])}, REPO_ROOT)
            assert saved.parent == (state / "settings").resolve()

            fake_package = base / "fake_package"
            (fake_package / "01 Code" / "CONNECTOR_SWEEP").mkdir(parents=True)
            (fake_package / "inputs").mkdir()
            (fake_package / "run_tenderfinder_demo.bat").write_text("@echo off\n", encoding="utf-8")
            (fake_package / "verify_package.bat").write_text("@echo off\n", encoding="utf-8")
            legacy_inbox = fake_package / "user_data" / "email_alerts" / "inbox"
            legacy_inbox.mkdir(parents=True)
            (legacy_inbox / "founder_owned.eml").write_text("Subject: preserved\n", encoding="utf-8")
            migration_state = base / "migration_state"
            os.environ[STATE_ROOT_ENV_VAR] = str(migration_state)
            migrated = load_user_config(fake_package)
            assert migrated["email_import_path"] == str(legacy_inbox)
            assert tenderfinder_user_config_path(fake_package).parent == (migration_state / "settings").resolve()
        finally:
            if previous_state is None:
                os.environ.pop(STATE_ROOT_ENV_VAR, None)
            else:
                os.environ[STATE_ROOT_ENV_VAR] = previous_state


def test_vancouver_snapshot_recomputes_and_legacy_is_explicit() -> None:
    base = {
        "source_id": "van_building_permits",
        "source_name": "Vancouver building permits",
        "project_id": "BP-2026-001",
        "app_no": "BP-2026-001",
        "title": "Watermain and roadwork permit",
        "municipality": "Vancouver",
        "address": "100 Main Street",
        "scope_summary": "Watermain and roadwork servicing",
        "fit_score": 0,
        "proposed_route": "Rejected_Archive",
        "hold_reason": "van_permit_noisy_interior_ti",
        "write_eligible": "N",
        "fit_reason": "van_permit:noisy",
    }
    current = demo.normalized_row({**base, "keyword_scoring_text": "watermain roadwork servicing"})
    event = current["_keyword_rescore"]
    assert event["old_tier"] == "noisy"
    assert event["new_tier"] == "strong"
    assert event["new_bucket"] == "Future_Projects"
    assert event["exception"] == ""

    legacy = demo.normalized_row(base)
    legacy_event = legacy["_keyword_rescore"]
    assert legacy_event["new_tier"] == "noisy"
    assert legacy_event["new_bucket"] == "Rejected_Archive"
    assert legacy_event["exception"] == "legacy_vancouver_scoring_text_unavailable"


def test_manual_triage_audit_and_weekly_log_survive() -> None:
    lead_id = "stable-lead-001"
    with tempfile.TemporaryDirectory() as temp:
        prior_path = Path(temp) / "prior_user_master.xlsx"
        prior = openpyxl.Workbook()
        future = prior.active
        future.title = "Future_Projects"
        future.append(demo.USER_FUTURE_PROJECTS_HEADERS)
        values = {header: "" for header in demo.USER_FUTURE_PROJECTS_HEADERS}
        values.update(
            {
                "Lead ID": lead_id,
                "Assigned To": "Founder",
                "Status": "In Progress",
                "Notes": "Preserve this manual decision.",
            }
        )
        future.append([values[header] for header in demo.USER_FUTURE_PROJECTS_HEADERS])
        prior_audit = prior.create_sheet("Keyword_Change_Audit")
        prior_audit.append(demo.USER_KEYWORD_AUDIT_HEADERS)
        prior_audit_values = {header: "" for header in demo.USER_KEYWORD_AUDIT_HEADERS}
        prior_audit_values.update(
            {
                "Lead ID": lead_id,
                "Old Score": 70,
                "New Score": 44,
                "Old Bucket": "Future_Projects",
                "New Bucket": "Run_Queue",
                "Assigned To": "Founder",
                "Status": "In Progress",
                "Notes": "Preserve this manual decision.",
            }
        )
        prior_audit.append([prior_audit_values[header] for header in demo.USER_KEYWORD_AUDIT_HEADERS])
        weekly = prior.create_sheet("Weekly_Review_Log")
        weekly.append(["Date", "Decision"])
        weekly.append(["2026-07-14", "Keep pursuing this lead"])
        prior.save(prior_path)
        prior.close()

        lookups = demo._load_prior_preserved_fields(prior_path)
        assert lookups["future_projects"][lead_id]["Status"] == "In Progress"
        assert lookups["keyword_audit"][lead_id]["Assigned To"] == "Founder"

        technical = openpyxl.Workbook()
        audit = technical.active
        audit.title = "Keyword_Change_Audit"
        audit.append(["Technical RESCORE_ALWAYS change events"])
        audit.append(
            [
                "lead_id", "source_id", "app_no", "old_score", "new_score", "score_delta",
                "old_tier", "new_tier", "old_bucket", "new_bucket", "changed", "exception",
            ]
        )
        audit.append(
            [lead_id, "source_one", "APP-1", 70, 44, -26, "", "", "Future_Projects", "Run_Queue", "YES", ""]
        )
        audit_rows = demo.build_user_keyword_audit_rows(
            technical,
            lookups["future_projects"],
            lookups["keyword_audit"],
        )
        audited = dict(zip(demo.USER_KEYWORD_AUDIT_HEADERS, audit_rows[0]))
        assert audited["Status"] == "In Progress"
        assert audited["Assigned To"] == "Founder"
        assert audited["Notes"] == "Preserve this manual decision."
        assert audited["Old Score"] == 70 and audited["New Score"] == 44
        assert audited["Old Bucket"] == "Future_Projects"
        assert audited["New Bucket"] == "Run_Queue"

        target = openpyxl.Workbook()
        target.active.title = "Weekly_Review_Log"
        assert demo._restore_prior_weekly_review_log(prior_path, target) is True
        assert target["Weekly_Review_Log"]["B2"].value == "Keep pursuing this lead"
        technical.close()
        target.close()


def test_engine_and_agent2_remain_isolated() -> None:
    engine_text = (CODE_DIR / "tenderfinder_engine.py").read_text(encoding="utf-8").casefold()
    launcher_text = (CODE_DIR / "tenderfinder_launcher_gui.py").read_text(encoding="utf-8").casefold()
    agent_text = (REPO_ROOT / "01 Code" / "tenderfinder_agent2.py").read_text(encoding="utf-8").casefold()
    assert "import tkinter" not in engine_text
    assert "from tkinter" not in engine_text
    assert "from tenderfinder_engine import" in launcher_text
    assert "import tenderfinder_agent2" not in engine_text
    assert "from tenderfinder_agent2" not in engine_text
    assert "import tenderfinder_agent2" not in launcher_text
    assert "from tenderfinder_agent2" not in launcher_text
    assert "tenderfinder_keywords_config" not in agent_text
    summary = registry_summary(root=REPO_ROOT)
    assert summary["path"] == str((REPO_ROOT / "config" / "sources.csv").resolve())
    assert summary["total"] == 39
    assert summary["enabled"] == 39
    assert summary["runtime_eligible"] == 27
    assert summary["verified_live"] == 1
    assert summary["ready_for_live_test"] == 26
    assert summary["tender"] == 21
    assert summary["development"] == 18


def test_master_template_selection_is_package_isolated() -> None:
    old_root = demo.ROOT
    old_history = demo.DEMO_HISTORY_DIR
    try:
        with tempfile.TemporaryDirectory() as temp:
            base = Path(temp)
            package = base / "current_package"
            other_install = base / "other_install"
            local_master = (
                package
                / "00 Master"
                / "TENDER_FINDER_Tender_Intelligence_Working_Master_TEMPLATE_v1.xlsx"
            )
            foreign_master = (
                other_install
                / "00 Master"
                / "TENDER_FINDER_Tender_Intelligence_Working_Master_TEMPLATE_v9.xlsx"
            )
            local_master.parent.mkdir(parents=True)
            foreign_master.parent.mkdir(parents=True)
            local_master.write_bytes(b"current package template")
            foreign_master.write_bytes(b"newer foreign install template")
            os.utime(local_master, (1_000, 1_000))
            os.utime(foreign_master, (2_000, 2_000))

            demo.ROOT = package
            demo.DEMO_HISTORY_DIR = package / "history"
            selected = demo.find_latest_master_workbook(extra_dirs=[other_install])
            assert selected == local_master.resolve()
    finally:
        demo.ROOT = old_root
        demo.DEMO_HISTORY_DIR = old_history


def main() -> int:
    tests = [
        test_editable_regex_is_bounded,
        test_source_registry_crud_fixture_preview_and_restore,
        test_runtime_isolation_and_manifest_identity,
        test_vancouver_snapshot_recomputes_and_legacy_is_explicit,
        test_manual_triage_audit_and_weekly_log_survive,
        test_engine_and_agent2_remain_isolated,
        test_master_template_selection_is_package_isolated,
    ]
    for test in tests:
        test()
        print(f"[PASS] {test.__name__}")
    print(f"Standalone weekly focused regressions: {len(tests)} passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
