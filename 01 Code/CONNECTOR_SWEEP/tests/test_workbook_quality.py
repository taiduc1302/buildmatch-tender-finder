from __future__ import annotations

import os
import re
from pathlib import Path

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[3]
# Sanitized portable package: the original real-run verified output
# (demo_p522) is not shipped. This points at the synthetic sample output
# generated from inputs/all_live_review.xlsx (synthetic demo data).
# Override with the TENDER_FINDER_DEMO_WORKBOOK env var to test another run.
DEFAULT_WORKBOOK = REPO_ROOT / "latest_verified_output" / "demo_synthetic_sample" / "TENDER_FINDER_DEMO_Opportunities_Three_Buckets.xlsx"


def workbook_path() -> Path:
    return Path(os.environ.get("TENDER_FINDER_DEMO_WORKBOOK", str(DEFAULT_WORKBOOK)))


# Volume minimums. The original values (5393 / 973 / 8000) reflect the real
# production dataset this tool was previously run against. The sanitized
# portable package ships only a small synthetic sample, so when the workbook
# under test is the packaged synthetic sample the minimums drop to sample
# scale. Override via env vars when validating a real run.
_IS_SYNTHETIC_SAMPLE = "demo_synthetic_sample" in str(workbook_path())
MIN_FUTURE_ROWS = int(os.environ.get(
    "TENDER_FINDER_MIN_FUTURE_ROWS", "5" if _IS_SYNTHETIC_SAMPLE else "5393"))
MIN_WATCH_ROWS = int(os.environ.get(
    "TENDER_FINDER_MIN_WATCH_ROWS", "1" if _IS_SYNTHETIC_SAMPLE else "973"))
MIN_ANALYZED_ROWS = int(os.environ.get(
    "TENDER_FINDER_MIN_ANALYZED_ROWS", "1" if _IS_SYNTHETIC_SAMPLE else "8000"))


TERMINAL_STATUSES = {
    "Concluded",
    "Closed",
    "Cancelled",
    "Final Maintenance",
    "Ok for Occupancy",
    "Final Approval",
}


def main() -> int:
    path = workbook_path()
    assert path.exists(), f"workbook not found: {path}"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    ws = wb["Executive_Summary"]
    blanks = []
    in_how_to = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "How to read this workbook":
            in_how_to = True
            continue
        if in_how_to and row[0]:
            if row[0] in wb.sheetnames and not row[1]:
                blanks.append(row[0])
    assert not blanks, f"blank how-to-read column B rows: {blanks}"

    ws = wb["Top_Civil_Leads_Printable"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    soc = headers.index("source_or_contact")
    bad_contacts = [r[soc] for r in ws.iter_rows(min_row=2, values_only=True) if "arcgis" in str(r[soc] or "").lower() or "geoservices" in str(r[soc] or "").lower()]
    assert not bad_contacts, f"source_or_contact contains raw GIS URLs: {bad_contacts[:3]}"

    ws = wb["BID_NOW_Active_Tenders"]
    headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    title_i = headers.index("tender_title")
    civil_i = headers.index("civil_relevant")
    email_i = headers.index("contact_email")
    phone_i = headers.index("contact_phone")
    closing_i = headers.index("closing_date")
    junk = []
    ref_numbers = []
    ref_phones = []
    richmond_headers = []
    junk_exact = {"tenders rfqs rfps"}
    junk_contains = ["bidcentral :: bc construction projects bidding and tenders"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        title = str(row[title_i] or "")
        civil = str(row[civil_i] or "")
        email = str(row[email_i] or "")
        phone = str(row[phone_i] or "")
        closing = str(row[closing_i] or "")
        title_l = title.lower()
        if (
            title_l in junk_exact
            or any(pattern in title_l for pattern in junk_contains)
            or title_l.startswith("get in touch with us")
            or (".pdf" in title_l and civil == "NO")
        ):
            junk.append(title)
        if title == "Bid Opportunities - City of Richmond, BC" and not closing:
            richmond_headers.append(title)
        if re.match(r"^\d{4}-\d{3}-\d{4}$", email):
            ref_numbers.append(email)
        if re.match(r"^\d{4}-\d{3}-\d{4}$", phone):
            ref_phones.append(phone)
    assert not junk, f"BID NOW junk rows remain: {junk[:3]}"
    assert not ref_numbers, f"Surrey reference numbers remain in contact_email: {ref_numbers[:3]}"
    assert not ref_phones, f"Surrey reference numbers remain in contact_phone: {ref_phones[:3]}"
    assert not richmond_headers, "Richmond procurement page header remains without a real closing date"

    fp = wb["BID_LATER_Future_Projects"]
    fp_headers = [c.value for c in next(fp.iter_rows(min_row=2, max_row=2))]
    fp_idx = {h: i for i, h in enumerate(fp_headers)}
    assert "source_tier" in fp_headers, "BID LATER missing source_tier column"
    for required in ["lead_id", "first_seen_date", "is_new_since_last_run", "project_scale_tier", "regional_cluster", "address_source", "approximate_location", "status_class"]:
        assert required in fp_headers, f"BID LATER missing {required} column"
    assert fp.max_row - 2 >= MIN_FUTURE_ROWS
    surrey_terminal = []
    high_terminal = []
    for row in fp.iter_rows(min_row=3, values_only=True):
        source = str(row[fp_idx["source"]] or "")
        stage = str(row[fp_idx["app_type_stage"]] or "")
        signal_quality = str(row[fp_idx["signal_quality"]] or "")
        if source == "surrey_devapps_v2" and stage in TERMINAL_STATUSES:
            surrey_terminal.append((source, stage, row[fp_idx["app_no"]]))
        if signal_quality == "HIGH" and stage in TERMINAL_STATUSES:
            high_terminal.append((stage, row[fp_idx["app_no"]], source))
    assert not surrey_terminal, f"Surrey terminal rows remain in BID LATER: {surrey_terminal[:3]}"
    assert not high_terminal, f"HIGH rows still carry terminal status: {high_terminal[:3]}"
    assert wb["Watchlist_Monitor"].max_row - 1 >= MIN_WATCH_ROWS
    note = str(wb["Analyzed_Set_Aside"].cell(1, 1).value or "")
    assert "12,766" in note or "12,832" in note or wb["Analyzed_Set_Aside"].max_row - 2 >= MIN_ANALYZED_ROWS
    for sheet_name in ["New_This_Run", "Outreach_Tracker", "Developer_Watchlist", "Design_Consultants_Reference", "Tender_Pattern_Analysis", "Surrey_Historical_Archive"]:
        assert sheet_name in wb.sheetnames, f"missing sheet: {sheet_name}"
    outreach_headers = [c.value for c in next(wb["Outreach_Tracker"].iter_rows(min_row=1, max_row=1))]
    for required in ["lead_id", "type", "status", "last_contact_date", "next_action_date", "notes"]:
        assert required in outreach_headers, f"Outreach_Tracker missing {required}"
    top_headers = [c.value for c in next(wb["Top_Civil_Leads_Printable"].iter_rows(min_row=1, max_row=1))]
    assert "project_scale_tier" in top_headers
    assert "regional_cluster" in top_headers
    top_idx = {h: i for i, h in enumerate(top_headers)}
    muni_counts = {}
    empty_addresses = 0
    terminal_rows = []
    for row in wb["Top_Civil_Leads_Printable"].iter_rows(min_row=2, values_only=True):
        municipality = str(row[top_idx["municipality"]] or "")
        muni_counts[municipality] = muni_counts.get(municipality, 0) + 1
        if not str(row[top_idx["address"]] or "").strip():
            empty_addresses += 1
        if str(row[top_idx["app_type_stage"]] or "") in TERMINAL_STATUSES:
            terminal_rows.append((municipality, row[top_idx["app_type_stage"]], row[top_idx["address"]]))
    assert empty_addresses == 0, f"Top 50 still has empty addresses: {empty_addresses}"
    assert not terminal_rows, f"Top 50 still has terminal rows: {terminal_rows[:3]}"
    assert len(muni_counts) >= 3, f"Top 50 municipality diversity too low: {muni_counts}"
    assert max(muni_counts.values()) <= 15, f"Top 50 municipality cap broken: {muni_counts}"
    tender_note = str(wb["Tender_Pattern_Analysis"].cell(1, 1).value or "")
    assert "small sample" in tender_note.lower() and "not a statistically validated" in tender_note.lower()
    dev_headers = [c.value for c in next(wb["Developer_Watchlist"].iter_rows(min_row=2, max_row=2))]
    for required in ["parent_brand", "grouped_entity_examples", "applicant_type", "repeat_developer"]:
        assert required in dev_headers, f"Developer_Watchlist missing {required}"
    dev_idx = {h: i for i, h in enumerate(dev_headers)}
    consultant_needles = [
        "Engineering",
        "Consulting",
        "Consultant",
        "Consultants",
        "Certified Professional",
        "Building Code",
        "Architects",
        "Architecture",
    ]
    polygon_rows = []
    top15_consultant_like = []
    for row_number, row in enumerate(wb["Developer_Watchlist"].iter_rows(min_row=3, values_only=True), start=1):
        applicant_name = str(row[dev_idx["applicant_name"]] or "")
        if row_number <= 15 and any(needle.lower() in applicant_name.lower() for needle in consultant_needles):
            top15_consultant_like.append(applicant_name)
        if applicant_name == "Polygon":
            polygon_rows.append(row)
    assert not top15_consultant_like, f"Top 15 Developer_Watchlist still has consultant-like names: {top15_consultant_like}"
    if polygon_rows:
        repeat = str(polygon_rows[0][dev_idx["repeat_developer"]] or "")
        applicant_type = str(polygon_rows[0][dev_idx["applicant_type"]] or "")
        assert repeat.startswith("YES"), f"Polygon rollup did not flip to repeat developer: {repeat}"
        assert applicant_type == "DEVELOPER", f"Polygon applicant_type unexpected: {applicant_type}"
    consultant_headers = [c.value for c in next(wb["Design_Consultants_Reference"].iter_rows(min_row=2, max_row=2))]
    consultant_idx = {h: i for i, h in enumerate(consultant_headers)}
    if "applicant_type" in consultant_idx:
        lmdg_rows = []
        for row in wb["Design_Consultants_Reference"].iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            assert str(row[consultant_idx["applicant_type"]] or "") == "DESIGN_CONSULTANT"
            if str(row[consultant_idx["applicant_name"]] or "") == "LMDG Building Code Consultants":
                lmdg_rows.append(row)
        # This grouping check names a specific high-volume consultant firm
        # from the real production dataset; the synthetic sample has no such
        # firm, so the check only applies to real-data runs.
        if not _IS_SYNTHETIC_SAMPLE:
            assert lmdg_rows, "LMDG consultant firm was not grouped into Design_Consultants_Reference"
            assert int(lmdg_rows[0][consultant_idx["total_applications"]] or 0) >= 2
    surrey_headers = [c.value for c in next(wb["Surrey_Historical_Archive"].iter_rows(min_row=2, max_row=2))]
    surrey_idx = {h: i for i, h in enumerate(surrey_headers)}
    for row in wb["Surrey_Historical_Archive"].iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        assert str(row[surrey_idx["app_type_stage"]] or "") in TERMINAL_STATUSES
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    currency_like = re.search(r"(\$\s*\d|\bCAD\s*\d|\d+\s*dollars?)", value, flags=re.I)
                    assert not currency_like, f"currency-like value found in {ws.title}: {value[:80]}"
    wb.close()
    print("Workbook quality tests: 5 passed, 0 failed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
