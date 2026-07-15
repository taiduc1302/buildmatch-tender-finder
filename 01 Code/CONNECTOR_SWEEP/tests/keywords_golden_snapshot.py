#!/usr/bin/env python3
"""Create/compare timestamp-free snapshots of an offline demo workbook."""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FUTURE_FIELDS = (
    "lead_id", "municipality", "app_no", "address", "address_source",
    "approximate_location", "app_type_stage", "applicant_name", "scope_summary",
    "fit_score", "signal_quality", "source", "source_url", "source_tier",
    "detail_available", "regional_cluster", "status_class", "project_scale_tier",
    "est_lots", "est_units", "est_storeys", "scale",
)
WATCH_FIELDS = (
    "municipality", "app_no", "address", "app_type_stage", "scope_summary",
    "fit_score", "source", "source_url",
)
ANALYZED_FIELDS = (
    "municipality", "app_no", "address", "app_type_stage", "fit_score",
    "source", "source_url", "route",
)
SOURCE_FIELDS = ("source_id", "final_status", "candidates", "civil_candidates")


def _records(ws: Any, required_header: str) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    header_index = next(
        index
        for index, row in enumerate(rows[:10])
        if required_header in {str(value or "").strip() for value in row}
    )
    headers = [str(value or "").strip() for value in rows[header_index]]
    records: list[dict[str, Any]] = []
    for values in rows[header_index + 1 :]:
        if not any(value not in (None, "") for value in values):
            continue
        record = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        records.append(record)
    return records


def _select(records: list[dict[str, Any]], fields: tuple[str, ...]) -> list[dict[str, Any]]:
    selected = [{field: record.get(field) for field in fields} for record in records]
    return sorted(selected, key=lambda row: json.dumps(row, sort_keys=True, default=str))


def build_snapshot(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        future = _select(_records(wb["BID_LATER_Future_Projects"], "fit_score"), FUTURE_FIELDS)
        watch = _select(_records(wb["Watchlist_Monitor"], "fit_score"), WATCH_FIELDS)
        analyzed = _select(_records(wb["Analyzed_Set_Aside"], "fit_score"), ANALYZED_FIELDS)
        source_statuses = _select(_records(wb["Source_Run_Log"], "source_id"), SOURCE_FIELDS)
    finally:
        wb.close()
    return {
        "counts": {
            "future": len(future),
            "watch": len(watch),
            "analyzed": len(analyzed),
            "sources": len(source_statuses),
        },
        "future": future,
        "watch": watch,
        "analyzed": analyzed,
        "source_statuses": source_statuses,
    }


def _canonical(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: _canonical(item) for key, item in sorted(value.items())}
    if isinstance(value, list):
        normalized = [_canonical(item) for item in value]
        return sorted(normalized, key=lambda item: json.dumps(item, sort_keys=True, default=str))
    return value


def compare_snapshots(actual: dict[str, Any], expected: dict[str, Any]) -> list[str]:
    actual_canonical = _canonical(actual)
    expected_canonical = _canonical(expected)
    if actual_canonical == expected_canonical:
        return []
    errors: list[str] = []
    for section in ("counts", "future", "watch", "analyzed", "source_statuses"):
        if actual_canonical.get(section) != expected_canonical.get(section):
            errors.append(f"{section} differs")
    return errors or ["snapshot differs"]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--compare", type=Path)
    args = parser.parse_args()
    snapshot = build_snapshot(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(snapshot, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8")
    print(f"snapshot={args.output}")
    print(f"counts={snapshot['counts']}")
    if args.compare:
        expected = json.loads(args.compare.read_text(encoding="utf-8"))
        errors = compare_snapshots(snapshot, expected)
        if errors:
            print("SNAPSHOT_COMPARE=FAIL " + "; ".join(errors))
            return 1
        print("SNAPSHOT_COMPARE=PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
