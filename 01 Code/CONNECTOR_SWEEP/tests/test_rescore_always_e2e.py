from __future__ import annotations

import hashlib
import os
import shutil
import sys
import tempfile
from pathlib import Path
from types import SimpleNamespace

import openpyxl


CODE_DIR = Path(__file__).resolve().parents[1]
PACKAGE_ROOT = CODE_DIR.parents[1]
CANONICAL = PACKAGE_ROOT / "config" / "keywords.xlsx"
sys.path.insert(0, str(CODE_DIR))

import tenderfinder_demo_three_buckets as demo  # noqa: E402
from tenderfinder_keywords_config import (  # noqa: E402
    CONFIG_ENV_VAR,
    KEYWORDS_STATE_ROOT_ENV_VAR,
    clear_keywords_cache,
    load_keywords_config,
)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _set_rule_active(path: Path, keyword: str, category: str, active: str) -> None:
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Keywords"]
    headers = {
        str(cell.value).strip().casefold(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }
    changed = 0
    for row in range(2, sheet.max_row + 1):
        if (
            str(sheet.cell(row, headers["keyword"]).value or "").strip().casefold()
            == keyword.casefold()
            and str(sheet.cell(row, headers["category"]).value or "").strip().casefold()
            == category.casefold()
        ):
            sheet.cell(row, headers["active"], active)
            changed += 1
    assert changed == 1, (keyword, category, changed)
    workbook.save(path)
    workbook.close()


def _synthetic_record() -> dict[str, object]:
    return {
        "source_id": "fixture_development_source",
        "source_name": "Fixture development source",
        "project_id": "RESCORE-001",
        "app_no": "RESCORE-001",
        "title": "Earthwork",
        "municipality": "Surrey",
        "scope_summary": "Earthwork",
        "fit_score": 52,
        "proposed_route": "Future_Projects",
        "write_eligible": "Y",
        "hold_reason": "",
    }


def test_rescore_always_changes_score_tier_bucket_and_writes_visible_audit() -> None:
    canonical_before = sha256(CANONICAL)
    old_config = os.environ.get(CONFIG_ENV_VAR)
    old_state = os.environ.get(KEYWORDS_STATE_ROOT_ENV_VAR)
    try:
        with tempfile.TemporaryDirectory() as temp:
            folder = Path(temp)
            working = folder / "keywords.xlsx"
            shutil.copy2(CANONICAL, working)
            pristine_copy_hash = sha256(working)
            os.environ[CONFIG_ENV_VAR] = str(working)
            os.environ[KEYWORDS_STATE_ROOT_ENV_VAR] = str(folder / "state")

            clear_keywords_cache()
            load_keywords_config(working, force_reload=True)
            baseline = demo.normalized_row(_synthetic_record())
            baseline_event = baseline["_keyword_rescore"]
            assert baseline["fit_score"] == 52
            assert baseline["signal_quality"] == "MEDIUM"
            assert baseline["route"] == "Future_Projects"
            assert baseline_event["changed"] is False

            _set_rule_active(working, "earthwork", "positive", "N")
            clear_keywords_cache()
            load_keywords_config(working, force_reload=True)
            changed = demo.normalized_row(_synthetic_record())
            event = changed["_keyword_rescore"]
            assert changed["fit_score"] == 43
            assert changed["signal_quality"] == "LOW"
            assert changed["route"] == "Run_Queue"
            assert changed["hold_reason"] == "rescore_fit_below_50"
            assert event["old_score"] == 52 and event["new_score"] == 43
            assert event["old_tier"] == "MEDIUM" and event["new_tier"] == "LOW"
            assert event["old_bucket"] == "Future_Projects"
            assert event["new_bucket"] == "Run_Queue"
            assert event["changed"] is True

            audit_path = folder / "rescore_audit.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Keyword_Change_Audit"
            rows = demo.keyword_audit_rows(
                SimpleNamespace(keyword_rescore_events=[event])
            )
            demo.write_rows(
                sheet,
                demo.KEYWORD_AUDIT_HEADERS,
                rows,
                "RESCORE_ALWAYS audit fixture",
            )
            workbook.save(audit_path)
            workbook.close()

            workbook = openpyxl.load_workbook(audit_path, data_only=False)
            sheet = workbook["Keyword_Change_Audit"]
            headers = [cell.value for cell in sheet[2]]
            values = [cell.value for cell in sheet[3]]
            audit = dict(zip(headers, values))
            assert audit["old_score"] == 52 and audit["new_score"] == 43
            assert audit["old_tier"] == "MEDIUM" and audit["new_tier"] == "LOW"
            assert audit["old_bucket"] == "Future_Projects"
            assert audit["new_bucket"] == "Run_Queue"
            workbook.close()

            shutil.copy2(CANONICAL, working)
            assert sha256(working) == pristine_copy_hash
            clear_keywords_cache()
            restored = demo.normalized_row(_synthetic_record())
            assert restored["fit_score"] == baseline["fit_score"]
            assert restored["signal_quality"] == baseline["signal_quality"]
            assert restored["route"] == baseline["route"]
    finally:
        if old_config is None:
            os.environ.pop(CONFIG_ENV_VAR, None)
        else:
            os.environ[CONFIG_ENV_VAR] = old_config
        if old_state is None:
            os.environ.pop(KEYWORDS_STATE_ROOT_ENV_VAR, None)
        else:
            os.environ[KEYWORDS_STATE_ROOT_ENV_VAR] = old_state
        clear_keywords_cache()
    assert sha256(CANONICAL) == canonical_before


def main() -> int:
    test_rescore_always_changes_score_tier_bucket_and_writes_visible_audit()
    print("[PASS] test_rescore_always_changes_score_tier_bucket_and_writes_visible_audit")
    print("RESCORE_ALWAYS end-to-end tests: 1 passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
