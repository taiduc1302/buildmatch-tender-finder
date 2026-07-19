"""Headless development-data refresh orchestration (Build Week Phase 2).

This is the GUI-independent controller behind the "Refresh Development Data"
action. It reads eligible sources from the truthful registry (excluding
manual_only / needs_configuration / wrong_source / blocked / disabled), acquires
records via a safe acquirer, normalizes and deduplicates them, validates the
result, writes a timestamped external dataset, promotes it atomically as the
active dataset, scores it into a ranked output workbook, writes a run manifest,
and returns a truthful ``RunMetrics`` object.

Failure behaviour is explicit: on total failure (or a failed dataset validation)
the previous active dataset is preserved, its last-successful timestamp is not
advanced, it is labelled cached/stale, and the packaged synthetic input under
``inputs/`` is never overwritten.

The heavy steps (acquire, write-dataset, score-and-build) are injectable so the
whole flow is testable headlessly with no network and no real workbook engine.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from tenderfinder_package_paths import detect_package_root
from tenderfinder_runtime import atomic_write_json, resolve_state_root, timestamp_now
import tenderfinder_data_modes as dm
from tenderfinder_presets import get_preset
from tenderfinder_source_registry import (
    load_source_rows,
    source_is_runtime_eligible,
    source_skip_reason,
)


# --------------------------------------------------------------------------- #
# Data classes
# --------------------------------------------------------------------------- #


@dataclass(frozen=True)
class SourceFetch:
    """What an acquirer returns for one source."""

    source_id: str
    ok: bool
    records: tuple[dict[str, Any], ...] = ()
    error: str = ""
    http_status: int = 0

    @property
    def count(self) -> int:
        return len(self.records)


@dataclass(frozen=True)
class ScoreResult:
    output_path: str
    scored: int
    bid_now: int
    bid_later: int
    watch: int
    skip: int


@dataclass(frozen=True)
class RefreshRequest:
    preset_id: str = "civil_contractor"
    source_ids: tuple[str, ...] = ()          # empty = all eligible development sources
    state_root: str | Path | None = None
    package_root: str | Path | None = None
    run_id: str = ""
    # Safe per-source cap for the FULL sweep acquirer (not a preview limit —
    # this is real paginated collection up to this many records per source).
    # 0 means "use the source's own default" (tenderfinder_raw_sweep's CLI
    # default of 20000, effectively unbounded for any real municipal feed).
    max_records_per_source: int = 2000


@dataclass
class RefreshResult:
    succeeded: bool
    message: str
    metrics: dm.RunMetrics
    provenance: dm.DatasetProvenance | None
    promotion: dm.PromotionResult | None
    source_outcomes: tuple[dict[str, Any], ...]
    output_paths: dict[str, Any]
    manifest_path: str
    stale: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "succeeded": self.succeeded,
            "message": self.message,
            "metrics": self.metrics.to_dict(),
            "provenance": self.provenance.to_dict() if self.provenance else None,
            "promotion": self.promotion.to_dict() if self.promotion else None,
            "source_outcomes": list(self.source_outcomes),
            "output_paths": self.output_paths,
            "manifest_path": self.manifest_path,
            "stale": self.stale,
        }


# --------------------------------------------------------------------------- #
# Eligible-source selection (truthful registry)
# --------------------------------------------------------------------------- #


def eligible_development_sources(
    *,
    package_root: str | Path | None = None,
    sources_path: str | Path | None = None,
    source_ids: tuple[str, ...] = (),
) -> list[dict[str, Any]]:
    """Return runnable development-track sources, excluding non-eligible ones.

    A source is excluded when it is disabled or its operational_status is not
    runtime-eligible (manual_only, needs_configuration, wrong_source, blocked,
    config_valid_only, deprecated).
    """
    rows = load_source_rows(sources_path, root=package_root, active_only=False, track="development")
    eligible = [row for row in rows if source_is_runtime_eligible(row)]
    if source_ids:
        wanted = {s.casefold() for s in source_ids}
        eligible = [row for row in eligible if row["source_id"].casefold() in wanted]
    return eligible


def source_health_rows(
    *,
    package_root: str | Path | None = None,
    sources_path: str | Path | None = None,
    track: str | None = None,
) -> list[dict[str, Any]]:
    """Per-source health snapshot for a display (never marks a failed source healthy)."""
    rows = load_source_rows(sources_path, root=package_root, active_only=False, track=track)
    health: list[dict[str, Any]] = []
    for row in rows:
        eligible = source_is_runtime_eligible(row)
        health.append(
            {
                "source_id": row["source_id"],
                "name": row.get("name", ""),
                "track": row.get("track", ""),
                "enabled": row.get("active", "") == "Y",
                "status": row.get("operational_status", ""),
                "runtime_eligible": eligible,
                "skip_reason": "" if eligible else source_skip_reason(row),
                "last_attempted": row.get("last_tested_at", ""),
                "last_result": row.get("last_test_result", ""),
                "last_http_status": row.get("last_http_status", ""),
                "last_error": row.get("last_error", ""),
                "last_candidate_count": row.get("last_candidate_count", ""),
                "last_normalized_count": row.get("last_normalized_count", ""),
            }
        )
    return health


# --------------------------------------------------------------------------- #
# Normalization / dedup / validation
# --------------------------------------------------------------------------- #


def _record_key(record: dict[str, Any]) -> tuple[str, str, str]:
    def norm(*keys: str) -> str:
        for key in keys:
            value = record.get(key)
            if value:
                return str(value).strip().casefold()
        return ""

    return (
        norm("source", "source_id"),
        norm("app_no", "lead_id", "record_id"),
        norm("address", "title", "tender_title", "source_url"),
    )


def _is_descriptive(record: dict[str, Any]) -> bool:
    return any(record.get(key) for key in ("address", "title", "tender_title", "scope_summary"))


def deduplicate_records(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    """Deduplicate by (source, app_no/lead_id, address/title/url) and drop
    non-descriptive stub records (no address/title/scope_summary) that carry
    no scoreable signal for the estimator.

    Real municipal open-data feeds routinely include a handful of thin/stub
    rows (e.g. an application number with every other field blank) alongside
    otherwise rich records; those stubs are noise, not a reason to reject an
    entire dataset (see ``validate_dataset``). Both duplicates and dropped
    stubs are reported together as ``removed`` — the truthful-metrics
    invariant ``normalized_records == records_before_dedup - duplicates_removed``
    holds either way.
    """
    seen: dict[tuple[str, str, str], dict[str, Any]] = {}
    for record in records:
        if not _is_descriptive(record):
            continue
        key = _record_key(record)
        if key not in seen:
            seen[key] = record
    deduped = list(seen.values())
    return deduped, len(records) - len(deduped)


def validate_dataset(records: list[dict[str, Any]]) -> dict[str, Any]:
    """Dataset-level validation. Per-record content quality (descriptive
    fields) is already enforced by ``deduplicate_records`` filtering out
    non-descriptive stubs before this runs — one thin record must never fail
    an otherwise-good dataset of hundreds of real records."""
    errors: list[str] = []
    if not records:
        errors.append("dataset is empty")
    return {"passed": not errors, "errors": errors}


# --------------------------------------------------------------------------- #
# Default heavy-step implementations (overridable / injectable)
# --------------------------------------------------------------------------- #


def _default_dataset_writer(records: list[dict[str, Any]], dest: Path) -> Path:
    """Write normalized records to a simple review-format workbook.

    Record values come from public source websites/feeds and are therefore
    untrusted for workbook-formula-injection purposes (a scope_summary or
    address field could contain a leading ``=``/``+``/``-``/``@``); they are
    written through ``append_untrusted_row`` so Excel always treats them as
    literal text, never as a formula.
    """
    import openpyxl

    from tenderfinder_excel_safety import append_untrusted_row

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Development_Review"
    headers = [
        "lead_id", "municipality", "app_no", "address", "app_type_stage",
        "status_class", "scope_summary", "applicant_name", "source", "source_url",
    ]
    ws.append(headers)
    for record in records:
        append_untrusted_row(ws, [record.get(h, "") for h in headers])
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def _read_records_for_scoring(dataset_path: Path) -> list[dict[str, Any]]:
    """Read back the records written by ``_default_dataset_writer`` (or a CSV
    with the same column names) so ``default_scorer`` can score them."""
    suffix = dataset_path.suffix.casefold()
    if suffix == ".csv":
        import csv

        with dataset_path.open(encoding="utf-8") as handle:
            return list(csv.DictReader(handle))
    import openpyxl

    wb = openpyxl.load_workbook(dataset_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


# Fit-score thresholds mirroring the deterministic routing used elsewhere in
# the product (tenderfinder_demo_three_buckets._rescore_route / GUI ranked
# opportunities): >=60 -> BID_LATER-equivalent, >=50 -> WATCH-equivalent, else
# SKIP. BID_NOW is a live-tender-track (Track B) concept; a development-only
# refresh never populates it, and reporting bid_now=0 here is truthful, not a
# bug — development-application leads are never live tenders.
_BID_LATER_MIN_FIT = 60
_WATCH_MIN_FIT = 50


def default_scorer(
    dataset_path: Path,
    out_dir: Path,
    *,
    preset_id: str = "civil_contractor",
    package_root: str | Path | None = None,
) -> ScoreResult:
    """Real deterministic scorer for the development-refresh dataset.

    Reads every record back from the written dataset, scores it under the
    active contractor preset with the same scoring engine used everywhere
    else in the product (``tenderfinder_guards.score_civil_fit_breakdown``),
    buckets it, and writes a ranked output workbook sorted by fit score
    descending. This is what makes "Refresh Development Data" truthfully
    report BID LATER / WATCH / SKIP counts instead of permanent zeros.
    """
    import os as _os

    import openpyxl

    import tenderfinder_guards as _guards
    from tenderfinder_excel_safety import append_untrusted_row
    from tenderfinder_keywords_config import clear_keywords_cache as _clear
    from tenderfinder_presets import resolve_preset_keywords_path

    records = _read_records_for_scoring(dataset_path)
    previous = _os.environ.get("TENDER_FINDER_KEYWORDS_CONFIG")
    _os.environ["TENDER_FINDER_KEYWORDS_CONFIG"] = str(
        resolve_preset_keywords_path(preset_id, root=package_root)
    )
    try:
        scored_rows: list[dict[str, Any]] = []
        bid_later = watch = skip = 0
        for record in records:
            text = " ".join(
                str(record.get(key, ""))
                for key in ("scope_summary", "app_type_stage", "address")
            )
            _clear()
            breakdown = _guards.score_civil_fit_breakdown(text, str(record.get("municipality", "")))
            fit = breakdown["fit_score"]
            if fit >= _BID_LATER_MIN_FIT:
                bucket = "BID_LATER"
                bid_later += 1
            elif fit >= _WATCH_MIN_FIT:
                bucket = "WATCH"
                watch += 1
            else:
                bucket = "SKIP"
                skip += 1
            scored_rows.append({**record, "fit_score": fit, "bucket": bucket})
    finally:
        if previous is None:
            _os.environ.pop("TENDER_FINDER_KEYWORDS_CONFIG", None)
        else:
            _os.environ["TENDER_FINDER_KEYWORDS_CONFIG"] = previous
        _clear()

    scored_rows.sort(key=lambda row: row["fit_score"], reverse=True)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "ranked_opportunities.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked_Opportunities"
    headers = [
        "rank", "fit_score", "bucket", "lead_id", "municipality", "app_no",
        "address", "app_type_stage", "scope_summary", "applicant_name",
        "source", "source_url",
    ]
    ws.append(headers)
    for rank, row in enumerate(scored_rows, start=1):
        # rank/fit_score/bucket are our own computed values (safe); the
        # remaining fields are public source data and go through the
        # untrusted-cell guard so a value like "=cmd|..." is never
        # interpreted as a formula.
        append_untrusted_row(ws, [rank] + [row.get(h, "") for h in headers[1:]])
    wb.save(out_path)

    return ScoreResult(
        output_path=str(out_path), scored=len(scored_rows), bid_now=0,
        bid_later=bid_later, watch=watch, skip=skip,
    )


def make_default_scorer(request: RefreshRequest) -> Callable[[Path, Path], ScoreResult]:
    """Build a ready-to-call scorer bound to this request's preset. Passed to
    ``refresh_development_data`` for a real production refresh."""

    def _score(dataset_path: Path, out_dir: Path) -> ScoreResult:
        return default_scorer(
            dataset_path, out_dir, preset_id=request.preset_id, package_root=request.package_root
        )

    return _score


def diagnostic_preview_acquirer(
    source: dict[str, Any], *, package_root: str | Path | None = None
) -> SourceFetch:
    """Bounded diagnostic-preview acquirer — source-health/validation use only.

    Uses the engine's guarded single-source live test, which enforces URL
    safety and readiness rules and persists per-source health, returning only a
    small bounded normalized sample (``normalized_preview``, typically ~5
    records). This is intentionally NOT the acquirer used by "Refresh
    Development Data" — a bounded preview must never power the production
    dataset. Use it for source-health checks / diagnostics /
    "Test Source (live)" only. See ``full_sweep_development_acquirer`` for the
    real production acquirer.
    """
    from tenderfinder_engine import test_source_definition

    source_id = source["source_id"]
    try:
        result = test_source_definition(
            source_id, root=package_root, allow_network=True, persist_result=True
        )
    except Exception as exc:  # noqa: BLE001 - any failure = a failed source
        return SourceFetch(source_id=source_id, ok=False, error=str(exc))
    records = tuple(result.get("normalized_preview", []) or ())
    ok = bool(result.get("passed")) and bool(records)
    return SourceFetch(
        source_id=source_id,
        ok=ok,
        records=records,
        error="; ".join(str(e) for e in (result.get("errors") or [])),
        http_status=int(result.get("http_status", 0) or 0),
    )


# Backwards-compatible alias for the old name (kept only for anything still
# importing it directly); new code should call diagnostic_preview_acquirer or
# full_sweep_development_acquirer explicitly by name.
default_development_acquirer = diagnostic_preview_acquirer


def _lead_to_record(source_id: str, source_name: str, lead: dict[str, Any]) -> dict[str, Any]:
    """Map one tenderfinder_raw_sweep normalized lead into a refresh-service record."""
    return {
        "source": source_id,
        "source_name": source_name,
        "lead_id": lead.get("project_id", ""),
        "app_no": lead.get("app_no", ""),
        "municipality": lead.get("municipality", ""),
        "address": lead.get("address", ""),
        "applicant_name": lead.get("owner", ""),
        "app_type_stage": lead.get("app_type_stage", ""),
        "scope_summary": lead.get("scope_summary", ""),
        "status_class": lead.get("proposed_route", ""),
        "source_url": lead.get("source_url", "") or lead.get("evidence_url", ""),
        "fit_score": lead.get("fit_score", ""),
        "date_found": lead.get("date_found", ""),
        "raw_hash": lead.get("raw_hash", ""),
        "dedupe_key": lead.get("dedupe_key", ""),
    }


# Statuses run_connector returns that mean "did not actually attempt a live
# fetch" (config/access gating, not a fetch failure) — never reported as ok.
_NON_FETCH_STATUSES = frozenset(
    {
        "skipped_source_status", "p3_extract_required", "access_test_required",
        "paid_or_login_skip", "disabled_wrong_layer", "needs_exact_url",
        "offline_no_fixture",
    }
)


def full_sweep_development_acquirer(
    source: dict[str, Any],
    *,
    max_records: int = 0,
    raw_dir: str | Path | None = None,
    package_root: str | Path | None = None,
) -> SourceFetch:
    """Real, unbounded (up to ``max_records``) full-sweep LIVE acquirer.

    This is the production acquirer for "Refresh Development Data". It calls
    ``tenderfinder_raw_sweep.run_connector`` directly — the same
    fully-paginated (ArcGIS/ODS batch queries with retry/backoff, continuing
    until the source is exhausted or ``max_records`` is reached), GUI-independent
    function used by the ``--review-only`` CLI sweep. It is NOT a bounded
    connector-test preview: every eligible lead the source returns (up to the
    cap) is collected, not a fixed small sample.

    Performs real network I/O against the source's public endpoint, so it is
    exercised only in live mode / the controlled live proof — never in the
    offline test suite (the orchestration around this acquirer is tested with
    injected fakes instead).
    """
    import tenderfinder_raw_sweep as raw_sweep

    source_id = source.get("source_id", "")
    source_name = source.get("name", "")
    conn = dict(source)
    conn.setdefault("runtime_skip_reason", source_skip_reason(source))
    cap = max_records if max_records > 0 else 20000  # tenderfinder_raw_sweep's own CLI default

    try:
        result = raw_sweep.run_connector(
            conn, cap, probe=False, raw_dir=str(raw_dir) if raw_dir else None,
        )
    except Exception as exc:  # noqa: BLE001 - any failure = a failed source
        return SourceFetch(source_id=source_id, ok=False, error=str(exc))

    status = str(result.get("status") or "")
    error = str(result.get("error") or "")
    leads = result.get("leads") or []
    records = tuple(_lead_to_record(source_id, source_name, lead) for lead in leads)
    ok = status not in _NON_FETCH_STATUSES and not error

    return SourceFetch(
        source_id=source_id,
        ok=ok,
        records=records,
        error=error or ("" if ok else f"non-fetch status: {status}"),
        http_status=int(result.get("http_status", 0) or 0),
    )


def make_full_sweep_acquirer(
    request: RefreshRequest, *, package_root: Path, run_id: str
) -> Callable[[dict[str, Any]], SourceFetch]:
    """Build a ready-to-call full-sweep acquirer bound to this request's raw-dir
    and per-source cap. This is what the GUI and CLI pass to
    ``refresh_development_data`` for a real production refresh."""
    raw_dir = dm.datasets_root(
        state_root=request.state_root, package_root=package_root, create=True
    ) / "raw" / run_id

    def _acquire(source: dict[str, Any]) -> SourceFetch:
        return full_sweep_development_acquirer(
            source,
            max_records=request.max_records_per_source,
            raw_dir=raw_dir,
            package_root=package_root,
        )

    return _acquire


def _empty_metrics(request: RefreshRequest, run_id: str, started: str) -> dm.RunMetrics:
    return dm.RunMetrics(
        run_id=run_id, run_started=started,
        active_preset_id=request.preset_id,
    )


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #


def refresh_development_data(
    request: RefreshRequest,
    *,
    acquirer: Callable[[dict[str, Any]], SourceFetch],
    dataset_writer: Callable[[list[dict[str, Any]], Path], Path] | None = None,
    scorer: Callable[[Path, Path], ScoreResult] | None = None,
    sources_path: str | Path | None = None,
    now: datetime | None = None,
) -> RefreshResult:
    """Run the full development-data refresh. See module docstring for behaviour.

    ``acquirer(source_row) -> SourceFetch`` is mandatory and must be safe (it
    must never contact blocked/login sources). ``dataset_writer`` and ``scorer``
    default to real implementations but are injectable for headless testing.
    """
    package_root = detect_package_root(
        Path(request.package_root).resolve() if request.package_root else None
    )
    dataset_writer = dataset_writer or _default_dataset_writer
    run_id = request.run_id or dm.new_dataset_id("refresh", now=now)
    started = timestamp_now()
    preset = get_preset(request.preset_id)

    try:
        eligible = eligible_development_sources(
            package_root=package_root, sources_path=sources_path,
            source_ids=request.source_ids,
        )
    except Exception as exc:  # registry error
        metrics = _empty_metrics(request, run_id, started)
        metrics.succeeded = False
        metrics.errors = (f"source registry error: {exc}",)
        metrics.run_ended = timestamp_now()
        return _failure(request, metrics, (), "source registry error", package_root)

    # Acquire per-source.
    outcomes: list[dict[str, Any]] = []
    all_records: list[dict[str, Any]] = []
    attempted = successful = failed = 0
    for source in eligible:
        attempted += 1
        try:
            fetch = acquirer(source)
        except Exception as exc:  # an acquirer that raises counts as a failed source
            fetch = SourceFetch(source_id=source["source_id"], ok=False, error=str(exc))
        outcomes.append(
            {
                "source_id": fetch.source_id,
                "ok": fetch.ok,
                "records": fetch.count,
                "error": fetch.error,
                "http_status": fetch.http_status,
            }
        )
        if fetch.ok:
            successful += 1
            all_records.extend(dict(r) for r in fetch.records)
        else:
            failed += 1

    metrics = dm.RunMetrics(
        run_id=run_id,
        run_started=started,
        sources_selected=len(eligible),
        sources_attempted=attempted,
        sources_successful=successful,
        sources_failed=failed,
        records_fetched=len(all_records),
        records_loaded=0,
        records_before_dedup=len(all_records),
        active_preset_id=preset.preset_id,
        active_preset_version=preset.version,
    )

    # Total failure: no source produced records -> keep last-known-good, mark stale.
    if successful == 0 or not all_records:
        metrics.succeeded = False
        metrics.duplicates_removed = 0
        metrics.normalized_records = 0
        metrics.data_mode = dm.MODE_UNKNOWN
        metrics.errors = ("no source produced records; previous dataset retained",)
        metrics.run_ended = timestamp_now()
        return _failure(request, metrics, tuple(outcomes),
                        "Refresh failed: no source produced records. Previous data retained.",
                        package_root)

    # Normalize + dedup + validate.
    deduped, duplicates_removed = deduplicate_records(all_records)
    validation = validate_dataset(deduped)
    metrics.duplicates_removed = duplicates_removed
    metrics.normalized_records = len(deduped)
    # records_live is set only once the dataset is actually promoted (below) -
    # records_fetched/normalized_records are truthful acquisition diagnostics
    # that must be reported even when a later step (validation, promotion)
    # fails, but records_live specifically means "live in the active dataset
    # as a result of THIS run" and must stay 0 until that is actually true.

    if not validation["passed"]:
        metrics.succeeded = False
        metrics.data_mode = dm.MODE_UNKNOWN
        metrics.errors = tuple(validation["errors"])
        metrics.run_ended = timestamp_now()
        return _failure(request, metrics, tuple(outcomes),
                        "Refresh failed dataset validation. Previous data retained.",
                        package_root)

    # Write the timestamped external dataset (never inside the package).
    datasets_dir = dm.datasets_root(
        state_root=request.state_root, package_root=package_root, create=True
    )
    dataset_path = datasets_dir / f"development_review_{run_id}.xlsx"
    dataset_writer(deduped, dataset_path)

    capture_ts = timestamp_now()
    provenance = dm.DatasetProvenance(
        dataset_id=run_id,
        path=str(dataset_path),
        data_mode=dm.MODE_LIVE,
        capture_timestamp=capture_ts,
        source_ids=tuple(o["source_id"] for o in outcomes if o["ok"]),
        source_results=tuple(outcomes),
        record_counts={"total": len(deduped), "live": len(deduped)},
        validation=validation,
        content_hash="",
        preset_id=preset.preset_id,
        preset_version=preset.version,
        last_successful_refresh=capture_ts,
    )

    promotion = dm.promote_dataset(
        provenance, state_root=request.state_root, package_root=package_root
    )
    if not promotion.promoted:
        metrics.succeeded = False
        metrics.data_mode = dm.MODE_UNKNOWN
        metrics.errors = tuple(promotion.errors)
        metrics.run_ended = timestamp_now()
        return _failure(request, metrics, tuple(outcomes),
                        "Dataset promotion refused. Previous data retained.", package_root)

    metrics.active_dataset_id = provenance.dataset_id
    metrics.active_dataset_path = provenance.path
    metrics.capture_timestamp = capture_ts
    metrics.data_mode = dm.MODE_LIVE
    metrics.stale = False
    metrics.records_live = len(deduped)

    # Score the promoted dataset into a ranked output workbook.
    output_paths: dict[str, Any] = {"dataset": str(dataset_path)}
    if scorer is not None:
        out_dir = datasets_dir / "runs" / run_id / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        score = scorer(dataset_path, out_dir)
        metrics.records_scored = score.scored
        metrics.bid_now = score.bid_now
        metrics.bid_later = score.bid_later
        metrics.watch = score.watch
        metrics.skip = score.skip
        output_paths["output_workbook"] = score.output_path

    metrics.succeeded = True
    metrics.run_ended = timestamp_now()

    manifest_path = _write_manifest(request, run_id, metrics, provenance, outcomes,
                                    output_paths, package_root)
    output_paths["manifest"] = str(manifest_path)

    return RefreshResult(
        succeeded=True,
        message=f"Refreshed {len(deduped)} development records from "
                f"{successful}/{attempted} sources.",
        metrics=metrics,
        provenance=provenance,
        promotion=promotion,
        source_outcomes=tuple(outcomes),
        output_paths=output_paths,
        manifest_path=str(manifest_path),
        stale=False,
    )


def _failure(
    request: RefreshRequest,
    metrics: dm.RunMetrics,
    outcomes: tuple[dict[str, Any], ...],
    message: str,
    package_root: Path,
) -> RefreshResult:
    """Build a failure result that preserves the previous active dataset."""
    previous = dm.load_active_dataset(state_root=request.state_root, package_root=package_root)
    stale_provenance = dm.mark_stale(previous) if previous else None
    if stale_provenance is not None:
        metrics.active_dataset_id = stale_provenance.dataset_id
        metrics.active_dataset_path = stale_provenance.path
        metrics.capture_timestamp = stale_provenance.capture_timestamp
        metrics.data_mode = stale_provenance.data_mode
        metrics.stale = True
    manifest_path = _write_manifest(request, metrics.run_id, metrics, stale_provenance,
                                    list(outcomes), {}, package_root)
    return RefreshResult(
        succeeded=False,
        message=message,
        metrics=metrics,
        provenance=stale_provenance,
        promotion=None,
        source_outcomes=outcomes,
        output_paths={},
        manifest_path=str(manifest_path),
        stale=True,
    )


def _write_manifest(
    request: RefreshRequest,
    run_id: str,
    metrics: dm.RunMetrics,
    provenance: dm.DatasetProvenance | None,
    outcomes: list[dict[str, Any]],
    output_paths: dict[str, Any],
    package_root: Path,
) -> Path:
    datasets_dir = dm.datasets_root(
        state_root=request.state_root, package_root=package_root, create=True
    )
    manifest_path = datasets_dir / "runs" / run_id / "run_manifest.json"
    manifest = {
        "schema_version": 1,
        "run_id": run_id,
        "kind": "development_refresh",
        "succeeded": metrics.succeeded,
        "preset_id": request.preset_id,
        "metrics": metrics.to_dict(),
        "provenance": provenance.to_dict() if provenance else None,
        "source_outcomes": outcomes,
        "output_paths": output_paths,
        "written_at": timestamp_now(),
    }
    atomic_write_json(manifest_path, manifest)
    return manifest_path
