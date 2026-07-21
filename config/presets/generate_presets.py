"""Deterministically regenerate the three contractor-preset keyword workbooks
from the canonical ``config/keywords.xlsx``.

Presets are expressed as alternate, fully-validated keyword workbooks — the
scoring engine itself is unchanged and fully data-driven. Selecting a preset is
just pointing ``load_keywords_config`` at a different workbook (see
``tenderfinder_presets.py``).

Run from anywhere:  ``python config/presets/generate_presets.py``

* ``civil_contractor.xlsx``       — byte-for-byte copy of the canonical
  workbook, so the existing Civil Contractor behaviour is preserved exactly as
  a versioned preset.
* ``multifamily_residential.xlsx`` — neutralizes the civil-oriented negative
  penalties on interior/mechanical/electrical/HVAC/suite scope and adds
  residential positives. It never penalizes residential building scope.
* ``general_contractor.xlsx``     — neutralizes the vertical-work penalties and
  adds broad commercial/institutional/residential positives, so both civil site
  work and vertical building scope score positively.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parents[1]
CANONICAL = REPO_ROOT / "config" / "keywords.xlsx"

# Negative rows (weight -12) that penalize vertical/interior building scope.
RESIDENTIAL_NEUTRALIZE = {
    "interior", "mechanical", "electrical", "hvac", "suite", "unit alteration",
    "kitchen", "bathroom", "tenant improvement", "renovation", "roofing",
    "facade", "cladding",
}
# A residential/GC preset still ignores non-construction noise.
KEEP_NEGATIVE = {"sign permit", "demolition only"}

RESIDENTIAL_POSITIVES = [
    "multifamily", "multi-family", "apartment", "townhouse", "townhome",
    "rowhouse", "condominium", "condo", "wood frame", "wood-frame",
    "mid-rise", "low-rise", "high-rise", "mixed-use", "residential",
    "dwelling unit", "purpose-built rental", "rental housing", "fourplex",
    "six-plex",
]

GENERAL_POSITIVES = [
    "commercial", "institutional", "renovation", "tenant improvement",
    "mixed-use", "multifamily", "structural", "building envelope",
    "mechanical", "electrical", "concrete structure", "steel structure",
    "school", "hospital", "recreation centre", "warehouse", "retail",
]

PROFILES = {
    "multifamily_residential.xlsx": {
        "company_name": "Multi-Family Residential Builder",
        "company_description": (
            "Multi-family residential builder delivering apartments, townhomes, "
            "condominiums, wood-frame and mid-rise housing across British Columbia."
        ),
    },
    "general_contractor.xlsx": {
        "company_name": "General Contractor",
        "company_description": (
            "General contractor delivering commercial, institutional, residential "
            "and light-civil construction across British Columbia."
        ),
    },
}


def _positive_row(keyword: str, note: str) -> list[object]:
    return [keyword, "contains", 9, "positive", note, "Y", ""]


def _rewrite(dest_name: str, *, neutralize: set[str], positives: list[str], note: str) -> None:
    wb = openpyxl.load_workbook(CANONICAL)
    profile = wb["Company_Profile"]
    meta = PROFILES[dest_name]
    profile.cell(row=2, column=1, value=meta["company_name"])
    profile.cell(row=2, column=2, value=meta["company_description"])

    keywords = wb["Keywords"]
    header = [cell.value for cell in keywords[1]]
    kw_col = header.index("keyword") + 1
    cat_col = header.index("category") + 1
    active_col = header.index("active") + 1
    existing_positive = set()
    for row in range(2, keywords.max_row + 1):
        kw = str(keywords.cell(row=row, column=kw_col).value or "").strip().casefold()
        cat = str(keywords.cell(row=row, column=cat_col).value or "").strip()
        if cat == "negative" and kw in neutralize and kw not in KEEP_NEGATIVE:
            keywords.cell(row=row, column=active_col, value="N")
        if cat == "positive":
            existing_positive.add(kw)

    next_row = keywords.max_row + 1
    for keyword in positives:
        if keyword.casefold() in existing_positive:
            continue
        for col, value in enumerate(_positive_row(keyword, note), start=1):
            keywords.cell(row=next_row, column=col, value=value)
        next_row += 1

    wb.save(HERE / dest_name)


def main() -> int:
    if not CANONICAL.exists():
        raise SystemExit(f"canonical keywords workbook missing: {CANONICAL}")
    # Civil = exact copy so scoring stays identical.
    shutil.copy2(CANONICAL, HERE / "civil_contractor.xlsx")
    _rewrite(
        "multifamily_residential.xlsx",
        neutralize=RESIDENTIAL_NEUTRALIZE,
        positives=RESIDENTIAL_POSITIVES,
        note="Residential building signal (multi-family preset).",
    )
    _rewrite(
        "general_contractor.xlsx",
        neutralize=RESIDENTIAL_NEUTRALIZE,
        positives=GENERAL_POSITIVES,
        note="General-contracting building signal (broad preset).",
    )
    print("Wrote 3 preset workbooks to", HERE)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
