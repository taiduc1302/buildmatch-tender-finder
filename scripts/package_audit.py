#!/usr/bin/env python3
"""Scan a repository or extracted clean release for secrets and private data.

Company/scoring profile text is intentional configuration and is not a secret.
Package mode additionally rejects forbidden build/runtime directories and
machine-specific paths. Repository mode skips the checkout's own `.git` and
`.venv`, while still scanning all source/configuration files.
"""
from __future__ import annotations

import argparse
import os
import re
from pathlib import Path


TEXT_EXTS = {
    ".py", ".md", ".txt", ".csv", ".bat", ".yml", ".yaml", ".json",
    ".command", ".example", ".cfg", ".ini", ".toml", ".html",
}
SECRET_PATTERNS = [
    ("Tavily API key", re.compile(r"tvly-(?!your-key-here)[A-Za-z0-9_-]{10,}")),
    ("OpenAI/Anthropic-style key", re.compile(r"sk-(?:ant-)?[A-Za-z0-9_-]{20,}")),
    ("AWS access key", re.compile(r"AKIA[0-9A-Z]{16}")),
    ("GitHub token", re.compile(r"gh[oprsu]_[A-Za-z0-9]{20,}")),
    ("Slack token", re.compile(r"xox[baprs]-[A-Za-z0-9-]{10,}")),
    ("private key", re.compile(r"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----")),
    (
        "assigned secret",
        re.compile(
            r"(?i)(?:password|passwd|secret|api[_-]?key|access[_-]?token|refresh[_-]?token)"
            r"\s*[:=]\s*['\"][^'\"\s]{8,}['\"]"
        ),
    ),
    ("cookie header", re.compile(r"(?i)(?:set-cookie|cookie)\s*:")),
]
PRIVATE_PATH_PATTERNS = [
    re.compile(r"(?i)C:[\\/]+Users[\\/]+(?!Public\b|ExampleUser\b|example\b|x\b|<)[A-Za-z0-9._-]+"),
    re.compile(r"(?i)OneDrive - (?!Example\b)"),
]
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
ALLOWED_EMAIL_DOMAINS = {
    "example.com", "example.org", "bidsandtenders.ca", "bcbid.gov.bc.ca",
    "gov.bc.ca", "surrey.ca", "bidcentral.ca", "buildingconnected.com",
    "civicinfo.bc.ca", "anthropic.com", "users.noreply.github.com",
}
SKIP_DIRS = {".git", ".venv", "venv", "__pycache__", ".pytest_cache", "node_modules"}
PACKAGE_FORBIDDEN_DIRS = SKIP_DIRS | {".codex_tmp", "user_data", "demo_history", "raw_runs"}
PACKAGE_FORBIDDEN_SUFFIXES = {
    ".pyc", ".pyo", ".log", ".tmp", ".eml", ".db", ".sqlite", ".sqlite3",
}


def iter_files(root: Path, *, package_mode: bool, findings: list[str]):
    for current, dirs, files in os.walk(root):
        current_path = Path(current)
        retained = []
        for name in dirs:
            folded = name.casefold()
            relative = (current_path / name).relative_to(root).as_posix()
            if package_mode and folded in PACKAGE_FORBIDDEN_DIRS:
                findings.append(f"FORBIDDEN_DIR {relative}")
                continue
            if folded in SKIP_DIRS:
                continue
            retained.append(name)
        dirs[:] = retained
        for name in files:
            path = current_path / name
            relative = path.relative_to(root).as_posix()
            if package_mode and (
                path.suffix.casefold() in PACKAGE_FORBIDDEN_SUFFIXES
                or path.name.startswith("~$")
            ):
                findings.append(f"FORBIDDEN_FILE {relative}")
                continue
            yield path


def check_text(relative: str, text: str, *, package_mode: bool, findings: list[str], warnings: list[str]) -> None:
    for label, pattern in SECRET_PATTERNS:
        for match in pattern.finditer(text):
            findings.append(f"SECRET {relative}: {label} matched {match.group(0)[:40]!r}")
    for pattern in PRIVATE_PATH_PATTERNS:
        for match in pattern.finditer(text):
            message = f"PRIVATE_PATH {relative}: {match.group(0)!r}"
            (findings if package_mode else warnings).append(message)
    for match in EMAIL_RE.finditer(text):
        address = match.group(0)
        domain = address.rsplit("@", 1)[1].casefold().rstrip(".")
        if domain not in ALLOWED_EMAIL_DOMAINS:
            findings.append(f"UNAPPROVED_EMAIL {relative}: {address}")


def audit(root: Path, *, mode: str) -> dict[str, object]:
    root = root.expanduser().resolve()
    package_mode = mode == "package"
    findings: list[str] = []
    warnings: list[str] = []
    text_count = 0
    workbook_count = 0
    this_script = Path(__file__).resolve()
    for path in iter_files(root, package_mode=package_mode, findings=findings):
        relative = path.relative_to(root).as_posix()
        if path.resolve() == this_script or relative.casefold() == "scripts/package_audit.py":
            continue
        if path.suffix.casefold() in TEXT_EXTS or path.name in {".gitignore", ".gitattributes"}:
            try:
                text = path.read_text(encoding="utf-8-sig")
            except UnicodeDecodeError:
                text = path.read_text(encoding="cp1252", errors="replace")
            check_text(relative, text, package_mode=package_mode, findings=findings, warnings=warnings)
            text_count += 1
        elif path.suffix.casefold() == ".xlsx" and not path.name.startswith("~$"):
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(path, read_only=True, data_only=True)
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        for value in row:
                            if isinstance(value, str):
                                check_text(
                                    f"{relative}::{sheet.title}",
                                    value,
                                    package_mode=package_mode,
                                    findings=findings,
                                    warnings=warnings,
                                )
                workbook.close()
                workbook_count += 1
            except Exception as exc:  # noqa: BLE001
                findings.append(f"XLSX_UNREADABLE {relative}: {type(exc).__name__}: {exc}")
    return {
        "mode": mode,
        "root": str(root),
        "text_files": text_count,
        "workbooks": workbook_count,
        "findings": findings,
        "warnings": warnings,
        "passed": not findings,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scan Tender Finder source/release content for secrets.")
    parser.add_argument("root", nargs="?", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--mode", choices=("repo", "package"), default="package")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = audit(args.root, mode=args.mode)
    print(
        f"scanned {result['text_files']} text files and {result['workbooks']} workbooks "
        f"under {result['root']}"
    )
    for warning in result["warnings"][:100]:
        print("WARNING: " + warning)
    if result["findings"]:
        print(f"PACKAGE AUDIT: FAIL ({len(result['findings'])} finding(s))")
        for finding in result["findings"][:200]:
            print("  " + finding)
        return 1
    print(f"PACKAGE AUDIT: PASS ({result['mode']} mode)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
